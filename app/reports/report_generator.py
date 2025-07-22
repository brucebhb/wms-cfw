#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
报表生成器 - 待重构
用于生成各类统计报表的Excel、PDF和CSV文件
"""

import os
import pandas as pd
import tempfile
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from flask import current_app

from app.reports.statistics_service import StatisticsService

class ReportGenerator:
    """报表生成器"""
    
    def __init__(self):
        self.statistics_service = StatisticsService()
        self.temp_dir = os.path.join(current_app.root_path, 'temp')
        
        # 确保临时目录存在
        if not os.path.exists(self.temp_dir):
            os.makedirs(self.temp_dir)
    
    def generate_report(self, report_type, format_type, user, start_date=None, end_date=None):
        """
        生成报表

        Args:
            report_type: 报表类型 (目前仅支持 'dashboard')
            format_type: 格式类型 (excel, pdf, csv)
            user: 当前用户
            start_date: 开始日期 (暂未使用)
            end_date: 结束日期 (暂未使用)

        Returns:
            str: 生成的报表文件路径
        """
        # 处理日期参数
        if start_date:
            start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
        else:
            start_date = (datetime.now() - timedelta(days=30)).date()
            
        if end_date:
            end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
        else:
            end_date = datetime.now().date()
        
        # 根据报表类型调用相应的生成方法
        if report_type == 'dashboard':
            return self._generate_dashboard_report(format_type, user)
        else:
            raise ValueError(f"不支持的报表类型: {report_type}。当前仅支持 'dashboard' 类型。")
    
    def _generate_dashboard_report(self, format_type, user):
        """生成仪表板报表"""
        # 获取仪表板数据
        data = self.statistics_service.get_dashboard_data(user)
        
        if format_type == 'excel':
            return self._generate_dashboard_excel(data, user)
        elif format_type == 'pdf':
            return self._generate_dashboard_pdf(data, user)
        elif format_type == 'csv':
            return self._generate_dashboard_csv(data, user)
        else:
            raise ValueError(f"不支持的格式类型: {format_type}")
    
    def _generate_dashboard_excel(self, data, user):
        """生成仪表板Excel报表"""
        wb = Workbook()
        
        # 创建概览工作表
        ws_overview = wb.active
        ws_overview.title = "运营概览"
        
        # 设置标题
        ws_overview.merge_cells('A1:H1')
        ws_overview['A1'] = "仓储管理系统运营概览"
        ws_overview['A1'].font = Font(size=16, bold=True)
        ws_overview['A1'].alignment = Alignment(horizontal='center')
        
        # 设置生成信息
        ws_overview['A2'] = f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws_overview['A3'] = f"生成用户: {user.username} ({user.real_name})"
        
        # 今日数据
        ws_overview['A5'] = "今日数据"
        ws_overview['A5'].font = Font(bold=True)
        
        headers = ['指标', '入库', '出库']
        for i, header in enumerate(headers):
            ws_overview.cell(row=6, column=i+1, value=header).font = Font(bold=True)
        
        metrics = [
            ['记录数', data['today_stats']['inbound']['count'], data['today_stats']['outbound']['count']],
            ['板数', data['today_stats']['inbound']['pallets'], data['today_stats']['outbound']['pallets']],
            ['件数', data['today_stats']['inbound']['packages'], data['today_stats']['outbound']['packages']],
            ['重量(kg)', data['today_stats']['inbound']['weight'], data['today_stats']['outbound']['weight']],
            ['体积(m³)', data['today_stats']['inbound']['volume'], data['today_stats']['outbound']['volume']]
        ]
        
        for i, row in enumerate(metrics):
            for j, value in enumerate(row):
                ws_overview.cell(row=7+i, column=j+1, value=value)
        
        # 仓库汇总
        ws_overview['A13'] = "仓库汇总"
        ws_overview['A13'].font = Font(bold=True)
        
        warehouse_headers = ['仓库', '入库记录数', '入库板数', '入库件数', '出库记录数', '出库板数', '出库件数']
        for i, header in enumerate(warehouse_headers):
            ws_overview.cell(row=14, column=i+1, value=header).font = Font(bold=True)
        
        for i, warehouse in enumerate(data['warehouse_summary']):
            row = [
                warehouse['warehouse_name'],
                warehouse['inbound_count'],
                warehouse['inbound_pallets'],
                warehouse['inbound_packages'],
                warehouse['outbound_count'],
                warehouse['outbound_pallets'],
                warehouse['outbound_packages']
            ]
            for j, value in enumerate(row):
                ws_overview.cell(row=15+i, column=j+1, value=value)
        
        # 调整列宽
        for col in range(1, 8):
            ws_overview.column_dimensions[get_column_letter(col)].width = 15
        
        # 创建库存工作表
        ws_inventory = wb.create_sheet("库存概览")
        
        ws_inventory['A1'] = "库存概览"
        ws_inventory['A1'].font = Font(size=16, bold=True)
        
        inventory_headers = ['仓库', '货物数量', '板数', '件数']
        for i, header in enumerate(inventory_headers):
            ws_inventory.cell(row=3, column=i+1, value=header).font = Font(bold=True)
        
        for i, warehouse in enumerate(data['inventory_overview']['by_warehouse']):
            row = [
                warehouse['warehouse_name'],
                warehouse['items'],
                warehouse['pallets'],
                warehouse['packages']
            ]
            for j, value in enumerate(row):
                ws_inventory.cell(row=4+i, column=j+1, value=value)
        
        # 总计行
        total_row = [
            '总计',
            data['inventory_overview']['total']['items'],
            data['inventory_overview']['total']['pallets'],
            data['inventory_overview']['total']['packages']
        ]
        for j, value in enumerate(total_row):
            cell = ws_inventory.cell(row=4+len(data['inventory_overview']['by_warehouse']), column=j+1, value=value)
            cell.font = Font(bold=True)
        
        # 创建客户工作表
        ws_customers = wb.create_sheet("客户TOP10")
        
        ws_customers['A1'] = "客户TOP10"
        ws_customers['A1'].font = Font(size=16, bold=True)
        
        customer_headers = ['客户名称', '入库记录数', '板数', '件数']
        for i, header in enumerate(customer_headers):
            ws_customers.cell(row=3, column=i+1, value=header).font = Font(bold=True)
        
        for i, customer in enumerate(data['top_customers']):
            row = [
                customer['customer_name'],
                customer['inbound_count'],
                customer['total_pallets'],
                customer['total_packages']
            ]
            for j, value in enumerate(row):
                ws_customers.cell(row=4+i, column=j+1, value=value)
        
        # 创建图表
        chart = BarChart()
        chart.title = "客户货物量TOP10"
        chart.y_axis.title = "件数"
        chart.x_axis.title = "客户"
        
        data_ref = Reference(ws_customers, min_col=4, min_row=3, max_row=3+len(data['top_customers']))
        cats_ref = Reference(ws_customers, min_col=1, min_row=4, max_row=3+len(data['top_customers']))
        
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        chart.height = 15
        chart.width = 20
        
        ws_customers.add_chart(chart, "A15")
        
        # 保存文件
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        file_path = os.path.join(self.temp_dir, f'dashboard_report_{timestamp}.xlsx')
        wb.save(file_path)
        
        return file_path
    
    def _generate_dashboard_pdf(self, data, user):
        """生成仪表板PDF报表"""
        # 先生成Excel，然后转换为PDF
        excel_path = self._generate_dashboard_excel(data, user)
        
        # 使用ReportLab生成PDF
        from reportlab.lib.pagesizes import letter, landscape
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib import colors
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        pdf_path = os.path.join(self.temp_dir, f'dashboard_report_{timestamp}.pdf')
        
        doc = SimpleDocTemplate(pdf_path, pagesize=landscape(letter))
        elements = []
        
        styles = getSampleStyleSheet()
        title_style = styles['Heading1']
        subtitle_style = styles['Heading2']
        normal_style = styles['Normal']
        
        # 添加标题
        elements.append(Paragraph("仓储管理系统运营概览", title_style))
        elements.append(Spacer(1, 12))
        elements.append(Paragraph(f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", normal_style))
        elements.append(Paragraph(f"生成用户: {user.username} ({user.real_name})", normal_style))
        elements.append(Spacer(1, 20))
        
        # 今日数据
        elements.append(Paragraph("今日数据", subtitle_style))
        elements.append(Spacer(1, 6))
        
        today_data = [['指标', '入库', '出库']]
        today_data.extend([
            ['记录数', data['today_stats']['inbound']['count'], data['today_stats']['outbound']['count']],
            ['板数', data['today_stats']['inbound']['pallets'], data['today_stats']['outbound']['pallets']],
            ['件数', data['today_stats']['inbound']['packages'], data['today_stats']['outbound']['packages']],
            ['重量(kg)', data['today_stats']['inbound']['weight'], data['today_stats']['outbound']['weight']],
            ['体积(m³)', data['today_stats']['inbound']['volume'], data['today_stats']['outbound']['volume']]
        ])
        
        today_table = Table(today_data, colWidths=[120, 100, 100])
        today_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        elements.append(today_table)
        elements.append(Spacer(1, 20))
        
        # 保存PDF
        doc.build(elements)
        
        return pdf_path
    
    def _generate_dashboard_csv(self, data, user):
        """生成仪表板CSV报表"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        file_path = os.path.join(self.temp_dir, f'dashboard_report_{timestamp}.csv')
        
        with open(file_path, 'w', encoding='utf-8') as f:
            # 写入标题
            f.write("仓储管理系统运营概览\n")
            f.write(f"生成时间,{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write(f"生成用户,{user.username} ({user.real_name})\n\n")
            
            # 今日数据
            f.write("今日数据\n")
            f.write("指标,入库,出库\n")
            f.write(f"记录数,{data['today_stats']['inbound']['count']},{data['today_stats']['outbound']['count']}\n")
            f.write(f"板数,{data['today_stats']['inbound']['pallets']},{data['today_stats']['outbound']['pallets']}\n")
            f.write(f"件数,{data['today_stats']['inbound']['packages']},{data['today_stats']['outbound']['packages']}\n")
            f.write(f"重量(kg),{data['today_stats']['inbound']['weight']},{data['today_stats']['outbound']['weight']}\n")
            f.write(f"体积(m³),{data['today_stats']['inbound']['volume']},{data['today_stats']['outbound']['volume']}\n\n")
            
            # 仓库汇总
            f.write("仓库汇总\n")
            f.write("仓库,入库记录数,入库板数,入库件数,出库记录数,出库板数,出库件数\n")
            for warehouse in data['warehouse_summary']:
                f.write(f"{warehouse['warehouse_name']},{warehouse['inbound_count']},{warehouse['inbound_pallets']},{warehouse['inbound_packages']},{warehouse['outbound_count']},{warehouse['outbound_pallets']},{warehouse['outbound_packages']}\n")
            
            f.write("\n库存概览\n")
            f.write("仓库,货物数量,板数,件数\n")
            for warehouse in data['inventory_overview']['by_warehouse']:
                f.write(f"{warehouse['warehouse_name']},{warehouse['items']},{warehouse['pallets']},{warehouse['packages']}\n")
            
            f.write(f"总计,{data['inventory_overview']['total']['items']},{data['inventory_overview']['total']['pallets']},{data['inventory_overview']['total']['packages']}\n\n")
            
            f.write("客户TOP10\n")
            f.write("客户名称,入库记录数,板数,件数\n")
            for customer in data['top_customers']:
                f.write(f"{customer['customer_name']},{customer['inbound_count']},{customer['total_pallets']},{customer['total_packages']}\n")
        
        return file_path
