from flask import Blueprint, render_template, flash, redirect, url_for, request, jsonify, current_app, send_file, abort, Response, g, make_response, session
from app import db, csrf
from app.models import InboundRecord, OutboundRecord, Inventory, Receiver, LabelCode, Warehouse, User, ReceiveRecord, TransitCargo
from sqlalchemy import func
from app.forms import InboundRecordForm, OutboundRecordForm, InboundRecordEditForm, InboundRecordEditForm2
from app.main import bp
from app.decorators import require_permission, require_any_permission, warehouse_data_filter, log_operation
from app.auth.decorators import register_page_permission, register_operation_permission
from datetime import datetime, timedelta
from flask_login import current_user, login_required
import pandas as pd
import os
import tempfile
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import time
import json
from app.utils import clean_dict_whitespace, strip_whitespace
from app.utils.identification_generator import IdentificationCodeGenerator
import csv
import io
from io import BytesIO

# 导入安全组件
try:
    from app.utils.exception_handler import (
        handle_exceptions, ValidationException, InventoryException,
        DataValidator, BusinessValidator
    )
    from app.utils.input_validator import (
        FormValidator, InputSanitizer, RequestValidator, BatchValidator
    )
    from app.utils.concurrency_control import (
        safe_inventory_update, with_inventory_lock, batch_inventory_update
    )
    from app.utils.sql_security import (
        SafeQueryBuilder, QueryOptimizer, validate_query_params
    )
    SECURITY_ENABLED = True
except ImportError as e:
    # 延迟记录错误，避免应用上下文问题
    SECURITY_ENABLED = False
    SECURITY_IMPORT_ERROR = str(e)

# 导入缓存和性能优化模块 - 暂时禁用
# from app.hot_data_cache import HotDataCacheService, cache_warmup
# from app.performance_monitor import performance_monitor, performance_metrics, perf_dashboard
# from app.cache_strategies import cache_invalidation
# from app.database_optimization import DatabaseOptimizer, QueryOptimizer

def determine_inventory_rollback_warehouse(outbound_record):
    """
    确定出库记录删除时库存应该回退到哪个仓库

    逻辑：
    1. 查找该识别编码的所有库存记录，按最后更新时间排序
    2. 如果有库存记录存在，回退到最新的库存记录所在仓库
    3. 如果没有库存记录，查找该识别编码是否被后端仓接收过
    4. 如果被后端仓接收过，回退到后端仓
    5. 否则回退到原始入库仓库

    Args:
        outbound_record: 出库记录对象

    Returns:
        int: 目标仓库ID
    """
    identification_code = outbound_record.identification_code

    if not identification_code:
        # 如果没有识别编码，回退到出库记录的操作仓库
        return outbound_record.operated_warehouse_id

    # 1. 查找该识别编码的所有库存记录，按最后更新时间排序
    existing_inventories = Inventory.query.filter_by(
        identification_code=identification_code
    ).order_by(Inventory.last_updated.desc()).all()

    if existing_inventories:
        # 找到有库存的仓库（库存数量大于0）
        for inventory in existing_inventories:
            if (inventory.pallet_count and inventory.pallet_count > 0) or \
               (inventory.package_count and inventory.package_count > 0):
                current_app.logger.info(f"库存回退：找到有库存的仓库 ID={inventory.operated_warehouse_id}, "
                                      f"识别编码={identification_code}, 库存={inventory.pallet_count}板{inventory.package_count}件")
                return inventory.operated_warehouse_id

        # 如果所有库存记录都是0，使用最新的库存记录所在仓库
        latest_inventory = existing_inventories[0]
        current_app.logger.info(f"库存回退：所有库存为0，使用最新库存记录仓库 ID={latest_inventory.operated_warehouse_id}, "
                              f"识别编码={identification_code}")
        return latest_inventory.operated_warehouse_id

    # 2. 如果没有库存记录，检查是否被后端仓接收过
    # 查找该识别编码的接收记录（InboundRecord中record_type='receive'或有batch_no的记录）
    backend_receive_record = InboundRecord.query.filter(
        InboundRecord.identification_code == identification_code,
        db.or_(
            InboundRecord.record_type == 'receive',
            db.and_(InboundRecord.batch_no.isnot(None), InboundRecord.batch_no != '')
        )
    ).first()

    if backend_receive_record:
        current_app.logger.info(f"库存回退：找到后端仓接收记录，回退到后端仓 ID={backend_receive_record.operated_warehouse_id}, "
                              f"识别编码={identification_code}")
        return backend_receive_record.operated_warehouse_id

    # 3. 最后回退到原始入库仓库
    original_inbound = InboundRecord.query.filter_by(
        identification_code=identification_code
    ).order_by(InboundRecord.inbound_time.asc()).first()

    if original_inbound:
        current_app.logger.info(f"库存回退：回退到原始入库仓库 ID={original_inbound.operated_warehouse_id}, "
                              f"识别编码={identification_code}")
        return original_inbound.operated_warehouse_id

    # 4. 如果都找不到，使用出库记录的操作仓库作为默认值
    current_app.logger.warning(f"库存回退：无法确定目标仓库，使用出库记录操作仓库 ID={outbound_record.operated_warehouse_id}, "
                             f"识别编码={identification_code}")
    return outbound_record.operated_warehouse_id

# 临时性能监控装饰器（替代被禁用的模块）
def performance_monitor(operation_name, slow_threshold=2.0):
    """临时性能监控装饰器"""
    def decorator(func):
        from functools import wraps
        @wraps(func)
        def wrapper(*args, **kwargs):
            return func(*args, **kwargs)
        return wrapper
    return decorator

def generate_real_label_pdf(print_data):
    """生成真实的标签PDF - 参考您的代码优化"""
    try:
        from reportlab.pdfgen import canvas
        from reportlab.lib.units import mm
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        import io
        import os

        current_app.logger.info("开始生成PDF...")

        # 注册中文字体（参考您的代码）
        try:
            # 尝试注册SimHei字体
            font_path = 'C:/Windows/Fonts/simhei.ttf'
            if os.path.exists(font_path):
                pdfmetrics.registerFont(TTFont('SimHei', font_path))
                chinese_font = 'SimHei'
                chinese_font_bold = 'SimHei'  # SimHei本身就是粗体
                current_app.logger.info("成功注册SimHei字体")
            else:
                # 尝试注册SimSun字体和粗体
                font_path = 'C:/Windows/Fonts/simsun.ttc'
                bold_font_path = 'C:/Windows/Fonts/simhei.ttf'  # 使用SimHei作为粗体
                if os.path.exists(font_path):
                    pdfmetrics.registerFont(TTFont('SimSun', font_path))
                    chinese_font = 'SimSun'
                    if os.path.exists(bold_font_path):
                        pdfmetrics.registerFont(TTFont('SimHei-Bold', bold_font_path))
                        chinese_font_bold = 'SimHei-Bold'
                    else:
                        chinese_font_bold = 'SimSun'
                    current_app.logger.info("成功注册SimSun字体")
                else:
                    chinese_font = 'Helvetica'
                    chinese_font_bold = 'Helvetica-Bold'
                    current_app.logger.warning("未找到中文字体，使用Helvetica")
        except Exception as font_error:
            chinese_font = 'Helvetica'
            chinese_font_bold = 'Helvetica-Bold'
            current_app.logger.warning(f"字体注册失败: {font_error}")

        # 获取纸张设置
        paper_width = float(print_data.get('paperWidth', 60)) * mm  # 默认60mm
        paper_height = float(print_data.get('paperHeight', 40)) * mm  # 默认40mm

        current_app.logger.info(f"纸张尺寸: {paper_width/mm}x{paper_height/mm}mm")

        # 创建PDF缓冲区
        buffer = io.BytesIO()

        # 创建Canvas对象
        c = canvas.Canvas(buffer, pagesize=(paper_width, paper_height))

        # 计算标签数量
        if print_data.get('isWhole', True):
            board_count = int(print_data.get('boardCount', 0))
            piece_count = int(print_data.get('pieceCount', 0))
            total_labels = board_count + piece_count
        else:
            total_labels = int(print_data.get('splitQuantity', 1))

        current_app.logger.info(f"需要生成 {total_labels} 个标签")

        # 确保至少生成一个标签
        if total_labels == 0:
            total_labels = 1

        # 生成每个标签
        for i in range(total_labels):
            current_app.logger.info(f"生成第 {i+1} 个标签")

            if i > 0:
                c.showPage()  # 新页面

            # 绘制标签内容
            try:
                draw_label_content_chinese(c, print_data, i, paper_width, paper_height, chinese_font, chinese_font_bold)
            except Exception as draw_error:
                current_app.logger.error(f"绘制第 {i+1} 个标签失败: {str(draw_error)}")
                # 绘制简单的错误标签
                c.setFont("Helvetica", 10)
                c.drawString(5*mm, paper_height - 15*mm, f"Label {i+1} - Error")
                c.drawString(5*mm, paper_height - 25*mm, str(draw_error)[:30])

        # 保存PDF
        current_app.logger.info("保存PDF...")
        c.save()

        # 获取PDF数据
        pdf_data = buffer.getvalue()
        buffer.close()

        current_app.logger.info(f"PDF生成完成，大小: {len(pdf_data)} 字节")

        # 验证PDF数据
        if len(pdf_data) < 100:
            raise Exception("PDF数据太小，可能生成失败")

        # 检查PDF头部
        if not pdf_data.startswith(b'%PDF'):
            raise Exception("PDF格式不正确，缺少PDF头部")

        return pdf_data

    except Exception as e:
        current_app.logger.error(f"PDF生成出错: {str(e)}")
        import traceback
        current_app.logger.error(f"错误详情: {traceback.format_exc()}")
        raise e

def draw_label_content_chinese(canvas_obj, print_data, index, width, height, chinese_font, chinese_font_bold):
    """绘制中文标签内容 - 参考您的代码格式"""
    from reportlab.lib.units import mm
    from datetime import datetime

    try:
        current_app.logger.info(f"开始绘制中文标签 {index + 1}")

        # 获取字体设置
        header_size = max(8, min(int(print_data.get('headerFontSize', 12)), 24))
        content_size = max(6, min(int(print_data.get('contentFontSize', 12)), 20))

        # 计算内容总高度（用于上下居中）
        line_height = 5*mm  # 每行间距
        total_lines = 4  # 标题、客户/日期、车牌、识别编码（可能多行）、订单类型/层数
        # 识别编码可能需要额外行数，预留空间
        id_code = print_data.get('identificationCode', '')
        if len(id_code) > 25:  # 如果识别编码较长，预留额外行
            total_lines += 1
        total_content_height = total_lines * line_height

        # 上下居中：从中心位置开始，向上偏移一半内容高度
        center_y = height / 2
        start_y = center_y + total_content_height / 2

        # 页边距和起始位置
        margin_left = float(print_data.get('marginLeft', 3)) * mm
        x_pos = margin_left + 2*mm
        y_pos = start_y

        # 确保位置在页面范围内
        if y_pos > height - 3*mm:
            y_pos = height - 3*mm
        if y_pos < total_content_height + 3*mm:
            y_pos = total_content_height + 3*mm
        if x_pos < 2*mm:
            x_pos = 2*mm

        current_app.logger.info(f"上下居中绘制位置: x={x_pos/mm:.1f}mm, y={y_pos/mm:.1f}mm, 内容高度={total_content_height/mm:.1f}mm")

        # 生成标题（参考您的格式）
        if print_data.get('isWhole', True):
            board_count = int(print_data.get('boardCount', 0))
            piece_count = int(print_data.get('pieceCount', 0))
            if index < board_count:
                current_text = f"第{index + 1}板"
            else:
                current_text = f"第{index - board_count + 1}件"
            header_text = f"共计{board_count}板{piece_count}件-{current_text}"
        else:
            split_num = int(print_data.get('splitBoardNum', 1))
            split_qty = int(print_data.get('splitQuantity', 1))
            header_text = f"共计{print_data.get('boardCount', 0)}板-拆第{split_num}板-{index + 1}/{split_qty}"

        # 绘制标题（第一行：14号字体，加粗）
        canvas_obj.setFont(chinese_font_bold, 14)
        canvas_obj.drawString(x_pos, y_pos, header_text)
        y_pos -= line_height

        # 获取当天日期（短格式）
        today_short = datetime.now().strftime('%m%d')

        # 绘制客户/日期（第二行：14号字体，加粗）
        canvas_obj.setFont(chinese_font_bold, 14)
        customer_date_text = f"{print_data.get('customerName', '')}/{today_short}"
        canvas_obj.drawString(x_pos, y_pos, customer_date_text)
        y_pos -= line_height

        # 绘制入库车牌
        canvas_obj.setFont(chinese_font, content_size - 1)
        plate_text = f"入库车牌: {print_data.get('plateNumber', '')}"
        canvas_obj.drawString(x_pos, y_pos, plate_text)
        y_pos -= line_height

        # 绘制识别编码（只显示内容，不显示标题，支持自动换行）
        if print_data.get('identificationCode'):
            id_text = print_data.get('identificationCode', '')

            # 计算可用宽度（标签宽度减去左右边距）
            available_width = width - x_pos - 3*mm

            # 检查文本是否需要换行
            text_width = canvas_obj.stringWidth(id_text, chinese_font, content_size)

            if text_width <= available_width:
                # 单行显示
                canvas_obj.drawString(x_pos, y_pos, id_text)
                y_pos -= line_height
            else:
                # 需要换行，按字符分割
                max_chars_per_line = int(len(id_text) * available_width / text_width)

                # 分割文本
                lines = []
                current_line = ""
                for char in id_text:
                    test_line = current_line + char
                    test_width = canvas_obj.stringWidth(test_line, chinese_font, content_size)
                    if test_width <= available_width:
                        current_line = test_line
                    else:
                        if current_line:
                            lines.append(current_line)
                        current_line = char

                if current_line:
                    lines.append(current_line)

                # 绘制多行文本
                for line in lines:
                    canvas_obj.drawString(x_pos, y_pos, line)
                    y_pos -= line_height

        # 绘制订单类型或层数
        if print_data.get('isWhole', True):
            if print_data.get('orderType'):
                order_text = f"订单类型: {print_data.get('orderType', '')}"
                canvas_obj.drawString(x_pos, y_pos, order_text)
                y_pos -= line_height
        else:
            if print_data.get('boardTiers'):
                tier_text = f"此板层数: {print_data.get('boardTiers', '')}"
                canvas_obj.drawString(x_pos, y_pos, tier_text)
                y_pos -= line_height



        # 绘制边框
        canvas_obj.rect(1*mm, 1*mm, width-2*mm, height-2*mm)

        current_app.logger.info(f"中文标签 {index + 1} 绘制完成")

    except Exception as e:
        current_app.logger.error(f"绘制中文标签内容时出错: {str(e)}")
        # 使用最简单的备用方案
        try:
            canvas_obj.setFont("Helvetica", 10)
            canvas_obj.drawString(5*mm, height - 15*mm, f"Label {index + 1}")
            canvas_obj.drawString(5*mm, height - 25*mm, print_data.get('identificationCode', 'ERROR'))
            canvas_obj.drawString(5*mm, height - 35*mm, f"Error: Drawing failed")
        except:
            pass

def draw_label_content(canvas_obj, print_data, index, width, height):
    """绘制单个标签内容"""
    from reportlab.lib.units import mm

    try:
        current_app.logger.info(f"开始绘制标签 {index + 1}")

        # 使用简单可靠的字体
        font_name = "Helvetica"

        # 获取字体设置
        header_size = max(8, min(int(print_data.get('headerFontSize', 16)), 24))
        content_size = max(6, min(int(print_data.get('contentFontSize', 12)), 20))

        # 页边距
        margin_top = float(print_data.get('marginTop', 5)) * mm
        margin_left = float(print_data.get('marginLeft', 3)) * mm

        # 起始位置
        y_pos = height - margin_top - 5*mm
        x_pos = margin_left + 2*mm

        # 确保位置在页面范围内
        if y_pos < 5*mm:
            y_pos = height - 10*mm
        if x_pos < 2*mm:
            x_pos = 2*mm

        current_app.logger.info(f"绘制位置: x={x_pos}, y={y_pos}")

        # 标题
        if print_data.get('isWhole', True):
            board_count = int(print_data.get('boardCount', 0))
            piece_count = int(print_data.get('pieceCount', 0))
            if index < board_count:
                title = f"Total {board_count}B {piece_count}P - Board {index + 1}"
            else:
                title = f"Total {board_count}B {piece_count}P - Piece {index - board_count + 1}"
        else:
            split_num = int(print_data.get('splitBoardNum', 1))
            split_qty = int(print_data.get('splitQuantity', 1))
            title = f"Total {print_data.get('boardCount', 0)}B - Split {split_num} - {index + 1}/{split_qty}"

        # 绘制标题
        canvas_obj.setFont(font_name, header_size)
        canvas_obj.drawString(x_pos, y_pos, title)
        y_pos -= 8*mm

        # 识别编码（只显示内容，不显示标题）
        canvas_obj.setFont(font_name, content_size)
        id_code = print_data.get('identificationCode', 'N/A')
        canvas_obj.drawString(x_pos, y_pos, id_code)
        y_pos -= 6*mm

        # 基本信息
        canvas_obj.setFont(font_name, content_size - 2)

        # 车牌
        plate = print_data.get('plateNumber', 'N/A')
        canvas_obj.drawString(x_pos, y_pos, f"Plate: {plate}")
        y_pos -= 5*mm

        # 客户
        customer = print_data.get('customerName', 'N/A')
        canvas_obj.drawString(x_pos, y_pos, f"Customer: {customer}")
        y_pos -= 5*mm

        # 订单类型或层数
        if print_data.get('isWhole', True):
            order_type = print_data.get('orderType', 'N/A')
            canvas_obj.drawString(x_pos, y_pos, f"Type: {order_type}")
        else:
            tiers = print_data.get('boardTiers', 'N/A')
            canvas_obj.drawString(x_pos, y_pos, f"Tiers: {tiers}")

        # 绘制边框
        canvas_obj.rect(1*mm, 1*mm, width-2*mm, height-2*mm)

        current_app.logger.info(f"标签 {index + 1} 绘制完成")

    except Exception as e:
        current_app.logger.error(f"绘制标签内容时出错: {str(e)}")
        # 使用最简单的备用方案
        try:
            canvas_obj.setFont("Helvetica", 10)
            canvas_obj.drawString(5*mm, height - 15*mm, f"Label {index + 1}")
            canvas_obj.drawString(5*mm, height - 25*mm, print_data.get('identificationCode', 'ERROR'))
            canvas_obj.drawString(5*mm, height - 35*mm, f"Error: Drawing failed")
        except:
            pass  # 如果连备用方案都失败，就跳过

def send_to_printer(pdf_data, printer_name, copies=1):
    """发送PDF到系统打印机"""
    import tempfile
    import os
    import platform

    # 验证PDF数据
    if not pdf_data or len(pdf_data) < 100:
        raise Exception("PDF数据无效或太小")

    # 创建临时文件，使用更安全的路径
    try:
        # 使用项目目录下的temp文件夹
        temp_dir = os.path.join(os.getcwd(), 'temp')
        if not os.path.exists(temp_dir):
            os.makedirs(temp_dir)

        # 生成安全的文件名
        import time
        safe_filename = f"label_print_{int(time.time())}.pdf"
        temp_path = os.path.join(temp_dir, safe_filename)

        # 写入PDF数据
        with open(temp_path, 'wb') as temp_file:
            temp_file.write(pdf_data)

        # 验证文件是否成功创建
        if not os.path.exists(temp_path) or os.path.getsize(temp_path) == 0:
            raise Exception("PDF文件创建失败")

        current_app.logger.info(f"PDF文件已创建: {temp_path} (大小: {os.path.getsize(temp_path)} 字节)")

        system = platform.system()
        current_app.logger.info(f"检测到操作系统: {system}")

        if system == "Windows":
            # Windows打印
            print_windows_pdf(temp_path, printer_name, copies)
        elif system == "Darwin":  # macOS
            print_mac_pdf(temp_path, printer_name, copies)
        elif system == "Linux":
            print_linux_pdf(temp_path, printer_name, copies)
        else:
            raise Exception(f"不支持的操作系统: {system}")

        # 打印成功后等待一段时间再删除文件
        import time
        time.sleep(3)  # 等待3秒让打印机处理文件

    except Exception as e:
        current_app.logger.error(f"打印过程出错: {str(e)}")
        # 保留文件用于调试
        if 'temp_path' in locals() and os.path.exists(temp_path):
            current_app.logger.info(f"打印失败，保留文件用于调试: {temp_path}")
        # 不抛出异常，因为PDF已经生成
    finally:
        # 延迟清理临时文件
        try:
            if 'temp_path' in locals() and os.path.exists(temp_path):
                # 再等待一段时间确保打印完成
                import time
                time.sleep(2)  # 再等待2秒
                os.unlink(temp_path)
                current_app.logger.info(f"临时文件已清理: {temp_path}")
        except Exception as e:
            current_app.logger.warning(f"清理临时文件失败: {str(e)}")

def print_windows_pdf(pdf_path, printer_name, copies):
    """Windows系统打印PDF"""
    import os
    import subprocess

    # 验证文件存在
    if not os.path.exists(pdf_path):
        raise Exception(f"PDF文件不存在: {pdf_path}")

    # 验证文件大小
    file_size = os.path.getsize(pdf_path)
    if file_size == 0:
        raise Exception(f"PDF文件为空: {pdf_path}")

    current_app.logger.info(f"准备打印PDF: {pdf_path} (大小: {file_size} 字节)")

    # 使用参考代码的方式：直接打开PDF文件
    try:
        current_app.logger.info(f"直接打开PDF文件...")

        # 方法1: 直接打开PDF文件（参考您的代码）
        try:
            import subprocess
            # 使用默认程序打开PDF
            subprocess.Popen([pdf_path], shell=True)
            current_app.logger.info(f"已打开PDF文件: {pdf_path}")
            return
        except Exception as e:
            current_app.logger.warning(f"subprocess打开失败: {str(e)}")

        # 方法2: 使用os.startfile
        try:
            os.startfile(pdf_path)
            current_app.logger.info(f"已通过os.startfile打开PDF文件")
            return
        except Exception as e:
            current_app.logger.warning(f"os.startfile失败: {str(e)}")

        # 方法3: 保存到桌面作为备用
        try:
            desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")
            desktop_pdf = os.path.join(desktop_path, f"label_print_{int(time.time())}.pdf")

            import shutil
            shutil.copy2(pdf_path, desktop_pdf)
            current_app.logger.info(f"PDF已保存到桌面: {desktop_pdf}")

            # 尝试打开桌面文件
            os.startfile(desktop_pdf)
            current_app.logger.info(f"已打开桌面PDF文件")
            return

        except Exception as e:
            current_app.logger.warning(f"桌面操作失败: {str(e)}")

        # 如果所有方法都失败
        current_app.logger.info(f"PDF文件已生成: {pdf_path}")
        current_app.logger.info(f"请手动打开文件进行查看和打印")

    except Exception as e:
        current_app.logger.error(f"PDF打开过程出错: {str(e)}")
        current_app.logger.info(f"PDF文件已生成: {pdf_path}")
        # 不抛出异常，因为PDF已经成功生成

def print_mac_pdf(pdf_path, printer_name, copies):
    """macOS系统打印PDF"""
    import subprocess

    cmd = [
        "lpr",
        "-P", printer_name,
        "-#", str(copies),
        pdf_path
    ]

    subprocess.run(cmd, check=True)
    current_app.logger.info(f"✅ 已发送到macOS打印机: {printer_name}")

def print_linux_pdf(pdf_path, printer_name, copies):
    """Linux系统打印PDF"""
    import subprocess

    cmd = [
        "lp",
        "-d", printer_name,
        "-n", str(copies),
        pdf_path
    ]

    subprocess.run(cmd, check=True)
    current_app.logger.info(f"✅ 已发送到Linux打印机: {printer_name}")

def generate_label_pdf(print_data):
    """生成标签PDF - 暂时返回错误消息，等reportlab库正确安装后再启用"""
    try:
        # 暂时返回一个简单的错误消息
        error_message = "PDF生成功能暂时不可用，请先正确安装reportlab库"
        current_app.logger.warning(error_message)

        # 创建一个简单的文本内容作为占位符
        buffer = io.BytesIO()
        content = f"""
PDF预览功能暂时不可用

原因：reportlab库未正确安装到当前虚拟环境

标签信息：
- 识别编码：{print_data.get('identificationCode', 'N/A')}
- 车牌号：{print_data.get('plateNumber', 'N/A')}
- 客户：{print_data.get('customerName', 'N/A')}
- 板数：{print_data.get('boardCount', 0)}
- 件数：{print_data.get('pieceCount', 0)}

请联系管理员安装reportlab库后重试。
        """.encode('utf-8')

        buffer.write(content)
        pdf_content = buffer.getvalue()
        buffer.close()

        return pdf_content

    except Exception as e:
        current_app.logger.error(f"生成PDF占位符时出错: {str(e)}")
        # 返回最简单的错误内容
        return b"PDF generation error"

def get_aggregated_inventory_direct():
    """直接从数据库获取聚合库存数据（替代缓存）"""
    try:
        from app.models import Inventory, Warehouse, InboundRecord, TransitCargo, OutboundRecord
        from sqlalchemy import func
        from datetime import datetime

        # 查询库存数据，按识别编码+仓库分组，分别显示不同状态
        inventory_data = db.session.query(
            Inventory.identification_code,
            # 从入库记录获取正确的客户名称和车牌号（因为入库记录是最原始和准确的）
            func.max(Inventory.customer_name).label('customer_name'),
            func.max(Inventory.plate_number).label('plate_number'),
            func.sum(Inventory.pallet_count).label('total_pallet_count'),
            func.sum(Inventory.package_count).label('total_package_count'),
            func.sum(Inventory.weight).label('total_weight'),
            func.sum(Inventory.volume).label('total_volume'),
            Inventory.operated_warehouse_id,
            Warehouse.warehouse_name,
            Warehouse.warehouse_type,
            # 业务字段聚合逻辑 - 优先获取非空值，处理None字符串
            func.coalesce(func.max(func.nullif(func.nullif(Inventory.order_type, ''), 'None')), '').label('order_type'),
            func.coalesce(func.max(func.nullif(func.nullif(Inventory.export_mode, ''), 'None')), '').label('export_mode'),
            func.coalesce(func.max(func.nullif(func.nullif(Inventory.customs_broker, ''), 'None')), '').label('customs_broker'),
            func.coalesce(func.max(func.nullif(func.nullif(Inventory.documents, ''), 'None')), '').label('documents'),
            func.coalesce(func.max(func.nullif(func.nullif(Inventory.service_staff, ''), 'None')), '').label('service_staff'),
            func.coalesce(func.max(func.nullif(func.nullif(Inventory.location, ''), 'None')), '').label('location'),
            # 添加入库时间字段
            func.max(Inventory.inbound_time).label('latest_inbound_time')
        ).join(
            Warehouse, Inventory.operated_warehouse_id == Warehouse.id
        ).filter(
            # 只查询有库存的记录
            db.or_(
                Inventory.pallet_count > 0,
                Inventory.package_count > 0
            )
        ).group_by(
            Inventory.identification_code,
            Inventory.operated_warehouse_id,
            Warehouse.warehouse_name,
            Warehouse.warehouse_type
        ).order_by(
            Inventory.identification_code,
            Warehouse.warehouse_type.desc(),  # 前端仓优先显示
            Warehouse.warehouse_name
        ).all()

        # 转换为字典格式，添加模板需要的字段
        result = []

        # 1. 处理库存记录
        for item in inventory_data:
            # 创建仓库对象
            class WarehouseObj:
                def __init__(self, id, warehouse_name, warehouse_type):
                    self.id = id
                    self.warehouse_name = warehouse_name
                    self.warehouse_type = warehouse_type

            warehouse_obj = WarehouseObj(
                item.operated_warehouse_id,
                item.warehouse_name,
                item.warehouse_type
            )

            # 从识别编码中提取正确的客户名称（格式：仓库前缀/客户名称/车牌/日期/序号）
            correct_customer_name = item.customer_name
            correct_plate_number = item.plate_number

            try:
                code_parts = item.identification_code.split('/')
                if len(code_parts) >= 3:
                    correct_customer_name = code_parts[1]  # 客户名称在第二部分
                    correct_plate_number = code_parts[2]   # 车牌号在第三部分
            except Exception as e:
                current_app.logger.warning(f"解析识别编码失败: {item.identification_code}, 错误: {e}")
                # 如果解析失败，尝试从入库记录获取正确信息
                inbound_record_for_name = InboundRecord.query.filter_by(
                    identification_code=item.identification_code
                ).first()
                if inbound_record_for_name:
                    correct_customer_name = inbound_record_for_name.customer_name
                    correct_plate_number = inbound_record_for_name.plate_number

            # 确定货物当前状态 - 库存记录显示实际在仓库的货物状态
            current_status = item.warehouse_type

            # 后端仓检查是否已出库到春疆货场或工厂
            if item.warehouse_type == 'backend':
                outbound_to_chunjiang = OutboundRecord.query.filter(
                    OutboundRecord.identification_code == item.identification_code,
                    OutboundRecord.operated_warehouse_id == item.operated_warehouse_id,
                    OutboundRecord.destination.in_(['春疆货场', '工厂'])
                ).first()

                if outbound_to_chunjiang:
                    current_status = 'shipped_to_chunjiang'

            # 使用库存表中的入库时间，如果没有则使用聚合的最新入库时间
            inbound_time = item.latest_inbound_time if hasattr(item, 'latest_inbound_time') and item.latest_inbound_time else datetime.now()

            # 获取原始重量和体积（从入库记录获取，不受分批影响）
            original_weight = item.total_weight or 0
            original_volume = item.total_volume or 0

            # 如果库存记录中的重量体积为0，从入库记录获取原始数据
            if original_weight == 0 or original_volume == 0:
                inbound_record_for_weight = InboundRecord.query.filter_by(
                    identification_code=item.identification_code
                ).first()

                if inbound_record_for_weight:
                    original_weight = inbound_record_for_weight.weight or 0
                    original_volume = inbound_record_for_weight.volume or 0

            # 业务字段优先级：库存表 -> 出库记录 -> 入库记录
            # 处理None字符串和空字符串
            order_type = item.order_type if item.order_type and item.order_type != 'None' else ''
            export_mode = item.export_mode if item.export_mode and item.export_mode != 'None' else ''
            customs_broker = item.customs_broker if item.customs_broker and item.customs_broker != 'None' else ''
            documents = item.documents if item.documents and item.documents != 'None' else ''
            service_staff = item.service_staff if item.service_staff and item.service_staff != 'None' else ''
            location = item.location if item.location and item.location != 'None' else ''

            # 如果库存表中的业务字段为空，尝试从出库记录获取
            if not export_mode or not customs_broker or not service_staff or not order_type:
                outbound_record = OutboundRecord.query.filter_by(
                    identification_code=item.identification_code,
                    operated_warehouse_id=item.operated_warehouse_id
                ).order_by(OutboundRecord.outbound_time.desc()).first()

                if outbound_record:
                    export_mode = export_mode or (outbound_record.export_mode if outbound_record.export_mode and outbound_record.export_mode != 'None' else '')
                    customs_broker = customs_broker or (outbound_record.customs_broker if outbound_record.customs_broker and outbound_record.customs_broker != 'None' else '')
                    service_staff = service_staff or (outbound_record.service_staff if outbound_record.service_staff and outbound_record.service_staff != 'None' else '')
                    order_type = order_type or (outbound_record.order_type if outbound_record.order_type and outbound_record.order_type != 'None' else '')

            # 如果仍然为空，尝试从入库记录获取
            if not export_mode or not customs_broker or not service_staff or not order_type:
                inbound_record = InboundRecord.query.filter_by(
                    identification_code=item.identification_code
                ).first()

                if inbound_record:
                    export_mode = export_mode or (inbound_record.export_mode if inbound_record.export_mode and inbound_record.export_mode != 'None' else '')
                    customs_broker = customs_broker or (inbound_record.customs_broker if inbound_record.customs_broker and inbound_record.customs_broker != 'None' else '')
                    service_staff = service_staff or (inbound_record.service_staff if inbound_record.service_staff and inbound_record.service_staff != 'None' else '')
                    order_type = order_type or (inbound_record.order_type if inbound_record.order_type and inbound_record.order_type != 'None' else '')

            result.append({
                'identification_code': item.identification_code,
                'customer_name': correct_customer_name,
                'plate_number': correct_plate_number,
                'total_pallet_count': int(item.total_pallet_count or 0),
                'total_package_count': int(item.total_package_count or 0),
                'total_weight': float(original_weight),
                'total_volume': float(original_volume),
                'operated_warehouse_id': item.operated_warehouse_id,
                # 添加筛选函数需要的字段
                'pallet_count': int(item.total_pallet_count or 0),
                'package_count': int(item.total_package_count or 0),
                'weight': float(original_weight),
                'volume': float(original_volume),
                # 添加模板需要的入库字段
                'inbound_pallet_count': int(item.total_pallet_count or 0),
                'inbound_package_count': int(item.total_package_count or 0),
                'current_warehouse_id': item.operated_warehouse_id,
                'current_warehouse': warehouse_obj,
                'current_status': current_status,
                'inbound_time': inbound_time,
                # 添加业务字段 - 使用聚合结果
                'order_type': order_type,
                'export_mode': export_mode,
                'customs_broker': customs_broker,
                'documents': documents,
                'service_staff': service_staff,
                'location': location
            })

        # 2. 处理在途记录 - 查询所有在途货物（只显示真正在途的货物）
        transit_data = db.session.query(
            TransitCargo.identification_code,
            TransitCargo.customer_name,
            TransitCargo.plate_number,
            func.sum(TransitCargo.pallet_count).label('total_pallet_count'),
            func.sum(TransitCargo.package_count).label('total_package_count'),
            func.sum(TransitCargo.weight).label('total_weight'),
            func.sum(TransitCargo.volume).label('total_volume'),
            TransitCargo.source_warehouse_id,
            TransitCargo.destination_warehouse_id,
            func.coalesce(func.max(func.nullif(TransitCargo.order_type, '')), '').label('order_type'),
            func.coalesce(func.max(func.nullif(TransitCargo.export_mode, '')), '').label('export_mode'),
            func.coalesce(func.max(func.nullif(TransitCargo.customs_broker, '')), '').label('customs_broker'),
            func.coalesce(func.max(func.nullif(TransitCargo.service_staff, '')), '').label('service_staff'),
            func.max(TransitCargo.departure_time).label('latest_departure_time')
        ).filter(
            TransitCargo.status == 'in_transit'  # 只查询真正在途的货物
        ).filter(
            db.or_(
                TransitCargo.pallet_count > 0,
                TransitCargo.package_count > 0
            )
        ).group_by(
            TransitCargo.identification_code,
            TransitCargo.customer_name,
            TransitCargo.plate_number,
            TransitCargo.source_warehouse_id,
            TransitCargo.destination_warehouse_id
        ).order_by(
            TransitCargo.identification_code
        ).all()

        # 收集已经在途的识别编码，避免在库存记录中重复显示
        in_transit_codes = set()
        for item in transit_data:
            in_transit_codes.add(item.identification_code)

        # 为在途记录创建虚拟的仓库对象和记录
        for item in transit_data:
            # 获取来源仓库信息
            source_warehouse = Warehouse.query.filter_by(id=item.source_warehouse_id).first()
            if not source_warehouse:
                continue

            # 创建仓库对象 - 显示来源仓库，但状态为在途
            class WarehouseObj:
                def __init__(self, id, warehouse_name, warehouse_type):
                    self.id = id
                    self.warehouse_name = warehouse_name
                    self.warehouse_type = warehouse_type

            warehouse_obj = WarehouseObj(
                item.source_warehouse_id,
                source_warehouse.warehouse_name,
                source_warehouse.warehouse_type
            )

            # 使用发货时间作为入库时间
            inbound_time = item.latest_departure_time if hasattr(item, 'latest_departure_time') and item.latest_departure_time else datetime.now()

            # 从识别编码中提取正确的客户名称和车牌号
            correct_customer_name = item.customer_name
            correct_plate_number = item.plate_number

            try:
                code_parts = item.identification_code.split('/')
                if len(code_parts) >= 3:
                    correct_customer_name = code_parts[1]  # 客户名称在第二部分
                    correct_plate_number = code_parts[2]   # 车牌号在第三部分
            except Exception as e:
                current_app.logger.warning(f"解析在途记录识别编码失败: {item.identification_code}, 错误: {e}")

            result.append({
                'identification_code': item.identification_code,
                'customer_name': correct_customer_name,
                'plate_number': correct_plate_number,
                'total_pallet_count': int(item.total_pallet_count or 0),
                'total_package_count': int(item.total_package_count or 0),
                'total_weight': float(item.total_weight or 0),
                'total_volume': float(item.total_volume or 0),
                'operated_warehouse_id': item.source_warehouse_id,
                # 添加筛选函数需要的字段
                'pallet_count': int(item.total_pallet_count or 0),
                'package_count': int(item.total_package_count or 0),
                'weight': float(item.total_weight or 0),
                'volume': float(item.total_volume or 0),
                # 添加模板需要的入库字段
                'inbound_pallet_count': int(item.total_pallet_count or 0),
                'inbound_package_count': int(item.total_package_count or 0),
                'current_warehouse_id': item.source_warehouse_id,
                'current_warehouse': warehouse_obj,
                'current_status': 'in_transit',  # 明确标记为在途状态
                'inbound_time': inbound_time,
                # 添加业务字段
                'order_type': item.order_type or '',
                'export_mode': item.export_mode or '',
                'customs_broker': item.customs_broker or '',
                'documents': '',
                'service_staff': item.service_staff or '',
                'location': ''
            })

        return result
    except Exception as e:
        current_app.logger.error(f"获取聚合库存数据失败: {e}")
        return []

def get_customer_list_direct(query_param=None):
    """直接从数据库获取客户列表（替代缓存）"""
    try:
        from app.models import InboundRecord

        # 构建查询
        query = db.session.query(InboundRecord.customer_name).distinct()

        if query_param:
            query = query.filter(InboundRecord.customer_name.like(f'%{query_param}%'))

        # 执行查询并获取结果
        customers = [row.customer_name for row in query.limit(20).all() if row.customer_name]

        return customers
    except Exception as e:
        current_app.logger.error(f"获取客户列表失败: {e}")
        return []

def get_cargo_status(inventory_record):
    """获取货物当前状态"""
    if not inventory_record.operated_warehouse:
        return {'status': 'unknown', 'label': '未知', 'class': 'secondary', 'icon': 'question'}

    warehouse_type = inventory_record.operated_warehouse.warehouse_type
    warehouse_name = inventory_record.operated_warehouse.warehouse_name
    identification_code = inventory_record.identification_code

    # 检查是否在途
    transit_record = TransitCargo.query.filter_by(
        identification_code=identification_code,
        status='in_transit'
    ).first()

    if transit_record:
        return {'status': 'in_transit', 'label': '在途', 'class': 'info', 'icon': 'truck'}

    if warehouse_type == 'frontend':
        return {'status': 'frontend', 'label': warehouse_name, 'class': 'primary', 'icon': 'warehouse'}
    elif warehouse_type == 'backend':
        # 检查是否已出库到春疆货场或工厂
        outbound_to_chunjiang = OutboundRecord.query.filter(
            OutboundRecord.identification_code == identification_code,
            OutboundRecord.destination.in_(['春疆货场', '工厂'])
        ).first()

        if outbound_to_chunjiang:
            return {'status': 'shipped_to_chunjiang', 'label': '已出库到春疆', 'class': 'success', 'icon': 'shipping-fast'}
        else:
            return {'status': 'backend', 'label': warehouse_name, 'class': 'warning', 'icon': 'building'}
    else:
        return {'status': 'other', 'label': '其他', 'class': 'secondary', 'icon': 'question'}
from io import StringIO, BytesIO
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl import Workbook
import win32print
from collections import defaultdict

def _get_operation_warehouse_id(is_backend_final_outbound=False, operation_type=None):
    """智能获取操作仓库ID"""
    # 如果是超级管理员，根据业务逻辑判断
    if hasattr(current_user, 'is_super_admin') and current_user.is_super_admin():
        if is_backend_final_outbound:
            # 后端仓最终出库属于后端仓操作
            return 4  # 凭祥北投仓
        else:
            # admin用户根据操作类型智能选择仓库
            if operation_type == 'backend_operation':
                return 4  # 后端仓操作
            elif operation_type == 'frontend_operation':
                # 前端仓操作，使用第一个前端仓库
                frontend_warehouse = Warehouse.query.filter_by(warehouse_type='frontend').first()
                return frontend_warehouse.id if frontend_warehouse else 1  # 默认平湖仓
            else:
                # 默认使用后端仓（因为admin通常管理后端仓）
                return 4
    else:
        # 普通用户使用自己的仓库ID
        return current_user.warehouse_id if hasattr(current_user, 'warehouse_id') else None

def get_admin_warehouse_id(operation_context='general'):
    """为admin用户获取合适的仓库ID"""
    if not (hasattr(current_user, 'is_super_admin') and current_user.is_super_admin()):
        return current_user.warehouse_id if hasattr(current_user, 'warehouse_id') else None

    # admin用户根据操作上下文确定仓库
    if operation_context == 'frontend':
        # 前端仓操作
        frontend_warehouse = Warehouse.query.filter_by(warehouse_type='frontend').first()
        return frontend_warehouse.id if frontend_warehouse else 1
    elif operation_context == 'backend':
        # 后端仓操作
        return 4  # 凭祥北投仓
    else:
        # 通用操作，默认使用后端仓
        return 4

def apply_inventory_filters(inventory_data, search_field, search_value, warehouse_id, start_date, end_date, stock_status, cargo_status):
    """应用库存筛选条件"""
    filtered_data = inventory_data

    # 搜索字段筛选
    if search_value:
        filtered_data = [
            item for item in filtered_data
            if search_value.lower() in str(item.get(search_field, '')).lower()
        ]

    # 仓库筛选
    if warehouse_id:
        filtered_data = [
            item for item in filtered_data
            if item.get('current_warehouse_id') == int(warehouse_id)
        ]

    # 日期范围筛选
    if start_date:
        try:
            start_datetime = datetime.strptime(start_date, '%Y-%m-%d')
            filtered_data = [
                item for item in filtered_data
                if item.get('inbound_time') and item['inbound_time'] >= start_datetime
            ]
        except ValueError:
            pass

    if end_date:
        try:
            end_datetime = datetime.strptime(end_date, '%Y-%m-%d') + timedelta(days=1)
            filtered_data = [
                item for item in filtered_data
                if item.get('inbound_time') and item['inbound_time'] < end_datetime
            ]
        except ValueError:
            pass

    # 库存状态筛选
    if stock_status == 'has_stock':
        filtered_data = [
            item for item in filtered_data
            if (item.get('pallet_count', 0) > 0) or (item.get('package_count', 0) > 0)
        ]
    elif stock_status == 'zero_stock':
        filtered_data = [
            item for item in filtered_data
            if (item.get('pallet_count', 0) == 0) and (item.get('package_count', 0) == 0)
        ]
    elif stock_status == 'low_stock':
        filtered_data = [
            item for item in filtered_data
            if ((item.get('pallet_count', 0) > 0) and (item.get('pallet_count', 0) < 5)) or
               ((item.get('package_count', 0) > 0) and (item.get('package_count', 0) < 10))
        ]

    # 货物状态筛选
    if cargo_status:
        filtered_data = [
            item for item in filtered_data
            if item.get('current_status') == cargo_status
        ]

    return filtered_data

def get_aggregated_inventory_data():
    """获取按货物当前状态聚合的库存数据 - 方案A：分状态显示"""

    # 1. 获取所有库存记录，按identification_code分组
    all_inventories = Inventory.query.all()
    inventory_groups = defaultdict(list)

    for inv in all_inventories:
        if inv.identification_code:
            inventory_groups[inv.identification_code].append(inv)

    aggregated_results = []

    # 2. 对每个identification_code进行状态分离处理
    for identification_code, inventories in inventory_groups.items():
        if not inventories:
            continue

        # 获取基础信息（使用最早的入库记录）
        base_inventory = min(inventories, key=lambda x: x.inbound_time or datetime.min)

        # 获取出库记录
        outbound_records = OutboundRecord.query.filter_by(identification_code=identification_code).all()

        # 计算出库到春疆货场/工厂的总数量
        chunjiang_outbound_pallet = sum(out.pallet_count or 0 for out in outbound_records
                                       if out.destination in ['春疆货场', '工厂'])
        chunjiang_outbound_package = sum(out.package_count or 0 for out in outbound_records
                                        if out.destination in ['春疆货场', '工厂'])

        # 获取原始入库数量（直接根据识别编码匹配第一条入库记录）
        original_inbound_record = InboundRecord.query.filter_by(identification_code=identification_code).first()
        if original_inbound_record:
            # 直接使用入库记录表中的数量
            original_pallet_count = original_inbound_record.pallet_count or 0
            original_package_count = original_inbound_record.package_count or 0
        else:
            # 如果没有入库记录，设为0（不应该出现这种情况）
            original_pallet_count = 0
            original_package_count = 0

        # 基础信息模板
        base_info = {
            'identification_code': identification_code,
            'customer_name': base_inventory.customer_name,
            'plate_number': base_inventory.plate_number or '',  # 添加入库车牌字段
            'order_type': base_inventory.order_type or '',
            'export_mode': base_inventory.export_mode or '',
            'customs_broker': base_inventory.customs_broker or '',
            'service_staff': base_inventory.service_staff or '',
            'documents': base_inventory.documents or '',
            'weight': base_inventory.weight or 0,
            'volume': base_inventory.volume or 0,
            'inbound_time': base_inventory.inbound_time,
            'last_updated': max(inv.last_updated for inv in inventories if inv.last_updated) or datetime.now(),
            'chunjiang_outbound_pallet': chunjiang_outbound_pallet,
            'chunjiang_outbound_package': chunjiang_outbound_package,
            'original_pallet_count': original_pallet_count,  # 添加原始入库数量
            'original_package_count': original_package_count
        }

        # 3. 处理在途状态
        in_transit_records = TransitCargo.query.filter_by(
            identification_code=identification_code,
            status='in_transit'
        ).all()

        if in_transit_records:
            total_transit_pallet = sum(transit.pallet_count or 0 for transit in in_transit_records)
            total_transit_package = sum(transit.package_count or 0 for transit in in_transit_records)

            if total_transit_pallet > 0 or total_transit_package > 0:
                transit_item = base_info.copy()
                transit_item.update({
                    'current_status': 'in_transit',
                    'current_warehouse_id': None,
                    'current_warehouse': None,
                    'pallet_count': total_transit_pallet,
                    'package_count': total_transit_package,
                    'inbound_pallet_count': base_info['original_pallet_count'],  # 使用原始入库数量
                    'inbound_package_count': base_info['original_package_count'],
                    'location': '在途中'
                })
                aggregated_results.append(transit_item)

        # 4. 处理前端仓状态
        frontend_inventories = [inv for inv in inventories
                               if inv.operated_warehouse and inv.operated_warehouse.warehouse_type == 'frontend']

        # 按仓库分组前端库存（只包含有库存的记录）
        frontend_warehouse_groups = defaultdict(list)
        for inv in frontend_inventories:
            if (inv.pallet_count or 0) > 0 or (inv.package_count or 0) > 0:
                frontend_warehouse_groups[inv.operated_warehouse_id].append(inv)

        for warehouse_id, warehouse_inventories in frontend_warehouse_groups.items():
            total_pallet = sum(inv.pallet_count or 0 for inv in warehouse_inventories)
            total_package = sum(inv.package_count or 0 for inv in warehouse_inventories)

            if total_pallet > 0 or total_package > 0:
                warehouse = warehouse_inventories[0].operated_warehouse
                location = warehouse_inventories[0].location or ''

                frontend_item = base_info.copy()
                frontend_item.update({
                    'current_status': 'frontend',
                    'current_warehouse_id': warehouse_id,
                    'current_warehouse': warehouse,
                    'pallet_count': total_pallet,
                    'package_count': total_package,
                    'inbound_pallet_count': base_info['original_pallet_count'],  # 使用原始入库数量
                    'inbound_package_count': base_info['original_package_count'],
                    'location': location
                })
                aggregated_results.append(frontend_item)

        # 5. 处理后端仓状态
        backend_inventories = [inv for inv in inventories
                              if inv.operated_warehouse and inv.operated_warehouse.warehouse_type == 'backend']

        # 按仓库分组后端库存（只包含有库存的记录）
        backend_warehouse_groups = defaultdict(list)
        for inv in backend_inventories:
            if (inv.pallet_count or 0) > 0 or (inv.package_count or 0) > 0:
                backend_warehouse_groups[inv.operated_warehouse_id].append(inv)

        for warehouse_id, warehouse_inventories in backend_warehouse_groups.items():
            total_pallet = sum(inv.pallet_count or 0 for inv in warehouse_inventories)
            total_package = sum(inv.package_count or 0 for inv in warehouse_inventories)

            # 显示所有记录，包括库存为0的记录（因为需要显示原始入库数量）
            if True:  # 移除库存大于0的限制
                warehouse = warehouse_inventories[0].operated_warehouse
                location = warehouse_inventories[0].location or ''

                backend_item = base_info.copy()
                backend_item.update({
                    'current_status': 'backend',
                    'current_warehouse_id': warehouse_id,
                    'current_warehouse': warehouse,
                    'pallet_count': total_pallet,
                    'package_count': total_package,
                    'inbound_pallet_count': base_info['original_pallet_count'],  # 使用原始入库数量
                    'inbound_package_count': base_info['original_package_count'],
                    'location': location
                })
                aggregated_results.append(backend_item)

    return aggregated_results



def _get_user_warehouse_info(required_type=None):
    """获取用户仓库信息，支持超级管理员"""
    if hasattr(current_user, 'is_super_admin') and current_user.is_super_admin():
        # 超级管理员根据需要的类型返回相应仓库
        if required_type:
            warehouse = Warehouse.query.filter_by(warehouse_type=required_type).first()
            if not warehouse:
                return None, f'系统中没有{required_type}仓库'
            return warehouse, None
        else:
            # 默认返回第一个前端仓库
            warehouse = Warehouse.query.filter_by(warehouse_type='frontend').first()
            if not warehouse:
                return None, '系统中没有前端仓库'
            return warehouse, None
    else:
        # 普通用户返回自己的仓库
        if not hasattr(current_user, 'warehouse') or not current_user.warehouse:
            return None, '用户仓库信息不完整'

        if required_type and current_user.warehouse.warehouse_type != required_type:
            return None, f'只有{required_type}仓库用户可以执行此操作'

        return current_user.warehouse, None

def _check_warehouse_access(warehouse_type=None):
    """检查用户是否有仓库访问权限"""
    if hasattr(current_user, 'is_super_admin') and current_user.is_super_admin():
        return True, None

    if not hasattr(current_user, 'warehouse') or not current_user.warehouse:
        return False, '用户仓库信息不完整'

    if warehouse_type and current_user.warehouse.warehouse_type != warehouse_type:
        return False, f'只有{warehouse_type}仓库用户可以访问此功能'

    return True, None

# 自定义简单分页类，替代flask_sqlalchemy的Pagination
class SimplePagination:
    def __init__(self, items=None, page=1, per_page=50, total=0):
        self.items = items or []
        self.page = page
        self.per_page = per_page
        self.total = total

    @property
    def has_prev(self):
        return self.page > 1

    @property
    def has_next(self):
        return self.page * self.per_page < self.total

    @property
    def prev_num(self):
        return self.page - 1 if self.has_prev else self.page

    @property
    def next_num(self):
        return self.page + 1 if self.has_next else self.page

    @property
    def pages(self):
        """计算总页数"""
        if self.per_page == 0:
            return 0
        return max(1, (self.total + self.per_page - 1) // self.per_page)

    def iter_pages(self, left_edge=1, right_edge=1, left_current=1, right_current=2):
        last = 0
        for num in range(1, int(self.total / self.per_page) + (1 if self.total % self.per_page else 0) + 1):
            if num <= left_edge or \
               (num > self.page - left_current - 1 and num < self.page + right_current) or \
               num > int(self.total / self.per_page) - right_edge:
                if last + 1 != num:
                    yield None
                yield num
                last = num

# 添加获取当前时间的函数给模板使用
@bp.context_processor
def utility_processor():
    """添加通用工具函数给模板使用"""
    def now():
        return datetime.now()

    return dict(now=now)

@bp.route('/')
@bp.route('/index')
@login_required  # 添加登录要求
def index():
    """首页"""
    # 检查用户是否已登录
    if not current_user.is_authenticated:
        return render_template('index.html', title='首页')

    # 根据用户类型显示不同的首页内容
    user_info = {
        'username': current_user.username,
        'real_name': getattr(current_user, 'real_name', ''),
        'warehouse_id': getattr(current_user, 'warehouse_id', None),
        'warehouse_name': '',
        'warehouse_type': '',
        'is_admin': current_user.username == 'admin' or getattr(current_user, 'is_admin', False)
    }

    # 获取仓库信息
    if user_info['warehouse_id']:
        try:
            warehouse = Warehouse.query.get(user_info['warehouse_id'])
            if warehouse:
                user_info['warehouse_name'] = warehouse.warehouse_name
                user_info['warehouse_type'] = warehouse.warehouse_type
        except:
            pass

    # 获取用户角色
    user_roles = []
    try:
        if hasattr(current_user, 'get_roles'):
            roles = current_user.get_roles()
            user_roles = [role.role_name for role in roles]
    except:
        pass

    user_info['roles'] = user_roles

    # 返回已登录用户的首页
    return render_template('index.html', title='仓储管理系统', user_info=user_info)

@bp.route('/test_permissions')
def test_permissions():
    """权限测试页面"""
    return render_template('test_permissions.html', title='权限测试')

    return render_template('index.html', title='首页', user_info=user_info)

@bp.route('/csrf_debug')
def csrf_debug():
    """CSRF调试页面"""
    return render_template('csrf_debug.html', title='CSRF调试')

@bp.route('/debug_permissions')
def debug_permissions():
    """权限调试页面"""
    return render_template('debug_permissions.html', title='权限调试')

@bp.route('/test_api')
def test_api():
    """API测试页面"""
    return '''
<!DOCTYPE html>
<html lang="zh-CN">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>API测试页面</title>
    <link href="https://cdn.bootcdn.net/ajax/libs/twitter-bootstrap/5.3.0/css/bootstrap.min.css" rel="stylesheet">
    <script src="https://cdn.bootcdn.net/ajax/libs/jquery/3.6.0/jquery.min.js"></script>
    <meta name="csrf-token" content="''' + csrf_token() + '''">
</head>
<body>
    <div class="container mt-4">
        <h2>后端库存API测试</h2>

        <div class="card">
            <div class="card-header">
                <h5>API响应测试</h5>
            </div>
            <div class="card-body">
                <button id="testBtn" class="btn btn-primary">测试API</button>
                <div id="result" class="mt-3"></div>
            </div>
        </div>

        <div class="card mt-4">
            <div class="card-header">
                <h5>数据表格</h5>
            </div>
            <div class="card-body">
                <div class="table-responsive">
                    <table class="table table-striped">
                        <thead>
                            <tr>
                                <th>ID</th>
                                <th>客户名称</th>
                                <th>识别编码</th>
                                <th>入库车牌</th>
                                <th>送货干线车</th>
                                <th>订单类型</th>
                                <th>出境模式</th>
                                <th>报关行</th>
                            </tr>
                        </thead>
                        <tbody id="dataTable">
                        </tbody>
                    </table>
                </div>
            </div>
        </div>
    </div>

    <script>
        $('#testBtn').click(function() {
            $('#result').html('<div class="spinner-border" role="status"><span class="visually-hidden">Loading...</span></div>');

            $.ajax({
                url: '/api/backend/inventory/search',
                method: 'GET',
                data: {},
                headers: {
                    'X-CSRFToken': $('meta[name=csrf-token]').attr('content')
                },
                success: function(response) {
                    console.log('API响应:', response);

                    if (response.success && response.data) {
                        let html = `
                            <div class="alert alert-success">
                                <h6>API调用成功</h6>
                                <p>返回记录数: ${response.data.length}</p>
                            </div>
                        `;
                        $('#result').html(html);

                        // 渲染表格
                        let tableHtml = '';
                        response.data.forEach(item => {
                            console.log('处理记录:', item);
                            tableHtml += `
                                <tr>
                                    <td>${item.id}</td>
                                    <td>${item.customer_name || '-'}</td>
                                    <td>${item.identification_code || '-'}</td>
                                    <td><strong style="color: blue;">${item.plate_number || '(空)'}</strong></td>
                                    <td><strong style="color: red;">${item.delivery_plate_number || '(空)'}</strong></td>
                                    <td>${item.order_type || '-'}</td>
                                    <td>${item.export_mode || '-'}</td>
                                    <td>${item.customs_broker || '-'}</td>
                                </tr>
                            `;
                        });
                        $('#dataTable').html(tableHtml);

                    } else {
                        $('#result').html(`
                            <div class="alert alert-danger">
                                <h6>API调用失败</h6>
                                <p>${response.message || '未知错误'}</p>
                            </div>
                        `);
                    }
                },
                error: function(xhr, status, error) {
                    console.error('API调用错误:', error);
                    $('#result').html(`
                        <div class="alert alert-danger">
                            <h6>API调用错误</h6>
                            <p>状态码: ${xhr.status}</p>
                            <p>错误信息: ${error}</p>
                            <p>响应: ${xhr.responseText}</p>
                        </div>
                    `);
                }
            });
        });
    </script>
</body>
</html>
    '''

@bp.route('/inbound', methods=['GET', 'POST'])
@require_permission('INBOUND_VIEW')
def inbound():
    """入库操作页面 - 批量导入页面"""
    # 获取当前用户的仓库信息
    warehouse_type = None
    warehouse_name = None
    if current_user.warehouse:
        warehouse_type = current_user.warehouse.warehouse_type
        warehouse_name = current_user.warehouse.warehouse_name

    # 根据仓库类型设置页面标题
    if warehouse_type == 'backend':
        title = '后端仓入库操作（批量录入）'
    else:
        title = '前端仓入库操作（批量录入）'

    return render_template('inbound_batch.html',
                         title=title,
                         warehouse_type=warehouse_type,
                         warehouse_name=warehouse_name)

@bp.route('/inbound/list')
@require_permission('INBOUND_VIEW')
def inbound_list():
    """入库记录列表"""
    # 获取搜索参数
    date_start = request.args.get('date_start', '')
    date_end = request.args.get('date_end', '')
    plate_number = request.args.get('plate_number', '')
    customer_name = request.args.get('customer_name', '')
    export_mode = request.args.get('export_mode', '')
    customs_broker = request.args.get('customs_broker', '')
    service_staff = request.args.get('service_staff', '')
    # 新增操作追踪筛选参数
    operated_warehouse_id = request.args.get('operated_warehouse_id', '')
    operated_user_id = request.args.get('operated_user_id', '')

    # 如果没有指定日期范围，默认使用最近一周的日期范围
    if not date_start and not date_end:
        today = datetime.now().strftime('%Y-%m-%d')
        one_week_ago = (datetime.now() - timedelta(days=7)).strftime('%Y-%m-%d')
        date_start = one_week_ago
        date_end = today

    # 确保搜索参数传递给模板，先定义
    search_params = {
        'date_start': date_start,
        'date_end': date_end,
        'plate_number': plate_number,
        'customer_name': customer_name,
        'export_mode': export_mode,
        'customs_broker': customs_broker,
        'service_staff': service_staff,
        'operated_warehouse_id': operated_warehouse_id,
        'operated_user_id': operated_user_id
    }

    try:
        page = request.args.get('page', 1, type=int)
        per_page = current_app.config.get('ITEMS_PER_PAGE', 50)

        # 构建查询 - 使用eager loading优化性能
        query = InboundRecord.query.options(
            db.joinedload(InboundRecord.operated_warehouse),
            db.joinedload(InboundRecord.operated_by_user)
        )

        # 按日期范围筛选
        if date_start:
            try:
                start_date = datetime.strptime(date_start, '%Y-%m-%d')
                query = query.filter(InboundRecord.inbound_time >= start_date)
            except ValueError:
                flash('开始日期格式无效', 'warning')

        if date_end:
            try:
                # 将结束日期设置为当天的23:59:59
                end_date = datetime.strptime(date_end, '%Y-%m-%d')
                end_date = end_date.replace(hour=23, minute=59, second=59)
                query = query.filter(InboundRecord.inbound_time <= end_date)
            except ValueError:
                flash('结束日期格式无效', 'warning')

        # 按车牌号筛选
        if plate_number:
            query = query.filter(InboundRecord.plate_number.like(f'%{plate_number}%'))

        # 按客户名称筛选
        if customer_name:
            query = query.filter(InboundRecord.customer_name.like(f'%{customer_name}%'))

        # 按出境模式筛选
        if export_mode:
            query = query.filter(InboundRecord.export_mode.like(f'%{export_mode}%'))

        # 按报关行筛选
        if customs_broker:
            query = query.filter(InboundRecord.customs_broker.like(f'%{customs_broker}%'))

        # 按跟单客服筛选
        if service_staff:
            query = query.filter(InboundRecord.service_staff.like(f'%{service_staff}%'))

        # 按操作仓库筛选
        if operated_warehouse_id:
            query = query.filter(InboundRecord.operated_warehouse_id == operated_warehouse_id)

        # 按操作用户筛选
        if operated_user_id:
            query = query.filter(InboundRecord.operated_by_user_id == operated_user_id)

        # 按入库时间升序排序
        query = query.order_by(InboundRecord.inbound_time.asc())

        # 获取总记录数
        total_count = query.count()

        # 分页
        items = query.offset((page - 1) * per_page).limit(per_page).all()

        # 使用自定义分页类
        records = SimplePagination(
            items=items,
            page=page,
            per_page=per_page,
            total=total_count
        )

        current_app.logger.info(f"分页信息: 页数={records.pages}, 总记录数={records.total}")

        # 获取仓库和用户选项数据
        from app.models import Warehouse, User
        warehouses = Warehouse.query.filter_by(status='active').all()
        users = User.query.filter_by(status='active').all()

        return render_template(
            'inbound_list.html',
            title='入库记录',
            records=records,
            search_params=search_params,
            warehouses=warehouses,
            users=users
        )
    except Exception as e:
        current_app.logger.error(f"Error in inbound_list: {str(e)}")
        flash(f"加载数据时出错: {str(e)}", "danger")

        # 在异常情况下也传递search_params
        empty_pagination = SimplePagination(items=[], page=1, per_page=50, total=0)

        return render_template(
            'inbound_list.html',
            title='入库记录',
            records=empty_pagination,
            search_params=search_params
        )

@bp.route('/export_inbound')
def export_inbound():
    """导出入库记录到Excel"""
    try:
        # 获取搜索参数（与inbound_list相同的参数处理）
        date_start = request.args.get('date_start', '')
        date_end = request.args.get('date_end', '')
        plate_number = request.args.get('plate_number', '')
        customer_name = request.args.get('customer_name', '')
        export_mode = request.args.get('export_mode', '')
        customs_broker = request.args.get('customs_broker', '')
        service_staff = request.args.get('service_staff', '')

        # 构建查询
        query = InboundRecord.query

        # 按日期范围筛选
        if date_start:
            try:
                start_date = datetime.strptime(date_start, '%Y-%m-%d')
                query = query.filter(InboundRecord.inbound_time >= start_date)
            except ValueError:
                flash('开始日期格式无效', 'warning')

        if date_end:
            try:
                end_date = datetime.strptime(date_end, '%Y-%m-%d')
                end_date = end_date.replace(hour=23, minute=59, second=59)
                query = query.filter(InboundRecord.inbound_time <= end_date)
            except ValueError:
                flash('结束日期格式无效', 'warning')

        # 按车牌号筛选
        if plate_number:
            query = query.filter(InboundRecord.plate_number.like(f'%{plate_number}%'))

        # 按客户名称筛选
        if customer_name:
            query = query.filter(InboundRecord.customer_name.like(f'%{customer_name}%'))

        # 按出境模式筛选
        if export_mode:
            query = query.filter(InboundRecord.export_mode.like(f'%{export_mode}%'))

        # 按报关行筛选
        if customs_broker:
            query = query.filter(InboundRecord.customs_broker.like(f'%{customs_broker}%'))

        # 按跟单客服筛选
        if service_staff:
            query = query.filter(InboundRecord.service_staff.like(f'%{service_staff}%'))

        # 按入库时间升序排序
        query = query.order_by(InboundRecord.inbound_time.asc())

        # 获取所有符合条件的记录
        records = query.all()

        # 如果没有记录，返回提示
        if not records:
            flash('没有找到符合条件的记录', 'warning')
            return redirect(url_for('main.inbound_list'))

        try:
            # 尝试导入pandas
            import pandas as pd
            from pandas import ExcelWriter
            from io import BytesIO

            # 构建DataFrame
            data = []
            for i, record in enumerate(records, 1):
                data.append({
                    '序号': i,
                    '入库时间': record.inbound_time.strftime('%Y-%m-%d') if record.inbound_time else '',
                    '入库车牌': record.plate_number,
                    '客户名称': record.customer_name,
                    '识别编码': record.identification_code or '',
                    '板数': record.pallet_count,
                    '件数': record.package_count,
                    '重量(kg)': record.weight,
                    '体积(m³)': record.volume,
                    '出境模式': record.export_mode,
                    '报关行': record.customs_broker,
                    '单据': record.documents,
                    '跟单客服': record.service_staff,
                    '创建时间': record.inbound_time.strftime('%Y-%m-%d %H:%M:%S') if record.inbound_time else ''
                })

            df = pd.DataFrame(data)

            # 创建内存文件对象
            output = BytesIO()

            # 使用ExcelWriter可以更好地控制格式
            with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
                # 写入数据
                df.to_excel(writer, sheet_name='入库记录', index=False)

                # 获取workbook和worksheet对象以进行格式调整
                workbook = writer.book
                worksheet = writer.sheets['入库记录']

                # 定义格式
                header_format = workbook.add_format({
                    'bold': True,
                    'text_wrap': True,
                    'valign': 'vcenter',
                    'align': 'center',
                    'bg_color': '#D9E1F2',  # 浅蓝色
                    'border': 1
                })

                # 为所有列设置宽度
                for i, col in enumerate(df.columns):
                    # 根据列名和内容设置适当的列宽
                    max_len = max(
                        df[col].astype(str).map(len).max(),  # 最长数据长度
                        len(str(col))  # 列名长度
                    ) + 2  # 添加一些额外空间

                    # 限制最大宽度
                    col_width = min(max_len, 30)
                    worksheet.set_column(i, i, col_width)

                # 设置表头格式
                for col_num, value in enumerate(df.columns.values):
                    worksheet.write(0, col_num, value, header_format)

                # 添加自动筛选
                worksheet.autofilter(0, 0, len(df), len(df.columns) - 1)

            # 设置文件指针到开始位置
            output.seek(0)

            # 生成下载文件名
            timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
            filename = f"入库记录导出_{timestamp}.xlsx"

            # 返回Excel文件
            return send_file(
                output,
                as_attachment=True,
                download_name=filename,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )

        except ImportError as e:
            # 如果pandas或xlsxwriter不可用，尝试使用openpyxl
            try:
                # 创建工作簿和工作表
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "入库记录"

                # 添加表头
                headers = ['序号', '入库时间', '入库车牌', '客户名称', '识别编码', '板数', '件数',
                          '重量(kg)', '体积(m³)', '出境模式', '报关行', '单据', '跟单客服', '创建时间']

                for col_idx, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_idx)
                    cell.value = header
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )

                # 添加数据
                for i, record in enumerate(records, 1):
                    row_idx = i + 1
                    ws.cell(row=row_idx, column=1).value = i
                    ws.cell(row=row_idx, column=2).value = record.outbound_time.strftime('%Y-%m-%d') if record.outbound_time else ''
                    ws.cell(row=row_idx, column=3).value = record.plate_number
                    ws.cell(row=row_idx, column=4).value = record.customer_name
                    ws.cell(row=row_idx, column=5).value = record.identification_code or ''
                    ws.cell(row=row_idx, column=6).value = record.pallet_count
                    ws.cell(row=row_idx, column=7).value = record.package_count
                    ws.cell(row=row_idx, column=8).value = record.weight
                    ws.cell(row=row_idx, column=9).value = record.volume
                    ws.cell(row=row_idx, column=10).value = record.export_mode
                    ws.cell(row=row_idx, column=11).value = record.customs_broker
                    ws.cell(row=row_idx, column=12).value = record.documents
                    ws.cell(row=row_idx, column=13).value = record.service_staff
                    ws.cell(row=row_idx, column=14).value = record.outbound_time.strftime('%Y-%m-%d %H:%M:%S') if record.outbound_time else ''

                # 调整列宽
                for col_idx, header in enumerate(headers, 1):
                    col_letter = get_column_letter(col_idx)
                    # 设置一个合理的默认宽度
                    ws.column_dimensions[col_letter].width = max(len(header) * 1.5, 10)

                # 创建临时文件
                temp_file = tempfile.NamedTemporaryFile(
                    suffix='.xlsx',
                    prefix='inbound_export_',
                    delete=False
                )

                # 保存工作簿
                wb.save(temp_file.name)

                # 返回文件下载响应
                return send_file(
                    temp_file.name,
                    as_attachment=True,
                    download_name=f"入库记录导出_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx",
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )

            except ImportError:
                # 如果openpyxl也不可用，返回错误信息
                flash("导出Excel需要pandas或openpyxl库，请安装: pip install pandas xlsxwriter 或 pip install openpyxl", "danger")
                return redirect(url_for('main.inbound_list'))
    except Exception as e:
        current_app.logger.error(f"导出数据时出错: {str(e)}")
        flash(f"导出数据时出错: {str(e)}", "danger")
        return redirect(url_for('main.inbound_list'))

@bp.route('/inbound/view/<int:id>')
@require_permission('INBOUND_VIEW')
def view_inbound(id):
    """查看入库记录详情"""
    try:
        # 查询记录
        record = InboundRecord.query.get_or_404(id)
        return render_template(
            'inbound_view.html',
            title='入库记录详情',
            record=record
        )
    except Exception as e:
        current_app.logger.error(f"查看入库记录详情时出错: {str(e)}")
        flash(f"查看入库记录详情时出错: {str(e)}", "danger")
        return redirect(url_for('main.inbound_list'))

@bp.route('/inbound/edit/<int:id>', methods=['GET', 'POST'])
@require_permission('INBOUND_EDIT')
@log_operation('inbound', 'edit', 'inbound_record')
def edit_inbound(id):
    """编辑入库记录"""
    try:
        # 查询记录
        record = InboundRecord.query.get_or_404(id)
        current_app.logger.info(f"找到记录: ID={record.id}, 单据={record.documents}")

        # 获取编辑前的值（用于后续库存调整）
        old_customer_name = record.customer_name
        old_pallet_count = record.pallet_count
        old_package_count = record.package_count
        old_weight = record.weight
        old_volume = record.volume
        old_identification_code = record.identification_code

        if request.method == 'POST':
            # 获取表单数据
            inbound_time = request.form.get('inbound_time', '')
            plate_number = request.form.get('plate_number', '')
            customer_name = request.form.get('customer_name', '')

            # 板数和件数的特殊处理，确保是整数
            try:
                pallet_value = request.form.get('pallet_count', '0') or 0
                package_value = request.form.get('package_count', '0') or 0

                # 检查是否为小数
                if isinstance(pallet_value, str) and '.' in pallet_value:
                    pallet_float = float(pallet_value)
                    if pallet_float != int(pallet_float):
                        flash("板数必须是整数，不能是小数", "danger")
                        return redirect(url_for('main.edit_inbound', id=id))
                    pallet_count = int(pallet_float)
                else:
                    pallet_count = int(pallet_value)

                if isinstance(package_value, str) and '.' in package_value:
                    package_float = float(package_value)
                    if package_float != int(package_float):
                        flash("件数必须是整数，不能是小数", "danger")
                        return redirect(url_for('main.edit_inbound', id=id))
                    package_count = int(package_float)
                else:
                    package_count = int(package_value)

                if pallet_count < 0 or package_count < 0:
                    flash("板数和件数不能为负数", "danger")
                    return redirect(url_for('main.edit_inbound', id=id))

            except ValueError:
                flash("板数和件数必须是有效的整数", "danger")
                return redirect(url_for('main.edit_inbound', id=id))

            # 判断板数和件数是否同时为0
            if pallet_count == 0 and package_count == 0:
                flash("板数和件数不能同时为0", "danger")
                return redirect(url_for('main.edit_inbound', id=id))

            # 重量和体积可以为空
            try:
                weight = float(request.form.get('weight', '0') or 0)
                volume = float(request.form.get('volume', '0') or 0)
            except ValueError:
                flash("重量和体积必须是有效数字", "danger")
                return redirect(url_for('main.edit_inbound', id=id))

            export_mode = request.form.get('export_mode', '')
            customs_broker = request.form.get('customs_broker', '')
            location = request.form.get('location', '')  # 获取库位信息
            documents = request.form.get('documents', '')
            service_staff = request.form.get('service_staff', '')

            # 验证必填字段
            if not inbound_time:
                flash("入库时间不能为空", "danger")
                return redirect(url_for('main.edit_inbound', id=id))
            if not plate_number:
                flash("入库车牌不能为空", "danger")
                return redirect(url_for('main.edit_inbound', id=id))
            if not customer_name:
                flash("客户名称不能为空", "danger")
                return redirect(url_for('main.edit_inbound', id=id))
            if not service_staff:
                flash("跟单客服不能为空", "danger")
                return redirect(url_for('main.edit_inbound', id=id))

            # 处理入库时间
            try:
                # 打印接收到的日期格式，用于调试
                current_app.logger.info(f"接收到的日期格式: {inbound_time}")

                # 尝试多种日期格式
                formats_to_try = [
                    '%Y-%m-%d',        # 标准日期格式: 2025-06-27
                    '%Y/%m/%d',        # 斜杠分隔: 2025/06/27
                    '%Y.%m.%d',        # 点分隔: 2025.06.27
                    '%Y年%m月%d日',     # 中文格式: 2025年06月27日
                    '%Y-%m-%dT%H:%M',  # HTML5日期时间格式: 2025-06-27T12:30
                    '%Y-%m-%d %H:%M:%S' # 完整日期时间: 2025-06-27 12:30:45
                ]

                inbound_time_obj = None
                for date_format in formats_to_try:
                    try:
                        inbound_time_obj = datetime.strptime(inbound_time, date_format)
                        break  # 如果成功解析，跳出循环
                    except ValueError:
                        continue  # 尝试下一个格式

                # 如果所有格式都失败
                if inbound_time_obj is None:
                    flash(f"入库时间格式不正确，请使用YYYY-MM-DD格式（如2025-06-27）", "danger")
                    return redirect(url_for('main.edit_inbound', id=id))

                # 设置时间为00:00:00，只保留日期部分
                inbound_time_obj = inbound_time_obj.replace(hour=0, minute=0, second=0)

            except Exception as e:
                flash(f"处理入库时间时出错: {str(e)}", "danger")
                return redirect(url_for('main.edit_inbound', id=id))

            # 更新记录
            record.outbound_time = inbound_time_obj
            record.plate_number = plate_number
            record.customer_name = customer_name
            record.pallet_count = pallet_count
            record.package_count = package_count
            record.weight = weight
            record.volume = volume
            record.export_mode = export_mode
            record.customs_broker = customs_broker
            record.location = location  # 更新库位
            record.documents = documents
            record.service_staff = service_staff

            # 判断是否需要更新识别编码
            identification_code_changed = False
            new_identification_code = old_identification_code

            # 如果客户名称、车牌号或日期变更了，则重新生成识别编码
            if (old_customer_name != customer_name or
                record.plate_number != plate_number or
                record.outbound_time.strftime('%Y-%m-%d') != inbound_time_obj.strftime('%Y-%m-%d')):

                # 生成新的识别编码 - 使用新规则：仓库前缀/客户全称/车牌/日期/序号
                new_identification_code = IdentificationCodeGenerator.generate_unique_identification_code(
                    warehouse_id=record.operated_warehouse_id,
                    customer_name=customer_name,
                    plate_number=plate_number,
                    operation_type='inbound',
                    inbound_date=inbound_time_obj
                )
                record.identification_code = new_identification_code
                identification_code_changed = True

            # 记录更新后的值
            current_app.logger.info(f"准备更新: documents={record.documents}, 识别编码={record.identification_code}")

            try:
                # 更新库存记录 - 按照识别编码管理
                # 查找旧的库存记录
                old_inventory = Inventory.query.filter_by(identification_code=old_identification_code).first()

                # 检查是否有出库记录，如果有则需要验证库存数量不能小于已出库数量
                outbound_records = OutboundRecord.query.filter_by(identification_code=old_identification_code).all()
                total_outbound_pallet = sum(record.pallet_count for record in outbound_records)
                total_outbound_package = sum(record.package_count for record in outbound_records)

                # 如果有出库记录，检查新的入库数量是否足够
                if outbound_records:
                    if pallet_count < total_outbound_pallet:
                        flash(f'修改失败：板数不能小于已出库数量 {total_outbound_pallet} 板', 'error')
                        return redirect(url_for('main.edit_inbound', id=id))
                    if package_count < total_outbound_package:
                        flash(f'修改失败：件数不能小于已出库数量 {total_outbound_package} 件', 'error')
                        return redirect(url_for('main.edit_inbound', id=id))

                # 如果识别编码变更了，需要创建新的库存记录
                if identification_code_changed:
                    # 如果有出库记录，识别编码变更会导致数据不一致，不允许修改
                    if outbound_records:
                        flash(f'修改失败：该货物已有出库记录，不能修改客户名称、车牌号或入库时间', 'error')
                        return redirect(url_for('main.edit_inbound', id=id))

                    # 检查新识别编码是否已存在库存记录
                    new_inventory = Inventory.query.filter_by(identification_code=new_identification_code).first()

                    if new_inventory:
                        # 如果新识别编码已存在库存记录，则更新它
                        new_inventory.inbound_pallet_count = pallet_count
                        new_inventory.inbound_package_count = package_count
                        # 由于识别编码变更，这是新的货物，没有出库记录
                        new_inventory.pallet_count = pallet_count
                        new_inventory.package_count = package_count
                        new_inventory.weight = weight
                        new_inventory.volume = volume
                        new_inventory.location = location
                        new_inventory.documents = documents
                        new_inventory.export_mode = export_mode
                        new_inventory.customs_broker = customs_broker
                        new_inventory.service_staff = service_staff
                    else:
                        # 如果新识别编码不存在库存记录，则创建
                        new_inventory = Inventory(
                            customer_name=customer_name,
                            identification_code=new_identification_code,
                            inbound_pallet_count=pallet_count,
                            inbound_package_count=package_count,
                            pallet_count=pallet_count,
                            package_count=package_count,
                            weight=weight,
                            volume=volume,
                            location=location,
                            documents=documents,
                            export_mode=export_mode,
                            customs_broker=customs_broker,
                            inbound_time=inbound_time_obj,
                            plate_number=plate_number,
                            service_staff=service_staff,
                            # 添加操作追踪信息
                            operated_by_user_id=current_user.id,
                            operated_warehouse_id=current_user.warehouse_id
                        )
                        db.session.add(new_inventory)

                    # 删除旧的库存记录
                    if old_inventory:
                        db.session.delete(old_inventory)
                else:
                    # 如果识别编码未变更，直接更新现有库存记录
                    if old_inventory:
                        old_inventory.customer_name = customer_name
                        # 更新入库数量
                        old_inventory.inbound_pallet_count = pallet_count
                        old_inventory.inbound_package_count = package_count
                        # 计算剩余库存 = 新入库数量 - 已出库数量
                        old_inventory.pallet_count = pallet_count - total_outbound_pallet
                        old_inventory.package_count = package_count - total_outbound_package
                        old_inventory.weight = weight
                        old_inventory.volume = volume
                        old_inventory.location = location
                        old_inventory.documents = documents
                        old_inventory.export_mode = export_mode
                        old_inventory.customs_broker = customs_broker
                        old_inventory.service_staff = service_staff
                    else:
                        # 如果旧的库存记录不存在，则创建
                        new_inventory = Inventory(
                            customer_name=customer_name,
                            identification_code=record.identification_code,
                            inbound_pallet_count=pallet_count,
                            inbound_package_count=package_count,
                            pallet_count=pallet_count,
                            package_count=package_count,
                            weight=weight,
                            volume=volume,
                            location=location,
                            documents=documents,
                            export_mode=export_mode,
                            customs_broker=customs_broker,
                            inbound_time=inbound_time_obj,
                            plate_number=plate_number,
                            service_staff=service_staff,
                            # 添加操作追踪信息
                            operated_by_user_id=current_user.id,
                            operated_warehouse_id=current_user.warehouse_id
                        )
                        db.session.add(new_inventory)

                db.session.commit()
                flash("入库记录更新成功", "success")
                return redirect(url_for('main.inbound_list'))
            except Exception as e:
                db.session.rollback()
                current_app.logger.error(f"更新入库记录时出错: {str(e)}")
                flash(f"更新入库记录时出错: {str(e)}", "danger")
                return redirect(url_for('main.edit_inbound', id=id))

        # GET请求，显示编辑表单
        form = InboundRecordEditForm2()
        form.inbound_time.data = record.outbound_time
        form.plate_number.data = record.plate_number
        form.customer_name.data = record.customer_name
        form.identification_code.data = record.identification_code
        form.pallet_count.data = record.pallet_count
        form.package_count.data = record.package_count
        form.weight.data = record.weight
        form.volume.data = record.volume
        form.export_mode.data = record.export_mode
        form.customs_broker.data = record.customs_broker
        form.location.data = record.location
        form.documents.data = record.documents
        form.service_staff.data = record.service_staff

        return render_template('inbound_edit.html', title='编辑入库记录', form=form, record=record)
    except Exception as e:
        current_app.logger.error(f"编辑入库记录时出错: {str(e)}")
        flash(f"编辑入库记录时出错: {str(e)}", "danger")
        return redirect(url_for('main.inbound_list'))

@bp.route('/inbound/delete/<int:id>', methods=['POST'])
@require_permission('INBOUND_DELETE')
@log_operation('inbound', 'delete', 'inbound_record')
def delete_inbound(id):
    """删除入库记录"""
    try:
        # 查询记录
        record = InboundRecord.query.get_or_404(id)

        # 获取记录信息，用于检查出库记录
        identification_code = record.identification_code
        customer_name = record.customer_name

        # 检查是否有对应的出库记录
        outbound_records = OutboundRecord.query.filter_by(identification_code=identification_code).all()

        if outbound_records:
            # 如果有出库记录，不允许删除
            flash(f'无法删除入库记录：客户 {customer_name} 的货物已有出库记录，请先处理相关出库记录', 'error')
            current_app.logger.warning(f"尝试删除有出库记录的入库记录: ID={id}, 识别编码={identification_code}")
            return redirect(url_for('main.inbound_list'))

        # 查找对应的库存记录
        inventory = Inventory.query.filter_by(identification_code=identification_code).first()

        # 删除库存记录
        if inventory:
            db.session.delete(inventory)
            current_app.logger.info(f"删除库存记录: {identification_code}")

        # 删除入库记录
        db.session.delete(record)
        db.session.commit()

        flash(f'已成功删除客户 {customer_name} 的入库记录及对应库存', 'success')
        current_app.logger.info(f"删除入库记录: ID={id}, 客户={customer_name}, 识别编码={identification_code}")

    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"删除入库记录时出错: {str(e)}")
        flash(f"删除入库记录时出错: {str(e)}", "danger")
    return redirect(url_for('main.inbound_list'))

@bp.route('/api/inbound/batch', methods=['POST'])
@csrf.exempt  # 豁免CSRF保护，因为这是API接口
@require_permission('INBOUND_CREATE')
@log_operation('inbound', 'batch_create', 'inbound_record')
def api_inbound_batch():
    """批量添加入库记录API"""
    if not request.is_json:
        return jsonify({'status': 'error', 'message': '请求内容必须是JSON格式'}), 400

    data = request.get_json()
    if not isinstance(data, list):
        return jsonify({'status': 'error', 'message': '数据必须是数组格式'}), 400

    # 处理每一项数据，去除所有字符串字段的空格
    cleaned_data = []
    for item in data:
        if isinstance(item, dict):
            # 清理字典中所有字符串值的空格
            cleaned_item = clean_dict_whitespace(item)
            cleaned_data.append(cleaned_item)
        else:
            cleaned_data.append(item)

    # 使用清理后的数据
    data = cleaned_data

    success_count = 0
    errors = []

    # 获取当前用户的仓库ID
    operated_warehouse_id = current_user.warehouse_id if current_user.warehouse_id else None

    # admin用户特殊处理：如果没有绑定仓库，需要从请求数据中确定仓库
    if not operated_warehouse_id:
        if current_user.is_super_admin():
            # admin用户可以根据业务逻辑自动匹配仓库
            # 这里可以根据具体需求来确定仓库ID，暂时允许继续执行
            current_app.logger.info(f"Admin用户 {current_user.username} 执行入库操作，将根据业务逻辑确定仓库")
        else:
            return jsonify({'status': 'error', 'message': '用户未绑定仓库，无法进行入库操作'}), 400

    # 按照客户名称和日期分组，用于生成识别编码序号
    date_customer_counts = {}

    # 预处理所有记录，准备生成序号
    for i, item in enumerate(data):
        try:
            # 验证和预处理入库时间
            if 'inbound_time' in item:
                if isinstance(item['inbound_time'], str):
                    try:
                        inbound_time = datetime.strptime(item['inbound_time'], '%Y-%m-%d %H:%M:%S')
                    except ValueError:
                        continue  # 日期格式错误，跳过
                elif isinstance(item['inbound_time'], datetime):
                    inbound_time = item['inbound_time']
                else:
                    continue  # 日期格式无效，跳过
            else:
                continue  # 缺少日期，跳过

            if 'customer_name' not in item or not item['customer_name']:
                continue  # 缺少客户名称，跳过

            if 'plate_number' not in item or not item['plate_number']:
                continue  # 缺少车牌号，跳过

            # 提取日期部分
            date_str = inbound_time.strftime('%Y-%m-%d')
            customer_name = item['customer_name']
            plate_number = item['plate_number']

            # 初始化计数器 - 基于日期+客户+车牌的组合
            date_customer_plate_key = f"{date_str}:{customer_name}:{plate_number}"
            if date_customer_plate_key not in date_customer_counts:
                # 查询数据库，获取当天该客户该车牌的现有记录数
                existing_count = InboundRecord.query.filter(
                    db.func.date(InboundRecord.inbound_time) == date_str,
                    InboundRecord.customer_name == customer_name,
                    InboundRecord.plate_number == plate_number
                ).count()
                date_customer_counts[date_customer_plate_key] = existing_count
        except:
            continue

    # 处理每条记录
    for i, item in enumerate(data):
        try:
            # 验证必填字段
            required_fields = ['inbound_time', 'plate_number', 'customer_name', 'service_staff', 'order_type', 'export_mode', 'customs_broker']
            missing_fields = []
            for field in required_fields:
                if field not in item or not item[field]:
                    missing_fields.append(field)

            if missing_fields:
                errors.append(f"第{i+1}条记录缺少必填字段: {', '.join(missing_fields)}")
                continue

            # 验证入库时间格式
            if isinstance(item['inbound_time'], str):
                try:
                    # 打印接收到的日期格式，用于调试
                    current_app.logger.info(f"批量导入第{i+1}条记录接收到的日期格式: {item['inbound_time']}")

                    # 尝试多种日期格式
                    formats_to_try = [
                        '%Y-%m-%d',        # 标准日期格式: 2025-06-27
                        '%Y/%m/%d',        # 斜杠分隔: 2025/06/27
                        '%Y.%m.%d',        # 点分隔: 2025.06.27
                        '%Y年%m月%d日',     # 中文格式: 2025年06月27日
                        '%Y-%m-%dT%H:%M',  # HTML5日期时间格式: 2025-06-27T12:30
                        '%Y-%m-%d %H:%M:%S' # 完整日期时间: 2025-06-27 12:30:45
                    ]

                    inbound_time = None
                    for date_format in formats_to_try:
                        try:
                            inbound_time = datetime.strptime(item['inbound_time'], date_format)
                            break  # 如果成功解析，跳出循环
                        except ValueError:
                            continue  # 尝试下一个格式

                    # 如果所有格式都失败
                    if inbound_time is None:
                        errors.append(f"第{i+1}条记录入库时间格式错误，请使用YYYY-MM-DD格式（如2025-06-27）")
                        continue

                    # 设置时间为00:00:00，只保留日期部分
                    inbound_time = inbound_time.replace(hour=0, minute=0, second=0)

                except Exception as e:
                    errors.append(f"第{i+1}条记录入库时间格式错误: {str(e)}")
                    continue
            elif isinstance(item['inbound_time'], datetime):
                inbound_time = item['inbound_time']
                # 设置时间为00:00:00，只保留日期部分
                inbound_time = inbound_time.replace(hour=0, minute=0, second=0)
            else:
                errors.append(f"第{i+1}条记录入库时间格式错误")
                continue

            # 验证件数和板数不能同时为空，且必须是整数
            try:
                pallet_value = item.get('pallet_count', 0) or 0
                package_value = item.get('package_count', 0) or 0

                # 检查是否为小数
                if isinstance(pallet_value, (int, float)) and pallet_value != int(pallet_value):
                    errors.append(f"第{i+1}条记录错误: 板数必须是整数，不能是小数")
                    continue
                if isinstance(package_value, (int, float)) and package_value != int(package_value):
                    errors.append(f"第{i+1}条记录错误: 件数必须是整数，不能是小数")
                    continue

                pallet_count = int(pallet_value)
                package_count = int(package_value)

                if pallet_count < 0 or package_count < 0:
                    errors.append(f"第{i+1}条记录错误: 板数和件数不能为负数")
                    continue

                if pallet_count == 0 and package_count == 0:
                    errors.append(f"第{i+1}条记录错误: 板数和件数不能同时为零")
                    continue
            except (ValueError, TypeError):
                errors.append(f"第{i+1}条记录错误: 板数和件数必须是有效的整数")
                continue

            # 处理重量和体积为空的情况
            try:
                weight = float(item.get('weight', 0) or 0)
                volume = float(item.get('volume', 0) or 0)
                if weight < 0 or volume < 0:
                    errors.append(f"第{i+1}条记录错误: 重量和体积必须为非负数")
                    continue
            except (ValueError, TypeError):
                errors.append(f"第{i+1}条记录错误: 重量和体积必须是有效数字")
                continue

            # 生成识别编码 - 使用新规则：仓库前缀/客户全称/车牌/日期/序号
            customer_name = item['customer_name']
            plate_number = item['plate_number']

            # 确定实际的操作仓库ID
            actual_warehouse_id = operated_warehouse_id or get_admin_warehouse_id('frontend')
            if not actual_warehouse_id:
                current_app.logger.error(f"无法确定操作仓库ID，用户: {current_user.username}")
                continue

            # 使用IdentificationCodeGenerator生成标准格式的识别编码
            identification_code = IdentificationCodeGenerator.generate_unique_identification_code(
                warehouse_id=actual_warehouse_id,
                customer_name=customer_name,
                plate_number=plate_number,
                operation_type='inbound',
                inbound_date=inbound_time
            )

            # 创建入库记录
            record = InboundRecord(
                inbound_time=inbound_time,
                delivery_plate_number=item.get('delivery_plate_number', ''),
                plate_number=plate_number,
                customer_name=customer_name,
                identification_code=identification_code,
                pallet_count=pallet_count,
                package_count=package_count,
                weight=weight,
                volume=volume,
                export_mode=item.get('export_mode', ''),
                order_type=item.get('order_type', ''),
                customs_broker=item.get('customs_broker', ''),
                location=item.get('location', ''),
                documents=item.get('documents', ''),
                service_staff=item.get('service_staff', ''),
                record_type='direct',  # 直接入库记录
                # 添加操作追踪信息
                operated_by_user_id=current_user.id,
                operated_warehouse_id=actual_warehouse_id or current_user.warehouse_id
            )

            db.session.add(record)
            success_count += 1

            # 创建或更新库存记录 - 按照识别编码管理库存
            inventory = Inventory.query.filter_by(identification_code=identification_code).first()

            # 如果该识别编码的库存记录不存在，则创建新记录
            if not inventory:
                inventory = Inventory(
                    customer_name=customer_name,
                    identification_code=identification_code,
                    inbound_pallet_count=pallet_count,
                    inbound_package_count=package_count,
                    pallet_count=pallet_count,
                    package_count=package_count,
                    weight=weight,
                    volume=volume,
                    location=item.get('location', ''),  # 使用提供的库位
                    documents=item.get('documents', ''),
                    export_mode=item.get('export_mode', ''),
                    order_type=item.get('order_type', ''),
                    customs_broker=item.get('customs_broker', ''),
                    inbound_time=inbound_time,
                    plate_number=plate_number,
                    service_staff=item.get('service_staff', ''),
                    # 添加操作追踪信息
                    operated_by_user_id=current_user.id,
                    operated_warehouse_id=actual_warehouse_id or current_user.warehouse_id
                )
                db.session.add(inventory)
            else:
                # 如果已存在，则更新库位等信息
                if item.get('location'):
                    inventory.location = item.get('location', '')
                if item.get('documents'):
                    inventory.documents = item.get('documents', '')
                if item.get('export_mode'):
                    inventory.export_mode = item.get('export_mode', '')
                if item.get('customs_broker'):
                    inventory.customs_broker = item.get('customs_broker', '')
                if item.get('service_staff'):
                    inventory.service_staff = item.get('service_staff', '')

        except Exception as e:
            errors.append(f"第{i+1}条记录处理错误: {str(e)}")

    if success_count > 0:
        try:
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            errors.append(f"提交数据库时出错: {str(e)}")
            return jsonify({
                'status': 'error',
                'message': f'提交数据库时出错',
                'errors': errors
            }), 500

    return jsonify({
        'status': 'success' if not errors else 'partial_success',
        'message': f'成功添加 {success_count} 条记录',
        'success_count': success_count,
        'total_count': len(data),
        'errors': errors
    })

@bp.route('/outbound', methods=['GET', 'POST'])
@require_permission('OUTBOUND_VIEW')
def outbound():
    """出库操作页面 - 根据用户仓库类型重定向"""
    # 检查用户仓库类型并重定向到对应页面
    if current_user.is_authenticated and current_user.warehouse:
        warehouse_type = current_user.warehouse.warehouse_type
        if warehouse_type == 'frontend':
            # 前端仓用户重定向到前端仓出库页面
            return redirect(url_for('main.frontend_outbound'))
        elif warehouse_type == 'backend':
            # 后端仓用户重定向到后端仓出库页面
            return redirect(url_for('main.backend_outbound'))

    # 默认情况下显示通用出库页面
    return render_template('outbound_batch.html', title='出库操作')

@bp.route('/outbound/list')
@require_permission('OUTBOUND_VIEW')
def outbound_list():
    """出库记录列表"""
    # 获取搜索参数和分页参数
    page = request.args.get('page', 1, type=int)
    per_page = current_app.config.get('ITEMS_PER_PAGE', 50)

    customer_name = request.args.get('customer_name', '')
    plate_number = request.args.get('plate_number', '')
    date_start = request.args.get('date_start', '')
    date_end = request.args.get('date_end', '')
    destination = request.args.get('destination', '')
    service_staff = request.args.get('service_staff', '')
    inbound_plate = request.args.get('inbound_plate', '')
    order_type = request.args.get('order_type', '')
    export_mode = request.args.get('export_mode', '')
    document_no = request.args.get('document_no', '')
    location = request.args.get('location', '')
    customs_broker = request.args.get('customs_broker', '')
    # 新增操作追踪筛选参数
    operated_warehouse_id = request.args.get('operated_warehouse_id', '')
    operated_user_id = request.args.get('operated_user_id', '')

    # 如果没有指定日期范围，默认使用最近一周的日期范围
    if not date_start and not date_end:
        today = datetime.now().strftime('%Y-%m-%d')
        one_week_ago = (datetime.now() - timedelta(days=7)).strftime('%Y-%m-%d')
        date_start = one_week_ago
        date_end = today

    # 添加调试代码：直接查询数据库中的一些记录，检查备注字段
    try:
        recent_records = OutboundRecord.query.order_by(OutboundRecord.id.desc()).limit(5).all()
        current_app.logger.info("===== 调试备注字段 =====")
        for rec in recent_records:
            current_app.logger.info(f"记录ID: {rec.id}, 客户: {rec.customer_name}")
            current_app.logger.info(f"备注字段: remarks='{rec.remarks}', remark1='{rec.remark1}', remark2='{rec.remark2}'")
            current_app.logger.info(f"备注字段类型: remarks={type(rec.remarks)}, remark1={type(rec.remark1)}, remark2={type(rec.remark2)}")
    except Exception as e:
        current_app.logger.error(f"调试备注字段时出错: {str(e)}")

    # 确保搜索参数传递给模板
    search_params = {
        'customer_name': customer_name,
        'plate_number': plate_number,
        'date_start': date_start,
        'date_end': date_end,
        'destination': destination,
        'service_staff': service_staff,
        'inbound_plate': inbound_plate,
        'order_type': order_type,
        'export_mode': export_mode,
        'document_no': document_no,
        'location': location,
        'customs_broker': customs_broker,
        'operated_warehouse_id': operated_warehouse_id,
        'operated_user_id': operated_user_id
    }

    try:
        # 查询构建
        query = OutboundRecord.query.options(
            db.joinedload(OutboundRecord.operated_warehouse),
            db.joinedload(OutboundRecord.operated_by_user)
        )

        # 按日期范围筛选
        if date_start:
            try:
                start_date = datetime.strptime(date_start, '%Y-%m-%d')
                query = query.filter(OutboundRecord.outbound_time >= start_date)
            except ValueError:
                flash('开始日期格式无效', 'warning')

        if date_end:
            try:
                # 将结束日期设置为当天的23:59:59
                end_date = datetime.strptime(date_end, '%Y-%m-%d')
                end_date = end_date.replace(hour=23, minute=59, second=59)
                query = query.filter(OutboundRecord.outbound_time <= end_date)
            except ValueError:
                flash('结束日期格式无效', 'warning')

        # 按车牌号筛选
        if plate_number:
            query = query.filter(OutboundRecord.plate_number.like(f'%{plate_number}%'))

        # 按客户名称筛选
        if customer_name:
            query = query.filter(OutboundRecord.customer_name.like(f'%{customer_name}%'))

        # 按目的地筛选
        if destination:
            query = query.filter(OutboundRecord.destination.like(f'%{destination}%'))

        # 按跟单客服筛选
        if service_staff:
            query = query.filter(OutboundRecord.service_staff.like(f'%{service_staff}%'))

        # 新增字段筛选
        if inbound_plate:
            query = query.filter(OutboundRecord.inbound_plate.like(f'%{inbound_plate}%'))

        if order_type:
            query = query.filter(OutboundRecord.order_type == order_type)

        if export_mode:
            query = query.filter(OutboundRecord.export_mode == export_mode)

        if document_no:
            query = query.filter(OutboundRecord.document_no.like(f'%{document_no}%'))

        if location:
            query = query.filter(OutboundRecord.location.like(f'%{location}%'))

        if customs_broker:
            query = query.filter(OutboundRecord.customs_broker.like(f'%{customs_broker}%'))

        # 按操作仓库筛选
        if operated_warehouse_id:
            query = query.filter(OutboundRecord.operated_warehouse_id == operated_warehouse_id)

        # 按操作用户筛选
        if operated_user_id:
            query = query.filter(OutboundRecord.operated_by_user_id == operated_user_id)

        # 按出库时间降序排序
        query = query.order_by(OutboundRecord.outbound_time.desc())

        # 获取总记录数
        total_count = query.count()

        # 分页
        items = query.offset((page - 1) * per_page).limit(per_page).all()

        # 使用自定义分页类
        records = SimplePagination(
            items=items,
            page=page,
            per_page=per_page,
            total=total_count
        )

        # 获取仓库和用户选项数据
        from app.models import Warehouse, User
        warehouses = Warehouse.query.filter_by(status='active').all()
        users = User.query.filter_by(status='active').all()

        return render_template(
            'outbound_list.html',
            title='出库记录',
            records=records,
            search_params=search_params,
            warehouses=warehouses,
            users=users
        )
    except Exception as e:
        current_app.logger.error(f"Error in outbound_list: {str(e)}")
        flash(f"加载数据时出错: {str(e)}", "danger")

        # 在异常情况下也传递search_params
        empty_pagination = SimplePagination(items=[], page=1, per_page=50, total=0)

        return render_template(
            'outbound_list.html',
            title='出库记录',
            records=empty_pagination,
            search_params=search_params
        )

@bp.route('/export_outbound')
def export_outbound():
    """导出出库记录"""
    # 获取搜索参数
    date_start = request.args.get('date_start', '')
    date_end = request.args.get('date_end', '')

    # 如果没有指定日期范围，默认导出最近一周的数据
    if not date_start and not date_end:
        today = datetime.now().strftime('%Y-%m-%d')
        week_ago = (datetime.now() - timedelta(days=7)).strftime('%Y-%m-%d')
        date_start = week_ago
        date_end = today

    plate_number = request.args.get('plate_number', '')
    customer_name = request.args.get('customer_name', '')
    destination = request.args.get('destination', '')
    service_staff = request.args.get('service_staff', '')

    # 新增字段
    inbound_plate = request.args.get('inbound_plate', '')
    order_type = request.args.get('order_type', '')
    export_mode = request.args.get('export_mode', '')
    document_no = request.args.get('document_no', '')
    location = request.args.get('location', '')
    customs_broker = request.args.get('customs_broker', '')

    # 构建查询
    query = OutboundRecord.query

    # 按日期范围筛选
    if date_start:
        try:
            start_date = datetime.strptime(date_start, '%Y-%m-%d')
            query = query.filter(OutboundRecord.outbound_time >= start_date)
        except ValueError:
            flash('开始日期格式无效', 'warning')

    if date_end:
        try:
            # 将结束日期设置为当天的23:59:59
            end_date = datetime.strptime(date_end, '%Y-%m-%d')
            end_date = end_date.replace(hour=23, minute=59, second=59)
            query = query.filter(OutboundRecord.outbound_time <= end_date)
        except ValueError:
            flash('结束日期格式无效', 'warning')

    # 按车牌号筛选
    if plate_number:
        query = query.filter(OutboundRecord.plate_number.like(f'%{plate_number}%'))

    # 按客户名称筛选
    if customer_name:
        query = query.filter(OutboundRecord.customer_name.like(f'%{customer_name}%'))

    # 按目的地筛选
    if destination:
        query = query.filter(OutboundRecord.destination.like(f'%{destination}%'))

    # 按跟单客服筛选
    if service_staff:
        query = query.filter(OutboundRecord.service_staff.like(f'%{service_staff}%'))

    # 新增字段筛选
    if inbound_plate:
        query = query.filter(OutboundRecord.inbound_plate.like(f'%{inbound_plate}%'))

    if order_type:
        query = query.filter(OutboundRecord.order_type == order_type)

    if export_mode:
        query = query.filter(OutboundRecord.export_mode == export_mode)

    if document_no:
        query = query.filter(OutboundRecord.document_no.like(f'%{document_no}%'))

    if location:
        query = query.filter(OutboundRecord.location.like(f'%{location}%'))

    if customs_broker:
        query = query.filter(OutboundRecord.customs_broker.like(f'%{customs_broker}%'))

    # 按出库时间降序排序
    records = query.order_by(OutboundRecord.outbound_time.desc()).all()

    # 创建DataFrame
    data = []
    for record in records:
        data.append({
            '出库时间': record.outbound_time.strftime('%Y-%m-%d') if record.outbound_time else '',
            '出库车牌': record.plate_number,
            '客户名称': record.customer_name,
            '入库车牌': record.inbound_plate or '',
            '订单类型': record.order_type or '',
            '板数': record.pallet_count,
            '件数': record.package_count,
            '重量(kg)': record.weight,
            '体积(m³)': record.volume,
            '出境模式': record.export_mode or '',
            '报关行': record.customs_broker or '',
            '库位': record.location or '',
            '单据': record.document_no or '',
            '目的地': record.destination or '',
            '跟单客服': record.service_staff or '',
            '创建时间': record.outbound_time.strftime('%Y-%m-%d %H:%M:%S')
        })

    df = pd.DataFrame(data)

    # 创建临时文件
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
        # 写入Excel
        df.to_excel(tmp.name, index=False, engine='openpyxl')
        tmp_name = tmp.name

    # 设置文件名
    now = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f"出库记录_{now}.xlsx"

    return send_file(
        tmp_name,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@bp.route('/outbound/view/<int:id>')
@require_permission('OUTBOUND_VIEW')
def view_outbound(id):
    """查看出库记录详情"""
    record = OutboundRecord.query.get_or_404(id)
    return render_template('outbound_view.html', title='出库记录详情', record=record)

@bp.route('/outbound/edit/<int:id>', methods=['GET', 'POST'])
@require_permission('OUTBOUND_EDIT')
@log_operation('outbound', 'edit', 'outbound_record')
def edit_outbound(id):
    """编辑出库记录"""
    record = OutboundRecord.query.get_or_404(id)
    form = OutboundRecordForm(obj=record)

    # 合并现有的备注字段到remarks字段中
    if not form.remarks.data:
        combined_remarks = []
        if record.remarks:
            combined_remarks.append(record.remarks)
        if record.remark1:
            combined_remarks.append(record.remark1)
        if record.remark2:
            combined_remarks.append(record.remark2)
        if combined_remarks:
            form.remarks.data = '; '.join(combined_remarks)

    if form.validate_on_submit():
        # 保存原始数据，用于后续更新库存
        original_identification_code = record.identification_code
        original_pallet_count = record.pallet_count
        original_package_count = record.package_count

        # 更新记录
        form.populate_obj(record)

        # 清空旧的备注字段，只使用新的remarks字段
        record.remark1 = ''
        record.remark2 = ''

        # 如果识别编码或数量发生变化，需要更新库存
        if (original_identification_code == record.identification_code and
            (original_pallet_count != record.pallet_count or original_package_count != record.package_count)):
            # 识别编码相同，但数量变化 - 需要调整库存
            inventory = Inventory.query.filter_by(identification_code=record.identification_code).first()
            if inventory:
                # 计算差值
                pallet_diff = original_pallet_count - record.pallet_count  # 正值表示减少出库数量，负值表示增加出库数量
                package_diff = original_package_count - record.package_count

                # 更新库存
                inventory.pallet_count = max(0, inventory.pallet_count + pallet_diff)
                inventory.package_count = max(0, inventory.package_count + package_diff)

                current_app.logger.info(f"编辑出库记录时更新库存: {record.identification_code}, "
                                       f"板数调整: {pallet_diff}, 新库存: {inventory.pallet_count}, "
                                       f"件数调整: {package_diff}, 新库存: {inventory.package_count}")

                # 清除库存缓存
                cache_invalidation.on_inventory_change(
                    warehouse_id=record.operated_warehouse_id
                )
                current_app.logger.info("已清除库存相关缓存")

        db.session.commit()

        # 如果更新了库存，刷新库存对象确保状态是最新的
        if (original_identification_code == record.identification_code and
            (original_pallet_count != record.pallet_count or original_package_count != record.package_count)):
            inventory = Inventory.query.filter_by(identification_code=record.identification_code).first()
            if inventory:
                db.session.refresh(inventory)
                current_app.logger.info(f"刷新后的库存状态: {inventory.identification_code}, 板数: {inventory.pallet_count}, 件数: {inventory.package_count}")

        flash('出库记录已更新', 'success')
        return redirect(url_for('main.view_outbound', id=record.id))

    return render_template('outbound_edit.html', title='编辑出库记录', form=form)

@bp.route('/outbound/delete/<int:id>', methods=['POST'])
@require_permission('OUTBOUND_DELETE')
@log_operation('outbound', 'delete', 'outbound_record')
def delete_outbound(id):
    """删除出库记录"""
    try:
        record = OutboundRecord.query.get_or_404(id)

        # 恢复库存 - 修复：必须指定仓库ID避免回滚到错误仓库
        if record.identification_code and record.operated_warehouse_id:
            inventory = Inventory.query.filter_by(
                identification_code=record.identification_code,
                operated_warehouse_id=record.operated_warehouse_id
            ).first()
            if inventory:
                # 将出库的数量加回库存
                inventory.pallet_count += record.pallet_count
                inventory.package_count += record.package_count

                # 对于PX开头的自行入库数据，需要恢复业务字段
                if record.identification_code and record.identification_code.startswith('PX/'):
                    # 查找原始入库记录来恢复业务字段
                    inbound_record = InboundRecord.query.filter_by(
                        identification_code=record.identification_code
                    ).first()

                    if inbound_record:
                        inventory.order_type = inbound_record.order_type
                        inventory.export_mode = inbound_record.export_mode
                        inventory.customs_broker = inbound_record.customs_broker
                        inventory.documents = inbound_record.documents

                        current_app.logger.info(f"恢复PX自行入库数据的业务字段: {record.identification_code}, "
                                               f"订单类型: {inbound_record.order_type}, "
                                               f"出境模式: {inbound_record.export_mode}, "
                                               f"报关行: {inbound_record.customs_broker}")

                current_app.logger.info(f"删除出库记录时恢复库存: {record.identification_code} (仓库ID: {record.operated_warehouse_id}), "
                                       f"板数恢复: +{record.pallet_count}, 新库存: {inventory.pallet_count}, "
                                       f"件数恢复: +{record.package_count}, 新库存: {inventory.package_count}")
            else:
                current_app.logger.warning(f"删除出库记录时未找到对应库存: {record.identification_code} (仓库ID: {record.operated_warehouse_id})")

                # 清除库存缓存
                cache_invalidation.on_inventory_change(
                    warehouse_id=record.operated_warehouse_id
                )
                current_app.logger.info("已清除库存相关缓存")

        # 处理对应的在途货物记录回退
        if record.identification_code and record.batch_no:
            # 查找对应的在途货物记录
            transit_cargos = TransitCargo.query.filter_by(
                identification_code=record.identification_code,
                batch_no=record.batch_no,
                status='in_transit'
            ).all()

            if transit_cargos:
                for transit_cargo in transit_cargos:
                    current_app.logger.info(f"删除出库记录时同时删除在途货物记录: {transit_cargo.identification_code}, 批次: {transit_cargo.batch_no}")
                    db.session.delete(transit_cargo)

                current_app.logger.info(f"共删除 {len(transit_cargos)} 条在途货物记录")

        # 删除记录
        db.session.delete(record)
        db.session.commit()

        # 如果更新了库存，刷新库存对象确保状态是最新的
        if record.identification_code:
            inventory = Inventory.query.filter_by(identification_code=record.identification_code).first()
            if inventory:
                db.session.refresh(inventory)
                current_app.logger.info(f"刷新后的库存状态: {inventory.identification_code}, 板数: {inventory.pallet_count}, 件数: {inventory.package_count}")

        flash('出库记录已删除', 'success')
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"删除出库记录时出错: {str(e)}")
        flash(f"删除出库记录时出错: {str(e)}", "danger")

    return redirect(url_for('main.outbound_list'))


# ==================== 性能监控和缓存管理API ====================

@bp.route('/api/performance/stats')
@csrf.exempt
@require_permission('ADMIN_VIEW')
def api_performance_stats():
    """获取性能统计API"""
    try:
        stats = perf_dashboard.get_performance_summary()
        return jsonify({
            'success': True,
            'data': stats
        })
    except Exception as e:
        current_app.logger.error(f"获取性能统计失败: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'获取性能统计失败: {str(e)}'
        }), 500


@bp.route('/api/cache/stats')
@csrf.exempt
@require_permission('ADMIN_VIEW')
def api_cache_stats():
    """获取缓存统计API"""
    try:
        from app.cache_config import get_cache_manager
        from app.cache_invalidation import cache_health

        cache_stats = get_cache_manager().get_cache_stats()
        health_status = cache_health.check_cache_health()

        return jsonify({
            'success': True,
            'data': {
                'redis_stats': cache_stats,
                'health_status': health_status,
                'app_metrics': {
                    'cache_hits': performance_metrics.cache_hits,
                    'cache_misses': performance_metrics.cache_misses,
                    'hit_rate': performance_metrics.get_cache_hit_rate()
                }
            }
        })
    except Exception as e:
        current_app.logger.error(f"获取缓存统计失败: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'获取缓存统计失败: {str(e)}'
        }), 500


@bp.route('/api/cache/clear', methods=['POST'])
@csrf.exempt
@require_permission('ADMIN_EDIT')
def api_clear_cache():
    """清除缓存API"""
    try:
        data = request.get_json() or {}
        cache_type = data.get('cache_type', 'all')

        if cache_type == 'all':
            deleted_count = cache_invalidation.clear_all_cache()
            message = f'已清除所有缓存，共 {deleted_count} 个键'
        elif cache_type == 'inventory':
            deleted_count = cache_invalidation.on_inventory_change()
            message = f'已清除库存缓存，共 {deleted_count} 个键'
        elif cache_type == 'user':
            user_id = data.get('user_id')
            if user_id:
                deleted_count = cache_invalidation.on_user_change(user_id)
                message = f'已清除用户 {user_id} 的缓存，共 {deleted_count} 个键'
            else:
                return jsonify({
                    'success': False,
                    'message': '清除用户缓存需要提供 user_id'
                }), 400
        else:
            return jsonify({
                'success': False,
                'message': f'不支持的缓存类型: {cache_type}'
            }), 400

        return jsonify({
            'success': True,
            'message': message
        })

    except Exception as e:
        current_app.logger.error(f"清除缓存失败: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'清除缓存失败: {str(e)}'
        }), 500


@bp.route('/api/cache/warmup', methods=['POST'])
@csrf.exempt
@require_permission('ADMIN_EDIT')
def api_warmup_cache():
    """缓存预热API"""
    try:
        data = request.get_json() or {}
        warmup_type = data.get('warmup_type', 'all')

        if warmup_type == 'all':
            cache_warmup.warmup_all_cache()
            message = '已完成所有缓存预热'
        elif warmup_type == 'inventory':
            cache_warmup.warmup_inventory_cache()
            message = '已完成库存缓存预热'
        elif warmup_type == 'basic':
            cache_warmup.warmup_basic_data_cache()
            message = '已完成基础数据缓存预热'
        else:
            return jsonify({
                'success': False,
                'message': f'不支持的预热类型: {warmup_type}'
            }), 400

        return jsonify({
            'success': True,
            'message': message
        })

    except Exception as e:
        current_app.logger.error(f"缓存预热失败: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'缓存预热失败: {str(e)}'
        }), 500


@bp.route('/api/database/optimize', methods=['POST'])
@csrf.exempt
@require_permission('ADMIN_EDIT')
def api_optimize_database():
    """数据库优化API"""
    try:
        data = request.get_json() or {}
        operation = data.get('operation', 'create_indexes')

        if operation == 'create_indexes':
            DatabaseOptimizer.create_indexes()
            message = '数据库索引创建完成'
        elif operation == 'analyze_tables':
            DatabaseOptimizer.analyze_tables()
            message = '数据库表分析完成'
        else:
            return jsonify({
                'success': False,
                'message': f'不支持的操作: {operation}'
            }), 400

        return jsonify({
            'success': True,
            'message': message
        })

    except Exception as e:
        current_app.logger.error(f"数据库优化失败: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'数据库优化失败: {str(e)}'
        }), 500

@bp.route('/api/outbound/batch', methods=['POST'])
@csrf.exempt  # 豁免CSRF保护，因为这是API接口
@require_permission('OUTBOUND_CREATE')
@log_operation('outbound', 'batch_create', 'outbound_record')
def api_outbound_batch():
    """处理批量出库数据提交"""
    if not request.is_json:
        return jsonify({'success': False, 'message': '请求必须是JSON格式'}), 400

    try:
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'message': '请求数据为空'}), 400

        items = data.get('items', [])
        if not items:
            return jsonify({'success': False, 'message': '未提供出库记录数据'}), 400

        # 添加调试日志
        current_app.logger.info("备注字段调试信息:")
        for idx, item in enumerate(items):
            current_app.logger.info(f"第{idx+1}条记录 - 客户: {item.get('customer_name', 'NA')}, 识别码: {item.get('identification_code', 'NA')}")
            current_app.logger.info(f"备注字段: remarks={item.get('remarks', 'NA')}, remark1={item.get('remark1', 'NA')}, remark2={item.get('remark2', 'NA')}")

        # 获取当前用户的仓库ID
        current_warehouse_id = current_user.warehouse_id if hasattr(current_user, 'warehouse_id') else None
        if not current_warehouse_id:
            if current_user.is_super_admin():
                # admin用户可以根据业务逻辑自动匹配仓库
                current_app.logger.info(f"Admin用户 {current_user.username} 执行出库操作，将根据业务逻辑确定仓库")
            else:
                return jsonify({'success': False, 'message': '无法确定当前仓库，请联系管理员'}), 400

        # 确定实际的仓库ID
        actual_warehouse_id = current_warehouse_id or get_admin_warehouse_id('frontend')

        # 使用统一的批次号生成器
        from app.utils.batch_generator import generate_batch_number
        batch_number = generate_batch_number(
            warehouse_id=actual_warehouse_id,
            destination_prefix=None,
            db_session=db.session
        )
        current_app.logger.info(f'批量出库生成批次号: {batch_number}，仓库ID: {current_warehouse_id}')

        # 设置共同的车牌号
        plate_number = data.get('common', {}).get('outbound_plate', '')
        if not plate_number:
            return jsonify({'success': False, 'message': '出库车牌不能为空'}), 400

        # 准备保存记录
        records_to_save = []
        total_items = len(items)

        # 辅助函数：解析日期时间字符串
        def parse_datetime(date_str):
            if not date_str:
                return None

            # 尝试多种日期格式
            formats_to_try = [
                '%Y-%m-%d %H:%M:%S',  # 标准日期时间格式: 2025-06-30 12:30:45
                '%Y-%m-%d',           # 仅日期格式: 2025-06-30
                '%Y/%m/%d',           # 斜杠分隔: 2025/06/30
                '%Y-%m-%dT%H:%M',     # HTML5日期时间格式: 2025-06-30T12:30
                '%Y-%m-%dT%H:%M:%S',  # ISO格式: 2025-06-30T12:30:45
                '%Y.%m.%d'            # 点分隔: 2025.06.30
            ]

            for fmt in formats_to_try:
                try:
                    return datetime.strptime(date_str, fmt)
                except ValueError:
                    continue

            # 如果所有格式都失败，抛出异常
            raise ValueError(f"无法解析日期时间: {date_str}")

        # 从common中获取公共字段
        common = data.get('common', {})

        # 解析日期时间字段
        arrival_time = parse_datetime(common.get('arrive_time')) if common.get('arrive_time') else None
        loading_start_time = parse_datetime(common.get('loading_start_time')) if common.get('loading_start_time') else None
        loading_end_time = parse_datetime(common.get('loading_end_time')) if common.get('loading_end_time') else None
        departure_time = parse_datetime(common.get('departure_time')) if common.get('departure_time') else None

        # 获取其他公共字段
        vehicle_type = common.get('vehicle_type', '')
        driver_name = common.get('driver_name', '')
        driver_phone = common.get('driver_phone', '')
        destination = common.get('destination', '')
        detailed_address = common.get('destination_address', '')
        contact_window = common.get('contact', '')
        receiver_id = common.get('receiver_id')
        large_layer = common.get('large_layer', 0)
        small_layer = common.get('small_layer', 0)
        pallet_board = common.get('pallet', 0)

        # 根据receiver_id获取目的仓库ID
        destination_warehouse_id = None
        if receiver_id:
            try:
                receiver = Receiver.query.get(receiver_id)
                if receiver and receiver.warehouse_name:
                    # 根据仓库名称查找对应的仓库ID
                    warehouse = Warehouse.query.filter_by(warehouse_name=receiver.warehouse_name).first()
                    if warehouse:
                        destination_warehouse_id = warehouse.id
            except Exception as e:
                current_app.logger.warning(f"获取目的仓库ID失败: {str(e)}")

        # 记录更新的库存
        updated_inventories = []

        # 先查询所有需要更新的库存记录，以减少数据库查询次数
        identification_codes = [item.get('identification_code') for item in items if item.get('identification_code')]
        inventory_dict = {}
        if identification_codes:
            inventories = Inventory.query.filter(Inventory.identification_code.in_(identification_codes)).all()
            for inv in inventories:
                inventory_dict[inv.identification_code] = inv

        for index, item in enumerate(items, 1):
            # 设置批次序号和总数
            item['batch_no'] = new_batch_no
            item['batch_sequence'] = index
            item['batch_total'] = total_items

            # 处理出库时间
            outbound_time = None
            if item.get('outbound_time'):
                outbound_time = parse_datetime(item['outbound_time'])
            else:
                outbound_time = now

            # 确保有出库车牌
            if not item.get('plate_number') and plate_number:
                item['plate_number'] = plate_number

            # 修改层板数量逻辑 - 只在第一条记录中保存层板数量，其他记录设为0
            # 注意：enumerate(items, 1) 从1开始，所以第一条记录的index是1
            current_large_layer = large_layer if index == 1 else 0
            current_small_layer = small_layer if index == 1 else 0
            current_pallet_board = pallet_board if index == 1 else 0

            # 获取入库日期
            inbound_date = None
            identification_code = item.get('identification_code')
            if identification_code:
                # 首先尝试从库存记录获取入库日期
                inventory = inventory_dict.get(identification_code)
                if inventory and inventory.inbound_time:
                    inbound_date = inventory.inbound_time
                else:
                    # 如果库存记录没有入库时间，尝试从入库记录获取
                    inbound_record = InboundRecord.query.filter_by(identification_code=identification_code).first()
                    if inbound_record and inbound_record.inbound_time:
                        inbound_date = inbound_record.inbound_time

            # 创建记录对象
            record = OutboundRecord(
                outbound_time=outbound_time,
                plate_number=item.get('plate_number', ''),
                customer_name=item.get('customer_name', ''),
                inbound_plate=item.get('inbound_plate', ''),
                identification_code=item.get('identification_code', ''),
                order_type=item.get('order_type', ''),
                pallet_count=item.get('pallet_count', 0),
                package_count=item.get('package_count', 0),
                weight=item.get('weight'),
                volume=item.get('volume'),
                export_mode=item.get('export_mode', ''),
                customs_broker=item.get('customs_broker', ''),
                location=item.get('location', ''),
                document_no=item.get('document_no', ''),
                document_count=item.get('document_count'),  # 只使用前端提供的单据份数，不设默认值
                documents=item.get('documents'),  # 只使用前端提供的单据信息，不设默认值
                service_staff=item.get('service_staff', ''),
                destination=destination,
                destination_warehouse_id=destination_warehouse_id,
                detailed_address=detailed_address,
                contact_window=contact_window,
                large_layer=current_large_layer,
                small_layer=current_small_layer,
                pallet_board=current_pallet_board,
                batch_no=new_batch_no,
                batch_sequence=index,
                batch_total=total_items,
                inbound_date=inbound_date,  # 添加入库日期
                # 备注字段
                remarks=item.get('remarks', '') or item.get('remark1', '') or '',
                remark1=item.get('remark1') if item.get('remark1') is not None else '',
                remark2=item.get('remark2') if item.get('remark2') is not None else '',
                # 新增字段
                vehicle_type=vehicle_type,
                driver_name=driver_name,
                driver_phone=driver_phone,
                arrival_time=arrival_time,
                loading_start_time=loading_start_time,
                loading_end_time=loading_end_time,
                departure_time=departure_time,
                receiver_id=receiver_id,
                # 添加操作追踪信息
                operated_by_user_id=current_user.id,
                operated_warehouse_id=actual_warehouse_id
            )
            records_to_save.append(record)

            # 更新库存 - 根据识别编码查找并减少库存
            identification_code = item.get('identification_code')
            if identification_code:
                inventory = inventory_dict.get(identification_code)
                if inventory:
                    # 获取出库数量，确保是整数
                    try:
                        pallet_value = item.get('pallet_count', 0) or 0
                        package_value = item.get('package_count', 0) or 0

                        # 检查是否为小数
                        if isinstance(pallet_value, (int, float)) and pallet_value != int(pallet_value):
                            current_app.logger.error(f"板数必须是整数: {pallet_value}")
                            continue
                        if isinstance(package_value, (int, float)) and package_value != int(package_value):
                            current_app.logger.error(f"件数必须是整数: {package_value}")
                            continue

                        outbound_pallet_count = int(pallet_value)
                        outbound_package_count = int(package_value)

                        if outbound_pallet_count < 0 or outbound_package_count < 0:
                            current_app.logger.error(f"板数和件数不能为负数: {outbound_pallet_count}, {outbound_package_count}")
                            continue

                    except (ValueError, TypeError):
                        current_app.logger.error(f"板数和件数必须是有效的整数: {item.get('pallet_count')}, {item.get('package_count')}")
                        continue

                    # 记录更新前的库存
                    before_pallet = inventory.pallet_count
                    before_package = inventory.package_count

                    # 更新库存数量
                    inventory.pallet_count = max(0, inventory.pallet_count - outbound_pallet_count)
                    inventory.package_count = max(0, inventory.package_count - outbound_package_count)

                    # 记录已更新的库存
                    updated_inventories.append({
                        'id': inventory.id,
                        'identification_code': inventory.identification_code,
                        'customer_name': inventory.customer_name,
                        'before_pallet': before_pallet,
                        'before_package': before_package,
                        'after_pallet': inventory.pallet_count,
                        'after_package': inventory.package_count,
                        'outbound_pallet': outbound_pallet_count,
                        'outbound_package': outbound_package_count
                    })

                    current_app.logger.info(f"更新库存: {identification_code}, 板数: {before_pallet} -> {inventory.pallet_count}, 件数: {before_package} -> {inventory.package_count}")
                else:
                    current_app.logger.warning(f"未找到库存记录: {identification_code}")

        # 批量保存记录
        db.session.add_all(records_to_save)

        # 创建在途货物记录（如果目的仓库是后端仓，但排除最终目的地）
        transit_cargo_records = []
        if destination_warehouse_id:  # 有目的仓库，说明是仓库间转运
            destination_warehouse = Warehouse.query.get(destination_warehouse_id)

            # 检查目的仓库是否是后端仓（假设发往后端仓的都是前端仓发出的）
            # 但排除最终目的地：凭祥保税仓和春疆货场不生成中转记录
            is_final_destination = False
            if destination_warehouse:
                warehouse_name = destination_warehouse.warehouse_name
                is_final_destination = (
                    warehouse_name == '凭祥保税仓' or
                    warehouse_name == '春疆货场' or
                    '保税仓' in warehouse_name or
                    '春疆' in warehouse_name
                )

            if (destination_warehouse and
                destination_warehouse.warehouse_type == 'backend' and
                not is_final_destination):

                current_app.logger.info(f"检测到发往后端仓的出库，创建在途货物记录")

                for record in records_to_save:
                    # 确定源仓库（优先使用用户绑定的仓库，否则使用第一个前端仓库）
                    source_warehouse = None
                    if current_user.warehouse_id:
                        source_warehouse = Warehouse.query.get(current_user.warehouse_id)

                    if not source_warehouse:
                        # 如果用户没有绑定仓库，使用第一个前端仓库作为默认源仓库
                        source_warehouse = Warehouse.query.filter_by(warehouse_type='frontend').first()

                    # 为每个出库记录创建对应的在途货物记录
                    transit_cargo = TransitCargo(
                        customer_name=record.customer_name,
                        identification_code=record.identification_code,
                        pallet_count=record.pallet_count,
                        package_count=record.package_count,
                        weight=record.weight,
                        volume=record.volume,
                        export_mode=record.export_mode,
                        order_type=record.order_type,
                        customs_broker=record.customs_broker,
                        documents=record.document_no,
                        service_staff=record.service_staff,
                        batch_no=record.batch_no,
                        batch_sequence=record.batch_sequence,
                        batch_total=record.batch_total,
                        source_warehouse_id=source_warehouse.id if source_warehouse else None,
                        destination_warehouse_id=destination_warehouse.id,
                        departure_time=record.departure_time or datetime.now(),
                        expected_arrival_time=None,  # 可以后续计算
                        plate_number=record.plate_number,
                        delivery_plate_number=record.delivery_plate_number,
                        inbound_plate=record.inbound_plate,
                        driver_name=driver_name,
                        driver_phone=driver_phone,
                        status='in_transit',
                        remark1=record.remark1,
                        remark2=record.remark2,
                        created_by_user_id=current_user.id,
                        created_at=datetime.now()
                    )
                    transit_cargo_records.append(transit_cargo)

                # 批量保存在途货物记录
                if transit_cargo_records:
                    db.session.add_all(transit_cargo_records)
                    current_app.logger.info(f"创建了 {len(transit_cargo_records)} 条在途货物记录")

        # 确保立即提交事务，更新库存和在途货物
        try:
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            # 检查是否是重复键错误
            if 'Duplicate entry' in str(e) and 'uk_outbound_identification' in str(e):
                current_app.logger.error(f"出库记录重复键冲突: {str(e)}")
                return jsonify({
                    'success': False,
                    'message': '出库记录保存失败：存在重复的识别编码和批次序号组合，请检查数据后重试'
                }), 400
            else:
                current_app.logger.error(f"出库记录保存失败: {str(e)}")
                return jsonify({'success': False, 'message': f'保存失败: {str(e)}'}), 500

        # 刷新所有更新过的库存对象，确保它们的状态是最新的
        for inventory in inventory_dict.values():
            db.session.refresh(inventory)
            current_app.logger.info(f"刷新后的库存状态: {inventory.identification_code}, 板数: {inventory.pallet_count}, 件数: {inventory.package_count}")

        # 清除库存缓存，确保下次获取库存时能获取到最新数据
        # 批量出库可能涉及多个仓库，清除所有库存缓存
        cache_invalidation.on_inventory_change()
        current_app.logger.info("已清除所有库存相关缓存")

        current_app.logger.info(f"成功保存 {len(records_to_save)} 条出库记录，批次号: {new_batch_no}")
        current_app.logger.info(f"更新了 {len(updated_inventories)} 条库存记录")

        # 构建成功消息
        message_parts = [f'成功保存 {len(records_to_save)} 条出库记录']
        message_parts.append(f'更新了 {len(updated_inventories)} 条库存记录')

        if transit_cargo_records:
            message_parts.append(f'创建了 {len(transit_cargo_records)} 条在途货物记录')

        return jsonify({
            'success': True,
            'saved_count': len(records_to_save),
            'batch_no': new_batch_no,
            'transit_cargo_count': len(transit_cargo_records),
            'message': '，'.join(message_parts)
        })
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"保存出库记录出错: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'保存出库记录出错: {str(e)}'
        }), 500

# 库存管理相关路由
@bp.route('/inventory')
@require_permission('INVENTORY_VIEW')
@performance_monitor('inventory_list', slow_threshold=2.0)
def inventory_list():
    """库存管理 - 统一库存查询（按货物当前状态聚合）- 自动应用双层缓存"""
    # 获取搜索参数
    search_field = request.args.get('search_field', 'customer_name')
    search_value = request.args.get('search_value', '')
    warehouse_id = request.args.get('warehouse_id', '')
    start_date = request.args.get('start_date', '')
    end_date = request.args.get('end_date', '')
    stock_status = request.args.get('stock_status', '')
    cargo_status = request.args.get('cargo_status', '')

    # 确保搜索参数传递给模板
    search_params = {
        'search_field': search_field,
        'search_value': search_value,
        'warehouse_id': warehouse_id,
        'start_date': start_date,
        'end_date': end_date,
        'stock_status': stock_status,
        'cargo_status': cargo_status
    }

    try:
        page = request.args.get('page', 1, type=int)
        per_page = current_app.config.get('ITEMS_PER_PAGE', 50)

        # 获取仓库列表（直接从数据库查询，缓存暂时禁用）
        from app.models import Warehouse
        warehouses = Warehouse.query.all()

        # 获取聚合库存数据（直接从数据库查询，缓存暂时禁用）
        from app.models import Inventory
        aggregated_inventory = get_aggregated_inventory_direct()

        # 应用搜索和筛选条件
        filtered_inventory = apply_inventory_filters(
            aggregated_inventory,
            search_field,
            search_value,
            warehouse_id,
            start_date,
            end_date,
            stock_status,
            cargo_status
        )

        # 按入库日期排序
        filtered_inventory.sort(key=lambda x: x.get('inbound_time') or datetime.min)

        # 获取总记录数
        total_count = len(filtered_inventory)

        # 分页处理
        start_index = (page - 1) * per_page
        end_index = start_index + per_page
        items = filtered_inventory[start_index:end_index]

        # 为每个库存记录添加状态信息和转换为对象格式
        processed_items = []
        for item in items:
            # 创建一个类似Inventory对象的字典，用于模板渲染
            inventory_obj = type('InventoryObj', (), {})()

            # 设置基本属性
            for key, value in item.items():
                setattr(inventory_obj, key, value)

            # 调试信息：记录重量和体积数据
            current_app.logger.info(f"全库存查询 - {item.get('customer_name', 'N/A')}: weight={item.get('weight', 'N/A')}, volume={item.get('volume', 'N/A')}")

            # 设置operated_warehouse属性
            if item.get('current_warehouse'):
                inventory_obj.operated_warehouse = item['current_warehouse']
            else:
                inventory_obj.operated_warehouse = None

            # 设置货物状态 - 显示具体仓库名称
            current_status = item.get('current_status')
            current_warehouse = item.get('current_warehouse')

            if current_status == 'frontend' and current_warehouse:
                status_label = current_warehouse.warehouse_name
            elif current_status == 'backend' and current_warehouse:
                status_label = current_warehouse.warehouse_name
            elif current_status == 'in_transit':
                status_label = '在途'
            elif current_status == 'shipped_to_chunjiang':
                status_label = '已出库到春疆'
            else:
                status_label = '未知'

            inventory_obj.cargo_status = {
                'status': current_status,
                'label': status_label,
                'class': 'primary' if current_status == 'frontend' else
                        'warning' if current_status == 'backend' else
                        'info' if current_status == 'in_transit' else
                        'success' if current_status == 'shipped_to_chunjiang' else 'secondary',
                'icon': 'warehouse' if current_status in ['frontend', 'backend'] else
                       'truck' if current_status == 'in_transit' else
                       'shipping-fast' if current_status == 'shipped_to_chunjiang' else 'question'
            }

            # 添加模板需要的其他字段
            if not hasattr(inventory_obj, 'id'):
                inventory_obj.id = f"virtual_{len(processed_items) + 1}"
            if not hasattr(inventory_obj, 'last_updated'):
                # 使用入库时间作为最后更新时间
                inventory_obj.last_updated = inventory_obj.inbound_time

            processed_items.append(inventory_obj)

        # 使用自定义分页类
        records = SimplePagination(
            items=processed_items,
            page=page,
            per_page=per_page,
            total=total_count
        )

        return render_template(
            'inventory_list.html',
            title='库存管理',
            records=records,
            search_params=search_params,
            warehouses=warehouses
        )
    except Exception as e:
        current_app.logger.error(f"Error in inventory_list: {str(e)}")
        flash(f"加载数据时出错: {str(e)}", "danger")

        # 在异常情况下也传递search_params
        empty_pagination = SimplePagination(items=[], page=1, per_page=50, total=0)

        return render_template(
            'inventory_list.html',
            title='库存管理',
            records=empty_pagination,
            search_params=search_params
        )

@bp.route('/api/customers')
@csrf.exempt  # 豁免CSRF保护，提高API响应速度
@performance_monitor('api_customers', slow_threshold=1.0)
def api_customers():
    """获取客户列表API - 缓存优化版本"""
    try:
        # 获取查询参数
        query_param = request.args.get('q', '')

        # 清理查询参数的空格
        query_param = strip_whitespace(query_param)

        # 直接从数据库查询客户列表（缓存暂时禁用）
        customers = get_customer_list_direct(query_param)

        if customers is not None:
            current_app.logger.debug(f"客户列表查询完成: query={query_param}")
            return jsonify({'customers': customers})

        # 缓存未命中，记录性能指标
        performance_metrics.record_cache_miss()
        current_app.logger.info(f"客户列表缓存未命中: query={query_param}")

        # 从库存表中获取不重复的客户名称
        customers_query = db.session.query(Inventory.customer_name).distinct()

        # 如果提供了查询参数，则进行筛选
        if query_param:
            customers_query = customers_query.filter(Inventory.customer_name.like(f'%{query_param}%'))

        # 执行查询并获取结果
        customers = [row[0] for row in customers_query.all()]

        # 如果库存表中没有足够的客户，则从入库记录表中补充
        if len(customers) < 20:
            inbound_customers_query = db.session.query(InboundRecord.customer_name).distinct()
            if query_param:
                inbound_customers_query = inbound_customers_query.filter(InboundRecord.customer_name.like(f'%{query_param}%'))
            inbound_customers = [row[0] for row in inbound_customers_query.all()]

            # 合并客户列表并去重
            all_customers = list(set(customers + inbound_customers))

            # 如果有查询参数，按相关性排序
            if query_param:
                all_customers.sort(key=lambda x: (0 if query_param.lower() in x.lower() else 1, x))
            else:
                all_customers.sort()

            # 限制返回数量
            customers = all_customers[:50]

        response_data = {'customers': customers}

        # 如果没有查询参数，保存到缓存
        if not query_param:
            setattr(g, 'customers_list_cache', {
                'timestamp': time.time(),
                'data': response_data
            })

        return jsonify(response_data)
    except Exception as e:
        current_app.logger.error(f"获取客户列表出错: {str(e)}")
        return jsonify({'customers': [], 'error': str(e)})

@bp.route('/export_inventory')
def export_inventory():
    """导出库存记录"""
    # 获取搜索参数
    customer_name = request.args.get('customer_name', '')
    location = request.args.get('location', '')

    # 构建查询
    query = Inventory.query

    # 只显示库存不为0的记录
    query = query.filter((Inventory.pallet_count > 0) | (Inventory.package_count > 0))

    # 按客户名称筛选
    if customer_name:
        query = query.filter(Inventory.customer_name.like(f'%{customer_name}%'))

    # 按库位筛选
    if location:
        query = query.filter(Inventory.location.like(f'%{location}%'))

    # 按入库日期降序排序
    query = query.order_by(Inventory.inbound_time.desc())

    # 执行查询
    records = query.all()

    # 创建工作簿
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "库存记录"

    # 添加表头
    headers = [
        '序号', '入库日期', '入库车牌', '客户名称', '识别编码',
        '入库板数', '入库件数', '库存板数', '库存件数',
        '重量(kg)', '体积(m³)', '出境模式', '报关行', '单据', '跟单客服', '库位', '最后更新时间'
    ]

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        # 设置表头样式
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # 添加数据
    for row_num, record in enumerate(records, 1):
        # 序号从1开始
        ws.cell(row=row_num+1, column=1).value = row_num

        # 入库日期
        if record.outbound_time:
            ws.cell(row=row_num+1, column=2).value = record.outbound_time.strftime('%Y-%m-%d')
        else:
            ws.cell(row=row_num+1, column=2).value = ''

        # 入库车牌
        ws.cell(row=row_num+1, column=3).value = record.plate_number or ''

        # 客户名称
        ws.cell(row=row_num+1, column=4).value = record.customer_name

        # 识别编码
        ws.cell(row=row_num+1, column=5).value = record.identification_code or ''

        # 入库板数
        ws.cell(row=row_num+1, column=6).value = record.inbound_pallet_count or record.pallet_count

        # 入库件数
        ws.cell(row=row_num+1, column=7).value = record.inbound_package_count or record.package_count

        # 库存板数
        ws.cell(row=row_num+1, column=8).value = record.pallet_count

        # 库存件数
        ws.cell(row=row_num+1, column=9).value = record.package_count

        # 重量(kg)
        ws.cell(row=row_num+1, column=10).value = record.weight if record.weight and record.weight > 0 else ''

        # 体积(m³)
        ws.cell(row=row_num+1, column=11).value = record.volume if record.volume and record.volume > 0 else ''

        # 出境模式
        ws.cell(row=row_num+1, column=12).value = record.export_mode or ''

        # 报关行
        ws.cell(row=row_num+1, column=13).value = record.customs_broker or ''

        # 单据
        ws.cell(row=row_num+1, column=14).value = record.documents or ''

        # 跟单客服
        ws.cell(row=row_num+1, column=15).value = record.service_staff or ''

        # 库位
        ws.cell(row=row_num+1, column=16).value = record.location or ''

        # 最后更新时间
        if record.last_updated:
            ws.cell(row=row_num+1, column=17).value = record.last_updated.strftime('%Y-%m-%d %H:%M:%S')
        else:
            ws.cell(row=row_num+1, column=17).value = ''

    # 调整列宽
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # 获取列字母
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width

    # 创建响应
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # 生成文件名
    now = datetime.now().strftime('%Y%m%d%H%M%S')
    filename = f"库存记录_{now}.xlsx"

    # 返回Excel文件
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        download_name=filename,
        as_attachment=True
    )

@bp.route('/inventory/edit/<int:id>', methods=['GET', 'POST'])
@require_permission('INVENTORY_EDIT')
@log_operation('inventory', 'edit', 'inventory_record')
def edit_inventory(id):
    """编辑库存记录"""
    try:
        # 查询记录
        inventory = Inventory.query.get_or_404(id)

        if request.method == 'POST':
            # 更新记录
            inventory.customer_name = request.form.get('customer_name', inventory.customer_name)
            inventory.pallet_count = int(request.form.get('pallet_count', inventory.pallet_count))
            inventory.package_count = int(request.form.get('package_count', inventory.package_count))
            inventory.weight = float(request.form.get('weight', inventory.weight))
            inventory.volume = float(request.form.get('volume', inventory.volume))
            inventory.location = request.form.get('location', inventory.location)

            db.session.commit()
            flash('库存记录已更新', 'success')
            return redirect(url_for('main.inventory_list'))

        return render_template(
            'inventory_edit.html',
            title='编辑库存记录',
            inventory=inventory
        )
    except Exception as e:
        current_app.logger.error(f"编辑库存记录时出错: {str(e)}")
        flash(f"编辑库存记录时出错: {str(e)}", "danger")
        return redirect(url_for('main.inventory_list'))

@bp.route('/update_documents', methods=['GET', 'POST'])
def update_documents():
    """临时功能：更新单据字段"""
    if request.method == 'POST':
        try:
            record_id = request.form.get('id')
            documents = request.form.get('documents', '')

            if record_id:
                record = InboundRecord.query.get(record_id)
                if record:
                    record.documents = documents
                    db.session.commit()
                    flash(f'记录 #{record_id} 的单据字段已更新为: {documents}', 'success')
                else:
                    flash(f'未找到ID为 {record_id} 的记录', 'warning')
            return redirect(url_for('main.update_documents'))
        except Exception as e:
            flash(f'更新单据字段时出错: {str(e)}', 'danger')

    # 获取所有记录
    records = InboundRecord.query.order_by(InboundRecord.inbound_time.asc()).all()

    return render_template(
        'update_documents.html',
        title='更新单据字段',
        records=records
    )

# 添加标签打印相关路由
@bp.route('/label_print')
def label_print():
    """标签打印页面"""
    return render_template('label_print.html', title='标签打印')

@bp.route('/performance_test')
def performance_test():
    """性能测试页面"""
    return render_template('performance_test.html', title='性能测试')

@bp.route('/api/inbound_records', methods=['GET'])
def get_inbound_records():
    """获取入库记录列表，用于标签打印选择"""
    try:
        current_app.logger.info(f"=== 标签打印API调用 ===")
        current_app.logger.info(f"请求用户: {current_user.username if current_user.is_authenticated else '未登录'}")

        # 检查用户是否已登录
        if not current_user.is_authenticated:
            current_app.logger.warning("用户未登录，拒绝访问")
            return jsonify({
                'success': False,
                'error': '请先登录',
                'records': []
            })

        # 获取查询参数
        keyword = request.args.get('keyword', '').strip()
        start_date = request.args.get('start_date', '')
        end_date = request.args.get('end_date', '')

        current_app.logger.info(f"查询参数: keyword={keyword}, start_date={start_date}, end_date={end_date}")

        # 构建查询
        query = InboundRecord.query

        # 根据用户权限过滤仓库
        # 只有admin和管理员可以查看所有仓库的数据
        # 其他账号只能查看自己仓库的数据
        if not (current_user.is_super_admin() or current_user.username == 'admin'):
            # 非管理员用户只能查看自己仓库的数据
            if hasattr(current_user, 'warehouse_id') and current_user.warehouse_id:
                query = query.filter(InboundRecord.operated_warehouse_id == current_user.warehouse_id)
                current_app.logger.info(f"标签打印权限控制: 用户 {current_user.username} 只能查看仓库ID {current_user.warehouse_id} 的数据")
            else:
                # 如果用户没有关联仓库，返回空结果
                current_app.logger.warning(f"标签打印权限控制: 用户 {current_user.username} 没有关联仓库，返回空结果")
                return jsonify({
                    'success': True,
                    'records': [],
                    'message': '您没有关联仓库，无法查看数据'
                })
        else:
            current_app.logger.info(f"标签打印权限控制: 管理员用户 {current_user.username} 可以查看所有仓库数据")

        # 关键词搜索
        if keyword:
            query = query.filter(
                db.or_(
                    InboundRecord.identification_code.like(f'%{keyword}%'),
                    InboundRecord.plate_number.like(f'%{keyword}%'),
                    InboundRecord.customer_name.like(f'%{keyword}%')
                )
            )

        # 日期范围过滤
        if start_date:
            try:
                start_datetime = datetime.strptime(start_date, '%Y-%m-%d')
                query = query.filter(InboundRecord.inbound_time >= start_datetime)
            except ValueError:
                pass

        if end_date:
            try:
                end_datetime = datetime.strptime(end_date, '%Y-%m-%d')
                end_datetime = end_datetime.replace(hour=23, minute=59, second=59)
                query = query.filter(InboundRecord.inbound_time <= end_datetime)
            except ValueError:
                pass

        # 只获取有识别编码的记录
        query = query.filter(InboundRecord.identification_code.isnot(None))
        query = query.filter(InboundRecord.identification_code != '')

        # 排序并限制数量
        records = query.order_by(InboundRecord.inbound_time.desc()).limit(100).all()

        current_app.logger.info(f"查询结果: 找到 {len(records)} 条记录")

        # 记录详细的查询结果（仅记录前5条）
        for i, record in enumerate(records[:5]):
            current_app.logger.info(f"记录 {i+1}: {record.identification_code} | 仓库ID: {record.operated_warehouse_id}")

        # 转换为字典格式
        records_data = []
        for record in records:
            records_data.append({
                'id': record.id,
                'identification_code': record.identification_code,
                'inbound_time': record.inbound_time.strftime('%Y-%m-%d') if record.inbound_time else '',
                'plate_number': record.plate_number,
                'customer_name': record.customer_name,
                'pallet_count': record.pallet_count or 0,
                'package_count': record.package_count or 0,
                'order_type': record.order_type or ''
            })

        current_app.logger.info(f"=== 标签打印API返回 {len(records_data)} 条记录 ===")

        return jsonify({
            'success': True,
            'records': records_data
        })

    except Exception as e:
        current_app.logger.error(f"获取入库记录失败: {str(e)}")
        return jsonify({
            'success': False,
            'error': str(e),
            'records': []
        })

@bp.route('/api/recent_codes', methods=['GET'])
def get_recent_codes():
    try:
        # 从数据库中获取最近100条记录
        recent_codes = LabelCode.query.order_by(LabelCode.created_at.desc()).limit(100).all()

        code_list = []
        for code_record in recent_codes:
            code_list.append({
                'id': code_record.id,
                'code': code_record.code,
                'plate_number': code_record.plate_number,
                'customer_name': code_record.customer_name,
                'pallet_count': code_record.pallet_count,
                'package_count': code_record.package_count,
                'order_type': code_record.order_type,
                'date': code_record.created_at.strftime('%Y-%m-%d'),
                'label_size': code_record.label_size,
                'label_format': code_record.label_format
            })

        return jsonify({'codes': code_list})
    except Exception as e:
        print(f"获取最近编码时出错: {str(e)}")
        return jsonify({'codes': [], 'error': str(e)})

@bp.route('/api/save_label_code', methods=['POST'])
@csrf.exempt  # 豁免CSRF保护，因为这是API接口
def save_label_code():
    """保存整板货物标签编码"""
    try:
        data = request.get_json()
        code = data.get('code', '')
        plate_number = data.get('plate_number', '')
        customer_name = data.get('customer_name', '')
        pallet_count = data.get('pallet_count', 0)
        package_count = data.get('package_count', 0)
        order_type = data.get('order_type', '')
        label_size = data.get('label_size', '40x60')
        label_format = data.get('label_format', 'standard')
        custom_width = data.get('custom_width', 40)
        custom_height = data.get('custom_height', 60)

        # 创建记录
        label_record = LabelCode(
            code=code,
            plate_number=plate_number,
            customer_name=customer_name,
            pallet_count=pallet_count,
            package_count=package_count,
            create_time=datetime.now(),
            order_type=order_type,
            label_size=label_size,
            label_format=label_format,
            custom_width=custom_width,
            custom_height=custom_height
        )

        db.session.add(label_record)
        db.session.commit()

        return jsonify({'status': 'success', 'message': '标签编码已保存'})
    except Exception as e:
        db.session.rollback()
        print(f"Error: {str(e)}")
        return jsonify({'status': 'error', 'message': str(e)})

@bp.route('/api/generate_pdf_preview', methods=['POST'])
@csrf.exempt  # 豁免CSRF保护，因为这是API接口
def api_generate_pdf_preview():
    """生成PDF预览"""
    try:
        # 获取打印数据
        print_data = request.get_json()
        if not print_data:
            return jsonify({'success': False, 'message': '没有接收到打印数据'}), 400

        current_app.logger.info(f"PDF预览请求: {print_data}")

        # 尝试生成PDF
        try:
            pdf_data = generate_real_label_pdf(print_data)

            # 保存PDF到临时文件
            import tempfile
            import os
            import time

            # 创建临时目录
            temp_dir = os.path.join(current_app.instance_path, 'temp')
            if not os.path.exists(temp_dir):
                os.makedirs(temp_dir)

            # 生成文件名
            timestamp = int(time.time())
            filename = f"label_preview_{timestamp}.pdf"
            temp_path = os.path.join(temp_dir, filename)

            # 写入PDF数据
            with open(temp_path, 'wb') as f:
                f.write(pdf_data)

            # 生成访问URL
            pdf_url = f"/static/temp/{filename}"

            # 计算标签数量
            if print_data.get('isWhole', True):
                board_count = int(print_data.get('boardCount', 0))
                piece_count = int(print_data.get('pieceCount', 0))
                label_count = board_count + piece_count
            else:
                label_count = int(print_data.get('splitQuantity', 1))

            # 确保至少有一个标签
            if label_count == 0:
                label_count = 1

            return jsonify({
                'success': True,
                'pdf_url': pdf_url,
                'label_count': label_count,
                'message': 'PDF生成成功'
            })

        except Exception as pdf_error:
            current_app.logger.error(f"PDF生成失败: {str(pdf_error)}")
            return jsonify({
                'success': False,
                'message': f'PDF生成失败: {str(pdf_error)}'
            }), 500

    except Exception as e:
        current_app.logger.error(f"PDF预览API出错: {str(e)}")
        return jsonify({'success': False, 'message': str(e)}), 500

@bp.route('/static/temp/<filename>')
def serve_temp_file(filename):
    """提供临时文件下载"""
    import os
    from flask import send_file

    temp_dir = os.path.join(current_app.instance_path, 'temp')
    file_path = os.path.join(temp_dir, filename)

    if os.path.exists(file_path):
        return send_file(file_path, as_attachment=False)
    else:
        return "文件不存在", 404

@bp.route('/api/print_labels', methods=['POST'])
@csrf.exempt  # 豁免CSRF保护，因为这是API接口
def api_print_labels():
    """处理标签打印请求"""
    try:
        # 获取打印数据
        print_data = request.get_json()
        if not print_data:
            return jsonify({'status': 'error', 'message': '缺少打印数据'}), 400

        # 验证必要的打印参数
        required_fields = ['identificationCode', 'plateNumber', 'customerName']
        for field in required_fields:
            if not print_data.get(field):
                return jsonify({'status': 'error', 'message': f'缺少必要字段: {field}'}), 400

        # 计算标签数量
        is_whole = print_data.get('isWhole', True)
        if is_whole:
            board_count = int(print_data.get('boardCount', 0))
            piece_count = int(print_data.get('pieceCount', 0))
            total_labels = board_count + piece_count
        else:
            total_labels = int(print_data.get('splitQuantity', 1))

        # 获取打印设置
        printer_name = print_data.get('printer', 'default')
        copies = int(print_data.get('copies', 1))
        paper_size = print_data.get('paperSize', '40x60')

        # 记录详细的打印信息
        current_app.logger.info(f"=== 标签打印请求 ===")
        current_app.logger.info(f"识别编码: {print_data.get('identificationCode')}")
        current_app.logger.info(f"车牌号: {print_data.get('plateNumber')}")
        current_app.logger.info(f"客户: {print_data.get('customerName')}")
        current_app.logger.info(f"标签类型: {'整板' if is_whole else '拆板'}")
        current_app.logger.info(f"标签数量: {total_labels}")
        current_app.logger.info(f"打印机: {printer_name}")
        current_app.logger.info(f"打印份数: {copies}")
        current_app.logger.info(f"纸张大小: {paper_size}")
        current_app.logger.info(f"纸张尺寸: {print_data.get('paperWidth', 40)}×{print_data.get('paperHeight', 60)}mm")
        current_app.logger.info(f"页边距: 上{print_data.get('marginTop', 0)}mm 下{print_data.get('marginBottom', 0)}mm 左{print_data.get('marginLeft', 0)}mm 右{print_data.get('marginRight', 0)}mm")

        # 生成任务ID
        job_id = f"PRINT_{int(time.time())}"

        # 尝试真实打印
        try:
            # 检查是否安装了reportlab
            try:
                import reportlab
                pdf_content = generate_real_label_pdf(print_data)
                current_app.logger.info(f"PDF生成成功，大小: {len(pdf_content)} 字节")

                # 发送到打印机
                try:
                    send_to_printer(pdf_content, printer_name, copies)
                    current_app.logger.info(f"真实打印成功: {job_id}")
                    print_status = "真实打印"
                except Exception as print_error:
                    current_app.logger.warning(f"打印机发送失败: {str(print_error)}")
                    print_status = "PDF已生成（打印机发送失败）"

            except ImportError:
                current_app.logger.warning("reportlab未安装，使用模拟打印")
                print_status = "模拟打印（需要安装reportlab库）"

        except Exception as e:
            current_app.logger.error(f"PDF生成失败: {str(e)}")
            print_status = f"PDF生成失败: {str(e)}"

        # 返回详细的成功响应
        return jsonify({
            'status': 'success',
            'message': f'打印任务已成功提交到 {printer_name}',
            'job_id': job_id,
            'print_type': print_status,
            'details': {
                'printer': printer_name,
                'labels_count': total_labels,
                'copies': copies,
                'paper_size': paper_size,
                'total_pages': total_labels * copies,
                'customer': print_data.get('customerName'),
                'plate_number': print_data.get('plateNumber'),
                'print_status': print_status
            }
        })

    except Exception as e:
        current_app.logger.error(f"处理打印请求时出错: {str(e)}")
        return jsonify({'status': 'error', 'message': f'打印失败: {str(e)}'}), 500

@bp.route('/api/printers', methods=['GET'])
def get_printers():
    try:
        # 使用win32print获取系统所有打印机列表
        import win32print

        printers = []
        try:
            # 获取默认打印机
            default_printer = win32print.GetDefaultPrinter()
            current_app.logger.info(f"默认打印机: {default_printer}")

            # 获取所有打印机 (本地和网络)
            printer_list = win32print.EnumPrinters(win32print.PRINTER_ENUM_LOCAL | win32print.PRINTER_ENUM_CONNECTIONS)

            for printer_info in printer_list:
                # 根据测试结果，打印机名称在元组的第三个元素 (索引2)
                printer_name = printer_info[2]
                is_default = (printer_name == default_printer)

                printers.append({
                    'name': printer_name,
                    'isDefault': is_default
                })

            current_app.logger.info(f"已获取 {len(printers)} 台打印机")
        except Exception as e:
            current_app.logger.error(f"使用win32print获取打印机列表时出错: {str(e)}")
            # 尝试备用方法
            import os
            import subprocess

            try:
                # 使用wmic命令获取打印机列表
                output = subprocess.check_output('wmic printer get name', shell=True).decode('utf-8', errors='ignore')
                lines = output.strip().split('\n')[1:]  # 跳过标题行

                for line in lines:
                    printer_name = line.strip()
                    if printer_name:
                        is_default = (printer_name == default_printer)
                        printers.append({
                            'name': printer_name,
                            'isDefault': is_default
                        })

                current_app.logger.info(f"使用wmic命令获取了 {len(printers)} 台打印机")
            except Exception as cmd_error:
                current_app.logger.error(f"使用wmic命令获取打印机列表时出错: {str(cmd_error)}")

        # 如果还是没有找到打印机，返回默认列表
        if not printers:
            printers = [
                {'name': '默认打印机', 'isDefault': True},
                {'name': 'HP LaserJet Pro', 'isDefault': False},
                {'name': 'Brother标签打印机', 'isDefault': False},
                {'name': '条码标签打印机', 'isDefault': False}
            ]
            current_app.logger.warning("未能获取到系统打印机，返回默认列表")

        return jsonify({'printers': printers, 'success': True})
    except Exception as e:
        current_app.logger.error(f"获取打印机列表总体出错: {str(e)}")
        # 出错时返回默认打印机列表
        default_printers = [
            {'name': '默认打印机', 'isDefault': True},
            {'name': 'HP LaserJet Pro', 'isDefault': False},
            {'name': 'Brother标签打印机', 'isDefault': False},
            {'name': '条码标签打印机', 'isDefault': False}
        ]
        return jsonify({'printers': default_printers, 'success': False, 'error': str(e)})

@bp.route('/api/inventory/list')
@csrf.exempt  # 豁免CSRF保护，提高API响应速度
@require_permission('INVENTORY_VIEW')
def api_inventory_list():
    """获取库存列表API"""
    try:
        # 检查是否需要强制刷新缓存
        force_refresh = request.args.get('refresh', 'false').lower() == 'true'

        # 使用缓存提高性能
        cache_key = 'inventory_list_cache'
        cached_data = getattr(g, cache_key, None) if not force_refresh else None

        # 如果有缓存且缓存时间不超过5分钟，直接返回缓存数据
        if cached_data and cached_data.get('timestamp') > (time.time() - 300):
            current_app.logger.info("从缓存返回库存列表数据")
            return jsonify(cached_data['data'])

        # 获取可能的ID参数，用于筛选特定库存项
        ids_param = request.args.get('ids')

        # 使用with_hint提示数据库使用索引
        query = db.session.query(
            Inventory.id,
            Inventory.customer_name,
            Inventory.identification_code,
            Inventory.inbound_time,
            Inventory.plate_number,
            Inventory.pallet_count,
            Inventory.package_count,
            Inventory.inbound_pallet_count,
            Inventory.inbound_package_count,
            Inventory.weight,
            Inventory.volume,
            Inventory.location,
            Inventory.documents,
            Inventory.export_mode,
            Inventory.customs_broker,
            Inventory.order_type,
            Inventory.service_staff
        ).with_hint(Inventory, '')

        # 如果提供了ID参数，则按ID筛选
        if ids_param:
            ids = [int(id_str) for id_str in ids_param.split(',') if id_str.strip().isdigit()]
            if ids:
                query = query.filter(Inventory.id.in_(ids))

        # 只获取有库存的记录
        query = query.filter(
            (Inventory.pallet_count > 0) |
            (Inventory.package_count > 0)
        )

        # 按ID倒序排序
        inventory_records = query.order_by(Inventory.id.desc()).all()

        # 格式化数据
        inventory_data = []
        for record in inventory_records:
            # 格式化时间
            inbound_time_str = record.outbound_time.strftime("%Y-%m-%d") if record.outbound_time else ""

            inventory_data.append({
                'id': record.id,
                'customer_name': record.customer_name,
                'identification_code': record.identification_code,
                'inbound_time': inbound_time_str,
                'plate_number': record.plate_number,
                'inbound_plate': record.plate_number, # 添加别名方便前端使用
                'pallet_count': record.pallet_count,
                'package_count': record.package_count,
                'inbound_pallet_count': record.inbound_pallet_count,
                'inbound_package_count': record.inbound_package_count,
                'weight': record.weight,
                'volume': record.volume,
                'location': record.location,
                'documents': record.documents,
                'document_no': record.documents, # 添加别名方便前端使用
                'export_mode': record.export_mode,
                'customs_broker': record.customs_broker,
                'order_type': record.order_type,
                'service_staff': record.service_staff
            })

        # 构建响应数据
        response_data = {
            'success': True,
            'inventory': inventory_data,
            'items': inventory_data,
            'cached': False,
            'timestamp': time.time()
        }

        # 保存到缓存
        setattr(g, cache_key, {
            'timestamp': time.time(),
            'data': response_data
        })

        current_app.logger.info(f"获取库存列表成功，共 {len(inventory_data)} 条记录")

        # 返回成功响应
        return jsonify(response_data)
    except Exception as e:
        current_app.logger.error(f"获取库存列表出错: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'获取库存列表出错: {str(e)}',
            'inventory': [],
            'items': []
        }), 500

@bp.route('/api/inventory/search')
@csrf.exempt  # 豁免CSRF保护，提高API响应速度
@require_permission('INVENTORY_VIEW')
def api_inventory_search():
    """搜索库存API"""
    # 获取搜索关键词
    search_term = request.args.get('term', '')
    detailed = request.args.get('detailed', 'false').lower() == 'true'

    if not search_term:
        return jsonify({'success': False, 'message': '请提供搜索关键词'}), 400

    try:
        # 构建查询 - 在客户名称、识别编码、入库车牌中进行搜索
        query = Inventory.query.filter(
            (Inventory.customer_name.like(f'%{search_term}%')) |
            (Inventory.identification_code.like(f'%{search_term}%')) |
            (Inventory.plate_number.like(f'%{search_term}%'))
        )

        # 只返回有库存的记录
        query = query.filter(
            (Inventory.pallet_count > 0) |
            (Inventory.package_count > 0)
        )

        # 获取结果
        items = query.all()

        # 转换为JSON格式返回
        result = []
        for item in items:
            if detailed:
                # 详细信息
                result.append({
                    'id': item.id,
                    'customer_name': item.customer_name,
                    'identification_code': item.identification_code,
                    'plate_number': item.plate_number,
                    'inbound_time': item.inbound_time.strftime('%Y-%m-%d') if item.inbound_time else '',
                    'pallet_count': item.pallet_count,
                    'package_count': item.package_count,
                    'weight': item.weight,
                    'volume': item.volume,
                    'export_mode': item.export_mode,
                    'location': item.location,
                    'order_type': item.order_type,
                    'customs_broker': item.customs_broker,
                    'service_staff': item.service_staff
                })
            else:
                # 简略信息
                result.append({
                    'id': item.id,
                    'customer_name': item.customer_name,
                    'identification_code': item.identification_code,
                    'plate_number': item.plate_number,
                    'pallet_count': item.pallet_count,
                    'package_count': item.package_count,
                    'weight': item.weight,
                    'volume': item.volume,
                    'export_mode': item.export_mode
                })

        return jsonify({
            'success': True,
            'items': result,
            'count': len(result)
        })
    except Exception as e:
        current_app.logger.error(f"Error in api_inventory_search: {str(e)}")
        return jsonify({'success': False, 'message': f'搜索库存数据失败: {str(e)}'}), 500

@bp.route('/api/inventory/refresh', methods=['POST'])
@csrf.exempt  # 豁免CSRF保护，因为这是API接口
def api_refresh_inventory():
    """刷新库存API"""
    if not request.is_json:
        return jsonify({'success': False, 'message': '请求内容必须是JSON格式'}), 400

    data = request.get_json()

    # 清理所有字符串字段的空格
    data = clean_dict_whitespace(data)

    # 获取客户名称
    customer_name = data.get('customer_name')
    if not customer_name:
        return jsonify({'success': False, 'message': '缺少客户名称参数'}), 400

    # 获取库存信息
    inventory = Inventory.query.filter_by(customer_name=customer_name).first()
    if not inventory:
        return jsonify({'success': False, 'message': f'未找到客户 {customer_name} 的库存记录'}), 404

    # 获取入库记录
    inbound_records = InboundRecord.query.filter_by(customer_name=customer_name).all()

    # 获取出库记录
    outbound_records = OutboundRecord.query.filter_by(customer_name=customer_name).all()

    # 计算总入库数量
    total_inbound_pallet = sum(record.pallet_count for record in inbound_records)
    total_inbound_package = sum(record.package_count for record in inbound_records)
    total_inbound_weight = sum(record.weight for record in inbound_records)
    total_inbound_volume = sum(record.volume for record in inbound_records)

    # 计算总出库数量
    total_outbound_pallet = sum(record.pallet_count for record in outbound_records)
    total_outbound_package = sum(record.package_count for record in outbound_records)
    total_outbound_weight = sum(record.weight for record in outbound_records)
    total_outbound_volume = sum(record.volume for record in outbound_records)

    # 计算剩余库存
    remaining_pallet = total_inbound_pallet - total_outbound_pallet
    remaining_package = total_inbound_package - total_outbound_package
    remaining_weight = total_inbound_weight - total_outbound_weight
    remaining_volume = total_inbound_volume - total_outbound_volume

    # 确保不会出现负数
    remaining_pallet = max(0, remaining_pallet)
    remaining_package = max(0, remaining_package)
    remaining_weight = max(0, remaining_weight)
    remaining_volume = max(0, remaining_volume)

    try:
        # 更新库存记录（注意：不修改inbound_pallet_count和inbound_package_count，它们应该保持原始入库数量）
        # inventory.inbound_pallet_count = total_inbound_pallet  # 移除：这会破坏原始入库数量
        # inventory.inbound_package_count = total_inbound_package  # 移除：这会破坏原始入库数量
        inventory.outbound_pallet_count = total_outbound_pallet
        inventory.outbound_package_count = total_outbound_package
        inventory.pallet_count = remaining_pallet
        inventory.package_count = remaining_package
        inventory.weight = remaining_weight
        inventory.volume = remaining_volume

        # 更新其他字段（如果提供）
        if 'location' in data:
            inventory.location = data['location']
        if 'documents' in data:
            inventory.documents = data['documents']
        if 'export_mode' in data:
            inventory.export_mode = data['export_mode']
        if 'order_type' in data:
            inventory.order_type = data['order_type']
        if 'customs_broker' in data:
            inventory.customs_broker = data['customs_broker']

        # 保存更新
        db.session.commit()

        return jsonify({
            'success': True,
            'message': '库存刷新成功',
            'data': {
                'customer_name': inventory.customer_name,
                'inbound_pallet_count': inventory.inbound_pallet_count,
                'inbound_package_count': inventory.inbound_package_count,
                'outbound_pallet_count': inventory.outbound_pallet_count,
                'outbound_package_count': inventory.outbound_package_count,
                'pallet_count': inventory.pallet_count,
                'package_count': inventory.package_count,
                'weight': inventory.weight,
                'volume': inventory.volume,
                'location': inventory.location,
                'documents': inventory.documents,
                'export_mode': inventory.export_mode,
                'order_type': inventory.order_type,
                'customs_broker': inventory.customs_broker
            }
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'刷新库存时出错: {str(e)}'}), 500

# 收货人API测试页面
@bp.route('/receiver/test')
@require_permission('RECEIVER_LIST')
def receiver_test():
    """收货人API测试页面"""
    return render_template('receiver_test.html', title='收货人API测试')

# 收货人信息列表页
@bp.route('/receiver')
@register_page_permission('RECEIVER_LIST', '收货人信息管理', 'FRONTEND_OUTBOUND', '/receiver', '收货人信息管理页面')
def receiver_list():
    """收货人信息列表"""
    # 获取搜索参数
    warehouse_name = request.args.get('warehouse_name', '')

    search_params = {
        'warehouse_name': warehouse_name
    }

    try:
        page = request.args.get('page', 1, type=int)
        per_page = current_app.config.get('ITEMS_PER_PAGE', 50)

        # 构建查询
        query = Receiver.query

        # 按目的仓筛选
        if warehouse_name:
            query = query.filter(Receiver.warehouse_name.like(f'%{warehouse_name}%'))

        # 按创建时间降序排序
        query = query.order_by(Receiver.created_at.desc())

        # 获取总记录数
        total_count = query.count()

        # 分页
        items = query.offset((page - 1) * per_page).limit(per_page).all()

        # 使用自定义分页类
        records = SimplePagination(
            items=items,
            page=page,
            per_page=per_page,
            total=total_count
        )

        current_app.logger.info(f"收货人信息分页: 页数={records.pages}, 总记录数={records.total}")

        return render_template(
            'receiver_list.html',
            title='收货人信息',
            records=records,
            search_params=search_params
        )
    except Exception as e:
        current_app.logger.error(f"Error in receiver_list: {str(e)}")
        flash(f'加载收货人信息时出错: {str(e)}', 'danger')
        return render_template(
            'receiver_list.html',
            title='收货人信息',
            records=SimplePagination(items=[], page=1, per_page=per_page, total=0),
            search_params=search_params
        )

# 添加收货人信息API
@bp.route('/api/receiver/add', methods=['POST'])
@csrf.exempt  # 豁免CSRF保护，因为这是API接口
@require_permission('RECEIVER_LIST')
def api_receiver_add():
    """添加收货人信息API"""
    try:
        # 获取JSON数据
        data = request.get_json()
        if not data:
            return jsonify(success=False, message='请求中未包含JSON数据')

        # 数据验证
        if not data.get('warehouse_name'):
            return jsonify(success=False, message='目的仓名称不能为空')

        if not data.get('address'):
            return jsonify(success=False, message='详细地址不能为空')

        if not data.get('contact'):
            return jsonify(success=False, message='联络窗口不能为空')

        # 检查是否已存在相同名称的记录
        existing_receiver = Receiver.query.filter_by(warehouse_name=data['warehouse_name']).first()
        if existing_receiver:
            return jsonify(success=False, message=f'已存在名称为 "{data["warehouse_name"]}" 的目的仓')

        # 创建新记录
        receiver = Receiver(
            warehouse_name=data['warehouse_name'],
            address=data['address'],
            contact=data['contact']
        )

        # 保存到数据库
        db.session.add(receiver)
        db.session.commit()

        return jsonify(success=True, message=f'成功添加目的仓 "{data["warehouse_name"]}"', id=receiver.id)

    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"添加收货人信息时出错: {str(e)}")
        return jsonify(success=False, message=f'添加收货人信息时出错: {str(e)}')

# 更新收货人信息API
@bp.route('/api/receiver/update/<int:id>', methods=['POST'])
@csrf.exempt  # 豁免CSRF保护，因为这是API接口
@require_permission('RECEIVER_LIST')
def api_receiver_update(id):
    """更新收货人信息API"""
    try:
        # 查找记录
        receiver = Receiver.query.get(id)
        if not receiver:
            return jsonify(success=False, message=f'未找到ID为 {id} 的收货人信息')

        # 获取JSON数据
        data = request.get_json()
        if not data:
            return jsonify(success=False, message='请求中未包含JSON数据')

        # 数据验证
        if not data.get('warehouse_name'):
            return jsonify(success=False, message='目的仓名称不能为空')

        if not data.get('address'):
            return jsonify(success=False, message='详细地址不能为空')

        if not data.get('contact'):
            return jsonify(success=False, message='联络窗口不能为空')

        # 检查是否已存在相同名称的其他记录
        if data['warehouse_name'] != receiver.warehouse_name:
            existing_receiver = Receiver.query.filter_by(warehouse_name=data['warehouse_name']).first()
            if existing_receiver and existing_receiver.id != id:
                return jsonify(success=False, message=f'已存在名称为 "{data["warehouse_name"]}" 的目的仓')

        # 更新记录
        receiver.warehouse_name = data['warehouse_name']
        receiver.address = data['address']
        receiver.contact = data['contact']

        # 保存到数据库
        db.session.commit()

        return jsonify(success=True, message=f'成功更新目的仓 "{data["warehouse_name"]}"')

    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"更新收货人信息时出错: {str(e)}")
        return jsonify(success=False, message=f'更新收货人信息时出错: {str(e)}')

# 删除收货人信息API
@bp.route('/api/receiver/delete/<int:id>', methods=['POST'])
@csrf.exempt  # 豁免CSRF保护，因为这是API接口
@require_permission('RECEIVER_LIST')
def api_receiver_delete(id):
    """删除收货人信息API"""
    try:
        # 查找记录
        receiver = Receiver.query.get(id)
        if not receiver:
            return jsonify(success=False, message=f'未找到ID为 {id} 的收货人信息')

        # 删除记录
        warehouse_name = receiver.warehouse_name
        db.session.delete(receiver)
        db.session.commit()

        return jsonify(success=True, message=f'成功删除目的仓 "{warehouse_name}"')

    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"删除收货人信息时出错: {str(e)}")
        return jsonify(success=False, message=f'删除收货人信息时出错: {str(e)}')

# 获取收货人信息列表API
@bp.route('/api/receivers')
@csrf.exempt  # 豁免CSRF保护，提高API响应速度
@require_permission('RECEIVER_LIST')
def api_receivers():
    """获取收货人信息列表API"""
    try:
        # 可选的搜索参数
        warehouse_name = request.args.get('warehouse_name', '')
        # 新增：业务场景参数，用于过滤目的仓选项
        scenario = request.args.get('scenario', '')  # 'frontend_to_backend' 表示前端仓发运后端仓

        # 构建查询
        query = Receiver.query

        # 应用筛选条件
        if warehouse_name:
            query = query.filter(Receiver.warehouse_name.like(f'%{warehouse_name}%'))

        # 根据业务场景过滤目的仓选项
        if scenario == 'frontend_to_backend':
            # 前端仓发运后端仓：只显示后端仓（凭祥）选项
            query = query.filter(Receiver.warehouse_name.like('%凭祥%'))
            current_app.logger.info("前端仓发运后端仓场景：只显示凭祥相关目的仓")

        # 按创建时间降序排序
        receivers = query.order_by(Receiver.created_at.desc()).all()

        # 转换为字典列表
        receivers_list = [receiver.to_dict() for receiver in receivers]

        current_app.logger.info(f"获取收货人列表：场景={scenario}, 搜索={warehouse_name}, 结果数量={len(receivers_list)}")
        return jsonify(success=True, receivers=receivers_list, count=len(receivers_list))

    except Exception as e:
        current_app.logger.error(f"获取收货人信息列表时出错: {str(e)}")
        return jsonify(success=False, message=f'获取收货人信息列表时出错: {str(e)}', receivers=[], count=0)

# 获取单个收货人信息API
@bp.route('/api/receiver/<int:id>')
@csrf.exempt  # 豁免CSRF保护，提高API响应速度
@require_permission('RECEIVER_LIST')
def api_receiver(id):
    """获取单个收货人信息API"""
    try:
        receiver = Receiver.query.get(id)
        if not receiver:
            return jsonify(success=False, message=f'未找到ID为 {id} 的收货人信息')

        return jsonify(success=True, receiver=receiver.to_dict())

    except Exception as e:
        current_app.logger.error(f"获取收货人信息时出错: {str(e)}")
        return jsonify(success=False, message=f'获取收货人信息时出错: {str(e)}')

# 根据仓库名称获取收货人信息API
@bp.route('/api/receiver/by_warehouse/<warehouse_name>')
@csrf.exempt  # 豁免CSRF保护，提高API响应速度
@require_permission('RECEIVER_LIST')
def api_receiver_by_warehouse(warehouse_name):
    """根据仓库名称获取收货人信息API"""
    try:
        receiver = Receiver.query.filter_by(warehouse_name=warehouse_name).first()
        if not receiver:
            return jsonify(success=False, message=f'未找到仓库 "{warehouse_name}" 的收货人信息')

        return jsonify(success=True, receiver=receiver.to_dict())

    except Exception as e:
        current_app.logger.error(f"根据仓库名称获取收货人信息时出错: {str(e)}")
        return jsonify(success=False, message=f'根据仓库名称获取收货人信息时出错: {str(e)}')

@bp.route('/download_inbound_template_original')
def download_inbound_template_original():
    """下载带有下拉列表的入库模板 - 原始实现"""
    try:
        # 创建临时文件
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            temp_path = tmp.name

        # 使用openpyxl创建工作簿
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "入库数据模板"

        # 添加表头
        headers = [
            '入库时间 *',
            '入库车牌 *',
            '客户名称 *',
            '板数 *',
            '件数 *',
            '重量(kg)',
            '体积(m³)',
            '出境模式 *',
            '订单类型 *',
            '报关行 *',
            '库位',
            '单据',
            '跟单客服 *'
        ]

        # 写入表头
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

        # 添加示例数据
        example_data = [
            datetime.now().strftime('%Y-%m-%d'),
            'A12345',
            '示例客户',
            10,
            100,
            1000,
            5,
            '保税',
            '零担',  # 订单类型示例值
            '示例报关行',
            'A区-01-01',
            '示例单据',
            '示例跟单客服'
        ]

        # 写入示例数据到第2行
        for col_num, value in enumerate(example_data, 1):
            cell = ws.cell(row=2, column=col_num)
            cell.value = value

        # 调整列宽
        for col in range(1, len(headers) + 1):
            column_letter = get_column_letter(col)
            ws.column_dimensions[column_letter].width = 15

        # 创建数据验证 - 使用最简单的方式
        dv = DataValidation(
            type="list",
            formula1='"零担,原车出境,换车出境,套牌订单"',
            allow_blank=True
        )

        # 应用到订单类型列(第9列)
        dv.add('I2:I1048576')

        # 添加数据验证到工作表
        ws.add_data_validation(dv)

        # 保存工作簿
        wb.save(temp_path)

        # 创建一个新的工作簿，从刚保存的文件读取，然后再次保存
        # 这样可以确保所有的临时数据和提示都被清除
        wb2 = openpyxl.load_workbook(temp_path)
        ws2 = wb2.active

        # 确保没有任何提示文字
        # 遍历所有行，清除除了表头和示例数据以外的所有内容
        for row in range(3, 11):  # 清除第3行到第10行的所有内容
            for col in range(1, len(headers) + 1):
                cell = ws2.cell(row=row, column=col)
                cell.value = None

        # 重新创建数据验证
        dv2 = DataValidation(
            type="list",
            formula1='"零担,原车出境,换车出境,套牌订单"',
            allow_blank=True
        )

        # 应用到订单类型列
        dv2.add('I2:I1048576')

        # 添加数据验证到工作表
        ws2.add_data_validation(dv2)

        # 再次保存工作簿
        wb2.save(temp_path)

        # 返回文件
        return send_file(
            temp_path,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            download_name='入库数据导入模板.xlsx',
            as_attachment=True
        )
    except Exception as e:
        current_app.logger.error(f"生成入库模板时出错: {str(e)}")
        flash(f'生成入库模板时出错: {str(e)}', 'danger')
        return redirect(url_for('main.inbound'))

@bp.route('/download_inbound_template')
def download_inbound_template():
    """下载入库模板 - 使用CSV格式，避免Excel提示问题"""
    try:
        # 创建临时文件
        with tempfile.NamedTemporaryFile(suffix='.csv', delete=False) as tmp:
            temp_path = tmp.name

        # 添加表头
        headers = [
            '入库时间 *',
            '入库车牌 *',
            '客户名称 *',
            '板数 *',
            '件数 *',
            '重量(kg)',
            '体积(m³)',
            '出境模式 *',
            '订单类型 *',
            '报关行 *',
            '库位',
            '单据',
            '跟单客服 *'
        ]

        # 添加示例数据
        example_data = [
            datetime.now().strftime('%Y-%m-%d'),
            'A12345',
            '示例客户',
            10,
            100,
            1000,
            5,
            '保税',
            '零担',  # 订单类型示例值
            '示例报关行',
            'A区-01-01',
            '示例单据',
            '示例跟单客服'
        ]

        # 写入CSV文件
        with open(temp_path, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.writer(f)
            writer.writerow(headers)
            writer.writerow(example_data)

        # 返回文件
        return send_file(
            temp_path,
            mimetype='text/csv',
            download_name='入库数据导入模板.csv',
            as_attachment=True
        )
    except Exception as e:
        current_app.logger.error(f"生成入库模板时出错: {str(e)}")
        flash(f'生成入库模板时出错: {str(e)}', 'danger')
        return redirect(url_for('main.inbound'))

@bp.route('/api/export/inbound_template', methods=['GET'])
@csrf.exempt  # 豁免CSRF保护，因为这是API接口
def export_inbound_template():
    """导出入库数据模板，包含下拉框"""
    # 创建工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "入库数据模板"

    # 表头
    headers = [
        '入库时间 *',
        '入库车牌 *',
        '客户名称 *',
        '板数 *',
        '件数 *',
        '重量(kg)',
        '体积(m³)',
        '出境模式 *',
        '订单类型 *',
        '报关行 *',
        '库位',
        '单据',
        '跟单客服 *'
    ]

    # 写入表头
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # 示例数据
    now = datetime.now().strftime("%Y-%m-%d")  # 只使用日期部分，不包含时间
    example_data = [
        now, 'A12345', '示例客户', 10, 100, 1000, 5, '保税', '零担', '示例报关行', 'A区-01-01', '示例单据', '示例跟单客服'
    ]

    # 写入示例数据
    for col_idx, value in enumerate(example_data, 1):
        cell = ws.cell(row=2, column=col_idx, value=value)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # 设置列宽
    column_widths = [15, 12, 15, 8, 8, 10, 10, 12, 12, 12, 12, 10, 12]
    for col_idx, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + col_idx)].width = width

    # 添加数据验证 - 订单类型下拉列表
    order_type_options = '"零担,原车出境,换车出境,套牌订单"'
    dv = DataValidation(type="list", formula1=order_type_options, allow_blank=True)
    dv.add('I2:I1000')  # 应用到I列（订单类型列）
    ws.add_data_validation(dv)

    # 添加数据验证 - 出境模式下拉列表
    export_mode_options = '"保税,清关"'
    dv_export = DataValidation(type="list", formula1=export_mode_options, allow_blank=True)
    dv_export.add('H2:H1000')  # 应用到H列（出境模式列）
    ws.add_data_validation(dv_export)

    # 保存到内存中
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # 发送文件
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='前端仓入库数据导入模板.xlsx'
    )

@bp.route('/api/export/backend_inbound_template', methods=['GET'])
@csrf.exempt  # 豁免CSRF保护，因为这是API接口
def export_backend_inbound_template():
    """导出后端仓入库数据模板，包含下拉框"""
    # 创建工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "后端仓入库数据模板"

    # 表头 - 后端仓的出境模式/报关行/订单类型为非必填，板数和件数至少一项必填
    headers = [
        '入库时间 *',
        '入库车牌 *',
        '客户名称 *',
        '板数 *',        # 与件数至少一项必填
        '件数 *',        # 与板数至少一项必填
        '重量(kg)',
        '体积(m³)',
        '出境模式',      # 后端仓非必填
        '订单类型',      # 后端仓非必填
        '报关行',        # 后端仓非必填
        '库位',
        '单据',
        '跟单客服 *'
    ]

    # 写入表头
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # 示例数据（与前端仓保持一致）
    now = datetime.now().strftime("%Y-%m-%d")  # 只使用日期部分，不包含时间
    example_data = [
        now, 'A12345', '示例客户', 10, 100, 1000, 5, '保税', '零担', '示例报关行', 'A区-01-01', '示例单据', '示例跟单客服'
    ]

    # 写入示例数据
    for col_idx, value in enumerate(example_data, 1):
        cell = ws.cell(row=2, column=col_idx, value=value)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # 设置列宽
    column_widths = [15, 12, 15, 8, 8, 10, 10, 12, 12, 12, 12, 10, 12]
    for col_idx, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + col_idx)].width = width

    # 添加数据验证 - 订单类型下拉列表（与前端仓保持一致）
    order_type_options = '"零担,原车出境,换车出境,套牌订单"'
    dv = DataValidation(type="list", formula1=order_type_options, allow_blank=True)
    dv.add('I2:I1000')  # 应用到I列（订单类型列）
    ws.add_data_validation(dv)

    # 添加数据验证 - 出境模式下拉列表（与前端仓保持一致）
    export_mode_options = '"保税,清关"'
    dv_export = DataValidation(type="list", formula1=export_mode_options, allow_blank=True)
    dv_export.add('H2:H1000')  # 应用到H列（出境模式列）
    ws.add_data_validation(dv_export)

    # 保存到内存中
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # 发送文件
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='后端仓入库数据导入模板.xlsx'
    )

@bp.route('/outbound/print/<int:id>', methods=['GET'])
@require_permission('OUTBOUND_PRINT')
def print_outbound(id):
    """打印出库单"""
    record = OutboundRecord.query.get_or_404(id)
    now = datetime.now()

    # 如果该记录没有批次号，生成一个临时批次号
    if not record.batch_no:
        record.batch_no = f"TMP{now.strftime('%Y%m%d')}{id}"
        record.batch_sequence = 1
        record.batch_total = 1

    # 为出库记录查找对应的入库记录，获取入库车牌信息
    if record.identification_code:
        # 通过识别编码查找对应的入库记录
        inbound_record = InboundRecord.query.filter_by(
            identification_code=record.identification_code
        ).first()
        if inbound_record and inbound_record.plate_number:
            # 将入库记录的车牌信息设置到出库记录的inbound_plate字段
            record.inbound_plate = inbound_record.plate_number
            current_app.logger.debug(f"为出库记录 {record.id} 设置入库车牌: {inbound_record.plate_number}")
        else:
            current_app.logger.debug(f"出库记录 {record.id} 未找到对应的入库记录或入库记录无车牌信息")

    return render_template('outbound_record.html', record=record, now=now)

@bp.route('/outbound/print', methods=['GET'])
def outbound_print():
    """出库单打印页面 - 重定向到打印列表"""
    return outbound_print_list()

@bp.route('/outbound/print_list', methods=['GET'])
def outbound_print_list():
    """出库单打印列表"""
    # 获取分页参数
    page = request.args.get('page', 1, type=int)
    per_page = 15  # 每页显示15个批次记录

    # 获取搜索参数
    search_params = {
        'date_start': request.args.get('date_start', datetime.now().strftime('%Y-%m-%d')),
        'date_end': request.args.get('date_end', (datetime.now() + timedelta(days=1)).strftime('%Y-%m-%d')),
        'plate_number': request.args.get('plate_number', '').strip(),  # 去除前后空白字符
        'batch_no': request.args.get('batch_no', '').strip(),
        'customer_name': request.args.get('customer_name', '').strip()
    }

    # 记录搜索参数
    current_app.logger.info(f"出库单打印列表搜索参数: {search_params}")
    current_app.logger.info(f"原始车牌号参数: '{request.args.get('plate_number', '')}'")
    current_app.logger.info(f"处理后车牌号: '{search_params['plate_number']}'")
    current_app.logger.info(f"车牌号长度: {len(search_params['plate_number'])}")

    # 检查是否有查询参数
    has_search_params = any([
        search_params['plate_number'],
        search_params['batch_no'],
        search_params['customer_name']
    ])

    # 如果没有额外的搜索参数，默认只取最近2天的数据以提高速度
    if not has_search_params:
        start_date = datetime.strptime(search_params['date_start'], '%Y-%m-%d')
        end_date = datetime.strptime(search_params['date_end'], '%Y-%m-%d')
        date_diff = (end_date - start_date).days

        if date_diff > 2:
            current_app.logger.info("日期范围超过2天且无其他筛选条件，默认只取最近2天的数据")
            search_params['date_start'] = datetime.now().strftime('%Y-%m-%d')

    # 构建查询 - 使用JOIN优化性能
    try:
        # 限制最大记录数以防止加载过多数据导致页面卡顿
        max_records = 1000

        # 构建查询 - 只显示前端仓出库记录和后端仓流转到前端仓的记录
        # 排除后端仓出到末端的记录和后端仓返回前端仓的记录
        query = OutboundRecord.query.join(
            Warehouse, OutboundRecord.operated_warehouse_id == Warehouse.id
        ).filter(
            db.or_(
                # 前端仓的所有出库记录
                Warehouse.warehouse_type == 'frontend',
                # 后端仓流转到前端仓的记录（目的地是前端仓），但排除返回前端仓的记录
                db.and_(
                    Warehouse.warehouse_type == 'backend',
                    db.or_(
                        OutboundRecord.destination.like('%平湖仓%'),
                        OutboundRecord.destination.like('%昆山仓%'),
                        OutboundRecord.destination.like('%成都仓%'),
                        OutboundRecord.destination.like('%凭祥北投仓%')
                    ),
                    # 排除后端仓返回前端仓的记录
                    ~OutboundRecord.remarks.like('%后端仓返回前端仓%')
                )
            )
        )

        # 添加仓库权限过滤
        # 管理员和后端仓账号可以查看所有仓库的数据
        # 前端仓账号只能查看自身仓库操作的数据
        if current_user.is_authenticated:
            is_admin = getattr(current_user, 'is_admin', False) or getattr(current_user, 'user_type', '') == 'admin'

            if not is_admin and hasattr(current_user, 'warehouse') and current_user.warehouse:
                if current_user.warehouse.warehouse_type == 'frontend':
                    # 前端仓用户只能查看自己仓库的出库记录
                    query = query.filter(OutboundRecord.operated_warehouse_id == current_user.warehouse_id)
                    current_app.logger.info(f"前端仓用户 {current_user.username} 只能查看仓库 {current_user.warehouse_id} 的出库记录")
                # 后端仓用户可以查看所有仓库的出库记录
                elif current_user.warehouse.warehouse_type == 'backend':
                    current_app.logger.info(f"后端仓用户 {current_user.username} 可以查看所有仓库的出库记录")
            elif is_admin:
                current_app.logger.info(f"管理员用户 {current_user.username} 可以查看所有仓库的出库记录")
            else:
                current_app.logger.warning(f"用户 {current_user.username} 没有关联仓库，无法查看出库记录")
                # 如果用户没有关联仓库，返回空结果
                query = query.filter(OutboundRecord.id == -1)  # 永远不会匹配的条件

        # 应用过滤条件
        if search_params['date_start'] and search_params['date_end']:
            start_date = datetime.strptime(search_params['date_start'], '%Y-%m-%d')
            end_date = datetime.strptime(search_params['date_end'], '%Y-%m-%d')
            end_date = end_date.replace(hour=23, minute=59, second=59)
            query = query.filter(OutboundRecord.outbound_time.between(start_date, end_date))

        if search_params['plate_number']:
            query = query.filter(OutboundRecord.plate_number.like(f"%{search_params['plate_number']}%"))

        if search_params['batch_no']:
            query = query.filter(OutboundRecord.batch_no.like(f"%{search_params['batch_no']}%"))

        if search_params['customer_name']:
            query = query.filter(OutboundRecord.customer_name.like(f"%{search_params['customer_name']}%"))

        # 先按出库日期降序排序，再按批次号排序
        query = query.order_by(OutboundRecord.outbound_time.desc(), OutboundRecord.batch_no, OutboundRecord.batch_sequence)

        # 优先使用批次号和出库时间索引进行查询优化
        # 限制记录数量，避免超时
        records = query.limit(max_records).all()

        current_app.logger.info(f"查询到的记录数: {len(records)}")

        # 如果记录数超过限制，添加警告信息
        show_warning = len(records) >= max_records

        # 计算行合并数量
        date_rowspans = {}
        plate_rowspans = {}
        batch_rowspans = {}
        batch_stats = {}
        batch_nos = set()

        # 使用字典和集合提高处理效率
        batch_first_records = {}  # 记录每个批次的第一条记录

        for record in records:
            # 记录批次号
            batch_no = record.batch_no
            batch_nos.add(batch_no)

            # 保存每个批次的第一条记录
            if batch_no not in batch_first_records:
                batch_first_records[batch_no] = record

            # 统计每个批次的记录数量（用于rowspan）
            if batch_no in batch_rowspans:
                batch_rowspans[batch_no] += 1
            else:
                batch_rowspans[batch_no] = 1

            # 统计每个日期的记录数量
            date_key = record.outbound_time.strftime('%Y-%m-%d')
            date_rowspans[date_key] = date_rowspans.get(date_key, 0) + 1

            # 统计每个日期+车牌组合的记录数量
            plate_key = date_key + '_' + record.plate_number
            plate_rowspans[plate_key] = plate_rowspans.get(plate_key, 0) + 1

        # 使用每个批次的第一条记录的层板数作为批次总数
        for batch_no, first_record in batch_first_records.items():
            batch_stats[batch_no] = {
                'big_pallet': first_record.large_layer or 0,
                'small_pallet': first_record.small_layer or 0,
                'card_pallet': first_record.pallet_board or 0
            }
            # 添加调试信息
            current_app.logger.debug(f"批次 {batch_no}: 大层板={first_record.large_layer}, 小层板={first_record.small_layer}, 卡板={first_record.pallet_board}")

        # 记录统计信息
        current_app.logger.info(f"不同批次号数量: {len(batch_nos)}")

        # 按批次分页处理
        batch_list = list(batch_nos)
        total_batches = len(batch_list)

        # 计算分页
        start_idx = (page - 1) * per_page
        end_idx = start_idx + per_page
        paginated_batch_nos = batch_list[start_idx:end_idx]

        # 只保留当前页面的批次记录
        paginated_records = [record for record in records if record.batch_no in paginated_batch_nos]

        # 重新计算当前页面的统计信息
        paginated_batch_stats = {batch_no: batch_stats[batch_no] for batch_no in paginated_batch_nos if batch_no in batch_stats}

        # 创建分页对象
        class BatchPagination:
            def __init__(self, items, page, per_page, total):
                self.items = items
                self.page = page
                self.per_page = per_page
                self.total = total
                self.pages = (total + per_page - 1) // per_page
                self.has_prev = page > 1
                self.has_next = page < self.pages
                self.prev_num = page - 1 if self.has_prev else None
                self.next_num = page + 1 if self.has_next else None

        batch_pagination = BatchPagination(paginated_batch_nos, page, per_page, total_batches)

        current_app.logger.debug("开始渲染模板")

        return render_template('outbound_print_list.html',
                            batch_records=paginated_records,  # 提供正确的变量名
                            batch_groups=paginated_batch_stats,  # 兼容性保留
                            date_rowspans=date_rowspans,
                            plate_rowspans=plate_rowspans,
                            batch_rowspans=batch_rowspans,
                            batch_stats=paginated_batch_stats,
                            batch_pagination=batch_pagination,
                            search_params=search_params,
                            show_warning=show_warning,
                            max_records=max_records)
    except Exception as e:
        current_app.logger.error(f"渲染模板出错: {str(e)}")
        return f"渲染出库单打印列表时出错: {str(e)}", 500

@bp.route('/outbound/print_selected_records', methods=['POST'])
@csrf.exempt  # 豁免CSRF保护，确保打印功能可用
@login_required
def print_selected_records():
    """处理选中记录的打印功能"""
    record_ids = request.form.getlist('record_ids[]')
    current_app.logger.info(f"选中的记录ID: {record_ids}")

    if not record_ids:
        flash('请至少选择一条记录进行打印', 'warning')
        return redirect(url_for('main.outbound_print_list'))

    # 查询选中的记录，包含收货人信息
    records = OutboundRecord.query.options(
        db.joinedload(OutboundRecord.receiver)
    ).filter(OutboundRecord.id.in_(record_ids)).all()
    current_app.logger.info(f"查询到的记录数: {len(records)}")

    # 为每个出库记录查找对应的入库记录，获取入库车牌信息，并关联收货人信息
    for record in records:
        if record.identification_code:
            # 通过识别编码查找对应的入库记录
            inbound_record = InboundRecord.query.filter_by(
                identification_code=record.identification_code
            ).first()
            if inbound_record and inbound_record.plate_number:
                # 将入库记录的车牌信息设置到出库记录的inbound_plate字段
                record.inbound_plate = inbound_record.plate_number
                current_app.logger.debug(f"为出库记录 {record.id} 设置入库车牌: {inbound_record.plate_number}")
            else:
                current_app.logger.debug(f"出库记录 {record.id} 未找到对应的入库记录或入库记录无车牌信息")

        # 如果出库记录没有关联收货人信息，根据目的地自动关联
        if not record.receiver_id and record.destination:
            from app.models import Receiver
            receiver = Receiver.query.filter_by(warehouse_name=record.destination).first()
            if receiver:
                record.receiver_id = receiver.id
                record.detailed_address = receiver.address
                record.contact_window = receiver.contact
                # 重新设置receiver关系，确保模板可以访问
                record.receiver = receiver
                current_app.logger.debug(f"为出库记录 {record.id} 自动关联收货人: {receiver.warehouse_name}")
            else:
                current_app.logger.debug(f"出库记录 {record.id} 未找到匹配的收货人信息: {record.destination}")

    # 计算批次统计信息
    batch_stats = {}
    total_items = 0
    total_weight = 0
    total_volume = 0

    for record in records:
        total_items += record.package_count or 0
        total_weight += record.weight or 0
        total_volume += record.volume or 0

        # 统计每个批次的记录数
        if record.batch_no not in batch_stats:
            batch_stats[record.batch_no] = {
                'count': 0
            }
        batch_stats[record.batch_no]['count'] += 1

    # 获取出库记录的仓库信息作为始发仓
    source_warehouse = None
    if records:
        # 从第一条出库记录获取仓库ID
        first_record = records[0]
        current_app.logger.info(f"第一条记录的operated_warehouse_id: {first_record.operated_warehouse_id}")
        if hasattr(first_record, 'operated_warehouse_id') and first_record.operated_warehouse_id:
            from app.models import Warehouse
            source_warehouse = Warehouse.query.get(first_record.operated_warehouse_id)
            current_app.logger.info(f"获取到始发仓: {source_warehouse.warehouse_name if source_warehouse else 'None'}")
        elif current_user.warehouse_id:
            # 如果出库记录没有仓库ID，则使用当前用户的仓库作为备选
            from app.models import Warehouse
            source_warehouse = Warehouse.query.get(current_user.warehouse_id)
            current_app.logger.info(f"使用用户仓库作为始发仓: {source_warehouse.warehouse_name if source_warehouse else 'None'}")
        else:
            current_app.logger.warning("无法获取始发仓信息")

    # 获取批次号和批次记录
    batch_no = records[0].batch_no if records else ''
    batch_items = records  # 将records作为batch_items传递给模板

    # 渲染打印模板
    return render_template('outbound_print_template.html',
                          records=records,
                          batch_items=batch_items,
                          batch_no=batch_no,
                          batch_stats=batch_stats,
                          total_items=total_items,
                          total_weight=total_weight,
                          total_volume=total_volume,
                          source_warehouse=source_warehouse,
                          now=datetime.now())


@bp.route('/test/warehouse_display')
@login_required
def test_warehouse_display():
    """测试始发仓显示"""
    # 获取一些出库记录
    records = OutboundRecord.query.filter(
        db.not_(OutboundRecord.destination.like('%春疆%'))
    ).limit(3).all()

    # 获取始发仓信息
    source_warehouse = None
    if records:
        first_record = records[0]
        if hasattr(first_record, 'operated_warehouse_id') and first_record.operated_warehouse_id:
            from app.models import Warehouse
            source_warehouse = Warehouse.query.get(first_record.operated_warehouse_id)

    return render_template('test_warehouse_display.html',
                          records=records,
                          source_warehouse=source_warehouse)


@bp.route('/outbound/exit_plan')
@require_permission('OUTBOUND_EXIT_PLAN')
def outbound_exit_plan():
    """出境计划单页面"""
    # 获取分页参数
    page = request.args.get('page', 1, type=int)
    per_page = 15  # 每页显示15个批次记录

    # 获取搜索参数
    search_params = {
        'date_start': request.args.get('date_start', (datetime.now() - timedelta(days=1)).strftime('%Y-%m-%d')),
        'date_end': request.args.get('date_end', datetime.now().strftime('%Y-%m-%d')),
        'plate_number': request.args.get('plate_number', '').strip(),
        'batch_no': request.args.get('batch_no', '').strip(),
        'customer_name': request.args.get('customer_name', '').strip()
    }

    # 记录搜索参数
    current_app.logger.info(f"出境计划单搜索参数: {search_params}")

    # 检查是否有查询参数
    has_search_params = any([
        search_params['date_start'] != (datetime.now() - timedelta(days=1)).strftime('%Y-%m-%d'),
        search_params['date_end'] != datetime.now().strftime('%Y-%m-%d'),
        search_params['plate_number'],
        search_params['batch_no'],
        search_params['customer_name']
    ])

    show_warning = not has_search_params

    try:
        # 限制最大记录数以防止加载过多数据导致页面卡顿
        max_records = 1000

        # 构建查询 - 查询后端仓出库到凭祥保税仓/春疆货场的记录
        # 通过多种方式识别凭祥保税仓/春疆货场的记录，但排除后端仓返回前端仓的记录
        query = OutboundRecord.query.filter(
            db.and_(
                db.or_(
                    OutboundRecord.destination == '春疆货场',
                    OutboundRecord.destination == '凭祥保税仓',
                    OutboundRecord.destination.like('%春疆%'),
                    OutboundRecord.destination.like('%保税仓%'),
                    OutboundRecord.detailed_address == '谅山春疆货场',
                    OutboundRecord.detailed_address.like('%春疆%'),
                    OutboundRecord.detailed_address.like('%保税仓%')
                ),
                # 排除后端仓返回前端仓的记录
                ~OutboundRecord.remarks.like('%后端仓返回前端仓%')
            )
        ).options(db.joinedload(OutboundRecord.operated_warehouse))

        # 应用过滤条件
        if search_params['date_start'] and search_params['date_end']:
            start_date = datetime.strptime(search_params['date_start'], '%Y-%m-%d')
            end_date = datetime.strptime(search_params['date_end'], '%Y-%m-%d')
            end_date = end_date.replace(hour=23, minute=59, second=59)
            query = query.filter(OutboundRecord.outbound_time.between(start_date, end_date))

        if search_params['plate_number']:
            query = query.filter(OutboundRecord.plate_number.like(f"%{search_params['plate_number']}%"))

        if search_params['batch_no']:
            query = query.filter(OutboundRecord.batch_no.like(f"%{search_params['batch_no']}%"))

        if search_params['customer_name']:
            query = query.filter(OutboundRecord.customer_name.like(f"%{search_params['customer_name']}%"))

        # 按记录生成时间降序排序（最新创建的记录在前），再按批次号降序排序，最后按操作顺序降序排序
        query = query.order_by(OutboundRecord.created_at.desc(), OutboundRecord.batch_no.desc(), OutboundRecord.batch_sequence.desc())

        # 限制记录数量，避免超时
        records = query.limit(max_records).all()

        current_app.logger.info(f"查询到的出境计划单记录数: {len(records)}")

        # 自动填充缺失的字段信息
        for record in records:
            # 通过识别编码查找对应的入库记录，填充入库车牌信息
            if record.identification_code and not record.inbound_plate:
                inbound_record = InboundRecord.query.filter_by(
                    identification_code=record.identification_code
                ).first()
                if inbound_record and inbound_record.plate_number:
                    record.inbound_plate = inbound_record.plate_number
                    current_app.logger.debug(f"为出境计划单记录 {record.id} 设置入库车牌: {inbound_record.plate_number}")

            # 如果送货干线车为空，使用出库车牌
            if record.plate_number and not record.delivery_plate_number:
                record.delivery_plate_number = record.plate_number
                current_app.logger.debug(f"为出境计划单记录 {record.id} 设置送货干线车: {record.plate_number}")

            # 填充documents字段，如果为空但document_count有值，则使用document_count的值
            if not record.documents and record.document_count is not None:
                record.documents = str(record.document_count)
                current_app.logger.debug(f"为出境计划单记录 {record.id} 设置单据字段: {record.documents}")

            # 通过识别编码查找对应的库存记录，填充出境模式、报关行、订单类型信息
            if record.identification_code and (not record.export_mode or not record.customs_broker or not record.order_type):
                inventory_record = Inventory.query.filter_by(
                    identification_code=record.identification_code
                ).first()
                if inventory_record:
                    if not record.export_mode and inventory_record.export_mode:
                        record.export_mode = inventory_record.export_mode
                        current_app.logger.debug(f"为出境计划单记录 {record.id} 设置出境模式: {inventory_record.export_mode}")
                    if not record.customs_broker and inventory_record.customs_broker:
                        record.customs_broker = inventory_record.customs_broker
                        current_app.logger.debug(f"为出境计划单记录 {record.id} 设置报关行: {inventory_record.customs_broker}")
                    if not record.order_type and inventory_record.order_type:
                        record.order_type = inventory_record.order_type
                        current_app.logger.debug(f"为出境计划单记录 {record.id} 设置订单类型: {inventory_record.order_type}")

        # 按批次号分组统计
        batch_stats = {}
        batch_nos = set()

        for record in records:
            batch_no = record.batch_no or f"TEMP_{record.id}"
            batch_nos.add(batch_no)

            if batch_no not in batch_stats:
                batch_stats[batch_no] = {
                    'records': [],
                    'total_weight': 0,
                    'total_volume': 0,
                    'total_pallet': 0,
                    'total_package': 0
                }

            batch_stats[batch_no]['records'].append(record)
            batch_stats[batch_no]['total_weight'] += record.weight or 0
            batch_stats[batch_no]['total_volume'] += record.volume or 0
            batch_stats[batch_no]['total_pallet'] += record.pallet_count or 0
            batch_stats[batch_no]['total_package'] += record.package_count or 0

        # 记录统计信息
        current_app.logger.info(f"查询到的出境计划单记录数: {len(records)}")
        current_app.logger.info(f"不同批次号数量: {len(batch_nos)}")

        # 按批次分页处理
        batch_list = list(batch_nos)
        total_batches = len(batch_list)

        # 计算分页
        start_idx = (page - 1) * per_page
        end_idx = start_idx + per_page
        paginated_batch_nos = batch_list[start_idx:end_idx]

        # 只保留当前页面的批次记录
        paginated_records = [record for record in records if record.batch_no in paginated_batch_nos or (not record.batch_no and f"TEMP_{record.id}" in paginated_batch_nos)]

        # 重新计算当前页面的统计信息
        paginated_batch_stats = {}
        for batch_no in paginated_batch_nos:
            if batch_no in batch_stats:
                paginated_batch_stats[batch_no] = batch_stats[batch_no]

        # 创建分页对象
        class BatchPagination:
            def __init__(self, items, page, per_page, total):
                self.items = items
                self.page = page
                self.per_page = per_page
                self.total = total
                self.pages = (total + per_page - 1) // per_page
                self.has_prev = page > 1
                self.has_next = page < self.pages
                self.prev_num = page - 1 if self.has_prev else None
                self.next_num = page + 1 if self.has_next else None

        batch_pagination = BatchPagination(paginated_batch_nos, page, per_page, total_batches)

        # 查询收货人信息，用于根据目的地显示联系人和地址
        from app.models import Receiver
        receivers = Receiver.query.all()
        receivers_dict = {receiver.warehouse_name: receiver for receiver in receivers}

        return render_template('outbound_exit_plan.html',
                            batch_records=paginated_records,
                            batch_groups=paginated_batch_stats,
                            batch_pagination=batch_pagination,
                            title='出境计划单',
                            search_params=search_params,
                            show_warning=show_warning,
                            max_records=max_records,
                            receivers=receivers_dict)
    except Exception as e:
        current_app.logger.error(f"渲染出境计划单时出错: {str(e)}")
        return f"渲染出境计划单时出错: {str(e)}", 500


# 添加API路由用于测试
@bp.route('/api/outbound/print/<int:record_id>', methods=['GET'])
@csrf.exempt
def api_outbound_print(record_id):
    """出库单打印API"""
    try:
        record = OutboundRecord.query.get(record_id)
        if not record:
            return jsonify({'success': False, 'message': '找不到出库记录'}), 404

        return jsonify({
            'success': True,
            'message': '出库单打印成功',
            'data': {
                'record_id': record_id,
                'customer_name': record.customer_name,
                'identification_code': record.identification_code
            }
        })
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@bp.route('/api/labels/print/<int:record_id>', methods=['GET'])
@csrf.exempt
def api_labels_print(record_id):
    """标签打印API"""
    try:
        # 可以是入库记录或出库记录
        inbound_record = InboundRecord.query.get(record_id)
        outbound_record = OutboundRecord.query.get(record_id)

        if not inbound_record and not outbound_record:
            return jsonify({'success': False, 'message': '找不到记录'}), 404

        record = inbound_record or outbound_record

        return jsonify({
            'success': True,
            'message': '标签打印成功',
            'data': {
                'record_id': record_id,
                'customer_name': record.customer_name,
                'identification_code': record.identification_code
            }
        })
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@bp.route('/api/outbound/records', methods=['GET'])
@csrf.exempt
def api_outbound_records():
    """获取出库记录列表API"""
    try:
        # 获取当前用户的仓库出库记录
        if current_user.warehouse_id:
            records = OutboundRecord.query.filter_by(
                operated_warehouse_id=current_user.warehouse_id
            ).order_by(OutboundRecord.outbound_time.desc()).limit(10).all()
        else:
            records = OutboundRecord.query.order_by(
                OutboundRecord.outbound_time.desc()
            ).limit(10).all()

        data = []
        for record in records:
            data.append({
                'id': record.id,
                'customer_name': record.customer_name,
                'identification_code': record.identification_code,
                'outbound_time': record.outbound_time.strftime('%Y-%m-%d') if record.outbound_time else '',
                'pallet_count': record.pallet_count,
                'package_count': record.package_count
            })

        return jsonify({
            'success': True,
            'data': data,
            'message': f'获取到{len(data)}条出库记录'
        })

    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@bp.route('/api/inbound/records', methods=['GET'])
@csrf.exempt
def api_inbound_records():
    """获取入库记录列表API"""
    try:
        # 获取当前用户的仓库入库记录
        if current_user.warehouse_id:
            records = InboundRecord.query.filter_by(
                operated_warehouse_id=current_user.warehouse_id
            ).order_by(InboundRecord.inbound_time.desc()).limit(10).all()
        else:
            records = InboundRecord.query.order_by(
                InboundRecord.inbound_time.desc()
            ).limit(10).all()

        data = []
        for record in records:
            data.append({
                'id': record.id,
                'customer_name': record.customer_name,
                'identification_code': record.identification_code,
                'inbound_time': record.outbound_time.strftime('%Y-%m-%d') if record.outbound_time else '',
                'pallet_count': record.pallet_count,
                'package_count': record.package_count
            })

        return jsonify({
            'success': True,
            'data': data,
            'message': f'获取到{len(data)}条入库记录'
        })

    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@bp.route('/api/inventory/transit', methods=['GET'])
@csrf.exempt
def api_transit_inventory():
    """在途货物库存API"""
    try:
        # 查找已出库但未接收的货物（在途货物）
        # 前端仓出库到后端仓，但后端仓还未接收的货物
        transit_goods = []

        # 获取前端仓库的出库记录，目标是后端仓
        frontend_warehouses = Warehouse.query.filter_by(warehouse_type='frontend').all()
        backend_warehouse = Warehouse.query.filter_by(warehouse_type='backend').first()

        if backend_warehouse:
            for frontend_warehouse in frontend_warehouses:
                # 查找从前端仓发往后端仓的出库记录
                outbound_records = OutboundRecord.query.filter_by(
                    operated_warehouse_id=frontend_warehouse.id
                ).filter(
                    OutboundRecord.remark1.like('%后端仓%')
                ).all()

                for record in outbound_records:
                    # 检查是否已被后端仓接收
                    received = InboundRecord.query.filter_by(
                        identification_code=record.identification_code,
                        operated_warehouse_id=backend_warehouse.id
                    ).first()

                    if not received:
                        # 这是在途货物
                        transit_goods.append({
                            'id': record.id,
                            'identification_code': record.identification_code,
                            'customer_name': record.customer_name,
                            'pallet_count': record.pallet_count,
                            'package_count': record.package_count,
                            'weight': record.weight,
                            'volume': record.volume,
                            'outbound_time': record.outbound_time.strftime('%Y-%m-%d') if record.outbound_time else '',
                            'from_warehouse': frontend_warehouse.warehouse_name,
                            'to_warehouse': backend_warehouse.warehouse_name,
                            'status': 'in_transit',
                            'batch_no': record.batch_no,
                            'plate_number': record.plate_number
                        })

        return jsonify({
            'success': True,
            'message': f'成功获取在途货物，共{len(transit_goods)}条记录',
            'data': transit_goods
        })

    except Exception as e:
        current_app.logger.error(f"获取在途货物失败: {str(e)}")
        return jsonify({'success': False, 'message': str(e)}), 500

@bp.route('/outbound/print_exit_plan', methods=['GET', 'POST'])
@csrf.exempt
def print_exit_plan():
    """打印出境计划单"""
    # 支持GET和POST两种方式获取记录ID
    if request.method == 'POST':
        record_ids = request.form.getlist('record_ids[]')
    else:  # GET请求
        record_ids = request.args.getlist('record_ids[]')

    current_app.logger.info(f"选中的出境计划单记录ID: {record_ids}")

    if not record_ids:
        flash('请至少选择一条记录进行打印', 'warning')
        return redirect(url_for('main.outbound_exit_plan'))

    # 查询选中的记录
    records = OutboundRecord.query.filter(OutboundRecord.id.in_(record_ids)).all()
    current_app.logger.info(f"查询到的出境计划单记录数: {len(records)}")

    # 自动填充缺失的字段信息
    for record in records:
        # 通过识别编码查找对应的入库记录，填充入库车牌信息
        if record.identification_code and not record.inbound_plate:
            inbound_record = InboundRecord.query.filter_by(
                identification_code=record.identification_code
            ).first()
            if inbound_record and inbound_record.plate_number:
                record.inbound_plate = inbound_record.plate_number
                current_app.logger.debug(f"为出境计划单打印记录 {record.id} 设置入库车牌: {inbound_record.plate_number}")

        # 如果送货干线车为空，使用出库车牌
        if record.plate_number and not record.delivery_plate_number:
            record.delivery_plate_number = record.plate_number
            current_app.logger.debug(f"为出境计划单打印记录 {record.id} 设置送货干线车: {record.plate_number}")

        # 填充documents字段，如果为空但document_count有值，则使用document_count的值
        if not record.documents and record.document_count is not None:
            record.documents = str(record.document_count)
            current_app.logger.debug(f"为出境计划单打印记录 {record.id} 设置单据字段: {record.documents}")

        # 通过识别编码查找对应的库存记录，填充出境模式、报关行、订单类型信息
        if record.identification_code and (not record.export_mode or not record.customs_broker or not record.order_type):
            inventory_record = Inventory.query.filter_by(
                identification_code=record.identification_code
            ).first()
            if inventory_record:
                if not record.export_mode and inventory_record.export_mode:
                    record.export_mode = inventory_record.export_mode
                    current_app.logger.debug(f"为出境计划单打印记录 {record.id} 设置出境模式: {inventory_record.export_mode}")
                if not record.customs_broker and inventory_record.customs_broker:
                    record.customs_broker = inventory_record.customs_broker
                    current_app.logger.debug(f"为出境计划单打印记录 {record.id} 设置报关行: {inventory_record.customs_broker}")
                if not record.order_type and inventory_record.order_type:
                    record.order_type = inventory_record.order_type
                    current_app.logger.debug(f"为出境计划单打印记录 {record.id} 设置订单类型: {inventory_record.order_type}")

    # 按批次号分组
    batch_groups = {}
    for record in records:
        batch_no = record.batch_no or f"TEMP_{record.id}"
        if batch_no not in batch_groups:
            batch_groups[batch_no] = []
        batch_groups[batch_no].append(record)

    # 获取出库记录的仓库信息作为始发仓
    source_warehouse = None
    if records:
        # 从第一条出库记录获取仓库ID
        first_record = records[0]
        if hasattr(first_record, 'operated_warehouse_id') and first_record.operated_warehouse_id:
            from app.models import Warehouse
            source_warehouse = Warehouse.query.get(first_record.operated_warehouse_id)
        elif current_user.warehouse_id:
            # 如果出库记录没有仓库ID，则使用当前用户的仓库作为备选
            from app.models import Warehouse
            source_warehouse = Warehouse.query.get(current_user.warehouse_id)

    # 查询收货人信息，用于根据目的地显示联系人和地址
    from app.models import Receiver
    receivers = Receiver.query.all()
    receivers_dict = {receiver.warehouse_name: receiver for receiver in receivers}

    # 渲染出境计划单打印模板
    return render_template('outbound_exit_plan_print.html',
                          batch_groups=batch_groups,
                          records=records,
                          source_warehouse=source_warehouse,
                          receivers=receivers_dict,
                          now=datetime.now())


# ==================== 前端仓和后端仓分离路由 ====================

def check_warehouse_permission(warehouse_type, operation_type='view'):
    """检查仓库权限的辅助函数"""
    # 记录调试信息
    current_app.logger.info(f"检查仓库权限: warehouse_type={warehouse_type}, operation_type={operation_type}")

    if not hasattr(current_user, 'is_authenticated') or not current_user.is_authenticated:
        current_app.logger.warning("用户未认证")
        return False

    current_app.logger.info(f"用户已认证: {current_user.username}")

    # 如果用户有系统管理员角色，允许所有操作
    if hasattr(current_user, 'has_role') and current_user.has_role('SYSTEM_ADMIN'):
        current_app.logger.info("用户是系统管理员")
        return True

    # 如果用户有系统管理员权限，允许所有操作
    if hasattr(current_user, 'has_permission') and current_user.has_permission('SYSTEM_CONFIG'):
        current_app.logger.info("用户有系统配置权限")
        return True

    # 检查用户所属仓库类型
    if hasattr(current_user, 'warehouse') and current_user.warehouse:
        user_warehouse_type = current_user.warehouse.warehouse_type
        current_app.logger.info(f"用户仓库类型: {user_warehouse_type}")

        # 操作权限：前端仓用户可以操作所有前端仓数据，后端仓用户只能操作后端仓数据
        if operation_type in ['create', 'edit', 'delete', 'add']:
            if user_warehouse_type == 'frontend' and warehouse_type == 'frontend':
                current_app.logger.info("前端仓用户操作前端仓数据 - 允许")
                return True  # 前端仓用户可以操作所有前端仓数据
            elif user_warehouse_type == 'backend' and warehouse_type == 'backend':
                current_app.logger.info("后端仓用户操作后端仓数据 - 允许")
                return True  # 后端仓用户可以操作后端仓数据
            else:
                current_app.logger.warning(f"不同类型仓库操作被拒绝: {user_warehouse_type} -> {warehouse_type}")
                return False  # 不同类型仓库之间不能操作

        # 查看权限：可以查看所有仓库的数据
        if operation_type == 'view':
            current_app.logger.info("查看权限 - 允许")
            return True
    else:
        current_app.logger.warning(f"用户没有仓库信息: warehouse={getattr(current_user, 'warehouse', None)}")

    current_app.logger.info("默认允许所有操作")
    return True  # 临时允许所有操作，用于测试


# ==================== 前端仓入库路由 ====================

# 旧的入库操作路由已移除，使用整合路由 frontend_inbound_integrated


# ==================== 前端仓接收路由 ====================

@bp.route('/frontend/receive', methods=['GET', 'POST'])
@require_permission('INBOUND_VIEW')
def frontend_receive():
    """前端仓接收操作页面（原版）"""
    if not check_warehouse_permission('frontend', 'view'):
        flash('您没有权限访问前端仓接收功能', 'error')
        return redirect(url_for('main.index'))

    return render_template('frontend/receive_batch.html',
                         title='前端仓接收操作',
                         warehouse_type='frontend')



# 获取前端仓待接收批次数据API
@bp.route('/api/frontend/pending-receive-batches', methods=['GET'])
@csrf.exempt
def api_frontend_pending_receive_batches():
    """获取前端仓待接收的后端仓出库数据（按批次分组）"""
    try:
        # 获取当前用户的仓库信息
        if not current_user.is_authenticated:
            current_app.logger.error(f"用户未认证，current_user.is_authenticated: {current_user.is_authenticated}")
            return jsonify({'success': False, 'message': '用户未登录'}), 401

        # 获取用户仓库信息（支持超级管理员）
        current_warehouse, error_msg = _get_user_warehouse_info('frontend')
        if not current_warehouse:
            current_app.logger.error(f"获取仓库信息失败: {error_msg}")
            return jsonify({'success': False, 'message': error_msg}), 400

        current_app.logger.info(f"当前用户: {current_user.username}, 当前仓库: {current_warehouse.warehouse_name}, 仓库ID: {current_warehouse.id}")

        # 获取后端仓库的出库记录，这些是待前端仓接收的货物
        backend_warehouses = Warehouse.query.filter_by(warehouse_type='backend').all()
        backend_warehouse_ids = [w.id for w in backend_warehouses]

        if not backend_warehouse_ids:
            return jsonify({'success': True, 'batches': []})

        # 查询后端仓库的出库记录，筛选目的仓库为当前仓库的记录
        outbound_records = OutboundRecord.query.options(
            db.joinedload(OutboundRecord.operated_warehouse),
            db.joinedload(OutboundRecord.destination_warehouse)
        ).filter(
            OutboundRecord.operated_warehouse_id.in_(backend_warehouse_ids),
            OutboundRecord.destination_warehouse_id == current_warehouse.id
        ).order_by(OutboundRecord.outbound_time.desc()).all()

        current_app.logger.info(f"查询条件: 后端仓库IDs={backend_warehouse_ids}, 目的仓库ID={current_warehouse.id}")
        current_app.logger.info(f"查询到的出库记录数: {len(outbound_records)}")

        # 过滤掉已接收的记录（通过检查是否存在对应的入库记录）- 优化查询
        received_outbound_ids = set()

        # 批量查询已接收的记录，避免N+1查询问题
        if outbound_records:
            identification_codes = [record.identification_code for record in outbound_records if record.identification_code]
            batch_nos = [record.batch_no for record in outbound_records if record.batch_no]

            if identification_codes and batch_nos:
                # 一次性查询所有可能已接收的记录
                existing_inbounds = InboundRecord.query.filter(
                    InboundRecord.identification_code.in_(identification_codes),
                    InboundRecord.operated_warehouse_id == current_warehouse.id,
                    InboundRecord.batch_no.in_(batch_nos),
                    InboundRecord.record_type == 'receive'  # 确保是接收记录
                ).with_entities(
                    InboundRecord.identification_code,
                    InboundRecord.batch_no,
                    InboundRecord.id
                ).all()

                # 建立已接收记录的映射
                received_mapping = set()
                for inbound_code, inbound_batch, inbound_id in existing_inbounds:
                    received_mapping.add((inbound_code, inbound_batch))

                # 检查每个出库记录是否已被接收
                for record in outbound_records:
                    if (record.identification_code, record.batch_no) in received_mapping:
                        received_outbound_ids.add(record.id)
                        current_app.logger.debug(f"出库记录 {record.id} (识别编码: {record.identification_code}) 已被接收")
                    else:
                        current_app.logger.debug(f"出库记录 {record.id} (识别编码: {record.identification_code}) 未被接收，将显示在待接收列表中")

        # 过滤掉已接收的记录
        outbound_records = [record for record in outbound_records if record.id not in received_outbound_ids]

        current_app.logger.info(f"过滤前记录数: {len(outbound_records) + len(received_outbound_ids)}, 已接收: {len(received_outbound_ids)}, 待接收: {len(outbound_records)}")

        # 记录待接收的识别编码，便于调试
        pending_codes = [record.identification_code for record in outbound_records]
        current_app.logger.info(f"待接收的识别编码: {pending_codes}")

        # 按批次号/送货干线车/发货仓库/接收仓库分组
        batch_groups = {}
        for record in outbound_records:
            # 构建分组键：批次号 + 送货干线车 + 发货仓库 + 接收仓库
            # 使用不包含下划线的占位符，避免分割时出现问题
            group_key = f"{record.batch_no or 'NOBATCH'}_{record.delivery_plate_number or 'NOPLATE'}_{record.operated_warehouse_id}_{record.destination_warehouse_id}"

            if group_key not in batch_groups:
                batch_groups[group_key] = {
                    'batch_no': record.batch_no or '',
                    'delivery_plate_number': record.delivery_plate_number or '',
                    'source_warehouse': record.operated_warehouse.warehouse_name if record.operated_warehouse else '',
                    'destination_warehouse': record.destination_warehouse.warehouse_name if record.destination_warehouse else '',
                    'source_warehouse_id': record.operated_warehouse_id,
                    'destination_warehouse_id': record.destination_warehouse_id,
                    'outbound_time': record.outbound_time.strftime('%Y-%m-%d') if record.outbound_time else '',
                    'items': [],
                    'total_pallet_count': 0,
                    'total_package_count': 0,
                    'total_weight': 0,
                    'total_volume': 0,
                    'item_count': 0
                }

            # 添加货物明细到分组
            batch_sequence_display = ''
            if record.batch_sequence and record.batch_total:
                batch_sequence_display = f"{record.batch_sequence}/{record.batch_total}"

            item = {
                'id': record.id,
                'outbound_time': record.outbound_time.strftime('%Y-%m-%d') if record.outbound_time else '',
                'batch_sequence_display': batch_sequence_display,
                'plate_number': record.plate_number or '',
                'customer_name': record.customer_name or '',
                'identification_code': record.identification_code or '',
                'order_type': record.order_type or '',
                'customs_broker': record.customs_broker or '',
                'pallet_count': record.pallet_count or 0,
                'package_count': record.package_count or 0,
                'weight': record.weight or 0,
                'volume': record.volume or 0,
                'documents': record.documents or '',
                'service_staff': record.service_staff or '',
                'remark1': record.remark1 or '',
                'remark2': record.remark2 or '',
                'batch_sequence': record.batch_sequence or 0,
                'batch_total': record.batch_total or 0
            }

            batch_groups[group_key]['items'].append(item)

            # 累计统计数据
            batch_groups[group_key]['total_pallet_count'] += item['pallet_count']
            batch_groups[group_key]['total_package_count'] += item['package_count']
            batch_groups[group_key]['total_weight'] += item['weight']
            batch_groups[group_key]['total_volume'] += item['volume']
            batch_groups[group_key]['item_count'] += 1

        # 转换为列表格式
        data = list(batch_groups.values())

        # 按出库时间倒序排列
        data.sort(key=lambda x: x['outbound_time'], reverse=True)

        return jsonify({'success': True, 'batches': data})

    except Exception as e:
        current_app.logger.error(f"获取前端仓待接收数据失败: {str(e)}")
        return jsonify({'success': False, 'message': f'获取数据失败: {str(e)}'}), 500


# 前端仓单个货物接收API
@bp.route('/api/frontend/receive-item', methods=['POST'])
@csrf.exempt
def api_frontend_receive_item():
    """前端仓接收单个货物"""
    try:
        data = request.get_json()
        item_id = data.get('item_id')
        receive_time_str = data.get('receive_time')

        if not item_id:
            return jsonify({'success': False, 'message': '缺少货物ID'}), 400

        # 查找出库记录
        outbound_record = OutboundRecord.query.get(item_id)
        if not outbound_record:
            return jsonify({'success': False, 'message': '未找到出库记录'}), 404

        # 解析接收时间
        receive_time = datetime.now()
        if receive_time_str:
            try:
                receive_time = datetime.fromisoformat(receive_time_str.replace('Z', '+00:00'))
            except:
                pass

        # 获取当前用户的仓库信息
        if not current_user.is_authenticated:
            return jsonify({'success': False, 'message': '用户未登录'}), 401

        # 检查仓库访问权限（支持超级管理员）
        has_access, error_msg = _check_warehouse_access()
        if not has_access:
            return jsonify({'success': False, 'message': error_msg}), 400

        # 创建接收记录
        # 确保shipping_warehouse字段正确设置
        shipping_warehouse_name = ''
        if outbound_record.operated_warehouse and hasattr(outbound_record.operated_warehouse, 'warehouse_name'):
            shipping_warehouse_name = outbound_record.operated_warehouse.warehouse_name
        elif outbound_record.operated_warehouse_id:
            # 如果关联对象为空，尝试直接查询
            warehouse = Warehouse.query.get(outbound_record.operated_warehouse_id)
            if warehouse:
                shipping_warehouse_name = warehouse.warehouse_name

        receive_record = ReceiveRecord(
            receive_time=receive_time,
            batch_no=outbound_record.batch_no,
            shipping_warehouse=shipping_warehouse_name,
            customer_name=outbound_record.customer_name,
            identification_code=outbound_record.identification_code,  # 添加识别编码
            pallet_count=outbound_record.pallet_count,
            package_count=outbound_record.package_count,
            weight=outbound_record.weight,
            volume=outbound_record.volume,
            shipping_time=outbound_record.outbound_time,
            documents=str(outbound_record.document_count) if outbound_record.document_count and outbound_record.document_count > 0 else '',
            service_staff=outbound_record.service_staff,
            delivery_plate_number=outbound_record.plate_number,  # 送货干线车（前端仓→后端仓）
            inbound_plate=outbound_record.inbound_plate,  # 入库车牌（工厂→前端仓）
            # 添加缺失的业务字段
            order_type=outbound_record.order_type,  # 订单类型
            customs_broker=outbound_record.customs_broker,  # 报关行
            export_mode=outbound_record.export_mode,  # 出境模式
            storage_location=outbound_record.location,  # 库位（OutboundRecord中是location字段）
            batch_sequence=outbound_record.batch_sequence,  # 批次序号
            batch_total=outbound_record.batch_total,  # 批次总数
            remark1=outbound_record.remark1,  # 备注1
            remark2=outbound_record.remark2,  # 备注2
            operated_warehouse_id=user_warehouse.id,
            created_by=current_user.id
        )

        db.session.add(receive_record)
        db.session.commit()

        return jsonify({'success': True, 'message': '接收成功'})

    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"前端仓货物接收失败: {str(e)}")
        return jsonify({'success': False, 'message': f'接收失败: {str(e)}'}), 500


# 获取批次明细API
@bp.route('/api/frontend/batch-details/<batch_key>', methods=['GET'])
@csrf.exempt
def api_frontend_batch_details(batch_key):
    """获取指定批次的货物明细"""
    try:
        current_app.logger.info(f"接收到批次键: {batch_key}")

        # 解析批次键 - 处理包含下划线的占位符
        current_app.logger.info(f"原始批次键: {batch_key}")

        # 先处理已知的占位符模式
        batch_key_normalized = batch_key.replace('NO_PLATE', 'NOPLATE').replace('NO_BATCH', 'NOBATCH')
        current_app.logger.info(f"标准化后的批次键: {batch_key_normalized}")

        parts = batch_key_normalized.split('_')
        current_app.logger.info(f"批次键分割结果: {parts}, 长度: {len(parts)}")

        if len(parts) < 4:
            current_app.logger.error(f"批次键格式错误，期望至少4个部分，实际: {len(parts)}")
            return jsonify({'success': False, 'message': f'批次键格式错误，期望至少4个部分，实际: {len(parts)}'}), 400

        batch_no = parts[0] if parts[0] != 'NOBATCH' else None
        delivery_plate = parts[1] if parts[1] != 'NOPLATE' else None

        # 安全地转换仓库ID
        try:
            source_warehouse_id = int(parts[2])
            dest_warehouse_id = int(parts[3])
            current_app.logger.info(f"解析结果 - 批次号: {batch_no}, 送货干线车: {delivery_plate}, 源仓库ID: {source_warehouse_id}, 目标仓库ID: {dest_warehouse_id}")
        except ValueError as ve:
            current_app.logger.error(f"仓库ID格式错误: {ve}")
            return jsonify({'success': False, 'message': f'仓库ID格式错误: {ve}'}), 400

        # 获取当前用户的仓库信息
        current_warehouse, error_msg = _get_user_warehouse_info('frontend')
        if not current_warehouse:
            current_app.logger.error(f"获取用户仓库信息失败: {error_msg}")
            return jsonify({'success': False, 'message': error_msg}), 400

        current_app.logger.info(f"当前用户仓库: {current_warehouse.warehouse_name} (ID: {current_warehouse.id})")

        # 查询该批次的所有货物
        query = OutboundRecord.query.options(
            db.joinedload(OutboundRecord.operated_warehouse),
            db.joinedload(OutboundRecord.destination_warehouse)
        ).filter(
            OutboundRecord.operated_warehouse_id == source_warehouse_id,
            OutboundRecord.destination_warehouse_id == dest_warehouse_id
        )

        if batch_no:
            query = query.filter(OutboundRecord.batch_no == batch_no)
            current_app.logger.info(f"添加批次号筛选: {batch_no}")
        if delivery_plate:
            query = query.filter(OutboundRecord.delivery_plate_number == delivery_plate)
            current_app.logger.info(f"添加送货干线车筛选: {delivery_plate}")

        records = query.order_by(OutboundRecord.batch_sequence).all()
        current_app.logger.info(f"查询到 {len(records)} 条记录")

        # 转换为明细格式
        items = []
        for record in records:
            batch_sequence_display = ''
            if record.batch_sequence and record.batch_total:
                batch_sequence_display = f"{record.batch_sequence}/{record.batch_total}"

            items.append({
                'id': record.id,
                'batch_sequence_display': batch_sequence_display,
                'customer_name': record.customer_name or '',
                'identification_code': record.identification_code or '',
                'order_type': record.order_type or '',
                'customs_broker': record.customs_broker or '',
                'outbound_time': record.outbound_time.strftime('%Y-%m-%d') if record.outbound_time else '',
                'pallet_count': record.pallet_count or 0,
                'package_count': record.package_count or 0,
                'weight': record.weight or 0,
                'volume': record.volume or 0,
                'documents': record.documents or '',
                'service_staff': record.service_staff or '',
                'remark1': record.remark1 or '',
                'remark2': record.remark2 or '',
                # 接收时可编辑的字段
                'received_pallet_count': record.pallet_count or 0,
                'received_package_count': record.package_count or 0,
                'receive_notes': ''
            })

        current_app.logger.info(f"成功返回 {len(items)} 个货物明细")
        return jsonify({'success': True, 'items': items})

    except Exception as e:
        current_app.logger.error(f"获取批次明细失败: {str(e)}", exc_info=True)
        return jsonify({'success': False, 'message': f'获取明细失败: {str(e)}'}), 500


# 前端仓批次接收API
@bp.route('/api/frontend/batch-receive', methods=['POST'])
@csrf.exempt
def api_frontend_batch_receive():
    """前端仓批次接收货物，支持差异录入"""
    try:
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'message': '请求数据格式错误'}), 400

        batch_no = data.get('batch_no')
        receive_time = data.get('receive_time')
        items = data.get('items', [])

        if not batch_no or not receive_time or not items:
            return jsonify({'success': False, 'message': '批次号、接收时间和货物明细为必填项'}), 400

        # 转换时间格式
        try:
            receive_datetime = datetime.strptime(receive_time, '%Y-%m-%dT%H:%M')
        except ValueError:
            return jsonify({'success': False, 'message': '接收时间格式错误'}), 400

        # 获取当前用户的前端仓库
        user_warehouse, error_msg = _get_user_warehouse_info('frontend')
        if not user_warehouse:
            return jsonify({'success': False, 'message': error_msg}), 400

        # 为每个货物创建接收记录
        for item in items:
            outbound_id = item.get('id')
            received_pallet_count = item.get('received_pallet_count', 0)
            received_package_count = item.get('received_package_count', 0)
            receive_notes = item.get('receive_notes', '')

            # 查找原始出库记录
            outbound_record = OutboundRecord.query.get(outbound_id)
            if not outbound_record:
                db.session.rollback()
                return jsonify({'success': False, 'message': f'未找到出库记录ID: {outbound_id}'}), 400

            # 检查数量差异并生成异常备注
            original_pallet = outbound_record.pallet_count or 0
            original_package = outbound_record.package_count or 0

            discrepancy_notes = []
            if received_pallet_count != original_pallet:
                discrepancy_notes.append(f"板数差异：发出{original_pallet}，接收{received_pallet_count}")
            if received_package_count != original_package:
                discrepancy_notes.append(f"件数差异：发出{original_package}，接收{received_package_count}")

            # 构建最终备注
            final_remark = receive_notes
            if discrepancy_notes:
                discrepancy_text = "；".join(discrepancy_notes)
                if final_remark:
                    final_remark = f"{final_remark}；{discrepancy_text}"
                else:
                    final_remark = discrepancy_text

            # 创建接收记录（使用InboundRecord表）
            receive_record = InboundRecord(
                inbound_time=receive_datetime,
                plate_number=outbound_record.plate_number,
                delivery_plate_number=outbound_record.delivery_plate_number,  # 送货干线车
                customer_name=outbound_record.customer_name,
                identification_code=outbound_record.identification_code,
                pallet_count=received_pallet_count,
                package_count=received_package_count,
                weight=outbound_record.weight,  # 重量保持原值
                volume=outbound_record.volume,  # 体积保持原值
                export_mode=outbound_record.export_mode,
                order_type=outbound_record.order_type,
                customs_broker=outbound_record.customs_broker,
                documents=outbound_record.documents,
                service_staff=outbound_record.service_staff,
                batch_no=batch_no,
                batch_total=outbound_record.batch_total,
                batch_sequence=outbound_record.batch_sequence,
                remark1=outbound_record.remark1,
                remark2=final_remark,  # 包含差异信息的备注
                record_type='receive',  # 接收记录
                operated_by_user_id=current_user.id,
                operated_warehouse_id=user_warehouse.id
            )

            db.session.add(receive_record)

            # 创建或更新库存记录
            existing_inventory = Inventory.query.filter_by(
                customer_name=outbound_record.customer_name,
                identification_code=outbound_record.identification_code,
                operated_warehouse_id=user_warehouse.id
            ).first()

            if existing_inventory:
                # 更新现有库存
                existing_inventory.pallet_count = (existing_inventory.pallet_count or 0) + received_pallet_count
                existing_inventory.package_count = (existing_inventory.package_count or 0) + received_package_count
                existing_inventory.last_updated = receive_datetime
                existing_inventory.version += 1
            else:
                # 创建新库存记录
                inventory_record = Inventory(
                    customer_name=outbound_record.customer_name,
                    identification_code=outbound_record.identification_code,
                    inbound_pallet_count=received_pallet_count,
                    inbound_package_count=received_package_count,
                    pallet_count=received_pallet_count,
                    package_count=received_package_count,
                    weight=outbound_record.weight,
                    volume=outbound_record.volume,
                    export_mode=outbound_record.export_mode,
                    order_type=outbound_record.order_type,
                    customs_broker=outbound_record.customs_broker,
                    documents=outbound_record.documents,
                    inbound_time=receive_datetime,
                    plate_number=outbound_record.plate_number,
                    service_staff=outbound_record.service_staff,
                    operated_by_user_id=current_user.id,
                    operated_warehouse_id=user_warehouse.id,
                    last_updated=receive_datetime,
                    version=1
                )
                db.session.add(inventory_record)

        # 提交事务
        db.session.commit()

        return jsonify({'success': True, 'message': f'批次 {batch_no} 接收成功，共处理 {len(items)} 个货物'})

    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"前端仓批次接收失败: {str(e)}")
        return jsonify({'success': False, 'message': f'接收失败: {str(e)}'}), 500


@bp.route('/frontend/inbound', methods=['GET', 'POST'])
@require_permission('INBOUND_VIEW')
def frontend_inbound_integrated():
    """前端仓整合入库操作页面（接收+入库）"""
    if not check_warehouse_permission('frontend', 'view'):
        flash('您没有权限访问前端仓入库功能', 'error')
        return redirect(url_for('main.index'))

    from app.utils import render_ajax_aware
    return render_ajax_aware('frontend/inbound_integrated.html',
                           title='前端仓入库操作',
                           warehouse_type='frontend')


@bp.route('/api/frontend/add-inbound', methods=['POST'])
@require_permission('INBOUND_ADD')
@csrf.exempt  # 豁免CSRF保护，因为这是API接口
def api_frontend_add_inbound():
    """前端仓新增入库API"""
    try:
        if not check_warehouse_permission('frontend', 'add'):
            return jsonify({'success': False, 'message': '您没有权限添加前端仓入库记录'}), 403

        # 获取JSON数据
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'message': '请求数据为空'}), 400

        current_app.logger.info(f"前端仓新增入库API被调用，数据: {data}")

        # 基本验证
        required_fields = ['customer_name', 'customs_broker', 'service_staff']
        for field in required_fields:
            if not data.get(field):
                return jsonify({'success': False, 'message': f'缺少必填字段: {field}'}), 400

        # 验证板数和件数
        pallet_count = int(data.get('pallet_count', 0) or 0)
        package_count = int(data.get('package_count', 0) or 0)

        if pallet_count == 0 and package_count == 0:
            return jsonify({'success': False, 'message': '板数和件数至少需要填写一个'}), 400

        # 获取当前用户的仓库信息
        if not current_user.warehouse or current_user.warehouse.warehouse_type != 'frontend':
            return jsonify({'success': False, 'message': '只有前端仓用户可以使用此功能'}), 403

        warehouse_id = current_user.warehouse_id
        warehouse = current_user.warehouse

        # 生成识别编码
        from app.utils.identification_generator import IdentificationGenerator
        generator = IdentificationGenerator()

        identification_code = generator.generate_identification_code(
            warehouse_prefix=warehouse.warehouse_prefix,
            customer_name=data['customer_name'],
            plate_number=data.get('plate_number', ''),
            inbound_date=datetime.now().date()
        )

        # 处理入库时间
        inbound_time_str = data.get('inbound_time')
        if inbound_time_str:
            inbound_time = datetime.fromisoformat(inbound_time_str.replace('T', ' '))
        else:
            inbound_time = datetime.now()

        # 创建入库记录
        inbound_record = InboundRecord(
            inbound_time=inbound_time,
            plate_number=data.get('plate_number', ''),
            customer_name=data['customer_name'],
            identification_code=identification_code,
            order_type=data.get('order_type', ''),
            export_mode=data.get('export_mode', ''),
            customs_broker=data['customs_broker'],
            service_staff=data['service_staff'],
            pallet_count=pallet_count,
            package_count=package_count,
            weight=float(data.get('weight', 0) or 0),
            volume=float(data.get('volume', 0) or 0),
            documents=int(data.get('documents', 0) or 0),
            remark1=data.get('remark1', ''),
            remark2=data.get('remark2', ''),
            record_type='direct',  # 直接入库
            operated_warehouse_id=warehouse_id,
            operated_by=current_user.id,
            created_at=datetime.now(),
            updated_at=datetime.now()
        )

        db.session.add(inbound_record)

        # 创建对应的库存记录
        inventory = Inventory(
            identification_code=identification_code,
            customer_name=data['customer_name'],
            pallet_count=pallet_count,
            package_count=package_count,
            weight=float(data.get('weight', 0) or 0),
            volume=float(data.get('volume', 0) or 0),
            inbound_time=inbound_time,
            operated_warehouse_id=warehouse_id,
            operated_by=current_user.id,
            last_updated=datetime.now(),
            version=1
        )

        db.session.add(inventory)
        db.session.commit()

        current_app.logger.info(f"前端仓新增入库成功: 用户={current_user.username}, 客户={data['customer_name']}, 识别编码={identification_code}")

        return jsonify({
            'success': True,
            'message': f'成功添加 1 条入库记录',
            'identification_code': identification_code,
            'inbound_id': inbound_record.id
        })

    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"前端仓新增入库失败: {str(e)}")
        return jsonify({'success': False, 'message': f'保存失败: {str(e)}'}), 500


@bp.route('/test-messages')
def test_messages():
    """消息系统测试页面"""
    return render_template('message_test.html', title='消息系统测试')


@bp.route('/admin/performance')
@require_permission('ADMIN')
def performance_dashboard():
    """性能监控面板"""
    try:
        from app.runtime_performance_manager import runtime_performance_manager
        summary = runtime_performance_manager.get_performance_summary()

        # 获取缓存统计
        cache_hits = runtime_performance_manager.metrics['cache_hits']
        cache_misses = runtime_performance_manager.metrics['cache_misses']
        cache_total = cache_hits + cache_misses
        cache_hit_rate = (cache_hits / cache_total * 100) if cache_total > 0 else 0

        # 构建快速状态数据（适配现有模板）
        quick_status = {
            'cache_hit_rate': round(cache_hit_rate, 1),
            'avg_query_time': summary.get('avg_response_time', 0) if summary.get('status') != 'no_data' else 0,
            'slow_queries_count': summary.get('slow_requests_count', 0) if summary.get('status') != 'no_data' else 0,
            'system_memory_percent': 0  # 暂时设为0，可以后续添加系统内存监控
        }

        from app.utils import render_ajax_aware
        return render_ajax_aware('admin/performance_dashboard.html',
                               title='性能监控面板',
                               quick_status=quick_status)
    except Exception as e:
        current_app.logger.error(f'获取性能数据失败: {str(e)}')
        flash(f'获取性能数据失败: {str(e)}', 'error')
        return redirect(url_for('main.index'))








@bp.route('/test/contact-window')
@login_required
def test_contact_window():
    """测试联络窗口功能页面"""
    return render_template('test_contact_window.html', title='测试联络窗口功能')

@bp.route('/test/receive/data')
def test_receive_data():
    """测试接收记录数据 - 无需权限"""
    from app.models import ReceiveRecord, Warehouse

    frontend_warehouses = Warehouse.query.filter_by(warehouse_type='frontend').all()
    if not frontend_warehouses:
        return jsonify({'error': '没有前端仓'})

    frontend_warehouse_ids = [w.id for w in frontend_warehouses]
    records = ReceiveRecord.query.filter(
        ReceiveRecord.operated_warehouse_id.in_(frontend_warehouse_ids)
    ).limit(10).all()

    result = []
    for record in records:
        result.append({
            'id': record.id,
            'batch_no': record.batch_no,
            'customer_name': record.customer_name,
            'identification_code': record.identification_code,
            'inbound_plate': record.inbound_plate,
            'delivery_plate_number': record.delivery_plate_number,
            'warehouse': record.operated_warehouse.warehouse_name if record.operated_warehouse else '无'
        })

    return jsonify({'records': result, 'count': len(result), 'message': '入库车牌数据修复成功！'})

@bp.route('/frontend/receive/list')
@require_permission('INBOUND_VIEW')
def frontend_receive_list():
    """前端仓接收记录列表"""
    if not check_warehouse_permission('frontend', 'view'):
        flash('您没有权限访问前端仓接收记录', 'error')
        return redirect(url_for('main.index'))

    # 获取搜索参数
    page = request.args.get('page', 1, type=int)
    per_page = 20  # 每页显示20条记录

    # 构建查询，显示前端仓的接收记录（使用InboundRecord模型，只显示接收类型）
    query = InboundRecord.query.options(
        db.joinedload(InboundRecord.operated_warehouse),
        db.joinedload(InboundRecord.operated_by_user)
    ).filter(
        InboundRecord.record_type == 'receive'  # 只显示接收记录
    )

    # 首先过滤出前端仓的数据
    frontend_warehouses = Warehouse.query.filter_by(warehouse_type='frontend').all()
    if not frontend_warehouses:
        # 如果没有前端仓，返回空查询
        query = query.filter(InboundRecord.id == -1)
    else:
        frontend_warehouse_ids = [w.id for w in frontend_warehouses]
        query = query.filter(InboundRecord.operated_warehouse_id.in_(frontend_warehouse_ids))

    # 根据用户权限进一步过滤数据
    # 管理员和后端仓账号可以查看所有前端仓的数据
    # 前端仓账号只能查看自身仓库操作的数据
    if hasattr(current_user, 'warehouse') and current_user.warehouse:
        if current_user.warehouse.warehouse_type == 'frontend' and not current_user.is_admin:
            # 前端仓用户只能查看自己仓库的数据
            query = query.filter(InboundRecord.operated_warehouse_id == current_user.warehouse_id)

    # 获取日期参数，如果没有指定则使用默认值
    date_start = request.args.get('date_start', '')
    date_end = request.args.get('date_end', '')

    # 如果没有指定日期范围，默认使用最近一周的日期范围
    if not date_start and not date_end:
        today = datetime.now().strftime('%Y-%m-%d')
        one_week_ago = (datetime.now() - timedelta(days=7)).strftime('%Y-%m-%d')
        date_start = one_week_ago
        date_end = today

    # 搜索过滤 - 支持新的搜索方式
    search_params = {
        'date_start': date_start,
        'date_end': date_end
    }

    # 获取搜索字段和值
    search_field = request.args.get('search_field', '')
    search_value = request.args.get('search_value', '')
    search_condition = request.args.get('search_condition', 'contains')

    if search_field and search_value:
        search_params['search_field'] = search_field
        search_params['search_value'] = search_value
        search_params['search_condition'] = search_condition

        # 根据搜索条件应用过滤
        if search_condition == 'exact':
            filter_condition = getattr(InboundRecord, search_field) == search_value
        elif search_condition == 'startswith':
            filter_condition = getattr(InboundRecord, search_field).like(f'{search_value}%')
        elif search_condition == 'endswith':
            filter_condition = getattr(InboundRecord, search_field).like(f'%{search_value}')
        else:  # contains
            filter_condition = getattr(InboundRecord, search_field).like(f'%{search_value}%')

        query = query.filter(filter_condition)

    # 兼容旧的搜索参数
    # 客户名称搜索
    if request.args.get('customer_name'):
        query = query.filter(InboundRecord.customer_name.contains(request.args.get('customer_name')))
        search_params['customer_name'] = request.args.get('customer_name')

    # 车牌号搜索（支持入库车牌和送货车牌）
    if request.args.get('plate_number'):
        plate_search = request.args.get('plate_number')
        query = query.filter(
            db.or_(
                InboundRecord.plate_number.contains(plate_search),
                InboundRecord.delivery_plate_number.contains(plate_search)
            )
        )
        search_params['plate_number'] = plate_search

    # 识别编码搜索（兼容原来的batch_no参数）
    if request.args.get('batch_no') or request.args.get('identification_code'):
        search_term = request.args.get('batch_no') or request.args.get('identification_code')
        query = query.filter(InboundRecord.identification_code.contains(search_term))
        search_params['batch_no'] = search_term

    # 订单类型搜索
    if request.args.get('order_type'):
        query = query.filter(InboundRecord.order_type.contains(request.args.get('order_type')))
        search_params['order_type'] = request.args.get('order_type')

    # 出境模式搜索
    if request.args.get('export_mode'):
        query = query.filter(InboundRecord.export_mode.contains(request.args.get('export_mode')))
        search_params['export_mode'] = request.args.get('export_mode')

    # 报关行搜索
    if request.args.get('customs_broker'):
        query = query.filter(InboundRecord.customs_broker.contains(request.args.get('customs_broker')))
        search_params['customs_broker'] = request.args.get('customs_broker')

    # 跟单客服搜索
    if request.args.get('service_staff'):
        query = query.filter(InboundRecord.service_staff.contains(request.args.get('service_staff')))
        search_params['service_staff'] = request.args.get('service_staff')

    # 批次号搜索
    if request.args.get('batch_no_search'):
        query = query.filter(InboundRecord.batch_no.contains(request.args.get('batch_no_search')))
        search_params['batch_no_search'] = request.args.get('batch_no_search')

    # 注意：接收记录列表已经过滤了只显示有批次号的记录，所以不需要额外的入库类型过滤

    # 日期范围搜索
    if date_start:
        try:
            start_date = datetime.strptime(date_start, '%Y-%m-%d')
            query = query.filter(InboundRecord.inbound_time >= start_date)
        except ValueError:
            pass

    if date_end:
        try:
            end_date = datetime.strptime(date_end, '%Y-%m-%d')
            # 设置为当天的23:59:59
            end_date = end_date.replace(hour=23, minute=59, second=59)
            query = query.filter(InboundRecord.inbound_time <= end_date)
        except ValueError:
            pass

    # 获取所有记录，按批次号分组
    all_records = query.order_by(InboundRecord.batch_no.desc(), InboundRecord.inbound_time.desc()).all()

    # 按批次号分组数据
    from collections import defaultdict, OrderedDict
    batch_groups = OrderedDict()

    for record in all_records:
        batch_no = record.batch_no or '未分配批次'
        if batch_no not in batch_groups:
            batch_groups[batch_no] = {
                'batch_info': {
                    'batch_no': batch_no,
                    'total_pallet_count': 0,
                    'total_package_count': 0,
                    'total_weight': 0,
                    'total_volume': 0,
                    'record_count': 0,
                    'first_receive_time': record.outbound_time,
                    'last_receive_time': record.outbound_time
                },
                'records': []
            }

        # 更新批次汇总信息
        batch_info = batch_groups[batch_no]['batch_info']
        batch_info['total_pallet_count'] += record.pallet_count or 0
        batch_info['total_package_count'] += record.package_count or 0
        batch_info['total_weight'] += record.weight or 0
        batch_info['total_volume'] += record.volume or 0
        batch_info['record_count'] += 1

        # 更新时间范围
        if record.outbound_time:
            if not batch_info['first_receive_time'] or record.outbound_time < batch_info['first_receive_time']:
                batch_info['first_receive_time'] = record.outbound_time
            if not batch_info['last_receive_time'] or record.outbound_time > batch_info['last_receive_time']:
                batch_info['last_receive_time'] = record.outbound_time

        # 添加记录到批次组
        batch_groups[batch_no]['records'].append(record)

    # 分页处理（对批次进行分页）
    batch_list = list(batch_groups.items())
    total_batches = len(batch_list)

    # 计算分页
    start_idx = (page - 1) * per_page
    end_idx = start_idx + per_page
    paginated_batches = batch_list[start_idx:end_idx]

    # 创建分页对象
    class BatchPagination:
        def __init__(self, items, page, per_page, total):
            self.items = items
            self.page = page
            self.per_page = per_page
            self.total = total
            self.pages = (total + per_page - 1) // per_page
            self.has_prev = page > 1
            self.has_next = page < self.pages
            self.prev_num = page - 1 if self.has_prev else None
            self.next_num = page + 1 if self.has_next else None

    batch_pagination = BatchPagination(paginated_batches, page, per_page, total_batches)

    from app.utils import render_ajax_aware
    return render_ajax_aware('frontend/receive_list.html',
                           batch_groups=batch_pagination,
                           search_params=search_params,
                           title='前端仓接收记录',
                           warehouse_type='frontend')


@bp.route('/delete_batch_receive', methods=['POST'])
@csrf.exempt
@require_permission('INBOUND_DELETE')
def delete_batch_receive():
    """删除批次的所有接收记录"""
    try:
        data = request.get_json()
        batch_no = data.get('batch_no')

        if not batch_no:
            return jsonify({'success': False, 'message': '批次号不能为空'})

        # 根据请求来源确定仓库类型
        current_warehouse_id = None
        warehouse_type = None

        # 从请求头或参数中获取来源页面信息
        referer = request.headers.get('Referer', '')

        if '/backend/' in referer:
            # 来自后端仓页面的请求
            warehouse_type = 'backend'
            from app.models import Warehouse
            backend_warehouse = Warehouse.query.filter_by(warehouse_type='backend').first()
            if backend_warehouse:
                current_warehouse_id = backend_warehouse.id
        elif '/frontend/' in referer:
            # 来自前端仓页面的请求
            warehouse_type = 'frontend'
            if hasattr(current_user, 'warehouse') and current_user.warehouse:
                current_warehouse_id = current_user.warehouse_id
                warehouse_type = current_user.warehouse.warehouse_type
        else:
            # 默认使用用户的仓库信息
            if hasattr(current_user, 'warehouse') and current_user.warehouse:
                current_warehouse_id = current_user.warehouse_id
                warehouse_type = current_user.warehouse.warehouse_type

        # 如果是超级管理员或者没有仓库信息，需要从请求中获取仓库信息或使用默认逻辑
        if not current_warehouse_id:
            # 查找该批次对应的仓库
            sample_record_inbound = InboundRecord.query.filter(
                InboundRecord.batch_no == batch_no,
                InboundRecord.record_type == 'receive'
            ).first()

            sample_record_receive = ReceiveRecord.query.filter(
                ReceiveRecord.batch_no == batch_no
            ).first()

            if sample_record_inbound and sample_record_inbound.operated_warehouse:
                current_warehouse_id = sample_record_inbound.operated_warehouse_id
                warehouse_type = sample_record_inbound.operated_warehouse.warehouse_type
            elif sample_record_receive and sample_record_receive.operated_warehouse:
                current_warehouse_id = sample_record_receive.operated_warehouse_id
                warehouse_type = sample_record_receive.operated_warehouse.warehouse_type
            else:
                # 如果找不到记录，默认为后端仓操作
                from app.models import Warehouse
                backend_warehouse = Warehouse.query.filter_by(warehouse_type='backend').first()
                if backend_warehouse:
                    current_warehouse_id = backend_warehouse.id
                    warehouse_type = 'backend'

        # 检查权限
        if warehouse_type == 'frontend':
            if not check_warehouse_permission('frontend', 'delete'):
                return jsonify({'success': False, 'message': '您没有权限删除前端仓接收记录'})
        elif warehouse_type == 'backend':
            if not check_warehouse_permission('backend', 'delete'):
                return jsonify({'success': False, 'message': '您没有权限删除后端仓接收记录'})
        else:
            return jsonify({'success': False, 'message': '无法确定仓库类型，请联系管理员'})

        if not current_warehouse_id:
            return jsonify({'success': False, 'message': '无法确定操作仓库，请联系管理员'})

        # 根据仓库类型查找该批次的所有接收记录
        records = []
        if warehouse_type == 'frontend':
            records = InboundRecord.query.filter(
                InboundRecord.batch_no == batch_no,
                InboundRecord.record_type == 'receive',
                InboundRecord.operated_warehouse_id == current_warehouse_id
            ).all()
        elif warehouse_type == 'backend':
            # 后端仓可以删除任何批次的接收记录，不限制仓库ID
            records = ReceiveRecord.query.filter(
                ReceiveRecord.batch_no == batch_no
            ).all()

            current_app.logger.info(f"后端仓查找批次 {batch_no}，找到 {len(records)} 条接收记录")
            for record in records:
                warehouse_name = record.operated_warehouse.warehouse_name if record.operated_warehouse else '未知仓库'
                current_app.logger.info(f"记录ID: {record.id}, 仓库: {warehouse_name} (ID: {record.operated_warehouse_id}), 识别编码: {record.identification_code}")

        current_app.logger.info(f"查找批次 {batch_no} 在仓库 {current_warehouse_id} ({warehouse_type}) 的接收记录，找到 {len(records)} 条记录")
        for record in records:
            if hasattr(record, 'operated_warehouse_id'):
                current_app.logger.info(f"记录ID: {record.id}, 识别编码: {record.identification_code}, 仓库ID: {record.operated_warehouse_id}")
            else:
                current_app.logger.info(f"记录ID: {record.id}, 识别编码: {record.identification_code}")

        if not records:
            return jsonify({'success': False, 'message': f'未找到批次 {batch_no} 在当前仓库的接收记录'})

        # 检查批次中的货物是否已有出库记录
        outbound_check_failed = []
        for record in records:
            if record.identification_code:
                # 根据仓库类型查找相应的出库记录
                if warehouse_type == 'frontend':
                    # 前端仓：只检查前端仓的出库记录
                    outbound_records = OutboundRecord.query.join(
                        Warehouse, OutboundRecord.operated_warehouse_id == Warehouse.id
                    ).filter(
                        OutboundRecord.identification_code == record.identification_code,
                        Warehouse.warehouse_type == 'frontend'
                    ).all()
                elif warehouse_type == 'backend':
                    # 后端仓：只检查后端仓的出库记录
                    outbound_records = OutboundRecord.query.join(
                        Warehouse, OutboundRecord.operated_warehouse_id == Warehouse.id
                    ).filter(
                        OutboundRecord.identification_code == record.identification_code,
                        Warehouse.warehouse_type == 'backend'
                    ).all()
                else:
                    outbound_records = []

                if outbound_records:
                    # 如果在相同仓库类型有出库记录，不允许删除接收记录
                    outbound_warehouses = []
                    for out in outbound_records:
                        if out.operated_warehouse:
                            warehouse_info = f"{out.operated_warehouse.warehouse_name}({out.operated_warehouse.warehouse_type})"
                            if warehouse_info not in outbound_warehouses:
                                outbound_warehouses.append(warehouse_info)

                    outbound_check_failed.append({
                        'identification_code': record.identification_code,
                        'customer_name': record.customer_name,
                        'outbound_count': len(outbound_records),
                        'outbound_warehouses': outbound_warehouses
                    })

        # 如果有货物已出库，不允许删除
        if outbound_check_failed:
            failed_codes = [item['identification_code'] for item in outbound_check_failed]

            # 构建详细的错误信息
            error_details = []
            for item in outbound_check_failed[:3]:  # 只显示前3个详情
                warehouses_str = "、".join(item['outbound_warehouses'])
                error_details.append(f"• {item['identification_code']} (客户: {item['customer_name']}) - 已在{warehouses_str}出库")

            warehouse_type_name = '前端仓' if warehouse_type == 'frontend' else '后端仓'
            error_message = f'批次 {batch_no} 中有 {len(outbound_check_failed)} 票货物已在{warehouse_type_name}出库，不能删除接收记录。\n\n'
            error_message += f'只有在{warehouse_type_name}没有出库记录的货物才能删除接收记录，以确保数据一致性。\n\n'
            error_message += f'已在{warehouse_type_name}出库的货物详情：\n' + '\n'.join(error_details)

            if len(outbound_check_failed) > 3:
                error_message += f'\n... 还有 {len(outbound_check_failed) - 3} 票货物也已出库'

            return jsonify({
                'success': False,
                'message': error_message
            })

        # 记录更新的识别编码，用于验证
        updated_identification_codes = []

        # 删除接收记录和对应的库存记录
        updated_count = 0
        deleted_inventory_count = 0
        records_without_outbound = 0

        for record in records:
            try:
                updated_identification_codes.append(record.identification_code)

                # 检查是否有对应的出库记录（验证数据完整性）
                has_outbound = False
                if record.identification_code and record.batch_no:
                    outbound_record = OutboundRecord.query.filter_by(
                        identification_code=record.identification_code,
                        batch_no=record.batch_no
                    ).first()

                    if outbound_record:
                        has_outbound = True
                        current_app.logger.info(f"找到对应的出库记录: 接收记录ID={record.id}, 出库记录ID={outbound_record.id}")
                    else:
                        records_without_outbound += 1
                        current_app.logger.warning(f"接收记录没有对应的出库记录: ID={record.id}, 识别编码={record.identification_code}, 批次={record.batch_no}")

                # 处理对应的库存记录 - 修复库存回退逻辑
                if record.identification_code:
                    # 确定要查找库存的仓库ID
                    inventory_warehouse_id = current_warehouse_id
                    if warehouse_type == 'backend' and hasattr(record, 'operated_warehouse_id') and record.operated_warehouse_id:
                        # 对于后端仓，使用接收记录的实际操作仓库ID
                        inventory_warehouse_id = record.operated_warehouse_id

                    # 获取接收记录的数量
                    receive_pallet_count = getattr(record, 'pallet_count', 0) or 0
                    receive_package_count = getattr(record, 'package_count', 0) or 0
                    receive_weight = getattr(record, 'weight', 0) or 0
                    receive_volume = getattr(record, 'volume', 0) or 0

                    current_app.logger.info(f"准备回退库存: 识别编码={record.identification_code}, 仓库ID={inventory_warehouse_id}, 板数={receive_pallet_count}, 件数={receive_package_count}")

                    # 查找库存记录 - 根据仓库类型使用不同的查找策略
                    main_inventory = None

                    if warehouse_type == 'frontend':
                        # 前端仓：直接通过识别编码查找
                        main_inventory = Inventory.query.filter_by(
                            identification_code=record.identification_code,
                            operated_warehouse_id=inventory_warehouse_id
                        ).first()
                        current_app.logger.info(f"前端仓库存查找: 识别编码={record.identification_code}, 仓库ID={inventory_warehouse_id}, 找到={main_inventory is not None}")

                    elif warehouse_type == 'backend':
                        # 后端仓：优先通过original_identification_code查找（合并库存的情况）
                        main_inventory = Inventory.query.filter_by(
                            original_identification_code=record.identification_code,
                            operated_warehouse_id=inventory_warehouse_id
                        ).first()

                        # 如果没找到，再尝试直接匹配识别编码
                        if not main_inventory:
                            main_inventory = Inventory.query.filter_by(
                                identification_code=record.identification_code,
                                operated_warehouse_id=inventory_warehouse_id
                            ).first()

                        current_app.logger.info(f"后端仓库存查找: 接收记录识别编码={record.identification_code}, 仓库ID={inventory_warehouse_id}, 找到库存={main_inventory.identification_code if main_inventory else 'None'}")

                    if main_inventory:
                        # 删除接收记录时，必须减少对应的库存数量
                        # 这是因为接收记录被删除意味着这部分货物"没有被接收"

                        # 安全地减少库存数量，确保不会出现负数
                        old_pallet = main_inventory.pallet_count or 0
                        old_package = main_inventory.package_count or 0
                        old_weight = main_inventory.weight or 0
                        old_volume = main_inventory.volume or 0

                        new_pallet = max(0, old_pallet - receive_pallet_count)
                        new_package = max(0, old_package - receive_package_count)
                        new_weight = max(0, old_weight - receive_weight)
                        new_volume = max(0, old_volume - receive_volume)

                        main_inventory.pallet_count = new_pallet
                        main_inventory.package_count = new_package
                        main_inventory.weight = new_weight
                        main_inventory.volume = new_volume
                        main_inventory.last_updated = datetime.now()

                        current_app.logger.info(f"回退库存数量: 识别编码={main_inventory.identification_code}, 板数{old_pallet}-{receive_pallet_count}={new_pallet}, 件数{old_package}-{receive_package_count}={new_package}")

                        # 如果库存数量减少到0，删除库存记录
                        if new_pallet == 0 and new_package == 0:
                            current_app.logger.info(f"删除库存记录（数量为0）: 识别编码={main_inventory.identification_code}, 仓库ID={inventory_warehouse_id}")
                            db.session.delete(main_inventory)
                            deleted_inventory_count += 1
                        else:
                            # 库存记录已更新
                            deleted_inventory_count += 1  # 计数器表示处理的库存记录数
                    else:
                        current_app.logger.warning(f"未找到对应的库存记录: 识别编码={record.identification_code}, 仓库ID={inventory_warehouse_id}, 仓库类型={warehouse_type}")

                # 处理对应的在途货物记录回退
                if record.identification_code and hasattr(record, 'batch_no') and record.batch_no:
                    # 查找对应的在途货物记录（状态为 received）
                    transit_cargo = TransitCargo.query.filter_by(
                        identification_code=record.identification_code,
                        batch_no=record.batch_no,
                        status='received'
                    ).first()

                    if transit_cargo:
                        # 将在途货物状态改回 in_transit
                        transit_cargo.status = 'in_transit'
                        transit_cargo.actual_arrival_time = None
                        transit_cargo.received_time = None
                        transit_cargo.received_by_user_id = None
                        transit_cargo.last_updated = datetime.now()

                        current_app.logger.info(f"删除接收记录时回退在途货物状态: {transit_cargo.identification_code}, 批次: {transit_cargo.batch_no}, received → in_transit")
                    else:
                        current_app.logger.warning(f"删除接收记录时未找到对应的在途货物记录: {record.identification_code}, 批次: {getattr(record, 'batch_no', 'N/A')}")

                # 删除接收记录
                db.session.delete(record)
                updated_count += 1
                current_app.logger.info(f"删除接收记录: ID={record.id}, 识别编码={record.identification_code}, 有出库记录={has_outbound}")

            except Exception as e:
                current_app.logger.error(f"删除接收记录 {record.id} 时出错: {str(e)}")
                continue

        # 提交事务
        db.session.commit()

        current_app.logger.info(f"用户 {current_user.username} 删除批次 {batch_no} 的 {updated_count} 条接收记录和 {deleted_inventory_count} 条库存记录")
        current_app.logger.info(f"删除的识别编码: {updated_identification_codes}")

        if records_without_outbound > 0:
            current_app.logger.warning(f"发现 {records_without_outbound} 条接收记录没有对应的出库记录，可能是测试数据或数据不一致")

        # 验证删除是否成功
        if warehouse_type == 'frontend':
            remaining_receive_records = InboundRecord.query.filter(
                InboundRecord.batch_no == batch_no,
                InboundRecord.record_type == 'receive',
                InboundRecord.operated_warehouse_id == current_warehouse_id
            ).count()
        else:  # backend
            remaining_receive_records = ReceiveRecord.query.filter(
                ReceiveRecord.batch_no == batch_no,
                ReceiveRecord.operated_warehouse_id == current_warehouse_id
            ).count()

        current_app.logger.info(f"删除后剩余的接收记录数: {remaining_receive_records}, 待接收记录数: {len(updated_identification_codes) - remaining_receive_records}")

        # 构建返回消息
        message = f'成功删除批次 {batch_no} 的 {updated_count} 条接收记录和 {deleted_inventory_count} 条库存记录'
        if records_without_outbound > 0:
            message += f'（其中 {records_without_outbound} 条记录没有对应的出库记录，可能是测试数据）'
        message += '，有对应出库记录的货物将重新显示在待接收列表中'

        return jsonify({
            'success': True,
            'message': message,
            'deleted_receive_count': updated_count,
            'deleted_inventory_count': deleted_inventory_count,
            'records_without_outbound': records_without_outbound,
            'deleted_codes': updated_identification_codes
        })

    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"删除批次接收记录时出错: {str(e)}")
        return jsonify({'success': False, 'message': f'删除失败: {str(e)}'})

@bp.route('/export_frontend_receive')
@require_permission('INBOUND_VIEW')
def export_frontend_receive():
    """导出前端仓接收记录到Excel"""
    try:
        # 获取搜索参数（与frontend_receive_list相同的参数处理）
        date_start = request.args.get('date_start', '')
        date_end = request.args.get('date_end', '')
        customer_name = request.args.get('customer_name', '')
        plate_number = request.args.get('plate_number', '')
        batch_no = request.args.get('batch_no', '')

        # 如果没有指定日期范围，默认使用最近一周的日期范围
        if not date_start and not date_end:
            today = datetime.now().strftime('%Y-%m-%d')
            one_week_ago = (datetime.now() - timedelta(days=7)).strftime('%Y-%m-%d')
            date_start = one_week_ago
            date_end = today

        # 构建查询，只显示前端仓的接收记录（使用record_type字段精确过滤）
        query = InboundRecord.query.options(
            db.joinedload(InboundRecord.operated_warehouse),
            db.joinedload(InboundRecord.operated_by_user)
        ).filter(
            # 只显示接收记录类型
            InboundRecord.record_type == 'receive'
        )

        # 首先过滤出前端仓的数据
        frontend_warehouses = Warehouse.query.filter_by(warehouse_type='frontend').all()
        if not frontend_warehouses:
            # 如果没有前端仓，返回空查询
            query = query.filter(InboundRecord.id == -1)
        else:
            frontend_warehouse_ids = [w.id for w in frontend_warehouses]
            query = query.filter(InboundRecord.operated_warehouse_id.in_(frontend_warehouse_ids))

        # 根据用户权限进一步过滤数据
        if hasattr(current_user, 'warehouse') and current_user.warehouse:
            if current_user.warehouse.warehouse_type == 'frontend':
                # 前端仓用户只能看自己仓库的接收数据
                query = query.filter_by(operated_warehouse_id=current_user.warehouse_id)

        # 应用搜索过滤
        if customer_name:
            query = query.filter(InboundRecord.customer_name.contains(customer_name))
        if plate_number:
            query = query.filter(InboundRecord.plate_number.contains(plate_number))
        if batch_no:
            query = query.filter(InboundRecord.identification_code.contains(batch_no))

        # 日期范围过滤
        if date_start:
            try:
                start_date = datetime.strptime(date_start, '%Y-%m-%d')
                query = query.filter(InboundRecord.inbound_time >= start_date)
            except ValueError:
                pass
        if date_end:
            try:
                end_date = datetime.strptime(date_end, '%Y-%m-%d')
                end_date = end_date.replace(hour=23, minute=59, second=59)
                query = query.filter(InboundRecord.inbound_time <= end_date)
            except ValueError:
                pass

        # 按批次号排序，然后按接收时间排序
        query = query.order_by(InboundRecord.batch_no.desc(), InboundRecord.inbound_time.desc())

        # 获取所有符合条件的记录
        records = query.all()

        # 如果没有记录，返回提示
        if not records:
            flash('没有找到符合条件的接收记录', 'warning')
            return redirect(url_for('main.frontend_receive_list'))

        # 准备导出数据
        data = []
        for record in records:
            data.append({
                '接收时间': record.outbound_time.strftime('%Y-%m-%d %H:%M:%S') if record.outbound_time else '',
                '批次号': record.batch_no or '',
                '识别编码': record.identification_code or '',
                '客户名称': record.customer_name or '',
                '送货干线车': record.delivery_plate_number or '',
                '入库车牌': record.plate_number or '',
                '板数': record.pallet_count or 0,
                '件数': record.package_count or 0,
                '重量(kg)': record.weight or 0,
                '体积(m³)': record.volume or 0,
                '出境模式': record.export_mode or '',
                '报关行': record.customs_broker or '',
                '订单类型': record.order_type or '',
                '跟单客服': record.service_staff or '',
                '库位': record.location or '',
                '单据': record.documents or '',
                '操作仓库': record.operated_warehouse.warehouse_name if record.operated_warehouse else '',
                '操作用户': record.operated_by_user.username if record.operated_by_user else '',
                '创建时间': record.outbound_time.strftime('%Y-%m-%d %H:%M:%S') if record.outbound_time else ''
            })

        # 使用pandas导出Excel
        try:
            import pandas as pd
            from io import BytesIO

            df = pd.DataFrame(data)
            output = BytesIO()

            with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
                df.to_excel(writer, sheet_name='前端仓接收记录', index=False)

                # 获取workbook和worksheet对象以进行格式调整
                workbook = writer.book
                worksheet = writer.sheets['前端仓接收记录']

                # 定义格式
                header_format = workbook.add_format({
                    'bold': True,
                    'text_wrap': True,
                    'valign': 'top',
                    'fg_color': '#D7E4BC',
                    'border': 1
                })

                # 应用表头格式
                for col_num, value in enumerate(df.columns.values):
                    worksheet.write(0, col_num, value, header_format)

                # 设置列宽
                worksheet.set_column('A:S', 15)

            output.seek(0)

            # 生成下载文件名
            timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
            filename = f"前端仓接收记录导出_{timestamp}.xlsx"

            # 返回Excel文件
            return send_file(
                output,
                as_attachment=True,
                download_name=filename,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )

        except ImportError:
            # 如果pandas不可用，使用openpyxl
            try:
                import openpyxl
                from tempfile import NamedTemporaryFile

                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "前端仓接收记录"

                # 添加表头
                headers = list(data[0].keys()) if data else []
                for col, header in enumerate(headers, 1):
                    ws.cell(row=1, column=col, value=header)

                # 添加数据
                for row, record_data in enumerate(data, 2):
                    for col, value in enumerate(record_data.values(), 1):
                        ws.cell(row=row, column=col, value=value)

                # 保存到临时文件
                temp_file = NamedTemporaryFile(delete=False, suffix='.xlsx')
                wb.save(temp_file.name)
                temp_file.close()

                # 返回文件下载响应
                return send_file(
                    temp_file.name,
                    as_attachment=True,
                    download_name=f"前端仓接收记录导出_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx",
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )

            except ImportError:
                flash("导出Excel需要pandas或openpyxl库，请安装: pip install pandas xlsxwriter 或 pip install openpyxl", "danger")
                return redirect(url_for('main.frontend_receive_list'))

    except Exception as e:
        current_app.logger.error(f"导出前端仓接收记录时出错: {str(e)}")
        flash(f"导出数据时出错: {str(e)}", "danger")
        return redirect(url_for('main.frontend_receive_list'))


@bp.route('/export_frontend_inbound')
@require_permission('INBOUND_VIEW')
def export_frontend_inbound():
    """导出前端仓入库记录到Excel"""
    try:
        current_app.logger.info("开始导出前端仓入库记录")

        # 获取搜索参数（与frontend_inbound_list相同的参数处理）
        date_start = request.args.get('date_start', '')
        date_end = request.args.get('date_end', '')
        warehouse_id = request.args.get('warehouse_id', '')
        customer_name = request.args.get('customer_name', '')
        plate_number = request.args.get('plate_number', '')
        identification_code = request.args.get('identification_code', '')
        order_type = request.args.get('order_type', '')
        export_mode = request.args.get('export_mode', '')
        customs_broker = request.args.get('customs_broker', '')
        service_staff = request.args.get('service_staff', '')
        location = request.args.get('location', '')

        current_app.logger.info(f"导出参数: 开始日期={date_start}, 结束日期={date_end}, 仓库ID={warehouse_id}, 客户={customer_name}")

        # 获取前端仓
        frontend_warehouses = Warehouse.query.filter_by(warehouse_type='frontend').all()
        frontend_warehouse_ids = [w.id for w in frontend_warehouses]

        if not frontend_warehouse_ids:
            flash('没有找到前端仓库', 'warning')
            return redirect(url_for('main.frontend_inbound_list'))

        # 构建查询 - 只查询前端仓的记录
        query = InboundRecord.query.options(
            db.joinedload(InboundRecord.operated_warehouse),
            db.joinedload(InboundRecord.operated_by_user)
        )
        query = query.filter(InboundRecord.operated_warehouse_id.in_(frontend_warehouse_ids))

        # 前端仓入库记录导出只导出直接入库记录（没有批次号）
        query = query.filter(InboundRecord.record_type == 'direct')

        # 根据用户权限进一步过滤数据
        # 管理员和后端仓账号可以查看所有前端仓的数据
        # 前端仓账号只能查看自身仓库操作的数据
        if hasattr(current_user, 'warehouse') and current_user.warehouse:
            if current_user.warehouse.warehouse_type == 'frontend' and not current_user.is_admin:
                # 前端仓用户只能导出自己仓库的数据
                query = query.filter(InboundRecord.operated_warehouse_id == current_user.warehouse_id)

        # 仓库过滤
        if warehouse_id:
            query = query.filter(InboundRecord.operated_warehouse_id == warehouse_id)

        # 日期范围搜索
        if date_start:
            try:
                start_date = datetime.strptime(date_start, '%Y-%m-%d')
                query = query.filter(InboundRecord.inbound_time >= start_date)
            except ValueError:
                pass

        if date_end:
            try:
                end_date = datetime.strptime(date_end, '%Y-%m-%d')
                end_date = end_date.replace(hour=23, minute=59, second=59)
                query = query.filter(InboundRecord.inbound_time <= end_date)
            except ValueError:
                pass

        # 其他搜索条件
        if customer_name:
            query = query.filter(InboundRecord.customer_name.like(f'%{customer_name}%'))
        if plate_number:
            query = query.filter(InboundRecord.plate_number.like(f'%{plate_number}%'))
        if identification_code:
            query = query.filter(InboundRecord.identification_code.like(f'%{identification_code}%'))
        if order_type:
            query = query.filter(InboundRecord.order_type.like(f'%{order_type}%'))
        if export_mode:
            query = query.filter(InboundRecord.export_mode.like(f'%{export_mode}%'))
        if customs_broker:
            query = query.filter(InboundRecord.customs_broker.like(f'%{customs_broker}%'))
        if service_staff:
            query = query.filter(InboundRecord.service_staff.like(f'%{service_staff}%'))
        if location:
            query = query.filter(InboundRecord.location.like(f'%{location}%'))

        # 执行查询，限制记录数量避免超时
        current_app.logger.info("开始查询数据库记录")
        records = query.order_by(InboundRecord.inbound_time.desc()).limit(5000).all()
        current_app.logger.info(f"查询到 {len(records)} 条记录")

        # 如果没有记录，返回提示
        if not records:
            flash('没有找到符合条件的入库记录', 'warning')
            return redirect(url_for('main.frontend_inbound_list'))

        # 准备导出数据
        current_app.logger.info("开始处理导出数据")
        data = []
        for i, record in enumerate(records):
            try:
                data.append({
                    '入库时间': record.outbound_time.strftime('%Y-%m-%d %H:%M:%S') if record.outbound_time else '',
                    '批次号': record.batch_no or '',
                    '客户名称': record.customer_name or '',
                    '车牌号': record.plate_number or '',
                    '识别编码': record.identification_code or '',
                    '板数': record.pallet_count or 0,
                    '件数': record.package_count or 0,
                    '重量(kg)': record.weight or 0,
                    '体积(m³)': record.volume or 0,
                    '订单类型': record.order_type or '',
                    '出境模式': record.export_mode or '',
                    '报关行': record.customs_broker or '',
                    '库位': record.location or '',
                    '单据': record.documents or '',
                    '跟单客服': record.service_staff or '',
                    '操作仓库': record.operated_warehouse.warehouse_name if record.operated_warehouse else '',
                    '操作用户': record.operated_by_user.username if record.operated_by_user else '',
                    '创建时间': record.outbound_time.strftime('%Y-%m-%d %H:%M:%S') if record.outbound_time else ''
                })

                # 每处理100条记录输出一次日志
                if (i + 1) % 100 == 0:
                    current_app.logger.info(f"已处理 {i + 1}/{len(records)} 条记录")

            except Exception as e:
                current_app.logger.error(f"处理第 {i+1} 条记录时出错: {str(e)}")
                current_app.logger.error(f"记录详情: ID={getattr(record, 'id', 'N/A')}, 客户={getattr(record, 'customer_name', 'N/A')}")
                continue

        current_app.logger.info(f"数据处理完成，共 {len(data)} 条有效记录")

        # 使用pandas导出Excel
        try:
            current_app.logger.info("开始生成Excel文件")
            import pandas as pd
            from io import BytesIO

            if not data:
                flash('没有数据可以导出', 'warning')
                return redirect(url_for('main.frontend_inbound_list'))

            df = pd.DataFrame(data)

            # 创建Excel文件
            output = BytesIO()
            current_app.logger.info("开始写入Excel数据")
            with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
                df.to_excel(writer, sheet_name='前端仓入库记录', index=False)

                # 获取工作表和工作簿对象
                workbook = writer.book
                worksheet = writer.sheets['前端仓入库记录']

                # 定义格式
                header_format = workbook.add_format({
                    'bold': True,
                    'text_wrap': True,
                    'valign': 'top',
                    'fg_color': '#D7E4BC',
                    'border': 1
                })

                # 应用表头格式
                for col_num, value in enumerate(df.columns.values):
                    worksheet.write(0, col_num, value, header_format)

                # 设置列宽
                for i, col in enumerate(df.columns):
                    max_len = max(df[col].astype(str).map(len).max(), len(col))
                    worksheet.set_column(i, i, min(max_len + 2, 50))

            output.seek(0)
            current_app.logger.info("Excel文件生成完成")

            # 生成下载文件名
            timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
            filename = f"前端仓入库记录导出_{timestamp}.xlsx"

            # 返回Excel文件
            return send_file(
                output,
                as_attachment=True,
                download_name=filename,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )

        except ImportError:
            flash("导出Excel需要pandas或xlsxwriter库，请安装: pip install pandas xlsxwriter 或 pip install openpyxl", "danger")
            return redirect(url_for('main.frontend_inbound_list'))

    except Exception as e:
        current_app.logger.error(f"导出前端仓入库记录时出错: {str(e)}")
        flash(f"导出数据时出错: {str(e)}", "danger")
        return redirect(url_for('main.frontend_inbound_list'))


@bp.route('/frontend/inbound/list')
@require_permission('INBOUND_VIEW')
def frontend_inbound_list():
    """前端仓入库记录列表 - 性能优化版本"""
    import time
    start_time = time.time()

    if not check_warehouse_permission('frontend', 'view'):
        flash('您没有权限访问前端仓入库记录', 'error')
        return redirect(url_for('main.index'))

    # 获取搜索参数
    page = request.args.get('page', 1, type=int)
    per_page = 20

    # 性能优化：使用子查询获取前端仓ID，避免加载完整的仓库对象
    frontend_warehouse_subquery = db.session.query(Warehouse.id).filter_by(
        warehouse_type='frontend',
        status='active'
    ).subquery()

    # 构建优化后的查询
    query = InboundRecord.query.options(
        db.joinedload(InboundRecord.operated_warehouse)
    ).filter(
        InboundRecord.operated_warehouse_id.in_(frontend_warehouse_subquery),
        InboundRecord.record_type == 'direct'
    )

    # 根据用户权限进一步过滤数据
    # 管理员和后端仓账号可以查看所有前端仓的数据
    # 前端仓账号只能查看自身仓库操作的数据
    if hasattr(current_user, 'warehouse') and current_user.warehouse:
        if current_user.warehouse.warehouse_type == 'frontend' and not current_user.is_admin:
            # 前端仓用户只能查看自己仓库的数据
            query = query.filter(InboundRecord.operated_warehouse_id == current_user.warehouse_id)

    # 搜索过滤
    search_params = {}

    # 前端仓入库记录界面只显示直接入库记录（没有批次号）
    query = query.filter(InboundRecord.record_type == 'direct')

    # 仓库搜索
    warehouse_id = request.args.get('warehouse_id')
    if warehouse_id:
        query = query.filter(InboundRecord.operated_warehouse_id == warehouse_id)
        search_params['warehouse_id'] = warehouse_id

    # 客户名称搜索
    if request.args.get('customer_name'):
        query = query.filter(InboundRecord.customer_name.contains(request.args.get('customer_name')))
        search_params['customer_name'] = request.args.get('customer_name')

    # 车牌号搜索
    if request.args.get('plate_number'):
        query = query.filter(InboundRecord.plate_number.contains(request.args.get('plate_number')))
        search_params['plate_number'] = request.args.get('plate_number')

    # 识别编码搜索
    if request.args.get('identification_code'):
        query = query.filter(InboundRecord.identification_code.contains(request.args.get('identification_code')))
        search_params['identification_code'] = request.args.get('identification_code')

    # 日期范围搜索
    # 如果没有指定日期参数，默认显示昨天和今天的数据
    date_start = request.args.get('date_start')
    date_end = request.args.get('date_end')

    # 如果没有任何日期参数，设置默认值为一周前到今天
    if not date_start and not date_end and not any(request.args.get(key) for key in ['customer_name', 'plate_number', 'identification_code']):
        today = datetime.now().date()
        one_week_ago = today - timedelta(days=7)
        date_start = one_week_ago.strftime('%Y-%m-%d')
        date_end = today.strftime('%Y-%m-%d')

    if date_start:
        try:
            start_date = datetime.strptime(date_start, '%Y-%m-%d')
            query = query.filter(InboundRecord.inbound_time >= start_date)
            search_params['date_start'] = date_start
        except ValueError:
            pass

    if date_end:
        try:
            end_date = datetime.strptime(date_end, '%Y-%m-%d')
            # 设置为当天的23:59:59
            end_date = end_date.replace(hour=23, minute=59, second=59)
            query = query.filter(InboundRecord.inbound_time <= end_date)
            search_params['date_end'] = date_end
        except ValueError:
            pass

    # 计算汇总数据（在分页之前）
    summary_data = {}
    try:
        # 获取总记录数
        total_records = query.count()

        # 计算汇总统计
        summary_query = query.with_entities(
            db.func.count(InboundRecord.id).label('total_count'),
            db.func.sum(InboundRecord.pallet_count).label('total_pallets'),
            db.func.sum(InboundRecord.package_count).label('total_packages'),
            db.func.sum(InboundRecord.weight).label('total_weight'),
            db.func.sum(InboundRecord.volume).label('total_volume')
        ).first()

        summary_data = {
            'total_records': total_records,
            'total_pallets': int(summary_query.total_pallets or 0),
            'total_packages': int(summary_query.total_packages or 0),
            'total_weight': float(summary_query.total_weight or 0),
            'total_volume': float(summary_query.total_volume or 0)
        }

        current_app.logger.info(f"前端仓入库汇总数据: {summary_data}")

    except Exception as e:
        current_app.logger.error(f"计算前端仓入库汇总数据失败: {str(e)}")
        summary_data = {
            'total_records': 0,
            'total_pallets': 0,
            'total_packages': 0,
            'total_weight': 0.0,
            'total_volume': 0.0
        }

    # 排序和分页 - 按操作时间降序排序
    records = query.order_by(InboundRecord.created_at.desc()).paginate(
        page=page, per_page=per_page, error_out=False
    )

    # 性能优化：缓存仓库列表查询
    warehouses = Warehouse.query.filter_by(
        warehouse_type='frontend',
        status='active'
    ).order_by(Warehouse.warehouse_name).all()

    # 记录查询性能
    end_time = time.time()
    execution_time = end_time - start_time
    if execution_time > 1.0:
        current_app.logger.warning(f"Slow frontend_inbound_list query: {execution_time:.2f}s")
    else:
        current_app.logger.info(f"frontend_inbound_list query time: {execution_time:.3f}s")

    from app.utils import render_ajax_aware
    return render_ajax_aware('frontend/inbound_list.html',
                           records=records,
                           search_params=search_params,
                           warehouses=warehouses,
                           summary_data=summary_data,
                           title='前端仓入库记录',
                           warehouse_type='frontend')


@bp.route('/frontend/inbound/delete/<int:id>', methods=['POST'])
@require_permission('INBOUND_DELETE')
@log_operation('inbound', 'delete', 'inbound_record')
def delete_frontend_inbound(id):
    """删除前端仓入库记录"""
    try:
        # 查询记录
        record = InboundRecord.query.get_or_404(id)

        # 检查权限：确保是前端仓的记录
        if not check_warehouse_permission('frontend', 'delete'):
            flash('您没有权限删除前端仓入库记录', 'error')
            return redirect(url_for('main.frontend_inbound_list'))

        # 检查记录是否属于前端仓
        if record.operated_warehouse and record.operated_warehouse.warehouse_type != 'frontend':
            flash('只能删除前端仓的入库记录', 'error')
            return redirect(url_for('main.frontend_inbound_list'))

        # 如果用户绑定了特定仓库，检查是否有权限删除该记录
        if hasattr(current_user, 'warehouse') and current_user.warehouse:
            # 对于前端仓用户，允许删除所有前端仓的记录
            if current_user.warehouse.warehouse_type == 'frontend':
                # 检查记录是否属于前端仓
                from app.models import Warehouse
                record_warehouse = Warehouse.query.get(record.operated_warehouse_id)

                if record_warehouse and record_warehouse.warehouse_type != 'frontend':
                    flash('前端仓用户只能删除前端仓的入库记录', 'error')
                    current_app.logger.warning(f"前端仓用户 {current_user.username} 尝试删除后端仓记录: 记录ID={id}, 记录仓库类型={record_warehouse.warehouse_type}")
                    return redirect(url_for('main.frontend_inbound_list'))
            else:
                # 对于后端仓用户，只能删除自己仓库的记录
                if record.operated_warehouse_id != current_user.warehouse_id:
                    flash('您只能删除自己仓库的入库记录', 'error')
                    current_app.logger.warning(f"用户 {current_user.username} 尝试删除其他仓库的记录: 记录ID={id}, 记录仓库={record.operated_warehouse_id}, 用户仓库={current_user.warehouse_id}")
                    return redirect(url_for('main.frontend_inbound_list'))

        # 获取记录信息，用于检查出库记录
        identification_code = record.identification_code
        customer_name = record.customer_name

        # 检查是否有对应的出库记录
        outbound_records = OutboundRecord.query.filter_by(identification_code=identification_code).all()

        if outbound_records:
            # 如果有出库记录，不允许删除
            flash(f'无法删除入库记录：客户 {customer_name} 的货物已有出库记录，请先处理相关出库记录', 'error')
            current_app.logger.warning(f"尝试删除有出库记录的前端仓入库记录: ID={id}, 识别编码={identification_code}")
            return redirect(url_for('main.frontend_inbound_list'))

        # 查找对应的库存记录
        inventory = Inventory.query.filter_by(identification_code=identification_code).first()

        # 删除库存记录
        if inventory:
            db.session.delete(inventory)
            current_app.logger.info(f"删除库存记录: {identification_code}")

        # 删除入库记录
        db.session.delete(record)
        db.session.commit()

        flash(f'已成功删除客户 {customer_name} 的入库记录及对应库存', 'success')
        current_app.logger.info(f"删除前端仓入库记录: ID={id}, 客户={customer_name}, 识别编码={identification_code}")

    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"删除前端仓入库记录时出错: {str(e)}")
        flash(f"删除入库记录时出错: {str(e)}", "danger")

    return redirect(url_for('main.frontend_inbound_list'))


@bp.route('/api/check-outbound-records/<identification_code>')
@require_permission('INBOUND_VIEW')
def check_outbound_records(identification_code):
    """检查指定识别编码是否有出库记录的API"""
    try:
        # 查找出库记录
        outbound_records = OutboundRecord.query.filter_by(identification_code=identification_code).all()

        if outbound_records:
            # 计算总出库数量
            total_outbound_pallet = sum(record.pallet_count for record in outbound_records)
            total_outbound_package = sum(record.package_count for record in outbound_records)

            return jsonify({
                'has_outbound': True,
                'outbound_count': len(outbound_records),
                'outbound_pallet': total_outbound_pallet,
                'outbound_package': total_outbound_package,
                'outbound_records': [
                    {
                        'id': record.id,
                        'outbound_time': record.outbound_time.strftime('%Y-%m-%d %H:%M:%S') if record.outbound_time else '',
                        'destination': record.destination,
                        'pallet_count': record.pallet_count,
                        'package_count': record.package_count,
                        'batch_no': record.batch_no
                    } for record in outbound_records
                ]
            })
        else:
            return jsonify({
                'has_outbound': False,
                'outbound_count': 0,
                'outbound_pallet': 0,
                'outbound_package': 0,
                'outbound_records': []
            })

    except Exception as e:
        current_app.logger.error(f"检查出库记录失败: {str(e)}")
        return jsonify({
            'error': True,
            'message': f'检查出库记录失败: {str(e)}'
        }), 500


# ==================== 后端仓入库路由 ====================

@bp.route('/backend/receive', methods=['GET', 'POST'])
@require_permission('INBOUND_VIEW')
def backend_receive():
    """后端仓入库操作页面"""
    if not check_warehouse_permission('backend', 'view'):
        flash('您没有权限访问后端仓入库功能', 'error')
        return redirect(url_for('main.index'))

    if request.method == 'POST':
        try:
            # 获取表单数据
            receive_time = request.form.get('receive_time')
            plate_number = request.form.get('plate_number')
            source_warehouse = request.form.get('source_warehouse')
            driver_name = request.form.get('driver_name')
            receive_notes = request.form.get('receive_notes')

            # 验证必填字段
            if not receive_time or not plate_number:
                flash('接收时间和车牌号为必填项', 'error')
                return render_template('backend/receive_batch.html',
                                     title='后端仓接收操作',
                                     warehouse_type='backend')

            # 转换时间格式
            receive_datetime = datetime.strptime(receive_time, '%Y-%m-%dT%H:%M')

            # 创建接收记录（使用InboundRecord表，标记为后端仓接收）
            receive_record = InboundRecord(
                inbound_time=receive_datetime,
                plate_number=plate_number,
                customer_name=source_warehouse or '未知来源',
                documents=f'后端仓接收-{source_warehouse}',
                service_staff=current_user.username if hasattr(current_user, 'username') else '系统',
                record_type='receive',  # 接收记录
                create_time=datetime.now()
            )

            # 添加备注信息
            if receive_notes:
                receive_record.documents += f' 备注:{receive_notes}'
            if driver_name:
                receive_record.documents += f' 司机:{driver_name}'

            db.session.add(receive_record)
            db.session.commit()

            flash(f'货物接收记录已成功创建，记录ID: {receive_record.id}', 'success')
            return redirect(url_for('main.backend_receive_list'))

        except Exception as e:
            db.session.rollback()
            current_app.logger.error(f"后端仓接收操作失败: {str(e)}")
            flash(f'接收操作失败: {str(e)}', 'error')

    return render_template('backend/receive_batch.html',
                         title='后端仓接收操作',
                         warehouse_type='backend')


# 获取待接收货物数据API
@bp.route('/api/backend/pending-receive', methods=['GET'])
@csrf.exempt
def api_backend_pending_receive():
    """获取后端仓待接收的前端出库数据"""
    try:
        # 获取前端仓库的出库记录，这些是待后端仓接收的货物
        # 查询条件：前端仓库的出库记录，且目标是后端仓库
        frontend_warehouses = Warehouse.query.filter_by(warehouse_type='frontend').all()
        frontend_warehouse_ids = [w.id for w in frontend_warehouses]

        if not frontend_warehouse_ids:
            return jsonify({'success': True, 'data': []})

        # 获取前端仓库的出库记录
        outbound_records = OutboundRecord.query.options(
            db.joinedload(OutboundRecord.operated_warehouse)
        ).filter(
            OutboundRecord.operated_warehouse_id.in_(frontend_warehouse_ids)
        ).order_by(OutboundRecord.outbound_time.desc()).limit(100).all()

        # 转换为JSON格式
        data = []
        for record in outbound_records:
            data.append({
                'id': record.id,
                'outbound_time': record.outbound_time.strftime('%Y-%m-%d') if record.outbound_time else '',
                'plate_number': record.plate_number or '',
                'customer_name': record.customer_name or '',
                'order_type': record.order_type or '',
                'customs_broker': record.customs_broker or '',
                'pallet_count': record.pallet_count or 0,
                'package_count': record.package_count or 0,
                'weight': record.weight or 0,
                'volume': record.volume or 0,
                'documents': record.documents or '',
                'service_staff': record.service_staff or '',
                'source_warehouse': record.operated_warehouse.warehouse_name if record.operated_warehouse else '',
                'batch_no': record.batch_no or '',
                'batch_sequence': f"{record.batch_sequence}/{record.batch_total}" if record.batch_sequence and record.batch_total else '',
                'remark1': record.remark1 or '',
                'remark2': record.remark2 or ''
            })

        return jsonify({'success': True, 'data': data})

    except Exception as e:
        current_app.logger.error(f"获取待接收货物数据失败: {str(e)}")
        return jsonify({'success': False, 'message': f'获取数据失败: {str(e)}'}), 500


# 获取待接收货物数据API（按批次分组）
@bp.route('/api/backend/pending-receive-batches', methods=['GET'])
@csrf.exempt
def api_backend_pending_receive_batches():
    """获取后端仓待接收的前端出库数据（按批次分组）"""
    try:
        # 获取前端仓库的出库记录，这些是待后端仓接收的货物
        frontend_warehouses = Warehouse.query.filter_by(warehouse_type='frontend').all()
        frontend_warehouse_ids = [w.id for w in frontend_warehouses]

        if not frontend_warehouse_ids:
            return jsonify({'success': True, 'data': []})

        # 查询前端仓库的出库记录，按批次号分组
        # 同时过滤掉识别编码前缀为PX的记录（这些应该是后端仓的货物）
        outbound_records = OutboundRecord.query.options(
            db.joinedload(OutboundRecord.operated_warehouse)
        ).filter(
            OutboundRecord.operated_warehouse_id.in_(frontend_warehouse_ids),
            OutboundRecord.batch_no.isnot(None),
            OutboundRecord.batch_no != '',
            ~OutboundRecord.identification_code.like('PX/%')  # 排除PX前缀的记录
        ).order_by(OutboundRecord.outbound_time.desc()).all()

        # 获取已接收的记录组合（批次号+识别编码）- 优化查询
        received_combinations = set()
        # 只查询有批次号和识别编码的记录，并限制查询范围
        receive_records = ReceiveRecord.query.filter(
            ReceiveRecord.batch_no.isnot(None),
            ReceiveRecord.batch_no != '',
            ReceiveRecord.identification_code.isnot(None),
            ReceiveRecord.identification_code != ''
        ).with_entities(
            ReceiveRecord.batch_no,
            ReceiveRecord.identification_code
        ).all()

        for batch_no, identification_code in receive_records:
            received_combinations.add((batch_no, identification_code))

        # 过滤掉已接收的记录
        outbound_records = [
            record for record in outbound_records
            if (record.batch_no, record.identification_code) not in received_combinations
        ]

        # 按批次号分组
        batch_groups = {}
        for record in outbound_records:
            batch_no = record.batch_no
            if batch_no not in batch_groups:
                batch_groups[batch_no] = {
                    'batch_no': batch_no,
                    'items': [],
                    'total_pallet_count': 0,
                    'total_package_count': 0,
                    'total_weight': 0,
                    'total_volume': 0,
                    'outbound_time': record.outbound_time,
                    'source_warehouse': record.operated_warehouse.warehouse_name if record.operated_warehouse else '',
                    'customer_names': set(),
                    'plate_numbers': set(),  # 送货干线车（前端仓发货车牌）
                    'inbound_plate_numbers': set()  # 入库车牌（后端仓接收车牌）
                }

            # 添加到批次组
            batch_group = batch_groups[batch_no]
            batch_group['items'].append({
                'id': record.id,
                'outbound_time': record.outbound_time.strftime('%Y-%m-%d %H:%M') if record.outbound_time else '',
                'plate_number': record.plate_number or '',
                'inbound_plate': record.inbound_plate or '',
                'customer_name': record.customer_name or '',
                'identification_code': record.identification_code or '',
                'order_type': record.order_type or '',
                'export_mode': record.export_mode or '',
                'customs_broker': record.customs_broker or '',
                'pallet_count': record.pallet_count or 0,
                'package_count': record.package_count or 0,
                'weight': record.weight or 0,
                'volume': record.volume or 0,
                'documents': record.documents or '',
                'service_staff': record.service_staff or '',
                'batch_sequence': record.batch_sequence or 0,
                'batch_total': record.batch_total or 0,
                'remark1': record.remark1 or '',
                'remark2': record.remark2 or ''
            })

            # 累计统计
            batch_group['total_pallet_count'] += record.pallet_count or 0
            batch_group['total_package_count'] += record.package_count or 0
            batch_group['total_weight'] += record.weight or 0
            batch_group['total_volume'] += record.volume or 0
            batch_group['customer_names'].add(record.customer_name or '')
            batch_group['plate_numbers'].add(record.plate_number or '')  # 送货干线车
            batch_group['inbound_plate_numbers'].add(record.inbound_plate or '')  # 入库车牌

        # 转换为列表格式
        data = []
        for batch_no, batch_group in batch_groups.items():
            batch_group['customer_names'] = ', '.join(filter(None, batch_group['customer_names']))
            batch_group['plate_numbers'] = ', '.join(filter(None, batch_group['plate_numbers']))  # 送货干线车
            batch_group['inbound_plate_numbers'] = ', '.join(filter(None, batch_group['inbound_plate_numbers']))  # 入库车牌
            batch_group['item_count'] = len(batch_group['items'])
            batch_group['outbound_time_str'] = batch_group['outbound_time'].strftime('%Y-%m-%d') if batch_group['outbound_time'] else ''
            data.append(batch_group)

        # 按发货时间排序
        data.sort(key=lambda x: x['outbound_time'] or datetime.min, reverse=True)

        return jsonify({'success': True, 'data': data})

    except Exception as e:
        current_app.logger.error(f"获取待接收批次数据失败: {str(e)}")
        return jsonify({'success': False, 'message': f'获取数据失败: {str(e)}'}), 500


# 批次接收货物API
@bp.route('/api/backend/batch-receive', methods=['POST'])
@csrf.exempt
def api_backend_batch_receive():
    """批次接收货物，支持差异录入"""
    try:
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'message': '请求数据格式错误'}), 400

        batch_no = data.get('batch_no')
        receive_time = data.get('receive_time')
        items = data.get('items', [])

        if not batch_no or not receive_time or not items:
            return jsonify({'success': False, 'message': '批次号、接收时间和货物明细为必填项'}), 400

        # 转换时间格式
        try:
            receive_datetime = datetime.strptime(receive_time, '%Y-%m-%dT%H:%M')
        except ValueError:
            return jsonify({'success': False, 'message': '接收时间格式错误'}), 400

        # 获取后端仓库
        backend_warehouse = Warehouse.query.filter_by(warehouse_type='backend').first()
        if not backend_warehouse:
            return jsonify({'success': False, 'message': '未找到后端仓库'}), 400

        # 为每个货物创建接收记录
        for item in items:
            outbound_id = item.get('id')
            actual_pallet_count = item.get('actual_pallet_count', 0)
            actual_package_count = item.get('actual_package_count', 0)
            remark2 = item.get('remark2', '')  # 备注2字段
            storage_location = item.get('storage_location', '')  # 库位

            # 查找原始出库记录
            outbound_record = OutboundRecord.query.get(outbound_id)
            if not outbound_record:
                db.session.rollback()
                return jsonify({'success': False, 'message': f'未找到出库记录ID: {outbound_id}'}), 400

            # 使用原始出库记录的重量和体积，而不是用户输入的实际值
            actual_weight = outbound_record.weight or 0
            actual_volume = outbound_record.volume or 0

            # 创建接收记录（使用ReceiveRecord表）
            receive_record = ReceiveRecord(
                receive_time=receive_datetime,
                delivery_plate_number=outbound_record.plate_number,  # 送货干线车（前端仓→后端仓）
                inbound_plate=outbound_record.inbound_plate,  # 入库车牌（工厂→前端仓）
                customer_name=outbound_record.customer_name,
                identification_code=outbound_record.identification_code,
                pallet_count=actual_pallet_count,
                package_count=actual_package_count,
                weight=actual_weight,
                volume=actual_volume,
                export_mode=outbound_record.export_mode,
                order_type=outbound_record.order_type,
                customs_broker=outbound_record.customs_broker,
                storage_location=storage_location,  # 库位字段
                documents=outbound_record.documents,
                service_staff=outbound_record.service_staff,
                batch_no=batch_no,
                batch_total=outbound_record.batch_total,
                batch_sequence=outbound_record.batch_sequence,
                remark1=outbound_record.remark1,
                remark2=remark2,  # 备注2字段
                created_by=current_user.id,  # 修正字段名
                operated_warehouse_id=backend_warehouse.id
            )

            db.session.add(receive_record)

            # 创建或更新后端仓库存记录 - 使用识别编码进行合并
            # 直接使用识别编码（不再需要处理_B后缀）
            original_identification_code = outbound_record.identification_code

            # 查找是否已存在相同原始识别编码的库存记录
            existing_inventory = Inventory.query.filter_by(
                original_identification_code=original_identification_code,
                operated_warehouse_id=backend_warehouse.id,
                customer_name=outbound_record.customer_name
            ).first()

            if existing_inventory:
                # 更新现有库存 - 合并同一票货物
                existing_inventory.pallet_count = (existing_inventory.pallet_count or 0) + actual_pallet_count
                existing_inventory.package_count = (existing_inventory.package_count or 0) + actual_package_count
                existing_inventory.weight = (existing_inventory.weight or 0) + actual_weight  # 累加重量
                existing_inventory.volume = (existing_inventory.volume or 0) + actual_volume  # 累加体积
                existing_inventory.inbound_pallet_count = (existing_inventory.inbound_pallet_count or 0) + actual_pallet_count
                existing_inventory.inbound_package_count = (existing_inventory.inbound_package_count or 0) + actual_package_count

                # 更新库位信息（如果提供了新的库位）
                if storage_location:
                    existing_inventory.location = storage_location

                # 确保车牌信息正确
                if not existing_inventory.plate_number:
                    # 从原始入库记录获取车牌信息
                    if original_identification_code:
                        original_inbound = InboundRecord.query.filter_by(
                            identification_code=original_identification_code
                        ).first()
                        if original_inbound and original_inbound.plate_number:
                            existing_inventory.plate_number = original_inbound.plate_number
                existing_inventory.last_updated = receive_datetime
                existing_inventory.version += 1
                current_app.logger.info(f'批次接收-合并后端仓库存: {existing_inventory.customer_name} {existing_inventory.identification_code} '
                                      f'原始编码: {original_identification_code} 板数增加{actual_pallet_count}→{existing_inventory.pallet_count} '
                                      f'件数增加{actual_package_count}→{existing_inventory.package_count}')
            else:
                # 获取原始入库车牌信息
                original_plate_number = ''
                if outbound_record.identification_code:
                    original_inbound = InboundRecord.query.filter_by(
                        identification_code=outbound_record.identification_code
                    ).first()
                    if original_inbound:
                        original_plate_number = original_inbound.plate_number or ''

                # 创建新库存记录
                inventory_record = Inventory(
                    customer_name=outbound_record.customer_name,
                    identification_code=outbound_record.identification_code,
                    original_identification_code=original_identification_code,  # 设置原始识别编码
                    inbound_pallet_count=actual_pallet_count,
                    inbound_package_count=actual_package_count,
                    pallet_count=actual_pallet_count,
                    package_count=actual_package_count,
                    weight=actual_weight,
                    volume=actual_volume,
                    export_mode=outbound_record.export_mode,
                    order_type=outbound_record.order_type,
                    customs_broker=outbound_record.customs_broker,
                    documents=outbound_record.documents,
                    location=storage_location,  # 添加库位字段
                    inbound_time=receive_datetime,
                    plate_number=original_plate_number,  # 使用原始入库车牌
                    service_staff=outbound_record.service_staff,
                    operated_by_user_id=current_user.id,
                    operated_warehouse_id=backend_warehouse.id,
                    last_updated=receive_datetime,
                    version=1
                )
                db.session.add(inventory_record)
                current_app.logger.info(f'批次接收-创建后端仓库存: {inventory_record.customer_name} {inventory_record.identification_code} '
                                      f'原始编码: {original_identification_code} 板数{actual_pallet_count} 件数{actual_package_count}')

            # 更新对应的在途货物状态
            transit_cargo = TransitCargo.query.filter_by(
                customer_name=outbound_record.customer_name,
                identification_code=outbound_record.identification_code,
                batch_no=batch_no,
                status='in_transit'
            ).first()

            if transit_cargo:
                # 更新在途货物状态为已接收
                transit_cargo.status = 'received'
                transit_cargo.actual_arrival_time = receive_datetime
                transit_cargo.received_time = receive_datetime
                transit_cargo.received_by_user_id = current_user.id
                transit_cargo.last_updated = receive_datetime

                # 如果接收数量与发运数量不一致，记录差异信息
                if (actual_pallet_count != (outbound_record.pallet_count or 0) or
                    actual_package_count != (outbound_record.package_count or 0)):

                    discrepancy_info = []
                    if actual_pallet_count != (outbound_record.pallet_count or 0):
                        discrepancy_info.append(f"板数差异：发运{outbound_record.pallet_count}，接收{actual_pallet_count}")
                    if actual_package_count != (outbound_record.package_count or 0):
                        discrepancy_info.append(f"件数差异：发运{outbound_record.package_count}，接收{actual_package_count}")

                    if transit_cargo.remark2:
                        transit_cargo.remark2 += f"；{';'.join(discrepancy_info)}"
                    else:
                        transit_cargo.remark2 = ';'.join(discrepancy_info)

                current_app.logger.info(f'更新在途货物状态: {transit_cargo.customer_name} {transit_cargo.identification_code} '
                                      f'状态: in_transit → received')
            else:
                current_app.logger.warning(f'未找到对应的在途货物记录: {outbound_record.customer_name} {outbound_record.identification_code} 批次{batch_no}')

        # 提交事务
        db.session.commit()

        return jsonify({
            'success': True,
            'message': f'批次 {batch_no} 接收成功，共处理 {len(items)} 个货物，后端仓库存已更新'
        })

    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"批次接收失败: {str(e)}")
        return jsonify({'success': False, 'message': f'接收失败: {str(e)}'}), 500


# 快速接收货物API
@bp.route('/api/backend/quick-receive', methods=['POST'])
@csrf.exempt
def api_backend_quick_receive():
    """快速接收单个货物"""
    try:
        current_app.logger.info(f"快速接收API被调用，请求数据: {request.get_data()}")

        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'message': '请求数据格式错误'}), 400

        item_id = data.get('item_id')
        receive_time = data.get('receive_time')

        current_app.logger.info(f"解析到的参数: item_id={item_id}, receive_time={receive_time}")

        if not item_id:
            return jsonify({'success': False, 'message': '缺少货物ID'}), 400

        # 获取出库记录
        outbound_record = OutboundRecord.query.get(item_id)
        if not outbound_record:
            return jsonify({'success': False, 'message': '找不到对应的出库记录'}), 404

        # 检查是否已经接收（通过批次号判断）
        if outbound_record.batch_no:
            existing_inbound = InboundRecord.query.filter_by(
                batch_no=outbound_record.batch_no,
                batch_sequence=outbound_record.batch_sequence
            ).first()
            if existing_inbound:
                return jsonify({'success': False, 'message': '该货物已经被接收'}), 400

        # 获取后端仓库
        backend_warehouse = Warehouse.query.filter_by(warehouse_type='backend').first()
        if not backend_warehouse:
            return jsonify({'success': False, 'message': '找不到后端仓库'}), 500

        # 创建入库记录
        inbound_record = InboundRecord(
            inbound_time=datetime.now(),
            delivery_plate_number=outbound_record.delivery_plate_number,  # 送货干线车
            plate_number='',  # 入库车牌（需要后续填写）
            customer_name=outbound_record.customer_name,
            identification_code=outbound_record.identification_code,
            order_type=outbound_record.order_type,
            customs_broker=outbound_record.customs_broker,
            pallet_count=outbound_record.pallet_count,
            package_count=outbound_record.package_count,
            weight=outbound_record.weight,
            volume=outbound_record.volume,
            documents=outbound_record.documents,
            service_staff=outbound_record.service_staff,
            batch_no=outbound_record.batch_no,
            batch_sequence=outbound_record.batch_sequence,
            batch_total=outbound_record.batch_total,
            remark1=outbound_record.remark1,
            remark2=outbound_record.remark2,
            record_type='receive',  # 接收记录（从前端仓接收到后端仓）
            operated_warehouse_id=backend_warehouse.id,
            operated_by_user_id=current_user.id if hasattr(current_user, 'id') else None
        )

        db.session.add(inbound_record)

        # 创建或更新后端仓库存记录 - 使用识别编码进行合并
        # 直接使用识别编码（不再需要处理_B后缀）
        original_identification_code = outbound_record.identification_code

        # 查找是否已存在相同原始识别编码的库存记录
        existing_inventory = Inventory.query.filter_by(
            original_identification_code=original_identification_code,
            operated_warehouse_id=backend_warehouse.id,
            customer_name=outbound_record.customer_name
        ).first()

        if existing_inventory:
            # 更新现有库存 - 合并同一票货物
            existing_inventory.pallet_count = (existing_inventory.pallet_count or 0) + (outbound_record.pallet_count or 0)
            existing_inventory.package_count = (existing_inventory.package_count or 0) + (outbound_record.package_count or 0)
            existing_inventory.weight = (existing_inventory.weight or 0) + (outbound_record.weight or 0)
            existing_inventory.volume = (existing_inventory.volume or 0) + (outbound_record.volume or 0)
            existing_inventory.inbound_pallet_count = (existing_inventory.inbound_pallet_count or 0) + (outbound_record.pallet_count or 0)
            existing_inventory.inbound_package_count = (existing_inventory.inbound_package_count or 0) + (outbound_record.package_count or 0)

            # 确保车牌信息正确
            if not existing_inventory.plate_number:
                # 从原始入库记录获取车牌信息
                if original_identification_code:
                    original_inbound = InboundRecord.query.filter_by(
                        identification_code=original_identification_code
                    ).first()
                    if original_inbound and original_inbound.plate_number:
                        existing_inventory.plate_number = original_inbound.plate_number

            existing_inventory.last_updated = datetime.now()
            existing_inventory.version += 1
            current_app.logger.info(f'快速接收-合并后端仓库存: {existing_inventory.customer_name} {existing_inventory.identification_code} '
                                  f'原始编码: {original_identification_code} 新增板数{outbound_record.pallet_count} 件数{outbound_record.package_count}')
        else:
            # 获取原始入库车牌信息
            original_plate_number = ''
            if outbound_record.identification_code:
                original_inbound = InboundRecord.query.filter_by(
                    identification_code=outbound_record.identification_code
                ).first()
                if original_inbound:
                    original_plate_number = original_inbound.plate_number or ''

            # 创建新库存记录
            inventory_record = Inventory(
                customer_name=outbound_record.customer_name,
                identification_code=outbound_record.identification_code,
                original_identification_code=original_identification_code,  # 设置原始识别编码
                inbound_pallet_count=outbound_record.pallet_count,
                inbound_package_count=outbound_record.package_count,
                pallet_count=outbound_record.pallet_count,
                package_count=outbound_record.package_count,
                weight=outbound_record.weight,
                volume=outbound_record.volume,
                export_mode=outbound_record.export_mode,
                order_type=outbound_record.order_type,
                customs_broker=outbound_record.customs_broker,
                documents=outbound_record.documents,
                inbound_time=datetime.now(),
                plate_number=original_plate_number,  # 使用原始入库车牌
                service_staff=outbound_record.service_staff,
                operated_by_user_id=current_user.id if hasattr(current_user, 'id') else None,
                operated_warehouse_id=backend_warehouse.id,
                last_updated=datetime.now(),
                version=1
            )
            db.session.add(inventory_record)
            current_app.logger.info(f'快速接收-创建后端仓库存: {inventory_record.customer_name} {inventory_record.identification_code} '
                                  f'原始编码: {original_identification_code}')

        # 更新对应的在途货物状态
        transit_cargo = TransitCargo.query.filter_by(
            customer_name=outbound_record.customer_name,
            identification_code=outbound_record.identification_code,
            batch_no=outbound_record.batch_no,
            status='in_transit'
        ).first()

        if transit_cargo:
            # 更新在途货物状态为已接收
            transit_cargo.status = 'received'
            transit_cargo.actual_arrival_time = datetime.now()
            transit_cargo.received_time = datetime.now()
            transit_cargo.received_by_user_id = current_user.id if hasattr(current_user, 'id') else None
            transit_cargo.last_updated = datetime.now()

            current_app.logger.info(f'快速接收-更新在途货物状态: {transit_cargo.customer_name} {transit_cargo.identification_code} '
                                  f'状态: in_transit → received')
        else:
            current_app.logger.warning(f'快速接收-未找到对应的在途货物记录: {outbound_record.customer_name} {outbound_record.identification_code}')

        db.session.commit()

        current_app.logger.info(f"快速接收成功，创建入库记录ID: {inbound_record.id}，后端仓库存已更新")
        return jsonify({'success': True, 'message': '快速接收成功，后端仓库存已更新', 'inbound_id': inbound_record.id})

    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"快速接收失败: {str(e)}", exc_info=True)
        return jsonify({'success': False, 'message': f'接收失败: {str(e)}'}), 500


# 获取货物详情API
@bp.route('/api/backend/cargo-detail/<int:item_id>', methods=['GET'])
@csrf.exempt
def api_backend_cargo_detail(item_id):
    """获取货物详情"""
    try:
        # 获取出库记录
        outbound_record = OutboundRecord.query.options(
            db.joinedload(OutboundRecord.operated_warehouse)
        ).get(item_id)

        if not outbound_record:
            return jsonify({'success': False, 'message': '找不到对应的货物记录'}), 404

        # 返回详细信息
        detail_data = {
            'id': outbound_record.id,
            'outbound_time': outbound_record.outbound_time.strftime('%Y-%m-%d %H:%M') if outbound_record.outbound_time else '',
            'plate_number': outbound_record.plate_number or '',
            'customer_name': outbound_record.customer_name or '',
            'order_type': outbound_record.order_type or '',
            'customs_broker': outbound_record.customs_broker or '',
            'pallet_count': outbound_record.pallet_count or 0,
            'package_count': outbound_record.package_count or 0,
            'weight': outbound_record.weight or 0,
            'volume': outbound_record.volume or 0,
            'documents': outbound_record.documents or '',
            'service_staff': outbound_record.service_staff or '',
            'source_warehouse': outbound_record.operated_warehouse.warehouse_name if outbound_record.operated_warehouse else '',
            'batch_no': outbound_record.batch_no or '',
            'batch_sequence': outbound_record.batch_sequence or 0,
            'batch_total': outbound_record.batch_total or 0,
            'remark1': outbound_record.remark1 or '',
            'remark2': outbound_record.remark2 or '',
            'create_time': outbound_record.outbound_time.strftime('%Y-%m-%d %H:%M') if outbound_record.outbound_time else ''
        }

        return jsonify({'success': True, 'data': detail_data})

    except Exception as e:
        current_app.logger.error(f"获取货物详情失败: {str(e)}")
        return jsonify({'success': False, 'message': f'获取详情失败: {str(e)}'}), 500


# 库存平移API
@bp.route('/api/inventory/transfer', methods=['POST'])
@csrf.exempt
def api_inventory_transfer():
    """库存平移API - 从前端仓转移到后端仓"""
    try:
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'message': '缺少请求数据'})

        source_warehouse = data.get('source_warehouse')  # 来源仓库
        target_warehouse = data.get('target_warehouse', '凭祥北投仓')  # 目标仓库
        inventory_ids = data.get('inventory_ids', [])  # 要转移的库存ID列表
        transfer_reason = data.get('transfer_reason', '前端仓到后端仓转运')

        if not inventory_ids:
            return jsonify({'success': False, 'message': '请选择要转移的库存'})

        transferred_count = 0
        transfer_details = []

        for inventory_id in inventory_ids:
            inventory = Inventory.query.get(inventory_id)
            if not inventory:
                continue

            # 记录转移前的信息
            transfer_detail = {
                'inventory_id': inventory_id,
                'customer_name': inventory.customer_name,
                'identification_code': inventory.identification_code,
                'pallet_count': inventory.pallet_count,
                'package_count': inventory.package_count,
                'source_location': inventory.location,
                'target_location': f'{target_warehouse}-待分配'
            }

            # 更新库存位置信息
            inventory.location = f'{target_warehouse}-待分配'
            inventory.last_updated = datetime.now()

            # 在documents字段中记录转移信息
            if inventory.documents:
                inventory.documents += f' | 转移:{source_warehouse}→{target_warehouse}'
            else:
                inventory.documents = f'转移:{source_warehouse}→{target_warehouse}'

            transfer_details.append(transfer_detail)
            transferred_count += 1

        db.session.commit()

        return jsonify({
            'success': True,
            'message': f'成功转移 {transferred_count} 条库存记录',
            'transferred_count': transferred_count,
            'transfer_details': transfer_details
        })

    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"库存平移失败: {str(e)}")
        return jsonify({'success': False, 'message': f'库存平移失败: {str(e)}'})


@bp.route('/backend/inbound', methods=['GET', 'POST'])
@require_permission('INBOUND_VIEW')
def backend_inbound():
    """后端仓直接入库操作页面"""
    if not check_warehouse_permission('backend', 'view'):
        flash('您没有权限访问后端仓入库功能', 'error')
        return redirect(url_for('main.index'))

    from app.utils import render_ajax_aware
    return render_ajax_aware('backend/receive_batch.html',
                           title='后端仓入库操作',
                           warehouse_type='backend')

@bp.route('/api/backend/inbound/direct/batch', methods=['POST'])
@csrf.exempt
@require_permission('INBOUND_CREATE')
def api_backend_inbound_direct_batch():
    """后端仓直接入库批量保存API"""
    try:
        data = request.get_json()
        if not data or 'records' not in data:
            return jsonify({'success': False, 'message': '请求数据格式错误'}), 400

        records = data['records']
        if not records:
            return jsonify({'success': False, 'message': '没有要保存的记录'}), 400

        # 检查权限
        if not check_warehouse_permission('backend', 'add'):
            return jsonify({'success': False, 'message': '您没有权限执行后端仓入库操作'}), 403

        # 获取后端仓库
        backend_warehouse = Warehouse.query.filter_by(warehouse_type='backend').first()
        if not backend_warehouse:
            return jsonify({'success': False, 'message': '找不到后端仓库'}), 500

        success_count = 0
        errors = []

        for i, record_data in enumerate(records):
            try:
                # 验证必填字段 - 统一与前端API的必填字段要求
                required_fields = ['inbound_time', 'plate_number', 'customer_name', 'service_staff']
                missing_fields = []
                for field in required_fields:
                    if field not in record_data or not record_data[field]:
                        missing_fields.append(field)

                if missing_fields:
                    errors.append(f"第{i+1}条记录缺少必填字段: {', '.join(missing_fields)}")
                    continue

                # 转换时间格式 - 支持多种格式
                if isinstance(record_data['inbound_time'], str):
                    try:
                        # 尝试多种日期格式
                        formats_to_try = [
                            '%Y-%m-%dT%H:%M',      # HTML5日期时间格式
                            '%Y-%m-%d %H:%M:%S',   # 完整日期时间
                            '%Y-%m-%d',            # 标准日期格式
                            '%Y/%m/%d',            # 斜杠分隔
                            '%Y.%m.%d',            # 点分隔
                        ]

                        inbound_time = None
                        for date_format in formats_to_try:
                            try:
                                inbound_time = datetime.strptime(record_data['inbound_time'], date_format)
                                break
                            except ValueError:
                                continue

                        if inbound_time is None:
                            errors.append(f"第{i+1}条记录入库时间格式错误")
                            continue

                    except Exception as e:
                        errors.append(f"第{i+1}条记录入库时间格式错误: {str(e)}")
                        continue
                elif isinstance(record_data['inbound_time'], datetime):
                    inbound_time = record_data['inbound_time']
                else:
                    errors.append(f"第{i+1}条记录入库时间格式错误")
                    continue

                # 验证件数和板数
                try:
                    # 处理空字符串的情况
                    pallet_count_str = record_data.get('pallet_count', '').strip()
                    package_count_str = record_data.get('package_count', '').strip()

                    pallet_count = int(pallet_count_str) if pallet_count_str else 0
                    package_count = int(package_count_str) if package_count_str else 0

                    if pallet_count < 0 or package_count < 0:
                        errors.append(f"第{i+1}条记录错误: 板数和件数不能为负数")
                        continue

                    # 如果用户填写了数据，至少需要填写板数或件数中的一项
                    if (pallet_count_str or package_count_str) and pallet_count == 0 and package_count == 0:
                        errors.append(f"第{i+1}条记录错误: 板数和件数至少需要填写一项且不能为零")
                        continue

                    # 如果都没填写，设置默认值
                    if not pallet_count_str and not package_count_str:
                        pallet_count = 1  # 默认1板

                except (ValueError, TypeError):
                    errors.append(f"第{i+1}条记录错误: 板数和件数必须是有效的整数")
                    continue

                # 处理重量和体积
                try:
                    weight = float(record_data.get('weight', 0) or 0)
                    volume = float(record_data.get('volume', 0) or 0)
                    if weight < 0 or volume < 0:
                        errors.append(f"第{i+1}条记录错误: 重量和体积必须为非负数")
                        continue
                except (ValueError, TypeError):
                    errors.append(f"第{i+1}条记录错误: 重量和体积必须是有效数字")
                    continue

                # 自动生成识别编码 - 统一使用IdentificationCodeGenerator
                customer_name = record_data['customer_name']
                plate_number = record_data['plate_number']

                identification_code = IdentificationCodeGenerator.generate_unique_identification_code(
                    warehouse_id=backend_warehouse.id,
                    customer_name=customer_name,
                    plate_number=plate_number,
                    operation_type='inbound',
                    inbound_date=inbound_time
                )

                # 创建入库记录
                inbound_record = InboundRecord(
                    inbound_time=inbound_time,
                    plate_number=plate_number,
                    customer_name=customer_name,
                    identification_code=identification_code,
                    inbound_plate=record_data.get('inbound_plate', ''),
                    order_type=record_data.get('order_type', ''),
                    export_mode=record_data.get('export_mode', ''),
                    customs_broker=record_data.get('customs_broker', ''),
                    pallet_count=pallet_count,
                    package_count=package_count,
                    weight=weight,
                    volume=volume,
                    location=record_data.get('location', ''),
                    documents=record_data.get('documents', ''),
                    service_staff=record_data['service_staff'],
                    record_type='direct',  # 直接入库记录
                    operated_warehouse_id=backend_warehouse.id,
                    operated_by_user_id=current_user.id
                )

                # 添加备注信息
                remarks = []
                if record_data.get('remark1'):
                    remarks.append(record_data['remark1'])
                if record_data.get('remark2'):
                    remarks.append(record_data['remark2'])
                if remarks:
                    inbound_record.documents = f"{inbound_record.documents} 备注: {'; '.join(remarks)}"

                db.session.add(inbound_record)

                # 创建或更新库存记录 - 按照识别编码管理库存
                inventory = Inventory.query.filter_by(identification_code=identification_code).first()

                if not inventory:
                    inventory = Inventory(
                        customer_name=customer_name,
                        identification_code=identification_code,
                        inbound_pallet_count=pallet_count,
                        inbound_package_count=package_count,
                        pallet_count=pallet_count,
                        package_count=package_count,
                        weight=weight,
                        volume=volume,
                        location=record_data.get('location', ''),
                        documents=record_data.get('documents', ''),
                        export_mode=record_data.get('export_mode', ''),
                        order_type=record_data.get('order_type', ''),
                        customs_broker=record_data.get('customs_broker', ''),
                        inbound_time=inbound_time,
                        plate_number=plate_number,
                        service_staff=record_data['service_staff'],
                        operated_warehouse_id=backend_warehouse.id,
                        operated_by_user_id=current_user.id
                    )
                    db.session.add(inventory)
                else:
                    # 如果已存在，累加数量
                    inventory.inbound_pallet_count += pallet_count
                    inventory.inbound_package_count += package_count
                    inventory.pallet_count += pallet_count
                    inventory.package_count += package_count
                    inventory.weight += weight
                    inventory.volume += volume

                success_count += 1

            except Exception as e:
                errors.append(f'第{i+1}条记录保存失败: {str(e)}')
                continue

        # 提交事务
        if success_count > 0:
            db.session.commit()
            current_app.logger.info(f"后端仓直接入库成功保存 {success_count} 条记录")
        else:
            db.session.rollback()

        return jsonify({
            'success': success_count > 0,
            'saved_count': success_count,
            'errors': errors,
            'message': f'成功保存 {success_count} 条记录' + (f'，{len(errors)} 条失败' if errors else '')
        })

    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"后端仓直接入库批量保存失败: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'批量入库失败: {str(e)}'
        }), 500




@bp.route('/backend/inbound/list')
@require_permission('INBOUND_VIEW')
def backend_inbound_list():
    """后端仓入库记录列表"""
    if not check_warehouse_permission('backend', 'view'):
        flash('您没有权限访问后端仓入库记录', 'error')
        return redirect(url_for('main.index'))

    # 获取搜索参数
    page = request.args.get('page', 1, type=int)
    per_page = 50

    # 构建查询，只显示后端仓的数据
    query = InboundRecord.query.options(
        db.joinedload(InboundRecord.operated_warehouse)
    )

    # 获取后端仓库，只显示直接入库的记录（没有批次号的）
    backend_warehouse = Warehouse.query.filter_by(warehouse_type='backend').first()
    if backend_warehouse:
        query = query.filter_by(operated_warehouse_id=backend_warehouse.id)
        # 只显示直接入库的记录（没有批次号或批次号为空）
        query = query.filter(db.or_(InboundRecord.batch_no.is_(None), InboundRecord.batch_no == ''))

    # 获取日期参数，如果没有指定则使用默认值
    date_start = request.args.get('date_start', '')
    date_end = request.args.get('date_end', '')

    # 如果没有指定日期范围，默认使用最近一周的日期范围
    if not date_start and not date_end:
        today = datetime.now().strftime('%Y-%m-%d')
        one_week_ago = (datetime.now() - timedelta(days=7)).strftime('%Y-%m-%d')
        date_start = one_week_ago
        date_end = today

    # 搜索过滤 - 支持新的搜索方式
    search_params = {
        'date_start': date_start,
        'date_end': date_end
    }

    # 获取搜索字段和值
    search_field = request.args.get('search_field', '')
    search_value = request.args.get('search_value', '')
    search_condition = request.args.get('search_condition', 'contains')

    if search_field and search_value:
        search_params['search_field'] = search_field
        search_params['search_value'] = search_value
        search_params['search_condition'] = search_condition

        # 根据搜索条件应用过滤
        if search_condition == 'exact':
            filter_condition = getattr(InboundRecord, search_field) == search_value
        elif search_condition == 'startswith':
            filter_condition = getattr(InboundRecord, search_field).like(f'{search_value}%')
        elif search_condition == 'endswith':
            filter_condition = getattr(InboundRecord, search_field).like(f'%{search_value}')
        else:  # contains
            filter_condition = getattr(InboundRecord, search_field).like(f'%{search_value}%')

        query = query.filter(filter_condition)

    # 兼容旧的搜索参数
    # 客户名称搜索
    if request.args.get('customer_name'):
        query = query.filter(InboundRecord.customer_name.contains(request.args.get('customer_name')))
        search_params['customer_name'] = request.args.get('customer_name')

    # 车牌号搜索
    if request.args.get('plate_number'):
        query = query.filter(InboundRecord.plate_number.contains(request.args.get('plate_number')))
        search_params['plate_number'] = request.args.get('plate_number')

    # 批次号搜索
    if request.args.get('batch_no'):
        query = query.filter(InboundRecord.batch_no.contains(request.args.get('batch_no')))
        search_params['batch_no'] = request.args.get('batch_no')

    # 识别编码搜索
    if request.args.get('identification_code'):
        query = query.filter(InboundRecord.identification_code.contains(request.args.get('identification_code')))
        search_params['identification_code'] = request.args.get('identification_code')

    # 订单类型搜索
    if request.args.get('order_type'):
        query = query.filter(InboundRecord.order_type == request.args.get('order_type'))
        search_params['order_type'] = request.args.get('order_type')

    # 出境模式搜索
    if request.args.get('export_mode'):
        query = query.filter(InboundRecord.export_mode.contains(request.args.get('export_mode')))
        search_params['export_mode'] = request.args.get('export_mode')

    # 报关行搜索
    if request.args.get('customs_broker'):
        query = query.filter(InboundRecord.customs_broker.contains(request.args.get('customs_broker')))
        search_params['customs_broker'] = request.args.get('customs_broker')

    # 跟单客服搜索
    if request.args.get('service_staff'):
        query = query.filter(InboundRecord.service_staff.contains(request.args.get('service_staff')))
        search_params['service_staff'] = request.args.get('service_staff')

    # 库位搜索
    if request.args.get('location'):
        query = query.filter(InboundRecord.location.contains(request.args.get('location')))
        search_params['location'] = request.args.get('location')

    # 入库类型搜索
    if request.args.get('inbound_type'):
        query = query.filter(InboundRecord.inbound_type.contains(request.args.get('inbound_type')))
        search_params['inbound_type'] = request.args.get('inbound_type')

    # 来源仓库搜索（暂时跳过，等添加source_outbound关系后再实现）
    if request.args.get('source_warehouse'):
        source_warehouse_name = request.args.get('source_warehouse')
        # TODO: 实现来源仓库搜索，需要添加source_outbound_id字段
        search_params['source_warehouse'] = source_warehouse_name

    # 日期范围搜索
    if date_start:
        try:
            start_date = datetime.strptime(date_start, '%Y-%m-%d')
            query = query.filter(InboundRecord.inbound_time >= start_date)
        except ValueError:
            pass

    if date_end:
        try:
            end_date = datetime.strptime(date_end, '%Y-%m-%d')
            # 设置为当天的23:59:59
            end_date = end_date.replace(hour=23, minute=59, second=59)
            query = query.filter(InboundRecord.inbound_time <= end_date)
        except ValueError:
            pass

    # 排序和分页
    records = query.order_by(InboundRecord.inbound_time.desc()).paginate(
        page=page, per_page=per_page, error_out=False
    )

    from app.utils import render_ajax_aware
    return render_ajax_aware('backend/inbound_records_list.html',
                           records=records,
                           search_params=search_params,
                           title='后端仓入库记录',
                           warehouse_type='backend')


@bp.route('/export_backend_inbound')
@require_permission('INBOUND_VIEW')
def export_backend_inbound():
    """导出后端仓入库记录到Excel"""
    try:
        # 记录导出操作开始
        current_app.logger.info(f"用户 {current_user.username} 开始导出后端仓入库记录")

        # 检查用户权限
        if not current_user.is_authenticated:
            current_app.logger.warning("未登录用户尝试导出数据")
            flash('请先登录后再导出数据', 'warning')
            return redirect(url_for('auth.login'))

        # 检查用户是否有查看权限
        if not current_user.has_permission('INBOUND_VIEW'):
            current_app.logger.warning(f"用户 {current_user.username} 没有入库查看权限")
            flash('您没有权限导出入库数据', 'error')
            return redirect(url_for('main.index'))
        # 获取搜索参数（与backend_inbound_list相同的参数处理）
        date_start = request.args.get('date_start', '')
        date_end = request.args.get('date_end', '')

        # 如果没有指定日期范围，默认使用最近一周的日期范围
        if not date_start and not date_end:
            today = datetime.now().strftime('%Y-%m-%d')
            one_week_ago = (datetime.now() - timedelta(days=7)).strftime('%Y-%m-%d')
            date_start = one_week_ago
            date_end = today

        # 构建查询，只显示后端仓的数据
        query = InboundRecord.query.options(
            db.joinedload(InboundRecord.operated_warehouse)
        )

        # 获取后端仓库，只显示直接入库的记录（没有批次号的）
        backend_warehouse = Warehouse.query.filter_by(warehouse_type='backend').first()
        if backend_warehouse:
            query = query.filter_by(operated_warehouse_id=backend_warehouse.id)
            # 只显示直接入库的记录（没有批次号或批次号为空）
            query = query.filter(db.or_(InboundRecord.batch_no.is_(None), InboundRecord.batch_no == ''))

        # 应用搜索过滤（与backend_inbound_list相同的逻辑）
        search_field = request.args.get('search_field', '')
        search_value = request.args.get('search_value', '')
        search_condition = request.args.get('search_condition', 'contains')

        if search_field and search_value:
            # 根据搜索条件应用过滤
            if search_condition == 'exact':
                filter_condition = getattr(InboundRecord, search_field) == search_value
            elif search_condition == 'startswith':
                filter_condition = getattr(InboundRecord, search_field).like(f'{search_value}%')
            elif search_condition == 'endswith':
                filter_condition = getattr(InboundRecord, search_field).like(f'%{search_value}')
            else:  # contains
                filter_condition = getattr(InboundRecord, search_field).like(f'%{search_value}%')

            query = query.filter(filter_condition)

        # 兼容旧的搜索参数
        if request.args.get('customer_name'):
            query = query.filter(InboundRecord.customer_name.contains(request.args.get('customer_name')))
        if request.args.get('plate_number'):
            query = query.filter(InboundRecord.plate_number.contains(request.args.get('plate_number')))
        if request.args.get('batch_no'):
            query = query.filter(InboundRecord.batch_no.contains(request.args.get('batch_no')))
        if request.args.get('identification_code'):
            query = query.filter(InboundRecord.identification_code.contains(request.args.get('identification_code')))
        if request.args.get('order_type'):
            query = query.filter(InboundRecord.order_type == request.args.get('order_type'))
        if request.args.get('export_mode'):
            query = query.filter(InboundRecord.export_mode.contains(request.args.get('export_mode')))
        if request.args.get('customs_broker'):
            query = query.filter(InboundRecord.customs_broker.contains(request.args.get('customs_broker')))
        if request.args.get('service_staff'):
            query = query.filter(InboundRecord.service_staff.contains(request.args.get('service_staff')))
        if request.args.get('location'):
            query = query.filter(InboundRecord.location.contains(request.args.get('location')))
        if request.args.get('inbound_type'):
            query = query.filter(InboundRecord.inbound_type.contains(request.args.get('inbound_type')))

        # 日期范围搜索
        if date_start:
            try:
                start_date = datetime.strptime(date_start, '%Y-%m-%d')
                query = query.filter(InboundRecord.inbound_time >= start_date)
            except ValueError:
                pass

        if date_end:
            try:
                end_date = datetime.strptime(date_end, '%Y-%m-%d')
                end_date = end_date.replace(hour=23, minute=59, second=59)
                query = query.filter(InboundRecord.inbound_time <= end_date)
            except ValueError:
                pass

        # 按入库时间升序排序
        query = query.order_by(InboundRecord.inbound_time.asc())

        # 获取所有符合条件的记录
        records = query.all()
        current_app.logger.info(f"查询到 {len(records)} 条后端仓入库记录")

        # 如果没有记录，返回提示
        if not records:
            current_app.logger.info("没有找到符合条件的后端仓入库记录")
            flash('没有找到符合条件的后端仓入库记录', 'warning')
            return redirect(url_for('main.backend_inbound_list'))

        try:
            # 尝试导入pandas
            import pandas as pd
            from pandas import ExcelWriter
            from io import BytesIO

            # 构建DataFrame
            data = []
            for i, record in enumerate(records, 1):
                data.append({
                    '序号': i,
                    '入库时间': record.outbound_time.strftime('%Y-%m-%d') if record.outbound_time else '',
                    '入库车牌': record.plate_number or '',
                    '客户名称': record.customer_name or '',
                    '识别编码': record.identification_code or '',
                    '板数': record.pallet_count or 0,
                    '件数': record.package_count or 0,
                    '重量(kg)': record.weight or 0,
                    '体积(m³)': record.volume or 0,
                    '订单类型': record.order_type or '',
                    '出境模式': record.export_mode or '',
                    '报关行': record.customs_broker or '',
                    '库位': record.location or '',
                    '单据': record.documents or '',
                    '跟单客服': record.service_staff or '',
                    '操作仓库': record.operated_warehouse.warehouse_name if record.operated_warehouse else '',
                    '创建时间': record.outbound_time.strftime('%Y-%m-%d %H:%M:%S') if record.outbound_time else ''
                })

            df = pd.DataFrame(data)

            # 创建内存文件对象
            output = BytesIO()

            # 使用ExcelWriter可以更好地控制格式
            with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
                # 写入数据
                df.to_excel(writer, sheet_name='后端仓入库记录', index=False)

                # 获取workbook和worksheet对象以进行格式调整
                workbook = writer.book
                worksheet = writer.sheets['后端仓入库记录']

                # 定义格式
                header_format = workbook.add_format({
                    'bold': True,
                    'text_wrap': True,
                    'valign': 'vcenter',
                    'align': 'center',
                    'bg_color': '#FFE6E6',  # 浅红色，符合后端仓主题
                    'border': 1
                })

                # 为所有列设置宽度
                for i, col in enumerate(df.columns):
                    # 根据列名和内容设置适当的列宽
                    max_len = max(
                        df[col].astype(str).map(len).max(),  # 最长数据长度
                        len(str(col))  # 列名长度
                    ) + 2  # 添加一些额外空间

                    # 限制最大宽度
                    col_width = min(max_len, 30)
                    worksheet.set_column(i, i, col_width)

                # 设置表头格式
                for col_num, value in enumerate(df.columns.values):
                    worksheet.write(0, col_num, value, header_format)

                # 添加自动筛选
                worksheet.autofilter(0, 0, len(df), len(df.columns) - 1)

            # 设置文件指针到开始位置
            output.seek(0)

            # 生成下载文件名
            timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
            filename = f"后端仓入库记录导出_{timestamp}.xlsx"

            # 记录成功导出
            current_app.logger.info(f"用户 {current_user.username} 成功导出 {len(records)} 条后端仓入库记录，文件名: {filename}")

            # 返回Excel文件
            return send_file(
                output,
                as_attachment=True,
                download_name=filename,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )

        except ImportError:
            flash("导出Excel需要pandas或xlsxwriter库，请安装: pip install pandas xlsxwriter 或 pip install openpyxl", "danger")
            return redirect(url_for('main.backend_inbound_list'))

    except Exception as e:
        current_app.logger.error(f"导出后端仓入库记录时出错: {str(e)}")
        flash(f"导出数据时出错: {str(e)}", "danger")
        return redirect(url_for('main.backend_inbound_list'))


@bp.route('/backend/receive/list')
@require_permission('INBOUND_VIEW')
def backend_receive_list():
    """后端仓接收记录列表（重定向到入库记录）"""
    return redirect(url_for('main.backend_inbound_list'))


# 后端仓接收记录列表（只显示从前端仓接收的货物）
@bp.route('/backend/receive/records')
@require_permission('INBOUND_VIEW')
def backend_receive_records():
    """后端仓接收记录列表（只显示从前端仓接收的货物）"""
    if not check_warehouse_permission('backend', 'view'):
        flash('您没有权限访问后端仓接收记录', 'error')
        return redirect(url_for('main.index'))

    # 获取搜索参数
    page = request.args.get('page', 1, type=int)
    per_page = 15  # 每页显示15个批次记录

    # 构建查询，显示ReceiveRecord表中的接收记录
    query = ReceiveRecord.query.options(
        db.joinedload(ReceiveRecord.operated_warehouse)
    )

    # 获取后端仓库，只显示后端仓的接收记录
    backend_warehouse = Warehouse.query.filter_by(warehouse_type='backend').first()
    if backend_warehouse:
        # 只显示后端仓自己的接收记录
        query = query.filter(ReceiveRecord.operated_warehouse_id == backend_warehouse.id)

    # 获取日期参数，如果没有指定则使用默认值
    date_start = request.args.get('date_start', '')
    date_end = request.args.get('date_end', '')

    # 如果没有指定日期范围，默认使用最近一周的日期范围
    if not date_start and not date_end:
        today = datetime.now().strftime('%Y-%m-%d')
        one_week_ago = (datetime.now() - timedelta(days=7)).strftime('%Y-%m-%d')
        date_start = one_week_ago
        date_end = today

    # 搜索过滤 - 支持新的搜索方式
    search_params = {
        'date_start': date_start,
        'date_end': date_end
    }

    # 获取搜索字段和值
    search_field = request.args.get('search_field', '')
    search_value = request.args.get('search_value', '')
    search_condition = request.args.get('search_condition', 'contains')

    if search_field and search_value:
        search_params['search_field'] = search_field
        search_params['search_value'] = search_value
        search_params['search_condition'] = search_condition

        # 根据搜索条件应用过滤
        if search_field == 'source_warehouse':
            # 来源仓库特殊处理
            if '平湖' in search_value or 'PH' in search_value.upper():
                filter_condition = ReceiveRecord.batch_no.like('PH%')
            elif '昆山' in search_value or 'KS' in search_value.upper():
                filter_condition = ReceiveRecord.batch_no.like('KS%')
            elif '成都' in search_value or 'CD' in search_value.upper():
                filter_condition = ReceiveRecord.batch_no.like('CD%')
            else:
                # 模糊匹配shipping_warehouse字段
                if search_condition == 'exact':
                    filter_condition = ReceiveRecord.shipping_warehouse == search_value
                elif search_condition == 'startswith':
                    filter_condition = ReceiveRecord.shipping_warehouse.like(f'{search_value}%')
                elif search_condition == 'endswith':
                    filter_condition = ReceiveRecord.shipping_warehouse.like(f'%{search_value}')
                else:  # contains
                    filter_condition = ReceiveRecord.shipping_warehouse.like(f'%{search_value}%')
        else:
            # 其他字段的正常处理
            if search_condition == 'exact':
                filter_condition = getattr(ReceiveRecord, search_field) == search_value
            elif search_condition == 'startswith':
                filter_condition = getattr(ReceiveRecord, search_field).like(f'{search_value}%')
            elif search_condition == 'endswith':
                filter_condition = getattr(ReceiveRecord, search_field).like(f'%{search_value}')
            else:  # contains
                filter_condition = getattr(ReceiveRecord, search_field).like(f'%{search_value}%')

        query = query.filter(filter_condition)

    # 兼容旧的搜索参数
    # 客户名称搜索
    if request.args.get('customer_name'):
        query = query.filter(ReceiveRecord.customer_name.contains(request.args.get('customer_name')))
        search_params['customer_name'] = request.args.get('customer_name')

    # 车牌号搜索（ReceiveRecord使用inbound_plate字段）
    if request.args.get('plate_number'):
        query = query.filter(ReceiveRecord.inbound_plate.contains(request.args.get('plate_number')))
        search_params['plate_number'] = request.args.get('plate_number')

    # 批次号搜索
    if request.args.get('batch_no'):
        query = query.filter(ReceiveRecord.batch_no.contains(request.args.get('batch_no')))
        search_params['batch_no'] = request.args.get('batch_no')

    # 跟单客服搜索
    if request.args.get('service_staff'):
        query = query.filter(ReceiveRecord.service_staff.contains(request.args.get('service_staff')))
        search_params['service_staff'] = request.args.get('service_staff')

    # 库位搜索（ReceiveRecord使用storage_location字段）
    if request.args.get('location'):
        query = query.filter(ReceiveRecord.storage_location.contains(request.args.get('location')))
        search_params['location'] = request.args.get('location')

    # 接收状态搜索
    if request.args.get('receive_status'):
        query = query.filter(ReceiveRecord.receive_status.contains(request.args.get('receive_status')))
        search_params['receive_status'] = request.args.get('receive_status')

    # 来源仓库搜索（根据批次号前缀判断）
    if request.args.get('source_warehouse'):
        source_warehouse_name = request.args.get('source_warehouse')
        search_params['source_warehouse'] = source_warehouse_name

        # 根据仓库名称确定批次号前缀
        if '平湖' in source_warehouse_name or 'PH' in source_warehouse_name.upper():
            query = query.filter(ReceiveRecord.batch_no.like('PH%'))
        elif '昆山' in source_warehouse_name or 'KS' in source_warehouse_name.upper():
            query = query.filter(ReceiveRecord.batch_no.like('KS%'))
        elif '成都' in source_warehouse_name or 'CD' in source_warehouse_name.upper():
            query = query.filter(ReceiveRecord.batch_no.like('CD%'))
        else:
            # 如果输入的是其他内容，尝试模糊匹配
            query = query.filter(ReceiveRecord.shipping_warehouse.contains(source_warehouse_name))

    # 日期范围搜索（ReceiveRecord使用receive_time字段）
    if date_start:
        try:
            start_date = datetime.strptime(date_start, '%Y-%m-%d')
            query = query.filter(ReceiveRecord.receive_time >= start_date)
        except ValueError:
            pass

    if date_end:
        try:
            end_date = datetime.strptime(date_end, '%Y-%m-%d')
            # 设置为当天的23:59:59
            end_date = end_date.replace(hour=23, minute=59, second=59)
            query = query.filter(ReceiveRecord.receive_time <= end_date)
        except ValueError:
            pass

    # 获取所有记录，按接收日期降序排序 - 限制查询数量以提高性能
    max_records = 200  # 减少查询数量以提高性能
    all_records = query.order_by(ReceiveRecord.receive_time.desc(), ReceiveRecord.batch_no.desc()).limit(max_records).all()

    # 优化：批量获取OutboundRecord数据，避免N+1查询问题
    identification_codes = [record.identification_code for record in all_records if record.identification_code]

    # 一次性查询所有相关的OutboundRecord
    outbound_records = {}
    if identification_codes:
        outbound_list = OutboundRecord.query.filter(
            OutboundRecord.identification_code.in_(identification_codes)
        ).all()

        # 建立识别编码到OutboundRecord的映射
        for outbound in outbound_list:
            outbound_records[outbound.identification_code] = outbound

    # 处理字段为空或'None'字符串的情况
    def get_valid_value(receive_value, outbound_value):
        """获取有效值，处理None和'None'字符串"""
        # 检查第一个值是否有效
        if receive_value and receive_value != 'None' and receive_value.strip():
            return receive_value
        # 检查第二个值是否有效
        if outbound_value and outbound_value != 'None' and outbound_value.strip():
            return outbound_value
        # 都无效则返回None
        return None

    # 为每个ReceiveRecord补充OutboundRecord数据和来源仓库信息
    for record in all_records:
        # 根据批次号前缀确定来源仓库
        if record.batch_no:
            if record.batch_no.startswith('PH'):
                record.source_warehouse_name = '平湖仓'
            elif record.batch_no.startswith('KS'):
                record.source_warehouse_name = '昆山仓'
            elif record.batch_no.startswith('CD'):
                record.source_warehouse_name = '成都仓'
            elif record.batch_no.startswith('PX'):
                record.source_warehouse_name = '凭祥北投仓'
            else:
                # 如果shipping_warehouse字段有值且不是'None'字符串，使用它
                if record.shipping_warehouse and record.shipping_warehouse != 'None' and record.shipping_warehouse.strip():
                    record.source_warehouse_name = record.shipping_warehouse
                else:
                    record.source_warehouse_name = '未知来源'
        else:
            # 如果shipping_warehouse字段有值且不是'None'字符串，使用它
            if record.shipping_warehouse and record.shipping_warehouse != 'None' and record.shipping_warehouse.strip():
                record.source_warehouse_name = record.shipping_warehouse
            else:
                record.source_warehouse_name = '未知来源'

        if record.identification_code and record.identification_code in outbound_records:
            outbound = outbound_records[record.identification_code]

            # 将OutboundRecord的数据添加到ReceiveRecord对象中（只在ReceiveRecord字段为空时补充）
            if not record.export_mode:
                record.export_mode = getattr(outbound, 'export_mode', None)
            if not record.customs_broker:
                record.customs_broker = getattr(outbound, 'customs_broker', None)
            if not record.order_type:
                record.order_type = getattr(outbound, 'order_type', None)

            # 如果ReceiveRecord中的字段为空或'None'，使用OutboundRecord的数据
            record.delivery_plate_number = get_valid_value(
                record.delivery_plate_number,
                getattr(outbound, 'delivery_plate_number', None)
            )
            record.service_staff = get_valid_value(
                record.service_staff,
                getattr(outbound, 'service_staff', None)
            )
            # 优先使用OutboundRecord的document_count作为单据份数
            outbound_document_count = getattr(outbound, 'document_count', None)
            outbound_documents = getattr(outbound, 'documents', None)

            if outbound_document_count and outbound_document_count > 0:
                record.documents = str(outbound_document_count)
            else:
                record.documents = get_valid_value(
                    record.documents,
                    outbound_documents
                )
            record.inbound_plate = get_valid_value(
                record.inbound_plate,
                getattr(outbound, 'inbound_plate', None)
            )
            # OutboundRecord使用location字段，ReceiveRecord使用storage_location字段
            record.storage_location = get_valid_value(
                record.storage_location,
                getattr(outbound, 'location', None)
            )
            # 补充重量和体积数据（只在ReceiveRecord字段为空或0时补充）
            if not record.weight or record.weight == 0:
                outbound_weight = getattr(outbound, 'weight', 0)
                if outbound_weight:
                    record.weight = outbound_weight
            if not record.volume or record.volume == 0:
                outbound_volume = getattr(outbound, 'volume', 0)
                if outbound_volume:
                    record.volume = outbound_volume

        else:
            # 如果没有找到对应的OutboundRecord，设置默认值
            record.export_mode = None
            record.customs_broker = None
            record.order_type = None

    # 按批次号分组数据
    from collections import defaultdict, OrderedDict
    batch_groups = OrderedDict()

    for record in all_records:
        batch_no = record.batch_no or '未分配批次'
        if batch_no not in batch_groups:
            batch_groups[batch_no] = {
                'batch_info': {
                    'batch_no': batch_no,
                    'total_pallet_count': 0,
                    'total_package_count': 0,
                    'total_weight': 0,
                    'total_volume': 0,
                    'record_count': 0,
                    'first_receive_time': record.receive_time,
                    'last_receive_time': record.receive_time,
                    'delivery_plate_number': record.delivery_plate_number,
                    'shipping_warehouse': record.shipping_warehouse,
                    'source_warehouse_name': record.source_warehouse_name  # 添加来源仓库名称
                },
                'records': []
            }

        # 更新批次汇总信息
        batch_info = batch_groups[batch_no]['batch_info']
        batch_info['total_pallet_count'] += record.pallet_count or 0
        batch_info['total_package_count'] += record.package_count or 0
        batch_info['total_weight'] += record.weight or 0
        batch_info['total_volume'] += record.volume or 0
        batch_info['record_count'] += 1

        # 更新时间范围
        if record.receive_time:
            if not batch_info['first_receive_time'] or record.receive_time < batch_info['first_receive_time']:
                batch_info['first_receive_time'] = record.receive_time
            if not batch_info['last_receive_time'] or record.receive_time > batch_info['last_receive_time']:
                batch_info['last_receive_time'] = record.receive_time

        # 添加记录到批次组
        batch_groups[batch_no]['records'].append(record)

    # 分页处理（对批次进行分页）
    batch_list = list(batch_groups.items())
    total_batches = len(batch_list)

    # 计算分页
    start_idx = (page - 1) * per_page
    end_idx = start_idx + per_page
    paginated_batches = batch_list[start_idx:end_idx]

    # 创建分页对象
    class BatchPagination:
        def __init__(self, items, page, per_page, total):
            self.items = items
            self.page = page
            self.per_page = per_page
            self.total = total
            self.pages = (total + per_page - 1) // per_page
            self.has_prev = page > 1
            self.has_next = page < self.pages
            self.prev_num = page - 1 if self.has_prev else None
            self.next_num = page + 1 if self.has_next else None

    batch_pagination = BatchPagination(paginated_batches, page, per_page, total_batches)

    from app.utils import render_ajax_aware
    return render_ajax_aware('backend/receive_records_list.html',
                           batch_groups=batch_pagination,
                           search_params=search_params,
                           title='后端仓接收记录',
                           warehouse_type='backend')

@bp.route('/api/debug/outbound_data', methods=['GET'])
@csrf.exempt
def debug_outbound_data():
    """调试API：查看最近的出库记录数据"""
    try:
        # 查询最近的前端仓出库记录
        from app.models import OutboundRecord, Warehouse

        frontend_warehouses = Warehouse.query.filter_by(warehouse_type='frontend').all()
        frontend_warehouse_ids = [w.id for w in frontend_warehouses]

        recent_outbound = OutboundRecord.query.filter(
            OutboundRecord.operated_warehouse_id.in_(frontend_warehouse_ids)
        ).order_by(OutboundRecord.departure_time.desc()).limit(5).all()

        data = []
        for record in recent_outbound:
            data.append({
                'id': record.id,
                'customer_name': record.customer_name,
                'identification_code': record.identification_code,
                'weight': record.weight,
                'volume': record.volume,
                'departure_time': record.departure_time.strftime('%Y-%m-%d %H:%M:%S') if record.departure_time else None
            })

        return jsonify({
            'success': True,
            'data': data
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'message': str(e)
        }), 500

@bp.route('/api/debug/backend_inventory', methods=['GET'])
@csrf.exempt
def debug_backend_inventory_api():
    """调试API：查看后端仓库存数据"""
    try:
        from app.models import Inventory, Warehouse

        backend_warehouse = Warehouse.query.filter_by(warehouse_name='凭祥北投仓').first()
        if not backend_warehouse:
            return jsonify({'success': False, 'message': '未找到后端仓库'})

        recent_inventory = Inventory.query.filter(
            Inventory.operated_warehouse_id == backend_warehouse.id
        ).order_by(Inventory.last_updated.desc()).limit(5).all()

        data = []
        for record in recent_inventory:
            data.append({
                'id': record.id,
                'customer_name': record.customer_name,
                'identification_code': record.identification_code,
                'weight': record.weight,
                'volume': record.volume,
                'last_updated': record.last_updated.strftime('%Y-%m-%d %H:%M:%S') if record.last_updated else None
            })

        return jsonify({
            'success': True,
            'data': data
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'message': str(e)
        }), 500

@bp.route('/api/fix_backend_inventory_data', methods=['POST'])
@csrf.exempt
@require_permission('ADMIN')
def fix_backend_inventory_data():
    """修复后端仓库存数据的重量和体积"""
    try:
        from app.models import Inventory, Warehouse, ReceiveRecord, OutboundRecord

        # 获取后端仓库
        backend_warehouse = Warehouse.query.filter_by(warehouse_name='凭祥北投仓').first()
        if not backend_warehouse:
            return jsonify({'success': False, 'message': '未找到后端仓库'})

        # 查询所有后端仓库存记录
        inventory_records = Inventory.query.filter(
            Inventory.operated_warehouse_id == backend_warehouse.id
        ).all()

        fixed_count = 0
        error_count = 0
        details = []

        for inventory in inventory_records:
            try:
                # 查找对应的接收记录
                receive_record = ReceiveRecord.query.filter_by(
                    identification_code=inventory.identification_code,
                    operated_warehouse_id=backend_warehouse.id
                ).first()

                if receive_record and (receive_record.weight or receive_record.volume):
                    # 从接收记录获取数据
                    old_weight = inventory.weight
                    old_volume = inventory.volume
                    inventory.weight = receive_record.weight or 0
                    inventory.volume = receive_record.volume or 0

                    details.append({
                        'id': inventory.id,
                        'customer': inventory.customer_name,
                        'identification_code': inventory.identification_code,
                        'source': 'receive_record',
                        'old_weight': old_weight,
                        'new_weight': inventory.weight,
                        'old_volume': old_volume,
                        'new_volume': inventory.volume
                    })
                    fixed_count += 1

                else:
                    # 查找原始出库记录
                    outbound_record = OutboundRecord.query.filter_by(
                        identification_code=inventory.identification_code
                    ).first()

                    if outbound_record and (outbound_record.weight or outbound_record.volume):
                        old_weight = inventory.weight
                        old_volume = inventory.volume
                        inventory.weight = outbound_record.weight or 0
                        inventory.volume = outbound_record.volume or 0

                        details.append({
                            'id': inventory.id,
                            'customer': inventory.customer_name,
                            'identification_code': inventory.identification_code,
                            'source': 'outbound_record',
                            'old_weight': old_weight,
                            'new_weight': inventory.weight,
                            'old_volume': old_volume,
                            'new_volume': inventory.volume
                        })
                        fixed_count += 1
                    else:
                        error_count += 1
                        details.append({
                            'id': inventory.id,
                            'customer': inventory.customer_name,
                            'identification_code': inventory.identification_code,
                            'source': 'not_found',
                            'error': '未找到对应的接收记录或出库记录'
                        })

            except Exception as e:
                error_count += 1
                details.append({
                    'id': inventory.id,
                    'customer': inventory.customer_name,
                    'identification_code': inventory.identification_code,
                    'error': str(e)
                })

        # 提交更改
        if fixed_count > 0:
            db.session.commit()
            current_app.logger.info(f"数据修复完成：成功修复 {fixed_count} 条记录，失败 {error_count} 条")

        return jsonify({
            'success': True,
            'message': f'数据修复完成：成功修复 {fixed_count} 条记录，失败 {error_count} 条',
            'fixed_count': fixed_count,
            'error_count': error_count,
            'details': details
        })

    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"数据修复失败: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'数据修复失败: {str(e)}'
        }), 500

@bp.route('/api/debug/inbound_data', methods=['GET'])
@csrf.exempt
def debug_inbound_data():
    """调试API：查看最近的入库记录数据"""
    try:
        from app.models import InboundRecord, Warehouse

        frontend_warehouses = Warehouse.query.filter_by(warehouse_type='frontend').all()
        frontend_warehouse_ids = [w.id for w in frontend_warehouses]

        recent_inbound = InboundRecord.query.filter(
            InboundRecord.operated_warehouse_id.in_(frontend_warehouse_ids)
        ).order_by(InboundRecord.inbound_time.desc()).limit(5).all()

        data = []
        for record in recent_inbound:
            data.append({
                'id': record.id,
                'customer_name': record.customer_name,
                'identification_code': record.identification_code,
                'weight': record.weight,
                'volume': record.volume,
                'inbound_time': record.outbound_time.strftime('%Y-%m-%d %H:%M:%S') if record.outbound_time else None
            })

        return jsonify({
            'success': True,
            'data': data
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'message': str(e)
        }), 500

@bp.route('/fix_data')
@require_permission('ADMIN')
def fix_data_page():
    """数据修复页面"""
    return render_template('fix_data.html')

@bp.route('/export_backend_receive')
@require_permission('INBOUND_VIEW')
def export_backend_receive():
    """导出后端仓接收记录到Excel"""
    try:
        current_app.logger.info("开始导出后端仓接收记录")

        # 获取搜索参数（与backend_receive_records相同的参数处理）
        date_start = request.args.get('date_start', '')
        date_end = request.args.get('date_end', '')
        customer_name = request.args.get('customer_name', '')
        plate_number = request.args.get('plate_number', '')
        batch_no = request.args.get('batch_no', '')
        source_warehouse = request.args.get('source_warehouse', '')

        # 如果没有指定日期范围，默认使用最近一周的日期范围
        if not date_start and not date_end:
            today = datetime.now().strftime('%Y-%m-%d')
            one_week_ago = (datetime.now() - timedelta(days=7)).strftime('%Y-%m-%d')
            date_start = one_week_ago
            date_end = today

        current_app.logger.info(f"导出参数: 开始日期={date_start}, 结束日期={date_end}, 客户={customer_name}, 车牌={plate_number}, 批次={batch_no}")

        # 构建查询，使用ReceiveRecord表显示后端仓接收记录
        query = ReceiveRecord.query.options(
            db.joinedload(ReceiveRecord.operated_warehouse),
            db.joinedload(ReceiveRecord.created_by_user)
        )

        # 获取后端仓库，只显示后端仓的接收记录
        backend_warehouse = Warehouse.query.filter_by(warehouse_type='backend').first()
        if backend_warehouse:
            query = query.filter_by(operated_warehouse_id=backend_warehouse.id)

        # 应用搜索过滤
        if customer_name:
            query = query.filter(ReceiveRecord.customer_name.contains(customer_name))
        if plate_number:
            query = query.filter(ReceiveRecord.inbound_plate.contains(plate_number))
        if batch_no:
            query = query.filter(ReceiveRecord.batch_no.contains(batch_no))

        # 日期范围过滤
        if date_start:
            try:
                start_date = datetime.strptime(date_start, '%Y-%m-%d')
                query = query.filter(ReceiveRecord.receive_time >= start_date)
            except ValueError:
                current_app.logger.warning(f"无效的开始日期格式: {date_start}")
        if date_end:
            try:
                end_date = datetime.strptime(date_end, '%Y-%m-%d')
                end_date = end_date.replace(hour=23, minute=59, second=59)
                query = query.filter(ReceiveRecord.receive_time <= end_date)
            except ValueError:
                current_app.logger.warning(f"无效的结束日期格式: {date_end}")

        # 按接收时间降序排序，限制记录数量避免超时
        query = query.order_by(ReceiveRecord.receive_time.desc()).limit(5000)

        # 获取所有符合条件的记录
        current_app.logger.info("开始查询数据库记录")
        records = query.all()
        current_app.logger.info(f"查询到 {len(records)} 条记录")

        # 如果没有记录，返回提示
        if not records:
            flash('没有找到符合条件的接收记录', 'warning')
            return redirect(url_for('main.backend_receive_records'))

        # 准备导出数据，添加来源仓库信息
        current_app.logger.info("开始处理导出数据")
        data = []
        for i, record in enumerate(records):
            try:
                # 确定来源仓库
                if record.batch_no:
                    if record.batch_no.startswith('PH'):
                        source_warehouse_name = '平湖仓'
                    elif record.batch_no.startswith('KS'):
                        source_warehouse_name = '昆山仓'
                    elif record.batch_no.startswith('CD'):
                        source_warehouse_name = '成都仓'
                    elif record.batch_no.startswith('PX'):
                        source_warehouse_name = '凭祥北投仓'
                    else:
                        if record.shipping_warehouse and record.shipping_warehouse != 'None' and record.shipping_warehouse.strip():
                            source_warehouse_name = record.shipping_warehouse
                        else:
                            source_warehouse_name = '未知来源'
                else:
                    if record.shipping_warehouse and record.shipping_warehouse != 'None' and record.shipping_warehouse.strip():
                        source_warehouse_name = record.shipping_warehouse
                    else:
                        source_warehouse_name = '未知来源'

                data.append({
                    '接收时间': record.receive_time.strftime('%Y-%m-%d %H:%M:%S') if record.receive_time else '',
                    '批次号': record.batch_no or '',
                    '来源仓库': source_warehouse_name,
                    '接收状态': record.receive_status or '已接收',
                    '识别编码': record.identification_code or '',
                    '客户名称': record.customer_name or '',
                    '送货干线车': record.delivery_plate_number or '',
                    '入库车牌': record.inbound_plate or '',
                    '板数': record.pallet_count or 0,
                    '件数': record.package_count or 0,
                    '重量(kg)': record.weight or 0,
                    '体积(m³)': record.volume or 0,
                    '出境模式': record.export_mode or '',
                    '报关行': record.customs_broker or '',
                    '订单类型': record.order_type or '',
                    '跟单客服': record.service_staff or '',
                    '库位': record.storage_location or '',
                    '单据': record.documents or '',
                    '批次序号': f"{record.batch_sequence}/{record.batch_total}" if record.batch_sequence and record.batch_total else '',
                    '备注1': record.remark1 or '',
                    '备注2': record.remark2 or '',
                    '操作仓库': record.operated_warehouse.warehouse_name if record.operated_warehouse else '',
                    '操作用户': record.created_by_user.username if record.created_by_user else '',
                    '创建时间': record.created_at.strftime('%Y-%m-%d %H:%M:%S') if record.created_at else ''
                })

                # 每处理100条记录输出一次日志
                if (i + 1) % 100 == 0:
                    current_app.logger.info(f"已处理 {i + 1}/{len(records)} 条记录")

            except Exception as e:
                current_app.logger.error(f"处理第 {i+1} 条记录时出错: {str(e)}")
                current_app.logger.error(f"记录详情: ID={getattr(record, 'id', 'N/A')}, 客户={getattr(record, 'customer_name', 'N/A')}")
                current_app.logger.error(f"operated_warehouse类型: {type(record.operated_warehouse) if hasattr(record, 'operated_warehouse') else 'N/A'}")
                if hasattr(record, 'operated_warehouse') and record.operated_warehouse:
                    current_app.logger.error(f"operated_warehouse属性: {dir(record.operated_warehouse)}")
                continue

        current_app.logger.info(f"数据处理完成，共 {len(data)} 条有效记录")

        # 使用pandas导出Excel
        try:
            current_app.logger.info("开始生成Excel文件")
            import pandas as pd
            from io import BytesIO

            if not data:
                flash('没有数据可以导出', 'warning')
                return redirect(url_for('main.backend_receive_records'))

            df = pd.DataFrame(data)
            output = BytesIO()

            current_app.logger.info("开始写入Excel数据")
            with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
                df.to_excel(writer, sheet_name='后端仓接收记录', index=False)

                # 获取workbook和worksheet对象以进行格式调整
                workbook = writer.book
                worksheet = writer.sheets['后端仓接收记录']

                # 定义格式
                header_format = workbook.add_format({
                    'bold': True,
                    'text_wrap': True,
                    'valign': 'top',
                    'fg_color': '#D7E4BC',
                    'border': 1
                })

                # 应用表头格式
                for col_num, value in enumerate(df.columns.values):
                    worksheet.write(0, col_num, value, header_format)

                # 设置列宽
                worksheet.set_column('A:S', 15)

            output.seek(0)
            current_app.logger.info("Excel文件生成完成")

            # 生成下载文件名
            timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
            filename = f"后端仓接收记录导出_{timestamp}.xlsx"

            # 返回Excel文件
            return send_file(
                output,
                as_attachment=True,
                download_name=filename,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )

        except ImportError:
            # 如果pandas不可用，使用openpyxl
            try:
                import openpyxl
                from tempfile import NamedTemporaryFile

                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "后端仓接收记录"

                # 添加表头
                headers = list(data[0].keys()) if data else []
                for col, header in enumerate(headers, 1):
                    ws.cell(row=1, column=col, value=header)

                # 添加数据
                for row, record_data in enumerate(data, 2):
                    for col, value in enumerate(record_data.values(), 1):
                        ws.cell(row=row, column=col, value=value)

                # 保存到临时文件
                temp_file = NamedTemporaryFile(delete=False, suffix='.xlsx')
                wb.save(temp_file.name)
                temp_file.close()

                # 返回文件下载响应
                return send_file(
                    temp_file.name,
                    as_attachment=True,
                    download_name=f"后端仓接收记录导出_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx",
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )

            except ImportError:
                flash("导出Excel需要pandas或openpyxl库，请安装: pip install pandas xlsxwriter 或 pip install openpyxl", "danger")
                return redirect(url_for('main.backend_receive_records'))

    except Exception as e:
        current_app.logger.error(f"导出后端仓接收记录时出错: {str(e)}")
        flash(f"导出数据时出错: {str(e)}", "danger")
        return redirect(url_for('main.backend_receive_records'))

    # 获取搜索参数
    page = request.args.get('page', 1, type=int)
    per_page = 50

    # 构建查询，显示后端仓的接收记录（暂时显示所有有批次号的记录）
    query = InboundRecord.query.options(db.joinedload(InboundRecord.operated_warehouse)).filter(
        # 只显示有批次号的记录（表示是接收其他仓库的货物）
        db.and_(InboundRecord.batch_no.isnot(None), InboundRecord.batch_no != '')
    )

    # 根据用户权限过滤数据
    if hasattr(current_user, 'warehouse') and current_user.warehouse:
        if current_user.warehouse.warehouse_type == 'backend':
            # 后端仓用户只能看自己仓库的接收数据
            query = query.filter_by(operated_warehouse_id=current_user.warehouse_id)
        else:
            # 前端仓或管理员用户可以看所有后端仓的接收数据
            # 进一步过滤：只显示后端仓库的接收记录
            backend_warehouses = Warehouse.query.filter_by(warehouse_type='backend').all()
            backend_warehouse_ids = [w.id for w in backend_warehouses]
            if backend_warehouse_ids:
                query = query.filter(InboundRecord.operated_warehouse_id.in_(backend_warehouse_ids))

    # 搜索过滤
    search_params = {}

    # 客户名称搜索
    if request.args.get('customer_name'):
        query = query.filter(InboundRecord.customer_name.contains(request.args.get('customer_name')))
        search_params['customer_name'] = request.args.get('customer_name')

    # 车牌号搜索
    if request.args.get('plate_number'):
        query = query.filter(InboundRecord.plate_number.contains(request.args.get('plate_number')))
        search_params['plate_number'] = request.args.get('plate_number')

    # 入库类型搜索
    if request.args.get('inbound_type'):
        inbound_type = request.args.get('inbound_type')
        if inbound_type == 'direct':
            # 客户直送：通常没有批次号或批次号为空
            query = query.filter(db.or_(InboundRecord.batch_no.is_(None), InboundRecord.batch_no == ''))
        elif inbound_type == 'transfer':
            # 接收其它仓库订单：有批次号
            query = query.filter(db.and_(InboundRecord.batch_no.isnot(None), InboundRecord.batch_no != ''))
        search_params['inbound_type'] = inbound_type

    # 日期范围搜索
    if request.args.get('date_start'):
        try:
            start_date = datetime.strptime(request.args.get('date_start'), '%Y-%m-%d')
            query = query.filter(InboundRecord.inbound_time >= start_date)
            search_params['date_start'] = request.args.get('date_start')
        except ValueError:
            pass

    if request.args.get('date_end'):
        try:
            end_date = datetime.strptime(request.args.get('date_end'), '%Y-%m-%d')
            # 设置为当天的23:59:59
            end_date = end_date.replace(hour=23, minute=59, second=59)
            query = query.filter(InboundRecord.inbound_time <= end_date)
            search_params['date_end'] = request.args.get('date_end')
        except ValueError:
            pass

    # 排序和分页
    records = query.order_by(InboundRecord.inbound_time.desc()).paginate(
        page=page, per_page=per_page, error_out=False
    )

    return render_template('backend/receive_list.html',
                         records=records,
                         search_params=search_params,
                         title='后端仓接收记录',
                         warehouse_type='backend')


# ==================== 前端仓出库路由 ====================

@bp.route('/frontend/outbound', methods=['GET', 'POST'])
@require_permission('OUTBOUND_VIEW')
def frontend_outbound():
    """前端仓出库操作主页面 - 选择出库类型"""
    if not check_warehouse_permission('frontend', 'view'):
        flash('您没有权限访问前端仓出库功能', 'error')
        return redirect(url_for('main.index'))

    return render_template('frontend/outbound_main.html',
                         title='前端仓出库操作',
                         warehouse_type='frontend')

@bp.route('/frontend/outbound/to_backend', methods=['GET', 'POST'])
@require_permission('OUTBOUND_VIEW')
def frontend_outbound_to_backend():
    """前端仓发货到后端仓（凭祥北投仓）"""
    if not check_warehouse_permission('frontend', 'view'):
        flash('您没有权限访问前端仓出库功能', 'error')
        return redirect(url_for('main.index'))

    # 获取当前用户的仓库信息
    user_warehouse = None
    user_warehouse_name = None
    is_admin = False

    if current_user.is_authenticated:
        is_admin = current_user.is_super_admin()
        if not is_admin and current_user.warehouse:
            user_warehouse = current_user.warehouse
            user_warehouse_name = user_warehouse.warehouse_name

    return render_template('frontend/outbound_to_backend.html',
                         title='前端仓发货到后端仓',
                         warehouse_type='frontend',
                         destination='backend',
                         user_warehouse_name=user_warehouse_name,
                         is_admin=is_admin)

@bp.route('/frontend/outbound/to_customer', methods=['GET', 'POST'])
@require_permission('OUTBOUND_VIEW')
def frontend_outbound_to_customer():
    """前端仓直接配送客户工厂"""
    if not check_warehouse_permission('frontend', 'view'):
        flash('您没有权限访问前端仓出库功能', 'error')
        return redirect(url_for('main.index'))

    # 为管理员用户获取所有仓库列表
    warehouses = []
    if current_user.is_super_admin():
        warehouses = Warehouse.query.all()

    return render_template('frontend/outbound_to_customer.html',
                         title='前端仓直接配送客户',
                         warehouse_type='frontend',
                         destination='customer',
                         warehouses=warehouses)





# 前端仓库存API
@bp.route('/api/inventory/frontend', methods=['GET'])
@csrf.exempt
def api_frontend_inventory():
    """获取前端仓库存数据 - 修复版本，与页面查询逻辑保持一致"""
    try:
        # 获取前端仓库ID列表
        frontend_warehouses = Warehouse.query.filter_by(warehouse_type='frontend').all()
        frontend_warehouse_ids = [w.id for w in frontend_warehouses]

        if not frontend_warehouse_ids:
            return jsonify({'success': True, 'inventory': []})

        # 根据用户权限过滤仓库 - 修复：与管理员API保持一致的权限逻辑
        # 超级管理员：不分配仓库，可以查看所有前端仓的库存
        # 普通用户：分配到具体仓库，只能查看自己仓库的库存
        if current_user.is_super_admin():
            # 超级管理员可以查看所有前端仓库存
            current_app.logger.info(f"超级管理员 {current_user.username} 可以查看所有前端仓库存")
        elif hasattr(current_user, 'warehouse') and current_user.warehouse:
            if current_user.warehouse.warehouse_type == 'frontend':
                # 前端仓用户只能查看自己仓库的库存
                frontend_warehouse_ids = [current_user.warehouse_id]
                current_app.logger.info(f"前端仓用户 {current_user.username} 只能查看仓库 {current_user.warehouse_id} 的库存")
            else:
                # 后端仓用户也可以查看所有前端仓库存（用于转移等操作）
                current_app.logger.info(f"后端仓用户 {current_user.username} 可以查看所有前端仓库存")
        else:
            # 未分配仓库的普通用户，不应该有权限访问
            current_app.logger.warning(f"用户 {current_user.username} 未分配仓库且非管理员，拒绝访问")
            return jsonify({
                'success': False,
                'message': '您未分配仓库，无法查看库存'
            })

        # 构建查询
        query = Inventory.query.options(db.joinedload(Inventory.operated_warehouse))
        query = query.filter(Inventory.operated_warehouse_id.in_(frontend_warehouse_ids))

        # 只显示库存不为0的记录
        query = query.filter((Inventory.pallet_count > 0) | (Inventory.package_count > 0))

        # 按入库日期升序排序
        query = query.order_by(Inventory.inbound_time.asc())

        # 获取所有记录（不分页，因为出库选择需要看到所有可用库存）
        inventory_records = query.all()

        # 转换为JSON格式
        inventory_data = []
        for record in inventory_records:
            # 使用统一的数据补充逻辑
            _enrich_inventory_record(record)

            # 获取完整的识别编码
            identification_code = record.identification_code or ''

            inventory_data.append({
                'id': record.id,
                'customer_name': record.customer_name or '',
                'inbound_plate': getattr(record, 'plate_number', '') or '',  # 修复字段名
                'identification_code': identification_code,
                'order_type': getattr(record, 'order_type', '') or '',
                'pallet_count': record.pallet_count or 0,  # 当前库存板数
                'package_count': record.package_count or 0,  # 当前库存件数
                'available_pallets': record.pallet_count or 0,  # 库存板数（与pallet_count相同）
                'available_packages': record.package_count or 0,  # 库存件数（与package_count相同）
                'weight': getattr(record, 'weight', 0) or 0,
                'volume': getattr(record, 'volume', 0) or 0,
                'export_mode': getattr(record, 'export_mode', '') or '',
                'customs_broker': getattr(record, 'customs_broker', '') or '',
                'service_staff': getattr(record, 'service_staff', '') or '',
                'location': record.location or '',  # 库位
                'documents': record.documents or '',  # 单据
                'document_count': record.documents or '',  # 添加单据份数字段
                'inbound_time': record.inbound_time.strftime('%Y-%m-%d') if record.inbound_time else '',
                'operated_warehouse_name': record.operated_warehouse.warehouse_name if record.operated_warehouse else ''
            })

        # 检查是否有重复的库存记录，如果有则进行清理
        duplicate_check = db.session.query(
            Inventory.identification_code,
            func.count(Inventory.id).label('count')
        ).filter(
            Inventory.operated_warehouse_id.in_(frontend_warehouse_ids)
        ).filter(
            (Inventory.pallet_count > 0) | (Inventory.package_count > 0)
        ).group_by(Inventory.identification_code).having(func.count(Inventory.id) > 1).all()

        if duplicate_check:
            current_app.logger.warning(f"发现 {len(duplicate_check)} 个识别编码有重复的库存记录，将进行自动清理")
            for dup in duplicate_check:
                current_app.logger.warning(f"重复识别编码: {dup.identification_code}, 记录数: {dup.count}")

        current_app.logger.info(f"前端仓库存API：通过SQL去重查询获取 {len(inventory_records)} 条记录，最终返回 {len(inventory_data)} 条记录")

        return jsonify({
            'success': True,
            'inventory': inventory_data,
            'total': len(inventory_data)
        })

    except Exception as e:
        current_app.logger.error(f"获取前端仓库存数据失败: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'获取库存数据失败: {str(e)}'
        }), 500


def _enrich_inventory_record(inventory):
    """为库存记录补充完整信息（使用与全库存查询相同的逻辑）"""
    try:
        # 业务字段优先级：库存表 -> 出库记录 -> 入库记录
        # 处理None字符串和空字符串
        order_type = getattr(inventory, 'order_type', '') if hasattr(inventory, 'order_type') and getattr(inventory, 'order_type') and getattr(inventory, 'order_type') != 'None' else ''
        export_mode = getattr(inventory, 'export_mode', '') if hasattr(inventory, 'export_mode') and getattr(inventory, 'export_mode') and getattr(inventory, 'export_mode') != 'None' else ''
        customs_broker = getattr(inventory, 'customs_broker', '') if hasattr(inventory, 'customs_broker') and getattr(inventory, 'customs_broker') and getattr(inventory, 'customs_broker') != 'None' else ''
        service_staff = getattr(inventory, 'service_staff', '') if hasattr(inventory, 'service_staff') and getattr(inventory, 'service_staff') and getattr(inventory, 'service_staff') != 'None' else ''
        plate_number = getattr(inventory, 'plate_number', '') if hasattr(inventory, 'plate_number') and getattr(inventory, 'plate_number') and getattr(inventory, 'plate_number') != 'None' else ''
        weight = getattr(inventory, 'weight', 0) or 0
        volume = getattr(inventory, 'volume', 0) or 0

        # 如果库存表中的业务字段为空，尝试从出库记录获取
        if not export_mode or not customs_broker or not service_staff or not order_type or not plate_number:
            outbound_record = OutboundRecord.query.filter_by(
                identification_code=inventory.identification_code,
                operated_warehouse_id=inventory.operated_warehouse_id
            ).order_by(OutboundRecord.outbound_time.desc()).first()

            if outbound_record:
                export_mode = export_mode or (outbound_record.export_mode if outbound_record.export_mode and outbound_record.export_mode != 'None' else '')
                customs_broker = customs_broker or (outbound_record.customs_broker if outbound_record.customs_broker and outbound_record.customs_broker != 'None' else '')
                service_staff = service_staff or (outbound_record.service_staff if outbound_record.service_staff and outbound_record.service_staff != 'None' else '')
                order_type = order_type or (outbound_record.order_type if outbound_record.order_type and outbound_record.order_type != 'None' else '')
                plate_number = plate_number or (outbound_record.plate_number if outbound_record.plate_number and outbound_record.plate_number != 'None' else '')

        # 如果仍然为空，尝试从入库记录获取
        if not export_mode or not customs_broker or not service_staff or not order_type or not plate_number or weight == 0 or volume == 0:
            inbound_record = InboundRecord.query.filter_by(
                identification_code=inventory.identification_code
            ).first()

            if inbound_record:
                export_mode = export_mode or (inbound_record.export_mode if inbound_record.export_mode and inbound_record.export_mode != 'None' else '')
                customs_broker = customs_broker or (inbound_record.customs_broker if inbound_record.customs_broker and inbound_record.customs_broker != 'None' else '')
                service_staff = service_staff or (inbound_record.service_staff if inbound_record.service_staff and inbound_record.service_staff != 'None' else '')
                order_type = order_type or (inbound_record.order_type if inbound_record.order_type and inbound_record.order_type != 'None' else '')
                plate_number = plate_number or (inbound_record.plate_number if inbound_record.plate_number and inbound_record.plate_number != 'None' else '')
                # 如果库存记录中的重量/体积为空，使用入库记录的
                if weight == 0 and inbound_record.weight:
                    weight = inbound_record.weight
                if volume == 0 and inbound_record.volume:
                    volume = inbound_record.volume

        # 设置属性
        inventory.order_type = order_type
        inventory.export_mode = export_mode
        inventory.customs_broker = customs_broker
        inventory.service_staff = service_staff
        inventory.plate_number = plate_number
        inventory.weight = weight
        inventory.volume = volume

    except Exception as e:
        current_app.logger.error(f"补充库存记录信息时出错: {str(e)}")
        # 设置默认值
        inventory.order_type = ''
        inventory.export_mode = ''
        inventory.customs_broker = ''
        inventory.service_staff = ''
        inventory.plate_number = ''


@bp.route('/api/inventory/cleanup_duplicates', methods=['POST'])
@csrf.exempt
@require_permission('ADMIN_MANAGE')
def api_cleanup_duplicate_inventory():
    """清理重复的库存记录"""
    try:
        # 查找所有重复的识别编码
        duplicate_codes = db.session.query(
            Inventory.identification_code,
            func.count(Inventory.id).label('count')
        ).group_by(Inventory.identification_code).having(func.count(Inventory.id) > 1).all()

        if not duplicate_codes:
            return jsonify({
                'success': True,
                'message': '没有发现重复的库存记录',
                'cleaned_count': 0
            })

        cleaned_count = 0
        for dup in duplicate_codes:
            identification_code = dup.identification_code

            # 获取该识别编码的所有库存记录，按ID排序
            records = Inventory.query.filter_by(
                identification_code=identification_code
            ).order_by(Inventory.id.desc()).all()

            if len(records) <= 1:
                continue

            # 保留最新的记录（ID最大的），合并其他记录的数据
            main_record = records[0]
            duplicate_records = records[1:]

            # 合并库存数量
            total_pallet_count = main_record.pallet_count or 0
            total_package_count = main_record.package_count or 0
            total_weight = main_record.weight or 0
            total_volume = main_record.volume or 0

            for dup_record in duplicate_records:
                total_pallet_count += (dup_record.pallet_count or 0)
                total_package_count += (dup_record.package_count or 0)
                total_weight += (dup_record.weight or 0)
                total_volume += (dup_record.volume or 0)

            # 更新主记录
            main_record.pallet_count = total_pallet_count
            main_record.package_count = total_package_count
            main_record.weight = total_weight
            main_record.volume = total_volume

            # 删除重复记录
            for dup_record in duplicate_records:
                current_app.logger.info(f"删除重复库存记录: ID={dup_record.id}, 识别编码={identification_code}")
                db.session.delete(dup_record)
                cleaned_count += 1

            current_app.logger.info(f"合并库存记录: 识别编码={identification_code}, 保留ID={main_record.id}, "
                                  f"合并后库存: 板数={total_pallet_count}, 件数={total_package_count}")

        db.session.commit()

        return jsonify({
            'success': True,
            'message': f'成功清理 {cleaned_count} 条重复的库存记录',
            'cleaned_count': cleaned_count
        })

    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"清理重复库存记录失败: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'清理失败: {str(e)}'
        }), 500


@bp.route('/api/outbound/batch_history', methods=['POST'])
@csrf.exempt
def api_outbound_batch_history():
    """获取分批出库历史"""
    try:
        data = request.get_json()
        identification_codes = data.get('identification_codes', [])

        if not identification_codes:
            return jsonify({'success': True, 'batch_history': {}})

        # 查询这些识别编码的出库历史
        batch_history = {}

        for code in identification_codes:
            # 获取该识别编码的所有出库记录，按时间排序
            outbound_records = OutboundRecord.query.filter_by(
                identification_code=code
            ).order_by(OutboundRecord.outbound_time.asc()).all()

            if outbound_records:
                history_list = []
                for record in outbound_records:
                    history_list.append({
                        'id': record.id,
                        'outbound_time': record.outbound_time.strftime('%Y-%m-%d %H:%M:%S'),
                        'plate_number': record.plate_number,
                        'pallet_count': record.pallet_count or 0,
                        'package_count': record.package_count or 0,
                        'remarks': record.remarks or ''
                    })

                batch_history[code] = history_list

        return jsonify({
            'success': True,
            'batch_history': batch_history
        })

    except Exception as e:
        current_app.logger.error(f"获取分批出库历史失败: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'获取分批历史失败: {str(e)}'
        }), 500


# ==================== 前端仓出库API ====================

# 前端仓发货到后端仓API - 安全增强版本
@bp.route('/api/frontend/outbound/to_backend', methods=['POST'])
@csrf.exempt
@login_required
@require_permission('OUTBOUND_CREATE')
def api_frontend_outbound_to_backend():
    """前端仓发货到后端仓API"""

    def _parse_datetime(time_str):
        """解析时间字符串，支持多种格式"""
        if not time_str:
            return None
        try:
            # 尝试ISO格式
            return datetime.strptime(time_str, '%Y-%m-%dT%H:%M')
        except ValueError:
            try:
                # 尝试空格分隔格式
                return datetime.strptime(time_str, '%Y-%m-%d %H:%M')
            except ValueError:
                current_app.logger.warning(f'无法解析时间格式: {time_str}')
                return None

    try:
        # 安全验证：检查用户认证和权限
        if not current_user.is_authenticated:
            current_app.logger.warning('前端仓出库请求被拒绝：用户未认证')
            return jsonify({'success': False, 'message': '用户未认证，请先登录'}), 401

        if not hasattr(current_user, 'warehouse_id') or not current_user.warehouse_id:
            current_app.logger.warning(f'前端仓出库请求被拒绝：用户{current_user.username}没有关联仓库')
            return jsonify({'success': False, 'message': '用户没有关联仓库，请联系管理员'}), 403

        # 安全验证：验证请求数据
        if SECURITY_ENABLED:
            try:
                data = RequestValidator.validate_json_request()
            except ValidationException as e:
                return jsonify({'success': False, 'message': f'请求验证失败: {str(e)}'}), 400
        else:
            data = request.get_json()
            if not data:
                current_app.logger.error('没有接收到数据')
                return jsonify({'success': False, 'message': '没有接收到数据'}), 400

        current_app.logger.info(f'收到前端仓出库请求，用户: {current_user.username}，仓库ID: {current_user.warehouse_id}')

        current_app.logger.info(f'接收到数据: {len(data.get("records", []))}条记录')

        records = data.get('records', [])
        common_data = data.get('commonData', {})

        current_app.logger.info(f'公共数据: {common_data}')
        current_app.logger.info(f'记录数据: {records}')

        if not records:
            return jsonify({'success': False, 'message': '没有出库记录'}), 400

        # 验证公共字段
        required_common_fields = ['trunkPlate', 'driverName', 'driverPhone', 'originWarehouse', 'destinationWarehouse']
        current_app.logger.info(f'开始验证公共字段: {required_common_fields}')

        for field in required_common_fields:
            field_value = common_data.get(field)
            current_app.logger.info(f'验证字段 {field}: {field_value}')
            if not field_value:
                current_app.logger.error(f'公共字段验证失败: {field} 为空')
                return jsonify({'success': False, 'message': f'公共字段{field}不能为空'}), 400

        current_app.logger.info('公共字段验证通过')

        success_count = 0
        errors = []

        # 使用公共字段中的干线车牌作为统一的车牌号
        trunk_plate = common_data.get('trunkPlate')

        # 根据目的仓设置destination_warehouse_id
        destination_warehouse = common_data.get('destinationWarehouse')
        destination_warehouse_id = None
        destination_prefix = None
        if destination_warehouse == '凭祥北投仓':
            destination_warehouse_id = 4  # 后端仓ID
            destination_prefix = None  # 使用默认仓库前缀
        elif destination_warehouse == '凭祥保税仓':
            destination_warehouse_id = 5  # 保税仓ID
            destination_prefix = 'PB'  # 保税仓专用前缀
        else:
            destination_warehouse_id = 4  # 默认为后端仓
            destination_prefix = None

        # 生成批次号
        from app.utils.batch_generator import generate_batch_number
        batch_number = generate_batch_number(
            warehouse_id=current_user.warehouse_id,
            destination_prefix=destination_prefix,
            db_session=db.session
        )
        current_app.logger.info(f'生成批次号: {batch_number}，目的仓: {destination_warehouse}')

        current_app.logger.info(f'开始处理{len(records)}条记录')
        for i, record_data in enumerate(records):
            try:
                current_app.logger.info(f'处理第{i+1}条记录: {record_data.get("customer_name", "未知客户")}')
                # 验证必填字段
                if not record_data.get('customer_name'):
                    errors.append(f'第{i+1}行：客户名称不能为空')
                    continue

                # 创建出库记录，使用公共字段数据
                # 解析离仓时间，支持多种格式
                departure_time = common_data.get('departureTime')
                outbound_time = _parse_datetime(departure_time) or datetime.now()

                # 处理识别编码，严格保持原始识别编码不变
                identification_code = record_data.get('identification_code', '').strip()

                # 验证识别编码格式和完整性
                if not identification_code:
                    errors.append(f'第{i+1}行：识别编码不能为空')
                    continue

                # 确保识别编码不包含任何后缀（如b2、_b2等）
                # 如果发现有后缀，移除后缀恢复原始编码
                if identification_code.endswith('b2') or identification_code.endswith('_b2') or identification_code.endswith('.b2'):
                    # 移除错误的后缀
                    for suffix in ['b2', '_b2', '.b2']:
                        if identification_code.endswith(suffix):
                            identification_code = identification_code[:-len(suffix)]
                            current_app.logger.warning(f"检测到识别编码有错误后缀，已修正: {record_data.get('identification_code')} -> {identification_code}")
                            break

                # 检查是否已存在相同识别编码的出库记录，用于确定分批序号
                existing_count = OutboundRecord.query.filter(
                    OutboundRecord.identification_code == identification_code
                ).count()

                # 分批序号从1开始，用于区分同一识别编码的多次出库
                split_sequence = existing_count + 1

                current_app.logger.info(f"识别编码: {identification_code}, 分批序号: {split_sequence}")

                # 层板数量逻辑 - 只在第一条记录中保存层板数量，其他记录设为0
                current_large_layer = int(common_data.get('largePallet', 0)) if common_data.get('largePallet') and i == 0 else 0
                current_small_layer = int(common_data.get('smallPallet', 0)) if common_data.get('smallPallet') and i == 0 else 0
                current_pallet_board = int(common_data.get('cardPallet', 0)) if common_data.get('cardPallet') and i == 0 else 0

                # 构建备注信息，如果是分批出库则添加分批信息
                base_remarks = record_data.get('remarks', '') or f'前端仓发货到后端仓 - 司机：{common_data.get("driverName")} - 车型：{common_data.get("vehicleType")} - 始发仓：{common_data.get("originWarehouse")}'

                # 不再自动添加分批信息到备注中

                outbound_record = OutboundRecord(
                    outbound_time=outbound_time,
                    delivery_plate_number=trunk_plate,
                    plate_number=trunk_plate,
                    customer_name=record_data.get('customer_name'),
                    identification_code=identification_code,  # 保持原始识别编码不变
                    pallet_count=int(record_data.get('pallet_count', 0)) if record_data.get('pallet_count') else 0,
                    package_count=int(record_data.get('package_count', 0)) if record_data.get('package_count') else 0,
                    weight=float(record_data.get('weight', 0)) if record_data.get('weight') else 0,
                    volume=float(record_data.get('volume', 0)) if record_data.get('volume') else 0,
                    destination=destination_warehouse,
                    destination_warehouse_id=destination_warehouse_id,
                    batch_no=batch_number,
                    batch_sequence=split_sequence,  # 使用分批序号作为批次序号
                    batch_total=0,  # 设置批次总数
                    remarks=base_remarks,
                    remark1=record_data.get('remarks', ''),  # 备注 -> 备注1
                    remark2=record_data.get('remarks2', ''), # 备注2 -> 备注2
                    operated_by_user_id=current_user.id,
                    operated_warehouse_id=current_user.warehouse_id,
                    # 添加运输相关信息
                    driver_name=common_data.get('driverName'),
                    driver_phone=common_data.get('driverPhone'),
                    vehicle_type=common_data.get('vehicleType'),
                    arrival_time=_parse_datetime(common_data.get('arrivalTime')),
                    loading_start_time=_parse_datetime(common_data.get('loadingStartTime')),
                    loading_end_time=_parse_datetime(common_data.get('loadingEndTime')),
                    # 添加其他字段
                    order_type=record_data.get('order_type'),
                    export_mode=record_data.get('export_mode'),
                    customs_broker=record_data.get('customs_broker'),
                    service_staff=record_data.get('service_staff'),
                    inbound_plate=record_data.get('inbound_plate'),
                    # 添加单据字段
                    documents=record_data.get('documents', ''),
                    document_count=None if not record_data.get('document_count') or record_data.get('document_count') == '' else (int(record_data.get('document_count')) if str(record_data.get('document_count')).isdigit() else None),
                    # 添加板材信息
                    large_layer=current_large_layer,
                    small_layer=current_small_layer,
                    pallet_board=current_pallet_board,
                    # 添加车挂和柜号信息（前端仓已移除这些字段）
                    trailer=common_data.get('trailer', ''),
                    container_number=common_data.get('container_number', '')
                )

                db.session.add(outbound_record)
                current_app.logger.info(f'创建出库记录：{record_data.get("customer_name")} - 批次号：{batch_number}')

                # 处理库存：减少相应数量
                inventory_id = record_data.get('inventory_id')
                if inventory_id:
                    inventory = Inventory.query.get(inventory_id)
                    if inventory:
                        out_pallet = int(record_data.get('pallet_count', 0))
                        out_package = int(record_data.get('package_count', 0))
                        out_weight = float(record_data.get('weight', 0))
                        out_volume = float(record_data.get('volume', 0))

                        inventory.pallet_count = max(0, inventory.pallet_count - out_pallet)
                        inventory.package_count = max(0, inventory.package_count - out_package)
                        inventory.weight = max(0, inventory.weight - out_weight)
                        inventory.volume = max(0, inventory.volume - out_volume)
                        inventory.last_updated = datetime.now()
                        current_app.logger.info(f'出库减少库存：库存ID {inventory_id}，目的仓：{destination_warehouse}')

                # 根据目的仓类型决定是否创建在途记录
                if destination_warehouse == '凭祥北投仓':
                    # 发往北投仓：创建在途记录，需要后续接收
                    transit_cargo = TransitCargo(
                        customer_name=record_data.get('customer_name'),
                        identification_code=record_data.get('identification_code', ''),
                        pallet_count=int(record_data.get('pallet_count', 0)),
                        package_count=int(record_data.get('package_count', 0)),
                        weight=float(record_data.get('weight', 0)),
                        volume=float(record_data.get('volume', 0)),
                        batch_no=batch_number,
                        source_warehouse_id=current_user.warehouse_id,
                        destination_warehouse_id=destination_warehouse_id,
                        departure_time=_parse_datetime(common_data.get('departureTime')) or datetime.now(),
                        plate_number=trunk_plate,
                        driver_name=common_data.get('driverName'),
                        driver_phone=common_data.get('driverPhone'),
                        status='in_transit',
                        created_by_user_id=current_user.id,
                        # 添加其他字段
                        export_mode=record_data.get('export_mode'),
                        order_type=record_data.get('order_type'),
                        customs_broker=record_data.get('customs_broker'),
                        service_staff=record_data.get('service_staff')
                    )
                    db.session.add(transit_cargo)
                    current_app.logger.info(f'创建在途记录：{record_data.get("customer_name")} -> 凭祥北投仓')
                elif destination_warehouse == '凭祥保税仓':
                    # 发往保税仓：不创建在途记录，直接完成
                    current_app.logger.info(f'保税仓出库完成：{record_data.get("customer_name")} -> 凭祥保税仓（无需接收）')

                success_count += 1

            except Exception as e:
                current_app.logger.error(f'处理第{i+1}行记录时出错: {str(e)}')
                current_app.logger.error(f'记录数据: {record_data}')
                errors.append(f'第{i+1}行：{str(e)}')
                continue

        if success_count > 0:
            db.session.commit()
            current_app.logger.info(f'前端仓发货到后端仓成功：{success_count}条记录，车牌：{trunk_plate}')

            return jsonify({
                'success': True,
                'message': f'成功保存{success_count}条记录到后端仓',
                'success_count': success_count,
                'total_count': len(records),
                'trunk_plate': trunk_plate,
                'batch_number': batch_number,
                'destination_warehouse': destination_warehouse,
                'errors': errors
            })
        else:
            current_app.logger.warning(f'前端仓发货到后端仓失败：没有成功保存任何记录，错误：{errors}')
            return jsonify({
                'success': False,
                'message': f'保存失败，没有成功保存任何记录',
                'success_count': success_count,
                'total_count': len(records),
                'errors': errors
            }), 400

    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f'前端仓发货到后端仓失败: {str(e)}')
        return jsonify({'success': False, 'message': str(e)}), 500

# 前端仓直接配送客户API
@bp.route('/api/frontend/outbound/to_customer', methods=['POST'])
@csrf.exempt
@require_permission('OUTBOUND_CREATE')
def api_frontend_outbound_to_customer():
    """前端仓直接配送客户API"""
    try:
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'message': '没有接收到数据'}), 400

        records = data.get('records', [])
        if not records:
            return jsonify({'success': False, 'message': '没有出库记录'}), 400

        # 生成批次号
        from app.utils.batch_generator import generate_batch_number
        batch_number = generate_batch_number(
            warehouse_id=current_user.warehouse_id,
            destination_prefix='CK',  # 客户配送专用前缀
            db_session=db.session
        )
        current_app.logger.info(f'前端仓直接配送客户生成批次号: {batch_number}')

        success_count = 0
        errors = []

        for i, record_data in enumerate(records):
            try:
                # 验证必填字段
                if not record_data.get('customer_name'):
                    errors.append(f'第{i+1}行：客户名称不能为空')
                    continue

                if not record_data.get('plate_number'):
                    errors.append(f'第{i+1}行：出库车牌不能为空')
                    continue

                if not record_data.get('destination_address'):
                    errors.append(f'第{i+1}行：目的地地址不能为空')
                    continue

                if not record_data.get('destination_contact'):
                    errors.append(f'第{i+1}行：目的地联系人不能为空')
                    continue

                if not record_data.get('driver_name'):
                    errors.append(f'第{i+1}行：司机姓名不能为空')
                    continue

                if not record_data.get('driver_phone'):
                    errors.append(f'第{i+1}行：司机联系方式不能为空')
                    continue

                # 解析出库日期
                outbound_date = datetime.now()
                if record_data.get('outbound_date'):
                    try:
                        outbound_date = datetime.strptime(record_data.get('outbound_date'), '%Y-%m-%d')
                    except ValueError:
                        outbound_date = datetime.now()

                # 确定操作仓库ID
                operated_warehouse_id = current_user.warehouse_id
                if current_user.is_super_admin() and record_data.get('source_warehouse'):
                    # 管理员用户根据选择的始发仓确定操作仓库
                    source_warehouse = Warehouse.query.filter_by(
                        warehouse_name=record_data.get('source_warehouse')
                    ).first()
                    if source_warehouse:
                        operated_warehouse_id = source_warehouse.id

                # 计算批次序号 - 同一次出库操作中的序号（从1开始递增）
                batch_sequence = i + 1  # 使用循环索引+1作为批次序号

                # 创建出库记录
                outbound_record = OutboundRecord(
                    outbound_time=outbound_date,
                    delivery_plate_number=record_data.get('plate_number'),
                    plate_number=record_data.get('plate_number'),
                    customer_name=record_data.get('customer_name'),
                    identification_code=record_data.get('identification_code', ''),
                    pallet_count=int(record_data.get('pallet_count', 0)) if record_data.get('pallet_count') else 0,
                    package_count=int(record_data.get('package_count', 0)) if record_data.get('package_count') else 0,
                    weight=float(record_data.get('weight', 0)) if record_data.get('weight') else 0,
                    volume=float(record_data.get('volume', 0)) if record_data.get('volume') else 0,
                    export_mode=record_data.get('export_mode', ''),
                    order_type=record_data.get('order_type', ''),
                    customs_broker=record_data.get('customs_broker', ''),
                    location=record_data.get('location', ''),
                    documents=record_data.get('documents', ''),
                    destination=record_data.get('destination', '客户工厂'),
                    detailed_address=record_data.get('destination_address'),
                    contact_window=record_data.get('destination_contact'),
                    driver_name=record_data.get('driver_name'),
                    driver_phone=record_data.get('driver_phone'),
                    batch_no=batch_number,
                    batch_sequence=batch_sequence,  # 同一次出库操作中的序号
                    batch_total=len(records),  # 同一次出库操作的总记录数
                    remarks=record_data.get('remarks', '前端仓直接配送客户'),
                    operated_by_user_id=current_user.id,
                    operated_warehouse_id=operated_warehouse_id
                )

                db.session.add(outbound_record)
                success_count += 1

                # 如果有识别编码，需要更新库存
                if record_data.get('identification_code'):
                    try:
                        # 修复：根据识别编码和操作仓库ID查找正确的库存记录
                        inventory_item = Inventory.query.filter_by(
                            identification_code=record_data.get('identification_code'),
                            operated_warehouse_id=operated_warehouse_id
                        ).first()

                        if inventory_item:
                            # 减少库存
                            pallet_out = int(record_data.get('pallet_count', 0)) if record_data.get('pallet_count') else 0
                            package_out = int(record_data.get('package_count', 0)) if record_data.get('package_count') else 0

                            inventory_item.pallet_count = max(0, inventory_item.pallet_count - pallet_out)
                            inventory_item.package_count = max(0, inventory_item.package_count - package_out)

                            current_app.logger.info(f'更新库存：{record_data.get("identification_code")} 仓库ID:{operated_warehouse_id} 板数-{pallet_out} 件数-{package_out}')
                        else:
                            current_app.logger.warning(f'未找到库存记录：{record_data.get("identification_code")} 仓库ID:{operated_warehouse_id}')
                    except Exception as inventory_error:
                        current_app.logger.warning(f'更新库存失败：{str(inventory_error)}')

            except Exception as e:
                errors.append(f'第{i+1}行：{str(e)}')
                continue

        if success_count > 0:
            db.session.commit()
            current_app.logger.info(f'前端仓直接配送客户成功：{success_count}条记录')

        return jsonify({
            'success': True,
            'message': f'成功保存{success_count}条记录',
            'success_count': success_count,
            'total_count': len(records),
            'errors': errors
        })

    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f'前端仓直接配送客户失败: {str(e)}')
        return jsonify({'success': False, 'message': str(e)}), 500

# 前端仓直接配送批量保存API
@bp.route('/api/frontend/outbound/direct/batch', methods=['POST'])
@csrf.exempt
@require_permission('OUTBOUND_CREATE')
def api_frontend_outbound_direct_batch():
    """前端仓直接配送批量出库（支持后端仓出库到凭祥保税仓/春疆货场）"""
    try:
        # 从请求数据中判断是否为后端仓最终出库
        data = request.get_json()
        if not data or 'records' not in data:
            return jsonify({'success': False, 'message': '请求数据格式错误'}), 400

        records = data['records']
        if not records:
            return jsonify({'success': False, 'message': '没有要保存的记录'}), 400

        # 检查是否为后端仓最终出库
        is_backend_final_outbound = False
        if records and len(records) > 0:
            first_record = records[0]
            factory_address = first_record.get('factory_address', '')
            # 判断是否为后端仓最终出库（春疆货场或凭祥保税仓）
            is_backend_final_outbound = factory_address in ['春疆货场', '凭祥保税仓']

        # 权限检查：前端仓用户可以发货到任何地方（包括后端仓）
        # 这是正常的业务流程：前端仓集拼货物后发往后端仓
        if not check_warehouse_permission('frontend', 'add'):
            return jsonify({'success': False, 'message': '您没有权限执行前端仓出库操作'}), 403

        # 生成批次号
        from app.utils.batch_generator import generate_batch_number

        # 根据目的地设置不同的前缀
        destination_prefix = None
        if is_backend_final_outbound:
            first_record = records[0]
            factory_address = first_record.get('factory_address', '')
            if factory_address == '春疆货场':
                destination_prefix = 'CJ'  # 春疆货场专用前缀
            elif factory_address == '凭祥保税仓':
                destination_prefix = 'PB'  # 保税仓专用前缀
        else:
            destination_prefix = 'CK'  # 客户配送专用前缀

        batch_number = generate_batch_number(
            warehouse_id=current_user.warehouse_id,
            destination_prefix=destination_prefix,
            db_session=db.session
        )
        current_app.logger.info(f'前端仓直接配送批量生成批次号: {batch_number}，目的地前缀: {destination_prefix}')

        success_count = 0
        errors = []

        for i, record_data in enumerate(records):
            try:
                # 验证必填字段
                required_fields = ['outbound_time', 'plate_number', 'customer_name',
                                 'identification_code', 'receiver_contact', 'factory_address']
                for field in required_fields:
                    if not record_data.get(field):
                        errors.append(f'第{i+1}行：{field}不能为空')
                        continue

                # 验证板数和件数至少一项必填
                pallet_count = record_data.get('pallet_count', 0)
                package_count = record_data.get('package_count', 0)
                if not pallet_count and not package_count:
                    errors.append(f'第{i+1}行：板数和件数至少需要填写一项')
                    continue

                # 解析出库时间
                outbound_time = datetime.fromisoformat(record_data['outbound_time'].replace('T', ' '))

                # 获取入库日期
                inbound_date = None
                identification_code = record_data.get('identification_code')
                if identification_code:
                    # 首先尝试从库存记录获取入库日期
                    inventory = Inventory.query.filter_by(identification_code=identification_code).first()
                    if inventory and inventory.inbound_time:
                        inbound_date = inventory.inbound_time
                    else:
                        # 如果库存记录没有入库时间，尝试从入库记录获取
                        inbound_record = InboundRecord.query.filter_by(identification_code=identification_code).first()
                        if inbound_record and inbound_record.inbound_time:
                            inbound_date = inbound_record.inbound_time

                # 创建出库记录
                outbound_record = OutboundRecord(
                    batch_no=batch_number,
                    outbound_time=outbound_time,
                    plate_number=record_data['plate_number'],
                    vehicle_type=record_data.get('vehicle_type', ''),
                    customer_name=record_data['customer_name'],
                    identification_code=record_data['identification_code'],
                    pallet_count=pallet_count,
                    package_count=package_count,
                    weight=record_data.get('weight', 0),
                    volume=record_data.get('volume', 0),
                    order_type=record_data.get('order_type', ''),
                    export_mode=record_data.get('export_mode', ''),  # 出境模式
                    customs_broker=record_data.get('customs_broker', ''),  # 报关行
                    delivery_plate_number=record_data.get('delivery_plate_number', '') or record_data.get('delivery_truck', ''),  # 送货干线车
                    inbound_plate=record_data.get('inbound_plate', ''),  # 入库车牌
                    service_staff=record_data.get('service_staff', ''),
                    documents=record_data.get('documents', ''),
                    destination=record_data['factory_address'],  # 目的地
                    detailed_address=record_data.get('detailed_address', ''),  # 详细地址
                    contact_window=record_data['receiver_contact'],  # 修正字段名
                    inbound_date=inbound_date,  # 添加入库日期
                    remark1=record_data.get('remark1', ''),
                    remark2=record_data.get('remark2', ''),
                    # 添加车挂和柜号信息（前端仓已移除这些字段）
                    trailer=record_data.get('trailer_number', '') or record_data.get('trailer', ''),
                    container_number=record_data.get('container_number', ''),
                    # 操作追踪
                    operated_by_user_id=current_user.id,
                    # 智能判断操作仓库ID
                    operated_warehouse_id=_get_operation_warehouse_id(is_backend_final_outbound)
                )

                db.session.add(outbound_record)

                # 扣减库存（后端仓最终出库需要扣减后端仓库存）
                if is_backend_final_outbound:
                    # 查找对应的库存记录进行扣减
                    inventory = Inventory.query.filter_by(
                        customer_name=record_data['customer_name'],
                        identification_code=record_data['identification_code']
                    ).first()

                    if inventory:
                        # 记录更新前的库存
                        before_pallet = inventory.pallet_count
                        before_package = inventory.package_count

                        # 更新库存（减少）
                        inventory.pallet_count = max(0, inventory.pallet_count - pallet_count)
                        inventory.package_count = max(0, inventory.package_count - package_count)

                        current_app.logger.info(f'库存扣减: {inventory.customer_name} {inventory.identification_code} '
                                              f'板数 {before_pallet}->{inventory.pallet_count} '
                                              f'件数 {before_package}->{inventory.package_count}')
                    else:
                        current_app.logger.warning(f'未找到对应库存记录: {record_data["customer_name"]} {record_data["identification_code"]}')

                success_count += 1

            except Exception as e:
                errors.append(f'第{i+1}行：保存失败 - {str(e)}')
                continue

        if success_count > 0:
            db.session.commit()

        # 构建详细的成功消息
        message_parts = []
        if success_count > 0:
            message_parts.append(f'✅ 成功保存 {success_count} 条出库记录')
            if is_backend_final_outbound:
                message_parts.append('📦 库存已自动扣减')
                message_parts.append('🚛 春疆货场/凭祥保税仓出库完成')

        if errors:
            message_parts.append(f'⚠️ {len(errors)} 条记录保存失败')

        return jsonify({
            'success': success_count > 0,
            'count': success_count,
            'errors': errors,
            'message': '，'.join(message_parts),
            'operation_type': 'chunjiang_outbound' if is_backend_final_outbound else 'direct_delivery',
            'inventory_updated': is_backend_final_outbound and success_count > 0
        })

    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"前端仓直接配送批量出库失败: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'批量出库失败: {str(e)}'
        }), 500


@bp.route('/api/generate/chunjiang/batch', methods=['POST'])
@csrf.exempt
@require_permission('OUTBOUND_ADD')
def api_generate_chunjiang_batch():
    """生成春疆货场批次号"""
    try:
        # 检查后端仓权限
        if not check_warehouse_permission('backend', 'add'):
            return jsonify({'success': False, 'message': '您没有权限执行后端仓出库操作'}), 403

        # 获取今天的日期前缀 (年月日格式，如25070401)
        today = datetime.now()
        date_prefix = today.strftime('%y%m%d')  # 年月日，如"250704"

        # 查询今天已有的春疆货场批次号
        today_start = today.replace(hour=0, minute=0, second=0, microsecond=0)
        today_end = today.replace(hour=23, minute=59, second=59, microsecond=999999)

        # 获取当前用户的仓库ID（应该是后端仓）
        current_warehouse_id = current_user.warehouse_id if hasattr(current_user, 'warehouse_id') else None

        # 查找今天该仓库最新的春疆货场批次号
        query = OutboundRecord.query.filter(
            OutboundRecord.batch_no.like(f'CJ{date_prefix}%'),
            OutboundRecord.outbound_time.between(today_start, today_end),
            OutboundRecord.destination == '谅山春疆货场'
        )

        # 如果能确定仓库ID，则限制为该仓库
        if current_warehouse_id:
            query = query.filter(OutboundRecord.operated_warehouse_id == current_warehouse_id)

        latest_batch = query.order_by(OutboundRecord.batch_no.desc()).first()

        if latest_batch and latest_batch.batch_no:
            # 如果今天已经有批次号，提取序号部分并加1
            try:
                # 批次号格式: CJ + 年月日(6位) + 序号(2位)，如 CJ2507040１, CJ25070402
                batch_seq = int(latest_batch.batch_no[8:]) + 1
                if batch_seq > 99:  # 序号最大99，超过则重置为1
                    batch_seq = 1
            except (ValueError, IndexError):
                # 如果提取失败，使用默认值
                batch_seq = 1
        else:
            # 如果今天还没有批次号，从1开始
            batch_seq = 1

        # 格式化批次号：CJ + 年月日(6位) + 序号(2位)，总共10个字符
        new_batch_number = f'CJ{date_prefix}{batch_seq:02d}'

        return jsonify({
            'success': True,
            'batch_number': new_batch_number,
            'message': f'成功生成春疆货场批次号: {new_batch_number}'
        })

    except Exception as e:
        current_app.logger.error(f"生成春疆货场批次号失败: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'生成批次号失败: {str(e)}'
        }), 500


@bp.route('/frontend/outbound/direct', methods=['GET', 'POST'])
@require_permission('OUTBOUND_VIEW')
def frontend_outbound_direct():
    """前端仓直接配送客户工厂"""
    # 获取目的地参数
    destination = request.args.get('destination', 'frontend')

    # 根据目的地确定仓库类型和权限检查
    if destination == 'chunjiang':
        # 前端仓发货到后端仓（凭祥保税仓/春疆货场）
        # 前端仓用户可以发货到后端仓，这是正常的业务流程
        if not check_warehouse_permission('frontend', 'view'):
            flash('您没有权限访问前端仓出库功能', 'error')
            return redirect(url_for('main.index'))
        warehouse_type = 'frontend'  # 保持为前端仓类型，因为是前端仓发起的出库
        title = '前端仓发货到后端仓（凭祥保税仓/春疆货场）'
    else:
        # 流转到前端仓，需要前端仓权限
        if not check_warehouse_permission('frontend', 'view'):
            flash('您没有权限访问前端仓出库功能', 'error')
            return redirect(url_for('main.index'))
        warehouse_type = 'frontend'
        title = '前端仓直接配送客户工厂'

    if request.method == 'POST':
        try:
            # 获取表单数据
            outbound_time = request.form.get('outbound_time')
            plate_number = request.form.get('plate_number', '').strip()
            customer_name = request.form.get('customer_name', '').strip()
            identification_code = request.form.get('identification_code', '').strip()
            pallet_count = request.form.get('pallet_count', 0, type=int)
            package_count = request.form.get('package_count', 0, type=int)
            weight = request.form.get('weight', 0, type=float)
            volume = request.form.get('volume', 0, type=float)
            order_type = request.form.get('order_type', '').strip()
            service_staff = request.form.get('service_staff', '').strip()
            remarks = request.form.get('remarks', '').strip()
            documents = request.form.get('documents', '').strip()

            # 直接配送特有字段
            receiver_contact = request.form.get('receiver_contact', '').strip()
            factory_address = request.form.get('factory_address', '').strip()

            # 验证必填字段
            if not all([outbound_time, plate_number, customer_name, service_staff, receiver_contact, factory_address]):
                flash('请填写所有必填字段', 'error')
                return redirect(request.url)

            # 解析时间
            try:
                outbound_datetime = datetime.strptime(outbound_time, '%Y-%m-%dT%H:%M')
            except ValueError:
                flash('出库时间格式不正确', 'error')
                return redirect(request.url)

            # 创建出库记录
            outbound_record = OutboundRecord(
                outbound_time=outbound_datetime,
                plate_number=plate_number,
                customer_name=customer_name,
                identification_code=identification_code,
                pallet_count=pallet_count,
                package_count=package_count,
                weight=weight,
                volume=volume,
                order_type=order_type,
                export_mode='',  # 添加出境模式字段
                customs_broker='',  # 添加报关行字段
                service_staff=service_staff,
                remarks=remarks,
                documents=documents,
                # 直接配送特有字段存储
                contact_window=receiver_contact,  # 复用联络窗口字段存储收货人联系方式
                detailed_address=factory_address,  # 使用详细地址字段存储工厂地址
                destination='直接配送客户工厂',  # 标识为直接配送
                # 操作追踪
                operated_by_user_id=current_user.id,
                operated_warehouse_id=current_user.warehouse_id
            )

            db.session.add(outbound_record)
            db.session.commit()

            flash('出库记录添加成功', 'success')
            return redirect(url_for('main.frontend_outbound_direct'))

        except Exception as e:
            db.session.rollback()
            current_app.logger.error(f"前端仓直接配送出库失败: {str(e)}")
            flash('出库操作失败，请重试', 'error')
            return redirect(request.url)

    # 获取当前用户的仓库信息
    user_warehouse = None
    user_warehouse_name = None
    is_admin = False

    if current_user.is_authenticated:
        is_admin = current_user.is_super_admin()
        if not is_admin and current_user.warehouse:
            user_warehouse = current_user.warehouse
            user_warehouse_name = user_warehouse.warehouse_name

    return render_template('frontend/outbound_direct.html',
                         title=title,
                         warehouse_type=warehouse_type,
                         destination=destination,
                         user_warehouse_name=user_warehouse_name,
                         is_admin=is_admin)


@bp.route('/export_frontend_outbound')
@require_permission('OUTBOUND_VIEW')
def export_frontend_outbound():
    """导出前端仓出库记录到Excel"""
    try:
        # 检查权限
        if not check_warehouse_permission('frontend', 'view'):
            flash('您没有权限导出前端仓出库记录', 'error')
            return redirect(url_for('main.frontend_outbound_list'))

        # 获取搜索参数（与frontend_outbound_list相同的参数处理）
        warehouse_id = request.args.get('warehouse_id', '')
        search_field = request.args.get('search_field', '')
        search_value = request.args.get('search_value', '')
        date_start = request.args.get('date_start', '')
        date_end = request.args.get('date_end', '')

        # 如果没有指定日期范围，默认使用最近一周的日期范围
        if not date_start and not date_end:
            today = datetime.now().strftime('%Y-%m-%d')
            one_week_ago = (datetime.now() - timedelta(days=7)).strftime('%Y-%m-%d')
            date_start = one_week_ago
            date_end = today

        # 构建查询，只显示前端仓的数据
        query = OutboundRecord.query.options(
            db.joinedload(OutboundRecord.operated_warehouse),
            db.joinedload(OutboundRecord.destination_warehouse)
        )

        # 获取前端仓库ID列表
        frontend_warehouses = Warehouse.query.filter_by(warehouse_type='frontend').all()
        frontend_warehouse_ids = [w.id for w in frontend_warehouses]

        # 只显示前端仓的出库记录
        query = query.filter(OutboundRecord.operated_warehouse_id.in_(frontend_warehouse_ids))

        # 权限过滤：如果用户有关联仓库，只显示其关联仓库的数据
        if current_user.warehouse_id:
            # 检查关联仓库是否为前端仓
            associated_warehouse = Warehouse.query.get(current_user.warehouse_id)
            if associated_warehouse and associated_warehouse.warehouse_type == 'frontend':
                query = query.filter(OutboundRecord.operated_warehouse_id == current_user.warehouse_id)
            else:
                # 如果关联的不是前端仓，则无权查看任何数据
                query = query.filter(False)

        # 应用搜索条件
        if warehouse_id:
            query = query.filter(OutboundRecord.operated_warehouse_id == warehouse_id)

        if search_field and search_value:
            if search_field == 'customer_name':
                query = query.filter(OutboundRecord.customer_name.like(f'%{search_value}%'))
            elif search_field == 'plate_number':
                query = query.filter(OutboundRecord.plate_number.like(f'%{search_value}%'))
            elif search_field == 'batch_no':
                query = query.filter(OutboundRecord.batch_no.like(f'%{search_value}%'))
            elif search_field == 'identification_code':
                query = query.filter(OutboundRecord.identification_code.like(f'%{search_value}%'))
            elif search_field == 'destination':
                query = query.filter(OutboundRecord.destination.like(f'%{search_value}%'))
            elif search_field == 'customs_broker':
                query = query.filter(OutboundRecord.customs_broker.like(f'%{search_value}%'))

        # 按日期范围筛选
        if date_start:
            try:
                start_date = datetime.strptime(date_start, '%Y-%m-%d')
                query = query.filter(OutboundRecord.outbound_time >= start_date)
            except ValueError:
                pass

        if date_end:
            try:
                end_date = datetime.strptime(date_end, '%Y-%m-%d')
                end_date = end_date.replace(hour=23, minute=59, second=59)
                query = query.filter(OutboundRecord.outbound_time <= end_date)
            except ValueError:
                pass

        # 获取所有记录
        records = query.order_by(OutboundRecord.outbound_time.desc()).all()

        # 准备导出数据
        data = []
        for record in records:
            try:
                data.append({
                    '批次号': record.batch_no or '',
                    '识别编码': record.identification_code or '',
                    '客户名称': record.customer_name or '',
                    '出库车牌': record.plate_number or '',
                    '目的地': record.destination or '',
                    '送货干线车': record.delivery_plate_number or '',
                    '报关行': record.customs_broker or '',
                    '出境模式': record.export_mode or '',
                    '大层数': record.large_layer or 0,
                    '小层数': record.small_layer or 0,
                    '托板数': record.pallet_board or 0,
                    '重量(KG)': record.weight or 0,
                    '体积(CBM)': record.volume or 0,
                    '单据': record.documents or '',
                    '入库日期': record.inbound_date.strftime('%Y-%m-%d') if record.inbound_date else '',
                    '出库时间': record.outbound_time.strftime('%Y-%m-%d %H:%M:%S') if record.outbound_time else '',
                    '出发时间': record.departure_time.strftime('%Y-%m-%d %H:%M:%S') if record.departure_time else '',
                    '备注1': record.remark1 or '',
                    '备注2': record.remark2 or '',
                    '操作仓库': record.operated_warehouse.warehouse_name if record.operated_warehouse else '',
                    '目标仓库': record.destination_warehouse.warehouse_name if record.destination_warehouse else '',
                    '操作用户': record.operated_by_user.username if record.operated_by_user else '',
                })
            except Exception as e:
                current_app.logger.error(f"处理出库记录时出错: {str(e)}")
                continue

        # 使用pandas导出Excel
        try:
            import pandas as pd
            from io import BytesIO

            if not data:
                flash('没有数据可以导出', 'warning')
                return redirect(url_for('main.frontend_outbound_list'))

            df = pd.DataFrame(data)
            output = BytesIO()

            with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
                df.to_excel(writer, sheet_name='前端仓出库记录', index=False)

                # 获取工作表和工作簿对象
                workbook = writer.book
                worksheet = writer.sheets['前端仓出库记录']

                # 设置列宽
                worksheet.set_column('A:A', 15)  # 批次号
                worksheet.set_column('B:B', 20)  # 识别编码
                worksheet.set_column('C:C', 15)  # 客户名称
                worksheet.set_column('D:D', 12)  # 出库车牌
                worksheet.set_column('E:E', 12)  # 目的地
                worksheet.set_column('F:F', 12)  # 送货干线车
                worksheet.set_column('G:G', 12)  # 报关行
                worksheet.set_column('H:H', 12)  # 出境模式
                worksheet.set_column('I:K', 8)   # 大层数、小层数、托板数
                worksheet.set_column('L:M', 10)  # 重量、体积
                worksheet.set_column('N:N', 10)  # 单据
                worksheet.set_column('O:Q', 18)  # 日期时间列
                worksheet.set_column('R:S', 15)  # 备注列
                worksheet.set_column('T:V', 12)  # 仓库、用户列

            output.seek(0)
            filename = f"前端仓出库记录导出_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"

            return send_file(
                output,
                as_attachment=True,
                download_name=filename,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )

        except ImportError:
            flash("导出Excel需要pandas或xlsxwriter库，请安装: pip install pandas xlsxwriter 或 pip install openpyxl", "danger")
            return redirect(url_for('main.frontend_outbound_list'))

    except Exception as e:
        current_app.logger.error(f"导出前端仓出库记录时出错: {str(e)}")
        flash(f"导出数据时出错: {str(e)}", "danger")
        return redirect(url_for('main.frontend_outbound_list'))






@bp.route('/frontend/outbound/list')
@require_permission('OUTBOUND_VIEW')
def frontend_outbound_list():
    """前端仓出库记录列表 - 按批次号分组显示"""



    if not check_warehouse_permission('frontend', 'view'):
        flash('您没有权限访问前端仓出库记录', 'error')
        return redirect(url_for('main.index'))

    # 获取搜索参数
    page = request.args.get('page', 1, type=int)
    per_page = 20  # 每页显示20条记录

    # 构建查询，只显示前端仓的数据
    # 优化：使用selectinload减少查询次数
    query = OutboundRecord.query.options(
        db.selectinload(OutboundRecord.operated_warehouse),
        db.selectinload(OutboundRecord.destination_warehouse)
    )

    # 获取前端仓库ID列表
    frontend_warehouses = Warehouse.query.filter_by(warehouse_type='frontend').all()
    frontend_warehouse_ids = [w.id for w in frontend_warehouses]
    query = query.filter(OutboundRecord.operated_warehouse_id.in_(frontend_warehouse_ids))

    # 根据用户权限进一步过滤数据
    # 管理员和后端仓账号可以查看所有前端仓的数据
    # 前端仓账号只能查看自身仓库操作的数据
    if hasattr(current_user, 'warehouse') and current_user.warehouse:
        if current_user.warehouse.warehouse_type == 'frontend' and not current_user.is_admin:
            # 前端仓用户只能查看自己仓库的数据
            query = query.filter(OutboundRecord.operated_warehouse_id == current_user.warehouse_id)

    # 搜索过滤
    search_params = {}

    # 仓库搜索
    warehouse_id = request.args.get('warehouse_id')
    if warehouse_id:
        query = query.filter(OutboundRecord.operated_warehouse_id == warehouse_id)
        search_params['warehouse_id'] = warehouse_id

    # 统一搜索框和字段选择器
    search_value = request.args.get('search_value', '').strip()
    search_field = request.args.get('search_field', 'customer_name')
    date_start = request.args.get('date_start')
    date_end = request.args.get('date_end')

    # 如果没有指定日期范围，默认显示两天的数据（今天和明天）
    if not date_start and not date_end:
        today = datetime.now()
        tomorrow = today + timedelta(days=1)
        date_start = today.strftime('%Y-%m-%d')
        date_end = tomorrow.strftime('%Y-%m-%d')

    # 日期范围过滤
    if date_start:
        try:
            start_date = datetime.strptime(date_start, '%Y-%m-%d')
            query = query.filter(OutboundRecord.outbound_time >= start_date)
            search_params['date_start'] = date_start
        except ValueError:
            pass

    if date_end:
        try:
            end_date = datetime.strptime(date_end + ' 23:59:59', '%Y-%m-%d %H:%M:%S')
            query = query.filter(OutboundRecord.outbound_time <= end_date)
            search_params['date_end'] = date_end
        except ValueError:
            pass

    # 根据选择的字段进行搜索
    if search_value:
        if search_field == 'customer_name':
            query = query.filter(OutboundRecord.customer_name.contains(search_value))
        elif search_field == 'plate_number':
            query = query.filter(OutboundRecord.plate_number.contains(search_value))
        elif search_field == 'batch_no':
            query = query.filter(OutboundRecord.batch_no.contains(search_value))
        elif search_field == 'identification_code':
            query = query.filter(OutboundRecord.identification_code.contains(search_value))
        elif search_field == 'destination':
            query = query.filter(OutboundRecord.destination.contains(search_value))
        elif search_field == 'customs_broker':
            query = query.filter(OutboundRecord.customs_broker.contains(search_value))

        search_params['search_value'] = search_value
        search_params['search_field'] = search_field

    # 优化：先获取批次信息进行分页，避免加载所有记录
    # 获取不重复的批次号和基本信息，按操作时间排序
    batch_query = db.session.query(
        OutboundRecord.batch_no,
        db.func.min(OutboundRecord.departure_time).label('min_departure_time'),
        db.func.max(OutboundRecord.departure_time).label('max_departure_time'),
        db.func.min(OutboundRecord.outbound_time).label('min_outbound_time'),
        db.func.max(OutboundRecord.outbound_time).label('max_outbound_time'),
        db.func.min(OutboundRecord.created_at).label('min_created_at'),
        db.func.max(OutboundRecord.created_at).label('max_created_at'),
        db.func.count(OutboundRecord.id).label('record_count')
    ).filter(
        OutboundRecord.operated_warehouse_id.in_(frontend_warehouse_ids)
    )

    # 应用相同的搜索过滤条件
    if hasattr(current_user, 'warehouse') and current_user.warehouse:
        if current_user.warehouse.warehouse_type == 'frontend' and not current_user.is_admin:
            batch_query = batch_query.filter(OutboundRecord.operated_warehouse_id == current_user.warehouse_id)

    if warehouse_id:
        batch_query = batch_query.filter(OutboundRecord.operated_warehouse_id == warehouse_id)

    if date_start:
        try:
            start_date = datetime.strptime(date_start, '%Y-%m-%d')
            batch_query = batch_query.filter(OutboundRecord.outbound_time >= start_date)
        except ValueError:
            pass

    if date_end:
        try:
            end_date = datetime.strptime(date_end + ' 23:59:59', '%Y-%m-%d %H:%M:%S')
            batch_query = batch_query.filter(OutboundRecord.outbound_time <= end_date)
        except ValueError:
            pass

    if search_value:
        if search_field == 'customer_name':
            batch_query = batch_query.filter(OutboundRecord.customer_name.contains(search_value))
        elif search_field == 'plate_number':
            batch_query = batch_query.filter(OutboundRecord.plate_number.contains(search_value))
        elif search_field == 'batch_no':
            batch_query = batch_query.filter(OutboundRecord.batch_no.contains(search_value))
        elif search_field == 'identification_code':
            batch_query = batch_query.filter(OutboundRecord.identification_code.contains(search_value))
        elif search_field == 'destination':
            batch_query = batch_query.filter(OutboundRecord.destination.contains(search_value))
        elif search_field == 'customs_broker':
            batch_query = batch_query.filter(OutboundRecord.customs_broker.contains(search_value))

    # 按批次号分组并按操作时间降序排序
    batch_info = batch_query.group_by(OutboundRecord.batch_no).order_by(
        db.func.min(OutboundRecord.created_at).desc()
    ).all()

    # 分页处理批次
    total_batches = len(batch_info)
    start_idx = (page - 1) * per_page
    end_idx = start_idx + per_page
    paginated_batch_info = batch_info[start_idx:end_idx]

    # 只获取当前页面需要的批次的详细记录
    current_page_batch_nos = [batch.batch_no for batch in paginated_batch_info if batch.batch_no]

    if current_page_batch_nos:
        all_records = query.filter(
            OutboundRecord.batch_no.in_(current_page_batch_nos)
        ).order_by(OutboundRecord.created_at.desc()).all()
    else:
        all_records = []

    # 批量获取入库记录以避免N+1查询问题
    if all_records:
        # 收集所有识别编码
        identification_codes = [r.identification_code for r in all_records if r.identification_code]

        # 批量查询入库记录
        inbound_records_dict = {}
        if identification_codes:
            inbound_records = InboundRecord.query.filter(
                InboundRecord.identification_code.in_(identification_codes)
            ).all()
            inbound_records_dict = {r.identification_code: r for r in inbound_records}

        # 为每个出库记录关联入库信息
        for record in all_records:
            if record.identification_code and record.identification_code in inbound_records_dict:
                inbound_record = inbound_records_dict[record.identification_code]

                # 设置入库车牌信息
                if inbound_record.plate_number and not record.inbound_plate:
                    record.inbound_plate = inbound_record.plate_number

                # 设置入库日期（如果出库记录中没有）
                if inbound_record.inbound_time and not record.inbound_date:
                    record.inbound_date = inbound_record.inbound_time

                # 设置订单类型（如果出库记录中没有）
                if inbound_record.order_type and not record.order_type:
                    record.order_type = inbound_record.order_type

                # 设置跟单客服（如果出库记录中没有）
                if inbound_record.service_staff and not record.service_staff:
                    record.service_staff = inbound_record.service_staff

                # 设置报关行（如果出库记录中没有）
                if inbound_record.customs_broker and not record.customs_broker:
                    record.customs_broker = inbound_record.customs_broker

    # 按批次号分组
    from collections import defaultdict
    batch_groups = defaultdict(list)
    for record in all_records:
        batch_no = record.batch_no or '未分配批次'
        batch_groups[batch_no].append(record)

    # 批量获取接收记录以避免N+1查询问题
    batch_nos = [batch_no for batch_no in batch_groups.keys() if batch_no and batch_no != '未分配批次']
    receive_records_dict = {}
    if batch_nos:
        receive_records = ReceiveRecord.query.filter(
            ReceiveRecord.batch_no.in_(batch_nos)
        ).all()

        # 按批次号分组接收记录
        from collections import defaultdict
        receive_records_by_batch = defaultdict(list)
        for record in receive_records:
            if record.batch_no:
                receive_records_by_batch[record.batch_no].append(record)
        receive_records_dict = dict(receive_records_by_batch)

    # 计算每个批次的汇总信息
    batch_summaries = {}
    for batch_no, records in batch_groups.items():
        if records:
            first_record = records[0]
            total_pallet = sum(r.pallet_count or 0 for r in records)
            total_package = sum(r.package_count or 0 for r in records)
            total_weight = sum(r.weight or 0 for r in records)
            total_volume = sum(r.volume or 0 for r in records)
            total_documents = sum(r.document_count or 0 for r in records)

            # 检查批次接收状态
            receive_status = 'unknown'
            received_count = 0
            total_count = len(records)

            if batch_no and batch_no != '未分配批次' and batch_no in receive_records_dict:
                # 使用预加载的接收记录
                received_records = receive_records_dict[batch_no]
                received_identification_codes = {r.identification_code for r in received_records if r.identification_code}

                # 统计已接收的记录数量
                for record in records:
                    if record.identification_code and record.identification_code in received_identification_codes:
                        received_count += 1

                # 判断接收状态
                if received_count == 0:
                    receive_status = 'not_received'  # 未接收
                elif received_count == total_count:
                    receive_status = 'fully_received'  # 全部接收
                else:
                    receive_status = 'partially_received'  # 部分接收
            elif batch_no and batch_no != '未分配批次':
                receive_status = 'not_received'  # 没有接收记录

            batch_summaries[batch_no] = {
                'batch_no': batch_no,
                'plate_number': first_record.plate_number,
                'inbound_plate': first_record.inbound_plate,
                'destination': first_record.destination,
                'destination_warehouse': first_record.destination_warehouse,
                'outbound_time': first_record.outbound_time,
                'departure_time': first_record.departure_time,
                'operated_warehouse': first_record.operated_warehouse,
                'total_pallet': total_pallet,
                'total_package': total_package,
                'total_weight': total_weight,
                'total_volume': total_volume,
                'total_documents': total_documents,
                'record_count': len(records),
                'records': records,
                'receive_status': receive_status,
                'received_count': received_count,
                'total_count': total_count
            }

    # 按出库时间排序批次
    from datetime import datetime as dt
    sorted_batches = sorted(batch_summaries.values(),
                          key=lambda x: x['outbound_time'] or dt.min,
                          reverse=True)

    # 创建分页对象（已经在前面完成了分页）
    class BatchPagination:
        def __init__(self, items, page, per_page, total):
            self.items = items
            self.page = page
            self.per_page = per_page
            self.total = total
            self.pages = (total + per_page - 1) // per_page
            self.has_prev = page > 1
            self.has_next = page < self.pages
            self.prev_num = page - 1 if self.has_prev else None
            self.next_num = page + 1 if self.has_next else None

    # 计算数据汇总（基于所有符合条件的记录，不仅仅是当前页）
    summary_data = {}
    try:
        # 获取所有符合条件的记录进行汇总计算
        all_summary_records = query.all()

        # 计算出库车次（按出库车牌去重）
        unique_plate_numbers = set(r.plate_number for r in all_summary_records if r.plate_number)
        total_vehicle_trips = len(unique_plate_numbers)

        # 计算总数量
        total_pallets = sum(r.pallet_count or 0 for r in all_summary_records)
        total_packages = sum(r.package_count or 0 for r in all_summary_records)

        # 计算总票数（按识别编码去重，考虑分批出库情况）
        unique_identification_codes = set(r.identification_code for r in all_summary_records if r.identification_code)
        total_shipments = len(unique_identification_codes)

        # 计算总重量和总体积
        total_weight = sum(r.weight or 0 for r in all_summary_records)
        total_volume = sum(r.volume or 0 for r in all_summary_records)

        summary_data = {
            'total_vehicle_trips': total_vehicle_trips,  # 出库车次
            'total_pallets': total_pallets,              # 总板数
            'total_packages': total_packages,            # 总件数
            'total_shipments': total_shipments,          # 总票数（按识别编码去重）
            'total_weight': round(total_weight, 1),      # 总重量
            'total_volume': round(total_volume, 2)       # 总体积
        }

        current_app.logger.info(f"前端仓出库汇总数据: {summary_data}")

    except Exception as e:
        current_app.logger.error(f"计算前端仓出库汇总数据失败: {str(e)}")
        summary_data = {
            'total_vehicle_trips': 0,  # 出库车次
            'total_pallets': 0,        # 总板数
            'total_packages': 0,       # 总件数
            'total_shipments': 0,      # 总票数
            'total_weight': 0.0,       # 总重量
            'total_volume': 0.0        # 总体积
        }

    # 使用已经排序的批次数据
    batch_pagination = BatchPagination(sorted_batches, page, per_page, total_batches)

    # 获取前端仓库列表用于搜索下拉框
    # 根据用户权限显示相应的仓库
    if hasattr(current_user, 'warehouse') and current_user.warehouse:
        if current_user.warehouse.warehouse_type == 'frontend' and not current_user.is_admin:
            # 前端仓用户只能看到自己的仓库
            warehouses = [current_user.warehouse]
        else:
            # 管理员和后端仓用户可以看到所有前端仓库
            warehouses = Warehouse.query.filter_by(warehouse_type='frontend').order_by(Warehouse.warehouse_name).all()
    else:
        # 没有关联仓库的用户（如管理员）可以看到所有前端仓库
        warehouses = Warehouse.query.filter_by(warehouse_type='frontend').order_by(Warehouse.warehouse_name).all()

    from app.utils import render_ajax_aware
    return render_ajax_aware('frontend/outbound_list.html',
                           batch_groups=batch_pagination,
                           search_params=search_params,
                           warehouses=warehouses,
                           summary_data=summary_data,
                           title='前端仓出库记录',
                           warehouse_type='frontend')


# 前端仓出库批次删除API
@bp.route('/api/frontend/outbound/delete-batch', methods=['POST'])
@csrf.exempt
def api_frontend_outbound_delete_batch():
    """删除前端仓整个批次的出库记录（仅在后端仓未接收时允许）"""
    try:
        # 检查用户权限
        if not current_user.is_authenticated:
            return jsonify({'success': False, 'message': '用户未认证'}), 401

        # 权限检查：超级管理员或前端仓用户
        if current_user.is_super_admin():
            # 超级管理员可以删除任何批次
            pass
        elif not hasattr(current_user, 'warehouse') or not current_user.warehouse:
            return jsonify({'success': False, 'message': '用户未分配仓库'}), 403
        elif current_user.warehouse.warehouse_type != 'frontend':
            return jsonify({'success': False, 'message': '只有前端仓用户可以删除前端仓出库记录'}), 403

        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'message': '请求数据格式错误'}), 400

        batch_no = data.get('batch_no')
        if not batch_no:
            return jsonify({'success': False, 'message': '缺少批次号'}), 400

        # 查找该批次的所有出库记录
        outbound_records = OutboundRecord.query.filter_by(batch_no=batch_no).all()
        if not outbound_records:
            return jsonify({'success': False, 'message': '未找到指定批次的出库记录'}), 404

        # 检查是否都是前端仓的记录，并且用户有权限删除
        frontend_warehouse_ids = [1, 2, 3]  # 平湖仓、昆山仓、成都仓
        user_warehouse_id = current_user.warehouse_id

        for record in outbound_records:
            # 检查是否为前端仓的记录
            if record.operated_warehouse_id not in frontend_warehouse_ids:
                return jsonify({'success': False, 'message': f'批次 {batch_no} 包含非前端仓的记录，无法删除'}), 403

            # 检查用户是否有权限删除此记录（只能删除自己仓库的记录，除非是管理员）
            if not current_user.is_super_admin() and record.operated_warehouse_id != user_warehouse_id:
                return jsonify({'success': False, 'message': '您只能删除自己仓库的出库记录'}), 403

        # 检查是否已被后端仓接收
        received_records = []
        for record in outbound_records:
            if record.identification_code:
                # 检查ReceiveRecord表中是否有对应的接收记录
                receive_record = ReceiveRecord.query.filter_by(
                    identification_code=record.identification_code,
                    batch_no=batch_no
                ).first()

                if receive_record:
                    received_records.append({
                        'customer_name': record.customer_name,
                        'identification_code': record.identification_code,
                        'receive_time': receive_record.receive_time
                    })

        if received_records:
            # 构建已接收记录的详细信息
            received_info = []
            for rec in received_records[:3]:  # 最多显示3条
                received_info.append(f"客户：{rec['customer_name']}，识别编码：{rec['identification_code']}")

            if len(received_records) > 3:
                received_info.append(f"等共 {len(received_records)} 条记录")

            return jsonify({
                'success': False,
                'message': f'批次 {batch_no} 中有 {len(received_records)} 条记录已被后端仓接收，无法删除：\n\n' + '\n'.join(received_info)
            }), 400

        # 开始批次删除操作（不需要显式begin，SQLAlchemy会自动管理事务）

        try:
            deleted_count = 0
            restored_inventory = []

            for record in outbound_records:
                # 1. 恢复库存 - 修复：确定正确的回退仓库
                if record.identification_code:
                    # 首先尝试找到当前应该存在库存的仓库
                    # 逻辑：查找该识别编码的最新库存记录，确定库存应该在哪个仓库
                    target_warehouse_id = determine_inventory_rollback_warehouse(record)

                    inventory = Inventory.query.filter_by(
                        identification_code=record.identification_code,
                        operated_warehouse_id=target_warehouse_id
                    ).first()

                    if inventory:
                        # 恢复库存数量
                        old_pallet = inventory.pallet_count or 0
                        old_package = inventory.package_count or 0

                        inventory.pallet_count = old_pallet + (record.pallet_count or 0)
                        inventory.package_count = old_package + (record.package_count or 0)
                        inventory.last_updated = datetime.now()

                        restored_inventory.append({
                            'identification_code': record.identification_code,
                            'customer_name': record.customer_name,
                            'pallet_restored': record.pallet_count or 0,
                            'package_restored': record.package_count or 0
                        })

                        current_app.logger.info(f"恢复库存：识别编码={record.identification_code}, 目标仓库ID={target_warehouse_id}, 板数+{record.pallet_count}, 件数+{record.package_count}")
                    else:
                        # 如果库存记录不存在，重新创建到正确的仓库
                        # 通过识别编码查找对应的入库记录
                        inbound_record = InboundRecord.query.filter_by(
                            identification_code=record.identification_code
                        ).first()

                        if inbound_record:
                            new_inventory = Inventory(
                                customer_name=record.customer_name,
                                identification_code=record.identification_code,
                                inbound_pallet_count=inbound_record.pallet_count or 0,
                                inbound_package_count=inbound_record.package_count or 0,
                                pallet_count=record.pallet_count or 0,
                                package_count=record.package_count or 0,
                                weight=record.weight,
                                volume=record.volume,
                                location=record.location,
                                documents=record.documents,
                                export_mode=record.export_mode,
                                order_type=record.order_type,
                                customs_broker=record.customs_broker,
                                inbound_time=inbound_record.inbound_time,
                                plate_number=inbound_record.plate_number,
                                service_staff=record.service_staff,
                                operated_by_user_id=current_user.id,
                                operated_warehouse_id=target_warehouse_id,  # 使用确定的目标仓库ID
                                last_updated=datetime.now()
                            )
                            db.session.add(new_inventory)

                            restored_inventory.append({
                                'identification_code': record.identification_code,
                                'customer_name': record.customer_name,
                                'pallet_restored': record.pallet_count or 0,
                                'package_restored': record.package_count or 0
                            })

                            current_app.logger.info(f"重新创建库存记录：识别编码={record.identification_code}, 目标仓库ID={target_warehouse_id}")

                # 2. 处理对应的在途货物记录回退
                if record.identification_code and record.batch_no:
                    # 查找对应的在途货物记录
                    transit_cargos = TransitCargo.query.filter_by(
                        identification_code=record.identification_code,
                        batch_no=record.batch_no,
                        status='in_transit'
                    ).all()

                    if transit_cargos:
                        for transit_cargo in transit_cargos:
                            current_app.logger.info(f"批次删除时同时删除在途货物记录: {transit_cargo.identification_code}, 批次: {transit_cargo.batch_no}")
                            db.session.delete(transit_cargo)

                        current_app.logger.info(f"批次 {batch_no} 共删除 {len(transit_cargos)} 条在途货物记录")

                # 3. 删除出库记录
                db.session.delete(record)
                deleted_count += 1

            # 3. 记录操作日志
            from app.models import AuditLog
            audit_log = AuditLog(
                user_id=current_user.id,
                warehouse_id=current_user.warehouse_id,
                module='frontend_outbound',
                action='delete_batch',
                resource_type='outbound_batch',
                resource_id=batch_no,
                old_values=json.dumps({
                    'batch_no': batch_no,
                    'deleted_count': deleted_count,
                    'restored_inventory': restored_inventory
                }),
                ip_address=request.remote_addr,
                user_agent=request.headers.get('User-Agent')
            )
            db.session.add(audit_log)

            db.session.commit()

            current_app.logger.info(f"前端仓批次删除成功：用户={current_user.username}, 批次={batch_no}, 删除记录数={deleted_count}")

            # 构建成功消息
            total_pallet = sum(item['pallet_restored'] for item in restored_inventory)
            total_package = sum(item['package_restored'] for item in restored_inventory)

            return jsonify({
                'success': True,
                'message': f'批次 {batch_no} 删除成功！\n\n删除记录：{deleted_count} 条\n恢复库存：板数 {total_pallet}，件数 {total_package}',
                'deleted_count': deleted_count,
                'restored_inventory': restored_inventory
            })

        except Exception as e:
            db.session.rollback()
            current_app.logger.error(f"删除批次时发生错误: {str(e)}")
            return jsonify({'success': False, 'message': f'批次删除失败: {str(e)}'}), 500

    except Exception as e:
        current_app.logger.error(f"前端仓批次删除API异常: {str(e)}")
        return jsonify({'success': False, 'message': f'系统错误: {str(e)}'}), 500





# ==================== 后端仓出库路由 ====================

@bp.route('/backend/outbound', methods=['GET', 'POST'])
@require_permission('OUTBOUND_VIEW')
def backend_outbound():
    """后端仓出库操作主页面 - 选择出库类型"""
    if not check_warehouse_permission('backend', 'view'):
        flash('您没有权限访问后端仓出库功能', 'error')
        return redirect(url_for('main.index'))

    return render_template('backend/outbound_main.html',
                         title='后端仓出库操作',
                         warehouse_type='backend')

@bp.route('/backend/outbound/to_final', methods=['GET', 'POST'])
@require_permission('OUTBOUND_VIEW')
def backend_outbound_to_final():
    """后端仓出库到末端（保税仓/春疆货场）"""
    if not check_warehouse_permission('backend', 'view'):
        flash('您没有权限访问后端仓出库功能', 'error')
        return redirect(url_for('main.index'))

    return render_template('backend/outbound_to_final.html',
                         title='后端仓出库到末端',
                         warehouse_type='backend',
                         destination='final')

@bp.route('/debug/backend/inventory')
@login_required
def debug_backend_inventory_page():
    """调试后端仓库存页面"""
    return render_template('debug_backend_inventory.html',
                         title='调试后端仓库存API')

@bp.route('/backend/outbound/return_frontend', methods=['GET', 'POST'])
@require_permission('OUTBOUND_VIEW')
def backend_outbound_return_frontend():
    """后端仓返回前端仓"""
    if not check_warehouse_permission('backend', 'view'):
        flash('您没有权限访问后端仓出库功能', 'error')
        return redirect(url_for('main.index'))

    # 获取收货人信息
    from app.models import Receiver
    receivers = Receiver.query.all()
    receivers_data = []
    for receiver in receivers:
        receivers_data.append({
            'warehouse_name': receiver.warehouse_name,
            'contact': receiver.contact,
            'address': receiver.address
        })

    return render_template('backend/outbound_return_frontend.html',
                         title='后端仓返回前端仓',
                         warehouse_type='backend',
                         destination='return_frontend',
                         receivers_data=receivers_data)


@bp.route('/backend/outbound/list')
@require_permission('OUTBOUND_VIEW')
def backend_outbound_list():
    """后端仓出库记录列表 - 按批次号分组显示"""
    if not check_warehouse_permission('backend', 'view'):
        flash('您没有权限访问后端仓出库记录', 'error')
        return redirect(url_for('main.index'))

    # 获取搜索参数
    page = request.args.get('page', 1, type=int)
    per_page = 15  # 每页显示15个批次记录

    # 获取日期参数，如果没有指定则使用默认值
    date_start = request.args.get('date_start', '')
    date_end = request.args.get('date_end', '')

    # 如果没有指定日期范围，默认使用最近一周的日期范围
    if not date_start and not date_end:
        today = datetime.now().strftime('%Y-%m-%d')
        one_week_ago = (datetime.now() - timedelta(days=7)).strftime('%Y-%m-%d')
        date_start = one_week_ago
        date_end = today

    # 获取后端仓库
    backend_warehouse = Warehouse.query.filter_by(warehouse_type='backend').first()
    if not backend_warehouse:
        flash('未找到后端仓库', 'error')
        return redirect(url_for('main.index'))

    # 构建查询，只显示后端仓的数据
    query = OutboundRecord.query.options(
        db.joinedload(OutboundRecord.operated_warehouse),
        db.joinedload(OutboundRecord.destination_warehouse)
    ).filter(OutboundRecord.operated_warehouse_id == backend_warehouse.id)

    # 根据用户权限过滤数据
    if hasattr(current_user, 'warehouse') and current_user.warehouse:
        if current_user.warehouse.warehouse_type == 'backend':
            # 后端仓用户只能看自己仓库的数据
            query = query.filter_by(operated_warehouse_id=current_user.warehouse_id)

    # 日期范围过滤
    if date_start:
        try:
            start_date = datetime.strptime(date_start, '%Y-%m-%d')
            query = query.filter(OutboundRecord.outbound_time >= start_date)
        except ValueError:
            pass
    if date_end:
        try:
            end_date = datetime.strptime(date_end, '%Y-%m-%d')
            end_date = end_date.replace(hour=23, minute=59, second=59)
            query = query.filter(OutboundRecord.outbound_time <= end_date)
        except ValueError:
            pass

    # 搜索过滤 - 支持新的搜索方式
    search_params = {
        'date_start': date_start,
        'date_end': date_end
    }

    # 获取搜索字段和值
    search_field = request.args.get('search_field', '')
    search_value = request.args.get('search_value', '')
    search_condition = request.args.get('search_condition', 'contains')

    if search_field and search_value:
        search_params['search_field'] = search_field
        search_params['search_value'] = search_value
        search_params['search_condition'] = search_condition

        # 根据搜索条件应用过滤
        if search_condition == 'exact':
            filter_condition = getattr(OutboundRecord, search_field) == search_value
        elif search_condition == 'startswith':
            filter_condition = getattr(OutboundRecord, search_field).like(f'{search_value}%')
        elif search_condition == 'endswith':
            filter_condition = getattr(OutboundRecord, search_field).like(f'%{search_value}')
        else:  # contains
            filter_condition = getattr(OutboundRecord, search_field).like(f'%{search_value}%')

        query = query.filter(filter_condition)

    # 兼容旧的搜索参数
    if request.args.get('plate_number'):
        query = query.filter(OutboundRecord.plate_number.contains(request.args.get('plate_number')))
        search_params['plate_number'] = request.args.get('plate_number')

    if request.args.get('batch_no'):
        query = query.filter(OutboundRecord.batch_no.contains(request.args.get('batch_no')))
        search_params['batch_no'] = request.args.get('batch_no')

    if request.args.get('customer_name'):
        query = query.filter(OutboundRecord.customer_name.contains(request.args.get('customer_name')))
        search_params['customer_name'] = request.args.get('customer_name')

    if request.args.get('destination'):
        query = query.filter(OutboundRecord.destination.contains(request.args.get('destination')))
        search_params['destination'] = request.args.get('destination')

    if request.args.get('identification_code'):
        query = query.filter(OutboundRecord.identification_code.contains(request.args.get('identification_code')))
        search_params['identification_code'] = request.args.get('identification_code')

    if request.args.get('service_staff'):
        query = query.filter(OutboundRecord.service_staff.contains(request.args.get('service_staff')))
        search_params['service_staff'] = request.args.get('service_staff')

    if request.args.get('customs_broker'):
        query = query.filter(OutboundRecord.customs_broker.contains(request.args.get('customs_broker')))
        search_params['customs_broker'] = request.args.get('customs_broker')

    if request.args.get('export_mode'):
        query = query.filter(OutboundRecord.export_mode.contains(request.args.get('export_mode')))
        search_params['export_mode'] = request.args.get('export_mode')

    if request.args.get('inbound_plate'):
        query = query.filter(OutboundRecord.inbound_plate.contains(request.args.get('inbound_plate')))
        search_params['inbound_plate'] = request.args.get('inbound_plate')

    # 获取所有符合条件的记录，按批次号和出库时间排序
    all_records = query.order_by(OutboundRecord.batch_no.desc(), OutboundRecord.outbound_time.asc()).all()

    # 按批次号分组
    from collections import defaultdict
    batch_groups = defaultdict(list)

    for record in all_records:
        # 为每个出库记录关联入库车牌信息和其他缺失字段
        if record.identification_code:
            # 通过识别编码查找对应的入库记录
            inbound_record = InboundRecord.query.filter_by(
                identification_code=record.identification_code
            ).first()
            if inbound_record:
                # 设置入库车牌信息
                if inbound_record.plate_number and not record.inbound_plate:
                    record.inbound_plate = inbound_record.plate_number

                # 设置入库日期（如果出库记录中没有）
                if inbound_record.inbound_time and not record.inbound_date:
                    record.inbound_date = inbound_record.inbound_time

                # 设置订单类型（如果出库记录中没有）
                if inbound_record.order_type and not record.order_type:
                    record.order_type = inbound_record.order_type

                # 设置跟单客服（如果出库记录中没有）
                if inbound_record.service_staff and not record.service_staff:
                    record.service_staff = inbound_record.service_staff

                # 设置报关行（如果出库记录中没有）
                if inbound_record.customs_broker and not record.customs_broker:
                    record.customs_broker = inbound_record.customs_broker

                # 设置出境模式（如果出库记录中没有）
                if inbound_record.export_mode and not record.export_mode:
                    record.export_mode = inbound_record.export_mode

        # 按批次号分组
        batch_no = record.batch_no or '未分批'
        batch_groups[batch_no].append(record)

    # 计算每个批次的汇总信息
    batch_summaries = {}
    for batch_no, records in batch_groups.items():
        if records:
            first_record = records[0]
            total_pallet_count = sum(r.pallet_count or 0 for r in records)
            total_package_count = sum(r.package_count or 0 for r in records)
            total_weight = sum(r.weight or 0 for r in records)
            total_volume = sum(r.volume or 0 for r in records)

            batch_summaries[batch_no] = {
                'batch_no': batch_no,
                'outbound_time': first_record.outbound_time,
                'plate_number': first_record.plate_number,
                'destination': first_record.destination,
                'export_mode': first_record.export_mode,
                'customs_broker': first_record.customs_broker,
                'total_pallet_count': total_pallet_count,
                'total_package_count': total_package_count,
                'total_weight': total_weight,
                'total_volume': total_volume,
                'record_count': len(records),
                'records': records
            }

    # 对批次进行分页
    batch_list = list(batch_summaries.values())
    batch_list.sort(key=lambda x: x['outbound_time'] or datetime.min, reverse=True)

    # 手动分页
    total_batches = len(batch_list)
    start_idx = (page - 1) * per_page
    end_idx = start_idx + per_page
    paginated_batches = batch_list[start_idx:end_idx]

    # 创建分页对象
    class BatchPagination:
        def __init__(self, items, page, per_page, total):
            self.items = items
            self.page = page
            self.per_page = per_page
            self.total = total
            self.pages = (total + per_page - 1) // per_page
            self.has_prev = page > 1
            self.has_next = page < self.pages
            self.prev_num = page - 1 if self.has_prev else None
            self.next_num = page + 1 if self.has_next else None

    batch_pagination = BatchPagination(paginated_batches, page, per_page, total_batches)

    from app.utils import render_ajax_aware
    return render_ajax_aware('backend/outbound_list.html',
                           batch_groups=batch_pagination,
                           search_params=search_params,
                           title='后端仓出库记录',
                           warehouse_type='backend')


@bp.route('/backend/outbound/return')
@require_permission('OUTBOUND_VIEW')
def backend_outbound_return():
    """后端仓返回前端仓页面"""
    if not check_warehouse_permission('backend', 'view'):
        flash('您没有权限访问后端仓返回前端仓功能', 'error')
        return redirect(url_for('main.index'))

    return render_template('backend/outbound_return.html',
                         title='后端仓返回前端仓',
                         warehouse_type='backend')


@bp.route('/backend/outbound/final')
@require_permission('OUTBOUND_VIEW')
def backend_outbound_final():
    """后端仓出库到凭祥保税仓/春疆货场页面"""
    if not check_warehouse_permission('backend', 'view'):
        flash('您没有权限访问后端仓出库到凭祥保税仓/春疆货场功能', 'error')
        return redirect(url_for('main.index'))

    return render_template('frontend/outbound_direct.html',
                         title='后端仓出库到凭祥保税仓/春疆货场',
                         warehouse_type='backend',
                         destination='chunjiang')


# ==================== 前端仓库存路由 ====================

@bp.route('/frontend/inventory/list')
@require_permission('INVENTORY_VIEW')
def frontend_inventory_list():
    """前端仓库存列表"""
    if not check_warehouse_permission('frontend', 'view'):
        flash('您没有权限访问前端仓库存', 'error')
        return redirect(url_for('main.index'))

    # 获取搜索参数
    customer_name = request.args.get('customer_name', '')
    location = request.args.get('location', '')

    # 确保搜索参数传递给模板
    search_params = {
        'customer_name': customer_name,
        'location': location
    }

    try:
        page = request.args.get('page', 1, type=int)
        per_page = 20  # 每页显示20条记录

        # 获取前端仓库
        frontend_warehouses = Warehouse.query.filter_by(warehouse_type='frontend').all()
        frontend_warehouse_ids = [w.id for w in frontend_warehouses]

        # 构建查询 - 只查询前端仓库的库存
        query = Inventory.query.options(db.joinedload(Inventory.operated_warehouse))
        query = query.filter(Inventory.operated_warehouse_id.in_(frontend_warehouse_ids))

        # 根据用户权限进一步过滤数据
        # 管理员和后端仓账号可以查看所有前端仓的数据
        # 前端仓账号只能查看自身仓库操作的数据
        if hasattr(current_user, 'warehouse') and current_user.warehouse:
            if current_user.warehouse.warehouse_type == 'frontend' and not current_user.is_admin:
                # 前端仓用户只能查看自己仓库的数据
                query = query.filter(Inventory.operated_warehouse_id == current_user.warehouse_id)

        # 只显示库存不为0的记录
        query = query.filter((Inventory.pallet_count > 0) | (Inventory.package_count > 0))

        # 按客户名称筛选
        if customer_name:
            query = query.filter(Inventory.customer_name.like(f'%{customer_name}%'))

        # 按库位筛选
        if location:
            query = query.filter(Inventory.location.like(f'%{location}%'))

        # 按入库日期升序排序
        query = query.order_by(Inventory.inbound_time.asc())

        # 获取总记录数
        total_count = query.count()

        # 分页
        items = query.offset((page - 1) * per_page).limit(per_page).all()

        # 为每个库存记录补充完整信息（使用与全库存查询相同的逻辑）
        enriched_items = []
        for inventory in items:
            # 使用与全库存查询相同的多表关联逻辑
            _enrich_inventory_record(inventory)
            enriched_items.append(inventory)

        # 使用自定义分页类
        records = SimplePagination(
            items=enriched_items,
            page=page,
            per_page=per_page,
            total=total_count
        )

        return render_template(
            'frontend/inventory_list.html',
            title='前端仓库存',
            records=records,
            search_params=search_params
        )
    except Exception as e:
        current_app.logger.error(f"Error in frontend_inventory_list: {str(e)}")
        flash('获取前端仓库存数据失败', 'error')
        return redirect(url_for('main.index'))


@bp.route('/export_frontend_inventory')
@require_permission('INVENTORY_VIEW')
def export_frontend_inventory():
    """导出前端仓库存记录"""
    if not check_warehouse_permission('frontend', 'view'):
        flash('您没有权限导出前端仓库存', 'error')
        return redirect(url_for('main.frontend_inventory_list'))

    # 获取搜索参数
    customer_name = request.args.get('customer_name', '')
    location = request.args.get('location', '')

    try:
        # 获取前端仓库
        frontend_warehouses = Warehouse.query.filter_by(warehouse_type='frontend').all()
        frontend_warehouse_ids = [w.id for w in frontend_warehouses]

        # 构建查询 - 只查询前端仓库的库存
        query = Inventory.query.options(db.joinedload(Inventory.operated_warehouse))
        query = query.filter(Inventory.operated_warehouse_id.in_(frontend_warehouse_ids))

        # 根据用户权限进一步过滤数据
        if hasattr(current_user, 'warehouse') and current_user.warehouse and not current_user.is_admin:
            if current_user.warehouse.warehouse_type == 'frontend':
                # 前端仓用户只能导出自己仓库的库存
                query = query.filter(Inventory.operated_warehouse_id == current_user.warehouse_id)

        # 只显示库存不为0的记录
        query = query.filter((Inventory.pallet_count > 0) | (Inventory.package_count > 0))

        # 按客户名称筛选
        if customer_name:
            query = query.filter(Inventory.customer_name.like(f'%{customer_name}%'))

        # 按库位筛选
        if location:
            query = query.filter(Inventory.location.like(f'%{location}%'))

        # 按入库日期升序排序
        query = query.order_by(Inventory.inbound_time.asc())

        # 执行查询
        inventory_records = query.all()

        # 为每个库存记录补充完整信息（使用与全库存查询相同的逻辑）
        for inventory in inventory_records:
            _enrich_inventory_record(inventory)

        records = inventory_records

        # 创建工作簿
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "前端仓库存记录"

        # 添加表头
        headers = [
            '序号', '入库日期', '入库车牌', '客户名称', '订单类型', '识别编码',
            '入库板数', '入库件数', '库存板数', '库存件数',
            '重量(kg)', '体积(m³)', '出境模式', '报关行', '单据', '跟单客服', '库位', '备注', '仓库'
        ]

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            # 设置表头样式
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

        # 添加数据
        for row_num, record in enumerate(records, 1):
            # 序号从1开始
            ws.cell(row=row_num+1, column=1).value = row_num

            # 入库日期
            if record.outbound_time:
                ws.cell(row=row_num+1, column=2).value = record.outbound_time.strftime('%Y-%m-%d')
            else:
                ws.cell(row=row_num+1, column=2).value = ''

            # 入库车牌
            ws.cell(row=row_num+1, column=3).value = record.plate_number or ''

            # 客户名称
            ws.cell(row=row_num+1, column=4).value = record.customer_name

            # 订单类型
            ws.cell(row=row_num+1, column=5).value = record.order_type or ''

            # 识别编码
            ws.cell(row=row_num+1, column=6).value = record.identification_code or ''

            # 入库板数
            ws.cell(row=row_num+1, column=7).value = record.inbound_pallet_count or record.pallet_count

            # 入库件数
            ws.cell(row=row_num+1, column=8).value = record.inbound_package_count or record.package_count

            # 库存板数
            ws.cell(row=row_num+1, column=9).value = record.pallet_count

            # 库存件数
            ws.cell(row=row_num+1, column=10).value = record.package_count

            # 重量(kg)
            ws.cell(row=row_num+1, column=11).value = record.weight if record.weight and record.weight > 0 else ''

            # 体积(m³)
            ws.cell(row=row_num+1, column=12).value = record.volume if record.volume and record.volume > 0 else ''

            # 出境模式
            ws.cell(row=row_num+1, column=13).value = record.export_mode or ''

            # 报关行
            ws.cell(row=row_num+1, column=14).value = record.customs_broker or ''

            # 单据
            ws.cell(row=row_num+1, column=15).value = record.documents or ''

            # 跟单客服
            ws.cell(row=row_num+1, column=16).value = record.service_staff or ''

            # 库位
            ws.cell(row=row_num+1, column=17).value = record.location or ''

            # 备注
            ws.cell(row=row_num+1, column=18).value = getattr(record, 'remark1', '') or ''

            # 仓库
            ws.cell(row=row_num+1, column=19).value = record.operated_warehouse.warehouse_name if record.operated_warehouse else ''

        # 调整列宽
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter  # 获取列字母
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column].width = adjusted_width

        # 创建响应
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # 生成文件名
        now = datetime.now().strftime('%Y%m%d%H%M%S')
        filename = f"前端仓库存记录_{now}.xlsx"

        # 返回Excel文件
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            download_name=filename,
            as_attachment=True
        )

    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        current_app.logger.error(f"导出前端仓库存记录失败: {str(e)}")
        current_app.logger.error(f"错误详情: {error_details}")
        flash(f'导出前端仓库存记录失败: {str(e)}', 'error')
        return redirect(url_for('main.frontend_inventory_list'))


# ==================== 后端仓库存路由 ====================

@bp.route('/backend/inventory/list')
@require_permission('INVENTORY_VIEW')
def backend_inventory_list():
    """后端仓库存列表"""
    if not check_warehouse_permission('backend', 'view'):
        flash('您没有权限访问后端仓库存', 'error')
        return redirect(url_for('main.index'))

    # 获取搜索参数
    customer_name = request.args.get('customer_name', '')
    location = request.args.get('location', '')

    # 确保搜索参数传递给模板
    search_params = {
        'customer_name': customer_name,
        'location': location
    }

    try:
        page = request.args.get('page', 1, type=int)
        per_page = current_app.config.get('ITEMS_PER_PAGE', 50)

        # 获取后端仓库
        backend_warehouse = Warehouse.query.filter_by(warehouse_type='backend').first()
        if not backend_warehouse:
            flash('未找到后端仓库', 'error')
            return redirect(url_for('main.index'))

        # 构建查询 - 只查询后端仓库的库存
        query = Inventory.query.options(db.joinedload(Inventory.operated_warehouse))
        query = query.filter(Inventory.operated_warehouse_id == backend_warehouse.id)

        # 只显示库存不为0的记录
        query = query.filter((Inventory.pallet_count > 0) | (Inventory.package_count > 0))

        # 按客户名称筛选
        if customer_name:
            query = query.filter(Inventory.customer_name.like(f'%{customer_name}%'))

        # 按库位筛选
        if location:
            query = query.filter(Inventory.location.like(f'%{location}%'))

        # 按入库日期升序排序
        query = query.order_by(Inventory.inbound_time.asc())

        # 获取总记录数
        total_count = query.count()

        # 分页
        items = query.offset((page - 1) * per_page).limit(per_page).all()

        # 使用自定义分页类
        records = SimplePagination(
            items=items,
            page=page,
            per_page=per_page,
            total=total_count
        )

        # 为模板添加辅助函数
        def get_frontend_inbound_record(identification_code):
            """获取前端仓入库记录"""
            if not identification_code:
                return None

            # 直接使用识别编码查找前端仓入库记录（不再需要处理_B后缀）
            frontend_inbound = InboundRecord.query.filter(
                InboundRecord.identification_code == identification_code,
                InboundRecord.record_type == 'direct'  # 只查找直接入库记录
            ).join(Warehouse).filter(
                Warehouse.warehouse_type == 'frontend'
            ).first()

            return frontend_inbound

        return render_template(
            'backend/inventory_list.html',
            title='后端仓库存',
            records=records,
            search_params=search_params,
            get_frontend_inbound_record=get_frontend_inbound_record
        )
    except Exception as e:
        current_app.logger.error(f"Error in backend_inventory_list: {str(e)}")
        flash('获取后端仓库存数据失败', 'error')
        return redirect(url_for('main.index'))


@bp.route('/export_backend_inventory')
@require_permission('INVENTORY_VIEW')
def export_backend_inventory():
    """导出后端仓库存记录"""
    if not check_warehouse_permission('backend', 'view'):
        flash('您没有权限导出后端仓库存', 'error')
        return redirect(url_for('main.backend_inventory_list'))

    # 获取搜索参数
    customer_name = request.args.get('customer_name', '')
    location = request.args.get('location', '')

    try:
        # 获取后端仓库
        backend_warehouse = Warehouse.query.filter_by(warehouse_type='backend').first()
        if not backend_warehouse:
            flash('未找到后端仓库', 'error')
            return redirect(url_for('main.backend_inventory_list'))

        # 构建查询 - 只查询后端仓库的库存
        query = Inventory.query.options(db.joinedload(Inventory.operated_warehouse))
        query = query.filter(Inventory.operated_warehouse_id == backend_warehouse.id)

        # 只显示库存不为0的记录
        query = query.filter((Inventory.pallet_count > 0) | (Inventory.package_count > 0))

        # 按客户名称筛选
        if customer_name:
            query = query.filter(Inventory.customer_name.like(f'%{customer_name}%'))

        # 按库位筛选
        if location:
            query = query.filter(Inventory.location.like(f'%{location}%'))

        # 按入库日期升序排序
        query = query.order_by(Inventory.inbound_time.asc())

        # 执行查询
        records = query.all()

        # 创建工作簿
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "后端仓库存记录"

        # 添加表头
        headers = [
            '序号', '入库日期', '入库车牌', '客户名称', '订单类型', '识别编码',
            '入库板数', '入库件数', '库存板数', '库存件数',
            '重量(kg)', '体积(m³)', '出境模式', '报关行', '单据', '跟单客服', '库位', '备注'
        ]

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            # 设置表头样式
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

        # 添加数据
        for row_num, record in enumerate(records, 1):
            # 序号从1开始
            ws.cell(row=row_num+1, column=1).value = row_num

            # 入库日期
            if record.outbound_time:
                ws.cell(row=row_num+1, column=2).value = record.outbound_time.strftime('%Y-%m-%d')
            else:
                ws.cell(row=row_num+1, column=2).value = ''

            # 入库车牌
            ws.cell(row=row_num+1, column=3).value = record.plate_number or ''

            # 客户名称
            ws.cell(row=row_num+1, column=4).value = record.customer_name

            # 订单类型
            ws.cell(row=row_num+1, column=5).value = record.order_type or ''

            # 识别编码
            ws.cell(row=row_num+1, column=6).value = record.identification_code or ''

            # 入库板数
            ws.cell(row=row_num+1, column=7).value = record.inbound_pallet_count or record.pallet_count

            # 入库件数
            ws.cell(row=row_num+1, column=8).value = record.inbound_package_count or record.package_count

            # 库存板数
            ws.cell(row=row_num+1, column=9).value = record.pallet_count

            # 库存件数
            ws.cell(row=row_num+1, column=10).value = record.package_count

            # 重量(kg)
            ws.cell(row=row_num+1, column=11).value = record.weight if record.weight and record.weight > 0 else ''

            # 体积(m³)
            ws.cell(row=row_num+1, column=12).value = record.volume if record.volume and record.volume > 0 else ''

            # 出境模式
            ws.cell(row=row_num+1, column=13).value = record.export_mode or ''

            # 报关行
            ws.cell(row=row_num+1, column=14).value = record.customs_broker or ''

            # 单据
            ws.cell(row=row_num+1, column=15).value = record.documents or ''

            # 跟单客服
            ws.cell(row=row_num+1, column=16).value = record.service_staff or ''

            # 库位
            ws.cell(row=row_num+1, column=17).value = record.location or ''

            # 备注
            ws.cell(row=row_num+1, column=18).value = getattr(record, 'remark1', '') or ''

        # 调整列宽
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter  # 获取列字母
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column].width = adjusted_width

        # 创建响应
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # 生成文件名
        now = datetime.now().strftime('%Y%m%d%H%M%S')
        filename = f"后端仓库存记录_{now}.xlsx"

        # 返回Excel文件
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            download_name=filename,
            as_attachment=True
        )

    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        current_app.logger.error(f"导出后端仓库存记录失败: {str(e)}")
        current_app.logger.error(f"错误详情: {error_details}")
        flash(f'导出后端仓库存记录失败: {str(e)}', 'error')
        return redirect(url_for('main.backend_inventory_list'))


# ==================== 后端仓API ====================

@bp.route('/api/backend/inventory')
@csrf.exempt
@require_permission('INVENTORY_VIEW')
def api_backend_inventory():
    """获取后端仓库存数据"""
    try:
        # 获取后端仓库
        backend_warehouse = Warehouse.query.filter_by(warehouse_type='backend').first()
        if not backend_warehouse:
            return jsonify({'success': False, 'message': '未找到后端仓库'}), 400

        # 查询后端仓库存（只显示有库存的记录）
        query = Inventory.query.filter(
            db.and_(
                Inventory.operated_warehouse_id == backend_warehouse.id,
                db.or_(
                    Inventory.pallet_count > 0,
                    Inventory.package_count > 0
                )
            )
        ).order_by(Inventory.inbound_time.desc())

        inventory_records = query.all()

        # 转换为JSON格式
        inventory_data = []
        for record in inventory_records:
            # 使用库存记录中的当前库存数量（这些已经是实时更新的可用库存）
            available_pallets = record.pallet_count or 0
            available_packages = record.package_count or 0

            # 只显示有库存的记录，且不显示负数库存（查询条件已经过滤了，这里再次确认）
            if available_pallets >= 0 and available_packages >= 0 and (available_pallets > 0 or available_packages > 0):
                # 清理和限制识别编码长度
                identification_code = record.identification_code or ''
                if len(identification_code) > 50:  # 限制识别编码长度
                    identification_code = identification_code[:50] + '...'

                # 查找对应的入库记录以获取送货干线车信息
                inbound_record = InboundRecord.query.filter_by(
                    identification_code=record.identification_code,
                    customer_name=record.customer_name
                ).order_by(InboundRecord.inbound_time.desc()).first()

                delivery_plate_number = ''
                if inbound_record:
                    delivery_plate_number = inbound_record.delivery_plate_number or ''

                inventory_data.append({
                    'id': record.id,
                    'customer_name': record.customer_name or '',
                    'plate_number': record.plate_number or '',
                    'identification_code': identification_code,
                    'order_type': record.order_type or '',
                    'inbound_pallet_count': record.inbound_pallet_count or 0,
                    'inbound_package_count': record.inbound_package_count or 0,
                    'available_pallets': available_pallets,  # 可用库存板数
                    'available_packages': available_packages,  # 可用库存件数
                    'pallet_count': available_pallets,  # 库存板数（兼容性）
                    'package_count': available_packages,  # 库存件数（兼容性）
                    'weight': record.weight or 0,
                    'volume': record.volume or 0,
                    'export_mode': record.export_mode or '',
                    'customs_broker': record.customs_broker or '',
                    'delivery_plate_number': delivery_plate_number,  # 从入库记录获取
                    'service_staff': record.service_staff or '',
                    'inbound_time': record.outbound_time.strftime('%Y-%m-%d') if record.outbound_time else '',
                    'location': record.location or '',
                    'documents': record.documents or ''
                })

        return jsonify({
            'success': True,
            'data': inventory_data,
            'total': len(inventory_data)
        })

    except Exception as e:
        current_app.logger.error(f"获取后端仓库存数据失败: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'获取库存数据失败: {str(e)}'
        }), 500


# ==================== 后端仓出库API ====================

# 后端仓出库到末端API
@bp.route('/api/backend/outbound/to_final', methods=['POST'])
@csrf.exempt
@require_permission('OUTBOUND_CREATE')
def api_backend_outbound_to_final():
    """后端仓出库到末端（保税仓/春疆货场）API"""
    try:
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'message': '没有接收到数据'}), 400

        records = data.get('records', [])
        final_destination = data.get('final_destination', '保税仓')

        if not records:
            return jsonify({'success': False, 'message': '没有出库记录'}), 400

        # 生成批次号
        from app.utils.batch_generator import generate_batch_number

        # 根据目的地设置前缀
        destination_prefix = None
        if final_destination == '保税仓' or final_destination == '凭祥保税仓':
            destination_prefix = 'PB'  # 保税仓专用前缀
        elif final_destination == '春疆货场':
            destination_prefix = 'CJ'  # 春疆货场专用前缀

        batch_number = generate_batch_number(
            warehouse_id=current_user.warehouse_id,
            destination_prefix=destination_prefix,
            db_session=db.session
        )
        current_app.logger.info(f'后端仓出库到末端生成批次号: {batch_number}，目的地: {final_destination}')

        success_count = 0
        errors = []

        for i, record_data in enumerate(records):
            try:
                # 验证必填字段
                if not record_data.get('customer_name'):
                    errors.append(f'第{i+1}行：客户名称不能为空')
                    continue

                if not record_data.get('delivery_plate_number'):
                    errors.append(f'第{i+1}行：运输车牌不能为空')
                    continue

                if not record_data.get('customs_broker'):
                    errors.append(f'第{i+1}行：报关行不能为空')
                    continue

                # 创建出库记录
                outbound_record = OutboundRecord(
                    outbound_time=datetime.strptime(record_data.get('outbound_time'), '%Y-%m-%d') if record_data.get('outbound_time') else datetime.now(),
                    delivery_plate_number=record_data.get('delivery_plate_number'),
                    plate_number=record_data.get('delivery_plate_number'),
                    customer_name=record_data.get('customer_name'),
                    identification_code=record_data.get('identification_code', ''),
                    pallet_count=int(record_data.get('pallet_count', 0)) if record_data.get('pallet_count') else 0,
                    package_count=int(record_data.get('package_count', 0)) if record_data.get('package_count') else 0,
                    weight=float(record_data.get('weight', 0)) if record_data.get('weight') else 0,
                    volume=float(record_data.get('volume', 0)) if record_data.get('volume') else 0,
                    destination=final_destination,
                    customs_broker=record_data.get('customs_broker'),
                    export_mode=record_data.get('export_mode', ''),
                    order_type=record_data.get('order_type', ''),
                    batch_no=batch_number,
                    remarks=record_data.get('remarks', ''),
                    operated_by_user_id=current_user.id,
                    operated_warehouse_id=current_user.warehouse_id
                )

                db.session.add(outbound_record)
                success_count += 1

            except Exception as e:
                errors.append(f'第{i+1}行：{str(e)}')
                continue

        if success_count > 0:
            db.session.commit()
            current_app.logger.info(f'后端仓出库到末端成功：{success_count}条记录')

        return jsonify({
            'success': True,
            'message': f'成功保存{success_count}条记录到{final_destination}',
            'success_count': success_count,
            'total_count': len(records),
            'errors': errors
        })

    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f'后端仓出库到末端失败: {str(e)}')
        return jsonify({'success': False, 'message': str(e)}), 500

# 后端仓返回前端仓API
@bp.route('/api/backend/outbound/return_frontend', methods=['POST'])
@csrf.exempt
@require_permission('OUTBOUND_CREATE')
def api_backend_outbound_return_frontend():
    """后端仓返回前端仓API"""
    print("=== 后端仓返回前端仓API被调用 ===")
    current_app.logger.info("=== 后端仓返回前端仓API被调用 ===")
    try:
        data = request.get_json()
        print(f"接收到的数据: {data}")
        current_app.logger.info(f"接收到的数据: {data}")

        if not data:
            return jsonify({'success': False, 'message': '没有接收到数据'}), 400

        current_app.logger.info(f"收到后端仓返回前端仓数据: {json.dumps(data, ensure_ascii=False)}")

        records = data.get('records', [])
        common_data = data.get('common_data', {})
        return_reason = data.get('return_reason', '客户要求')

        if not records:
            return jsonify({'success': False, 'message': '没有出库记录'}), 400

        # 获取公共数据
        delivery_plate_number = common_data.get('plate_number', '')
        driver_name = common_data.get('driver_name', '')
        driver_contact = common_data.get('driver_contact', '')
        target_warehouse = common_data.get('target_warehouse', '')
        return_time_str = common_data.get('return_time', '')

        if not delivery_plate_number:
            return jsonify({'success': False, 'message': '出库车牌不能为空'}), 400

        if not driver_name:
            return jsonify({'success': False, 'message': '司机姓名不能为空'}), 400

        if not driver_contact:
            return jsonify({'success': False, 'message': '司机联系方式不能为空'}), 400

        if not target_warehouse:
            return jsonify({'success': False, 'message': '目的仓不能为空'}), 400

        # 生成批次号
        from app.utils.batch_generator import generate_batch_number
        batch_number = generate_batch_number(
            warehouse_id=current_user.warehouse_id,
            destination_prefix='RT',  # 返回专用前缀
            db_session=db.session
        )
        current_app.logger.info(f'后端仓返回前端仓生成批次号: {batch_number}')

        success_count = 0
        errors = []

        for i, record_data in enumerate(records):
            try:
                current_app.logger.info(f"处理第{i+1}行数据: {record_data}, 类型: {type(record_data)}, 长度: {len(record_data) if isinstance(record_data, list) else 'N/A'}")

                # Handsontable数据是数组格式，需要转换为字典
                if isinstance(record_data, list):
                    # 根据列的顺序映射数据（移除了返回时间列）
                    # ['plate_number', 'customer_name', 'identification_code',
                    #  'pallet_count', 'package_count', 'weight', 'volume', 'documents', 'service_staff', 'remarks']
                    if len(record_data) >= 10:
                        record_dict = {
                            'plate_number': record_data[0],
                            'customer_name': record_data[1],
                            'identification_code': record_data[2],
                            'pallet_count': record_data[3],
                            'package_count': record_data[4],
                            'weight': record_data[5],
                            'volume': record_data[6],
                            'documents': record_data[7],
                            'service_staff': record_data[8],
                            'remarks': record_data[9]
                        }
                        current_app.logger.info(f"转换后的字典: {record_dict}")
                    else:
                        errors.append(f'第{i+1}行：数据格式不完整，只有{len(record_data)}列')
                        continue
                elif isinstance(record_data, dict):
                    record_dict = record_data
                else:
                    errors.append(f'第{i+1}行：不支持的数据类型 {type(record_data)}')
                    continue

                # 验证必填字段
                customer_name_raw = record_dict.get('customer_name', '')
                customer_name = str(customer_name_raw).strip() if customer_name_raw else ''
                if not customer_name:
                    errors.append(f'第{i+1}行：客户名称不能为空')
                    continue

                # 验证数量字段
                try:
                    pallet_count_raw = record_dict.get('pallet_count', 0)
                    package_count_raw = record_dict.get('package_count', 0)

                    # 处理空字符串和None值
                    pallet_count = int(pallet_count_raw) if pallet_count_raw and str(pallet_count_raw).strip() else 0
                    package_count = int(package_count_raw) if package_count_raw and str(package_count_raw).strip() else 0

                    current_app.logger.info(f"第{i+1}行数量字段: 板数={pallet_count}, 件数={package_count}")

                    if pallet_count <= 0 and package_count <= 0:
                        errors.append(f'第{i+1}行：板数和件数至少填写一项且大于0')
                        continue
                except (ValueError, TypeError) as e:
                    errors.append(f'第{i+1}行：数量字段格式错误 - {str(e)}')
                    continue

                # 查找对应的库存记录进行扣减
                identification_code = record_dict.get('identification_code', '')
                if identification_code:
                    # 查找该识别编码对应的有实际库存的记录（不限制仓库类型）
                    # 优先查找有库存数量的记录
                    inventory_record = Inventory.query.filter(
                        Inventory.identification_code == identification_code,
                        db.or_(
                            Inventory.pallet_count > 0,
                            Inventory.package_count > 0
                        )
                    ).first()

                    # 如果没有找到有库存的记录，则查找所有记录中的第一个
                    if not inventory_record:
                        inventory_record = Inventory.query.filter(
                            Inventory.identification_code == identification_code
                        ).first()

                    warehouse_info = ""
                    if inventory_record:
                        warehouse = Warehouse.query.get(inventory_record.operated_warehouse_id)
                        warehouse_info = f"{warehouse.warehouse_name}({warehouse.warehouse_type})" if warehouse else "未知仓库"

                    current_app.logger.info(f"查找库存: 识别编码={identification_code}, 找到记录: {inventory_record is not None}, 仓库: {warehouse_info}")

                    if inventory_record:
                        # 检查库存是否足够
                        current_pallet = inventory_record.pallet_count or 0
                        current_package = inventory_record.package_count or 0

                        if (current_pallet < pallet_count or current_package < package_count):
                            errors.append(f'第{i+1}行：库存不足，当前库存 {current_pallet}板 {current_package}件，需要 {pallet_count}板 {package_count}件')
                            continue

                        # 扣减库存
                        inventory_record.pallet_count = current_pallet - pallet_count
                        inventory_record.package_count = current_package - package_count
                        inventory_record.last_updated = datetime.now()

                        current_app.logger.info(f"扣减库存：{identification_code} 从 {current_pallet}板{current_package}件 减少到 {inventory_record.pallet_count}板{inventory_record.package_count}件")
                    else:
                        current_app.logger.warning(f"未找到库存记录：{identification_code}")
                        # 不阻止创建出库记录，但记录警告
                else:
                    current_app.logger.warning(f"第{i+1}行：识别编码为空，无法扣减库存")

                # 处理其他字段
                try:
                    weight_raw = record_dict.get('weight', 0)
                    volume_raw = record_dict.get('volume', 0)

                    weight = float(weight_raw) if weight_raw and str(weight_raw).strip() else 0
                    volume = float(volume_raw) if volume_raw and str(volume_raw).strip() else 0

                    # 使用公共数据中的返回时间
                    if return_time_str and str(return_time_str).strip():
                        outbound_time = datetime.strptime(str(return_time_str).strip(), '%Y-%m-%d')
                    else:
                        outbound_time = datetime.now()

                except (ValueError, TypeError) as e:
                    errors.append(f'第{i+1}行：日期或数值字段格式错误 - {str(e)}')
                    continue

                # 查找该识别编码已有的最大批次序号，为返回记录分配新的序号
                max_sequence = db.session.query(db.func.max(OutboundRecord.batch_sequence)).filter_by(
                    identification_code=identification_code
                ).scalar() or 0

                current_app.logger.info(f"识别编码 {identification_code} 的最大批次序号: {max_sequence}")

                # 为返回记录使用下一个序号
                return_batch_sequence = max_sequence + 1
                current_app.logger.info(f"为返回记录分配的批次序号: {return_batch_sequence}")

                # 确定操作仓库ID：如果是admin用户且没有仓库ID，使用库存记录的仓库ID
                operated_warehouse_id = current_user.warehouse_id
                if not operated_warehouse_id and inventory_record:
                    operated_warehouse_id = inventory_record.operated_warehouse_id
                    current_app.logger.info(f"Admin用户使用库存记录的仓库ID: {operated_warehouse_id}")

                # 查找对应的收货人信息
                receiver = Receiver.query.filter_by(warehouse_name=target_warehouse).first()
                receiver_id = receiver.id if receiver else None
                current_app.logger.info(f"目的地 {target_warehouse} 对应的收货人ID: {receiver_id}")

                # 创建出库记录，保持原识别编码不变
                outbound_record = OutboundRecord(
                    outbound_time=outbound_time,
                    delivery_plate_number=delivery_plate_number,
                    plate_number=record_dict.get('plate_number', delivery_plate_number),
                    customer_name=customer_name,
                    identification_code=identification_code,  # 保持原识别编码不变
                    pallet_count=pallet_count,
                    package_count=package_count,
                    weight=weight,
                    volume=volume,
                    destination=target_warehouse,
                    order_type='',  # 添加订单类型字段
                    export_mode='',  # 添加出境模式字段
                    customs_broker='',  # 添加报关行字段
                    batch_no=batch_number,
                    batch_sequence=return_batch_sequence,  # 使用新的批次序号避免冲突
                    driver_name=driver_name,  # 添加司机姓名
                    driver_phone=driver_contact,  # 添加司机联系方式
                    receiver_id=receiver_id,  # 设置收货人ID
                    remarks=record_dict.get('remarks', ''),  # 使用用户输入的备注，不自动添加标记
                    operated_by_user_id=current_user.id,
                    operated_warehouse_id=operated_warehouse_id  # 使用确定的仓库ID
                )

                db.session.add(outbound_record)
                success_count += 1
                current_app.logger.info(f"第{i+1}行处理成功: 客户={customer_name}, 板数={pallet_count}, 件数={package_count}")

            except Exception as e:
                current_app.logger.error(f'第{i+1}行处理失败: {str(e)}', exc_info=True)
                errors.append(f'第{i+1}行：{str(e)}')
                continue

        if success_count > 0:
            try:
                db.session.commit()
                current_app.logger.info(f'后端仓返回前端仓成功：{success_count}条记录')
            except Exception as commit_error:
                db.session.rollback()
                current_app.logger.error(f'提交事务失败: {str(commit_error)}', exc_info=True)
                return jsonify({'success': False, 'message': f'保存失败：{str(commit_error)}'}), 500

        return jsonify({
            'success': True,
            'message': f'成功保存{success_count}条返回记录',
            'success_count': success_count,
            'total_count': len(records),
            'errors': errors
        })

    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f'后端仓返回前端仓失败: {str(e)}', exc_info=True)
        return jsonify({'success': False, 'message': str(e)}), 500

@bp.route('/api/backend/outbound/return', methods=['POST'])
@csrf.exempt
@require_permission('OUTBOUND_CREATE')
def api_backend_outbound_return_new():
    """后端仓出库到春疆货场API"""
    try:
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'message': '未提供出库数据'}), 400

        # 如果data是列表，取第一个元素；如果是单个对象，直接使用
        if isinstance(data, list):
            if not data:
                return jsonify({'success': False, 'message': '出库数据为空'}), 400
            item = data[0]
        else:
            item = data

        # 获取后端仓库
        backend_warehouse = Warehouse.query.filter_by(warehouse_type='backend').first()
        if not backend_warehouse:
            return jsonify({'success': False, 'message': '找不到后端仓库'}), 500

        # 生成批次号
        from app.utils.batch_generator import generate_batch_number
        batch_number = generate_batch_number(
            warehouse_id=backend_warehouse.id,
            destination_prefix='CJ',  # 春疆货场专用前缀
            db_session=db.session
        )
        current_app.logger.info(f'后端仓出库到春疆货场生成批次号: {batch_number}')

        # 创建出库记录
        outbound_record = OutboundRecord(
            outbound_time=datetime.now(),
            plate_number=item.get('plate_number', ''),
            customer_name=item.get('customer_name', ''),
            identification_code='',  # 后端出库到春疆不需要识别编码
            pallet_count=item.get('pallet_count', 0),
            package_count=item.get('package_count', 0),
            weight=item.get('weight', 0),
            volume=item.get('volume', 0),
            export_mode=item.get('export_mode', ''),
            order_type=item.get('order_type', ''),
            customs_broker=item.get('customs_broker', ''),
            service_staff=item.get('service_staff', ''),
            destination=item.get('receiver_name', '春疆货场'),
            warehouse_address=item.get('receiver_address', '谅山春疆货场'),
            batch_no=batch_number,
            remark1='',
            operated_by_user_id=current_user.id,
            operated_warehouse_id=backend_warehouse.id
        )

        db.session.add(outbound_record)
        db.session.commit()

        return jsonify({
            'success': True,
            'message': '出库成功',
            'data': {
                'outbound_id': outbound_record.id,
                'destination': outbound_record.destination
            }
        })

    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"后端出库到春疆失败: {str(e)}")
        return jsonify({'success': False, 'message': str(e)}), 500

@bp.route('/api/backend/outbound/return_old', methods=['POST'])
@csrf.exempt
@require_permission('OUTBOUND_CREATE')
def api_backend_outbound_return():
    """后端仓返回前端仓API"""
    try:
        data = request.get_json()
        if not data or 'records' not in data:
            return jsonify({'success': False, 'message': '未提供出库记录数据'}), 400

        items = data['records']
        if not items:
            return jsonify({'success': False, 'message': '出库记录列表为空'}), 400

        # 获取当前用户的仓库ID（后端仓）
        current_warehouse_id = current_user.warehouse_id if hasattr(current_user, 'warehouse_id') else None
        if not current_warehouse_id:
            return jsonify({'success': False, 'message': '无法确定当前仓库，请联系管理员'}), 400

        # 使用统一的批次号生成器，不传递db.session避免事务冲突
        from app.utils.batch_generator import generate_batch_number
        batch_number = generate_batch_number(
            warehouse_id=current_warehouse_id,
            destination_prefix='RT'  # 返回专用前缀，不传递db_session
        )
        current_app.logger.info(f'后端仓返回前端仓(旧API)生成批次号: {batch_number}')

        # 记录更新的库存
        updated_inventories = []

        # 先查询所有需要更新的库存记录
        identification_codes = [item.get('identification_code') for item in items if item.get('identification_code')]
        inventory_dict = {}
        if identification_codes:
            inventories = Inventory.query.filter(Inventory.identification_code.in_(identification_codes)).all()
            for inv in inventories:
                inventory_dict[inv.identification_code] = inv

        records_to_save = []

        for index, item in enumerate(items, 1):
            # 设置批次序号和总数
            item['batch_no'] = batch_number
            item['batch_sequence'] = index
            item['batch_total'] = len(items)

            # 解析出库时间
            outbound_time_str = item.get('outbound_time')
            try:
                outbound_time = datetime.fromisoformat(outbound_time_str.replace('T', ' '))
            except (ValueError, AttributeError):
                outbound_time = datetime.now()

            # 获取目标仓库ID
            target_warehouse_name = item.get('target_warehouse', '')
            destination_warehouse_id = None
            if target_warehouse_name:
                target_warehouse = Warehouse.query.filter_by(warehouse_name=target_warehouse_name).first()
                if target_warehouse:
                    destination_warehouse_id = target_warehouse.id

            # 获取入库日期
            inbound_date = None
            identification_code = item.get('identification_code')
            if identification_code:
                # 首先尝试从库存记录获取入库日期
                inventory = inventory_dict.get(identification_code)
                if inventory and inventory.inbound_time:
                    inbound_date = inventory.inbound_time
                else:
                    # 如果库存记录没有入库时间，尝试从入库记录获取
                    inbound_record = InboundRecord.query.filter_by(identification_code=identification_code).first()
                    if inbound_record and inbound_record.inbound_time:
                        inbound_date = inbound_record.inbound_time

            # 创建出库记录
            record = OutboundRecord(
                outbound_time=outbound_time,
                delivery_plate_number=item.get('delivery_plate_number', ''),
                plate_number=item.get('plate_number', ''),
                inbound_plate=item.get('inbound_plate', ''),  # 新增入库车牌字段
                customer_name=item.get('customer_name', ''),
                identification_code=item.get('identification_code', ''),
                order_type=item.get('order_type', ''),
                export_mode=item.get('export_mode', ''),  # 添加出境模式字段
                customs_broker=item.get('customs_broker', ''),  # 添加报关行字段
                pallet_count=item.get('pallet_count', 0),
                package_count=item.get('package_count', 0),
                weight=item.get('weight'),
                volume=item.get('volume'),
                service_staff=item.get('service_staff', ''),
                documents=item.get('documents', ''),
                remark1=item.get('remark1', ''),
                remark2=item.get('remark2', ''),
                destination=target_warehouse_name,  # 目标前端仓
                destination_warehouse_id=destination_warehouse_id,  # 目标仓库ID
                warehouse_address=item.get('warehouse_address', ''),  # 新增仓库地址字段
                contact_window=item.get('contact_window', ''),  # 新增联络窗口字段
                transport_company=item.get('transport_company', ''),
                inbound_date=inbound_date,  # 添加入库日期
                batch_no=batch_number,
                batch_sequence=index,
                batch_total=len(items),
                operated_by_user_id=current_user.id,
                operated_warehouse_id=_get_operation_warehouse_id()
            )
            records_to_save.append(record)

            # 更新库存 - 根据识别编码查找并减少库存
            identification_code = item.get('identification_code')
            if identification_code:
                inventory = inventory_dict.get(identification_code)
                if inventory:
                    # 获取出库数量，确保是整数
                    try:
                        pallet_value = item.get('pallet_count', 0) or 0
                        package_value = item.get('package_count', 0) or 0

                        # 检查是否为小数
                        if isinstance(pallet_value, (int, float)) and pallet_value != int(pallet_value):
                            current_app.logger.error(f"后端仓出库板数必须是整数: {pallet_value}")
                            continue
                        if isinstance(package_value, (int, float)) and package_value != int(package_value):
                            current_app.logger.error(f"后端仓出库件数必须是整数: {package_value}")
                            continue

                        outbound_pallet_count = int(pallet_value)
                        outbound_package_count = int(package_value)

                        if outbound_pallet_count < 0 or outbound_package_count < 0:
                            current_app.logger.error(f"后端仓出库板数和件数不能为负数: {outbound_pallet_count}, {outbound_package_count}")
                            continue

                    except (ValueError, TypeError):
                        current_app.logger.error(f"后端仓出库板数和件数必须是有效的整数: {item.get('pallet_count')}, {item.get('package_count')}")
                        continue

                    # 记录更新前的库存
                    before_pallet = inventory.pallet_count
                    before_package = inventory.package_count

                    # 更新库存（减少）
                    inventory.pallet_count = max(0, inventory.pallet_count - outbound_pallet_count)
                    inventory.package_count = max(0, inventory.package_count - outbound_package_count)

                    # 记录已更新的库存
                    updated_inventories.append({
                        'id': inventory.id,
                        'identification_code': inventory.identification_code,
                        'customer_name': inventory.customer_name,
                        'before_pallet': before_pallet,
                        'before_package': before_package,
                        'after_pallet': inventory.pallet_count,
                        'after_package': inventory.package_count,
                        'outbound_pallet': outbound_pallet_count,
                        'outbound_package': outbound_package_count
                    })

                    current_app.logger.info(f"更新后端仓库存: {identification_code}, 板数: {before_pallet} -> {inventory.pallet_count}, 件数: {before_package} -> {inventory.package_count}")
                else:
                    current_app.logger.warning(f"未找到后端仓库存记录: {identification_code}")

        # 批量保存记录
        db.session.add_all(records_to_save)

        # 确保立即提交事务，更新库存
        db.session.commit()

        # 刷新所有更新过的库存对象
        for inventory in inventory_dict.values():
            db.session.refresh(inventory)
            current_app.logger.info(f"刷新后的后端仓库存状态: {inventory.identification_code}, 板数: {inventory.pallet_count}, 件数: {inventory.package_count}")

        return jsonify({
            'success': True,
            'message': f'成功保存 {len(records_to_save)} 条返回前端仓记录',
            'saved_count': len(records_to_save),
            'batch_no': batch_number,
            'updated_inventories': updated_inventories
        })

    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"保存返回前端仓记录失败: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'保存失败: {str(e)}'
        }), 500


# ==================== 在途货物管理API ====================

@bp.route('/api/transit/cargo/list', methods=['GET'])
@csrf.exempt
@login_required
def api_transit_cargo_list():
    """获取在途货物列表"""
    try:
        # 获取搜索参数
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 50, type=int)
        customer_name = request.args.get('customer_name', '').strip()
        identification_code = request.args.get('identification_code', '').strip()
        batch_no = request.args.get('batch_no', '').strip()
        plate_number = request.args.get('plate_number', '').strip()
        service_staff = request.args.get('service_staff', '').strip()
        order_type = request.args.get('order_type', '').strip()
        export_mode = request.args.get('export_mode', '').strip()
        customs_broker = request.args.get('customs_broker', '').strip()
        status = request.args.get('status', '').strip()
        source_warehouse_id = request.args.get('source_warehouse_id', type=int)
        destination_warehouse_id = request.args.get('destination_warehouse_id', type=int)

        # 构建查询
        query = TransitCargo.query

        # 应用搜索条件
        if customer_name:
            query = query.filter(TransitCargo.customer_name.like(f'%{customer_name}%'))
        if identification_code:
            query = query.filter(TransitCargo.identification_code.like(f'%{identification_code}%'))
        if batch_no:
            query = query.filter(TransitCargo.batch_no.like(f'%{batch_no}%'))
        if plate_number:
            query = query.filter(db.or_(
                TransitCargo.plate_number.like(f'%{plate_number}%'),
                TransitCargo.delivery_plate_number.like(f'%{plate_number}%'),
                TransitCargo.inbound_plate.like(f'%{plate_number}%')
            ))
        if service_staff:
            query = query.filter(TransitCargo.service_staff.like(f'%{service_staff}%'))
        if order_type:
            query = query.filter(TransitCargo.order_type == order_type)
        if export_mode:
            query = query.filter(TransitCargo.export_mode == export_mode)
        if customs_broker:
            query = query.filter(TransitCargo.customs_broker.like(f'%{customs_broker}%'))
        if status:
            query = query.filter(TransitCargo.status == status)
        else:
            # 如果没有指定状态，默认只显示在途中的货物
            query = query.filter(TransitCargo.status == 'in_transit')
        if source_warehouse_id:
            query = query.filter(TransitCargo.source_warehouse_id == source_warehouse_id)
        if destination_warehouse_id:
            query = query.filter(TransitCargo.destination_warehouse_id == destination_warehouse_id)

        # 按创建时间降序排序
        query = query.order_by(TransitCargo.created_at.desc())

        # 分页
        transit_cargos = query.paginate(page=page, per_page=per_page, error_out=False)

        # 转换为字典列表
        cargo_list = [cargo.to_dict() for cargo in transit_cargos.items]

        return jsonify({
            'success': True,
            'cargos': cargo_list,
            'total': transit_cargos.total,
            'page': page,
            'per_page': per_page,
            'pages': transit_cargos.pages
        })

    except Exception as e:
        current_app.logger.error(f"获取在途货物列表失败: {str(e)}")
        return jsonify({'success': False, 'message': f'获取在途货物列表失败: {str(e)}'}), 500


@bp.route('/api/transit/cargo/list/by-batch', methods=['GET'])
@csrf.exempt
@login_required
def api_transit_cargo_list_by_batch():
    """按批次获取在途货物列表"""
    try:
        current_app.logger.info("开始处理批次API请求")

        # 获取搜索参数
        customer_name = request.args.get('customer_name', '').strip()
        identification_code = request.args.get('identification_code', '').strip()
        batch_no = request.args.get('batch_no', '').strip()
        plate_number = request.args.get('plate_number', '').strip()
        service_staff = request.args.get('service_staff', '').strip()
        order_type = request.args.get('order_type', '').strip()
        export_mode = request.args.get('export_mode', '').strip()
        customs_broker = request.args.get('customs_broker', '').strip()
        status = request.args.get('status', '').strip()
        source_warehouse_id = request.args.get('source_warehouse_id', type=int)
        destination_warehouse_id = request.args.get('destination_warehouse_id', type=int)

        # 获取分页参数
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 20, type=int)

        current_app.logger.info(f"搜索参数: customer_name={customer_name}, status={status}")

        # 构建查询
        query = TransitCargo.query
        current_app.logger.info("创建基础查询成功")

        # 应用搜索条件
        if customer_name:
            query = query.filter(TransitCargo.customer_name.like(f'%{customer_name}%'))
        if identification_code:
            query = query.filter(TransitCargo.identification_code.like(f'%{identification_code}%'))
        if batch_no:
            query = query.filter(TransitCargo.batch_no.like(f'%{batch_no}%'))
        if plate_number:
            query = query.filter(db.or_(
                TransitCargo.plate_number.like(f'%{plate_number}%'),
                TransitCargo.delivery_plate_number.like(f'%{plate_number}%'),
                TransitCargo.inbound_plate.like(f'%{plate_number}%')
            ))
        if service_staff:
            query = query.filter(TransitCargo.service_staff.like(f'%{service_staff}%'))
        if order_type:
            query = query.filter(TransitCargo.order_type == order_type)
        if export_mode:
            query = query.filter(TransitCargo.export_mode == export_mode)
        if customs_broker:
            query = query.filter(TransitCargo.customs_broker.like(f'%{customs_broker}%'))
        if status:
            query = query.filter(TransitCargo.status == status)
        else:
            # 如果没有指定状态，默认只显示在途中的货物
            query = query.filter(TransitCargo.status == 'in_transit')
        if source_warehouse_id:
            query = query.filter(TransitCargo.source_warehouse_id == source_warehouse_id)
        if destination_warehouse_id:
            query = query.filter(TransitCargo.destination_warehouse_id == destination_warehouse_id)

        current_app.logger.info("应用搜索条件成功")

        # 按批次号和创建时间排序
        query = query.order_by(TransitCargo.batch_no.desc(), TransitCargo.created_at.desc())
        current_app.logger.info("添加排序成功")

        # 获取所有符合条件的货物
        all_cargos = query.all()
        current_app.logger.info(f"查询到 {len(all_cargos)} 条记录")

        # 按批次分组
        batches = {}
        for i, cargo in enumerate(all_cargos):
            current_app.logger.debug(f"处理第 {i+1} 条记录，ID: {cargo.id}")
            batch_key = cargo.batch_no or '未知批次'  # 处理空批次号

            if batch_key not in batches:
                batches[batch_key] = {
                    'batch_no': batch_key,
                    'items': [],
                    'total_pallets': 0,
                    'total_packages': 0,
                    'total_weight': 0,
                    'total_volume': 0,
                    'item_count': 0,
                    'departure_time': None,
                    'status_summary': {},
                    'source_warehouse': None,
                    'destination_warehouse': None
                }

            batch_info = batches[batch_key]

            # 转换为字典
            try:
                cargo_dict = cargo.to_dict()
                current_app.logger.debug(f"货物 {cargo.id} 转换为字典成功")
            except Exception as e:
                current_app.logger.error(f"货物 {cargo.id} 转换为字典失败: {e}")
                continue

            batch_info['items'].append(cargo_dict)
            batch_info['item_count'] += 1
            batch_info['total_pallets'] += cargo.pallet_count or 0
            batch_info['total_packages'] += cargo.package_count or 0
            batch_info['total_weight'] += cargo.weight or 0
            batch_info['total_volume'] += cargo.volume or 0

            # 记录出发时间（取最早的）
            # 使用一个临时变量存储datetime对象进行比较，最后再转换为字符串
            if 'departure_time_obj' not in batch_info:
                batch_info['departure_time_obj'] = None

            if not batch_info['departure_time_obj'] or (cargo.departure_time and cargo.departure_time < batch_info['departure_time_obj']):
                batch_info['departure_time_obj'] = cargo.departure_time
                batch_info['departure_time'] = cargo.departure_time.strftime('%Y-%m-%d %H:%M:%S') if cargo.departure_time else None

            # 统计状态
            status = cargo.status
            if status not in batch_info['status_summary']:
                batch_info['status_summary'][status] = 0
            batch_info['status_summary'][status] += 1

            # 记录仓库信息（取第一个）
            if not batch_info['source_warehouse']:
                batch_info['source_warehouse'] = cargo_dict.get('source_warehouse_name')
            if not batch_info['destination_warehouse']:
                batch_info['destination_warehouse'] = cargo_dict.get('destination_warehouse_name')

        current_app.logger.info(f"分组完成，共 {len(batches)} 个批次")

        # 转换为列表并按批次号排序
        batch_list = list(batches.values())

        # 清理临时字段
        for batch in batch_list:
            if 'departure_time_obj' in batch:
                del batch['departure_time_obj']

        batch_list.sort(key=lambda x: x['batch_no'], reverse=True)

        # 计算统计信息（基于所有数据，不受分页影响）
        unique_identification_codes = set()
        total_pallets_all = 0
        total_packages_all = 0

        for batch in batch_list:
            if batch['items']:
                for item in batch['items']:
                    if item.get('identification_code'):
                        unique_identification_codes.add(item['identification_code'])
            total_pallets_all += batch.get('total_pallets', 0)
            total_packages_all += batch.get('total_packages', 0)

        # 应用分页
        total_batches = len(batch_list)
        start_index = (page - 1) * per_page
        end_index = start_index + per_page
        paginated_batches = batch_list[start_index:end_index]

        return jsonify({
            'success': True,
            'batches': paginated_batches,
            'total_batches': total_batches,
            'total_items': len(all_cargos),
            'total_tickets': len(unique_identification_codes),  # 按识别编码去重的总票数
            'total_pallets_all': total_pallets_all,  # 所有批次的总板数
            'total_packages_all': total_packages_all,  # 所有批次的总件数
            'current_page': page,
            'per_page': per_page,
            'total_pages': (total_batches + per_page - 1) // per_page
        })

    except Exception as e:
        import traceback
        current_app.logger.error(f"按批次获取在途货物列表失败: {str(e)}")
        current_app.logger.error(f"错误详情: {traceback.format_exc()}")
        return jsonify({'success': False, 'message': f'获取列表失败: {str(e)}'}), 500





@bp.route('/api/transit/cargo/statistics', methods=['GET'])
@csrf.exempt
@login_required
def api_transit_cargo_statistics():
    """获取在途货物统计信息"""
    try:
        # 获取搜索参数
        customer_name = request.args.get('customer_name', '').strip()
        identification_code = request.args.get('identification_code', '').strip()
        batch_no = request.args.get('batch_no', '').strip()
        plate_number = request.args.get('plate_number', '').strip()
        service_staff = request.args.get('service_staff', '').strip()
        order_type = request.args.get('order_type', '').strip()
        export_mode = request.args.get('export_mode', '').strip()
        customs_broker = request.args.get('customs_broker', '').strip()

        # 构建基础查询
        query = db.session.query(
            TransitCargo.status,
            func.count(TransitCargo.id).label('count'),
            func.sum(TransitCargo.pallet_count).label('total_pallets'),
            func.sum(TransitCargo.package_count).label('total_packages')
        )

        # 应用搜索条件
        if customer_name:
            query = query.filter(TransitCargo.customer_name.like(f'%{customer_name}%'))
        if identification_code:
            query = query.filter(TransitCargo.identification_code.like(f'%{identification_code}%'))
        if batch_no:
            query = query.filter(TransitCargo.batch_no.like(f'%{batch_no}%'))
        if plate_number:
            query = query.filter(db.or_(
                TransitCargo.plate_number.like(f'%{plate_number}%'),
                TransitCargo.delivery_plate_number.like(f'%{plate_number}%'),
                TransitCargo.inbound_plate.like(f'%{plate_number}%')
            ))
        if service_staff:
            query = query.filter(TransitCargo.service_staff.like(f'%{service_staff}%'))
        if order_type:
            query = query.filter(TransitCargo.order_type == order_type)
        if export_mode:
            query = query.filter(TransitCargo.export_mode == export_mode)
        if customs_broker:
            query = query.filter(TransitCargo.customs_broker.like(f'%{customs_broker}%'))

        # 按状态统计
        status_stats = query.group_by(TransitCargo.status).all()

        # 统计在途车辆数量（去重车牌号）
        vehicle_query = db.session.query(
            func.count(func.distinct(TransitCargo.plate_number)).label('vehicle_count')
        )

        # 应用相同的搜索条件
        if customer_name:
            vehicle_query = vehicle_query.filter(TransitCargo.customer_name.like(f'%{customer_name}%'))
        if identification_code:
            vehicle_query = vehicle_query.filter(TransitCargo.identification_code.like(f'%{identification_code}%'))
        if batch_no:
            vehicle_query = vehicle_query.filter(TransitCargo.batch_no.like(f'%{batch_no}%'))
        if plate_number:
            vehicle_query = vehicle_query.filter(db.or_(
                TransitCargo.plate_number.like(f'%{plate_number}%'),
                TransitCargo.delivery_plate_number.like(f'%{plate_number}%'),
                TransitCargo.inbound_plate.like(f'%{plate_number}%')
            ))
        if service_staff:
            vehicle_query = vehicle_query.filter(TransitCargo.service_staff.like(f'%{service_staff}%'))
        if order_type:
            vehicle_query = vehicle_query.filter(TransitCargo.order_type == order_type)
        if export_mode:
            vehicle_query = vehicle_query.filter(TransitCargo.export_mode == export_mode)
        if customs_broker:
            vehicle_query = vehicle_query.filter(TransitCargo.customs_broker.like(f'%{customs_broker}%'))

        # 只统计在途状态的车辆
        vehicle_stats = vehicle_query.filter(TransitCargo.status.in_(['in_transit', 'arrived'])).first()

        # 按路线统计
        route_stats = db.session.query(
            TransitCargo.source_warehouse_id,
            TransitCargo.destination_warehouse_id,
            func.count(TransitCargo.id).label('count')
        ).filter(TransitCargo.status.in_(['in_transit', 'arrived'])).group_by(
            TransitCargo.source_warehouse_id,
            TransitCargo.destination_warehouse_id
        ).all()

        # 格式化统计结果
        status_summary = {}
        for stat in status_stats:
            status_summary[stat.status] = {
                'count': stat.count,
                'total_pallets': stat.total_pallets or 0,
                'total_packages': stat.total_packages or 0,
                'display_name': {
                    'in_transit': '运输中',
                    'arrived': '已到达',
                    'received': '已接收',
                    'cancelled': '已取消'
                }.get(stat.status, stat.status)
            }

        route_summary = []
        for stat in route_stats:
            source_warehouse = Warehouse.query.get(stat.source_warehouse_id)
            dest_warehouse = Warehouse.query.get(stat.destination_warehouse_id)
            route_summary.append({
                'source_warehouse': source_warehouse.warehouse_name if source_warehouse else '未知',
                'destination_warehouse': dest_warehouse.warehouse_name if dest_warehouse else '未知',
                'count': stat.count
            })

        return jsonify({
            'success': True,
            'status_summary': status_summary,
            'route_summary': route_summary,
            'vehicle_count': vehicle_stats.vehicle_count if vehicle_stats else 0
        })

    except Exception as e:
        current_app.logger.error(f"获取在途货物统计失败: {str(e)}")
        return jsonify({'success': False, 'message': f'获取统计信息失败: {str(e)}'}), 500


# ==================== 在途货物管理页面路由 ====================

@bp.route('/transit/cargo/list')
@login_required
def transit_cargo_list():
    """在途货物列表页面"""
    return render_template('transit/cargo_list.html',
                         title='在途货物管理',
                         page_type='transit_cargo')





# ==================== 测试数据API ====================

@bp.route('/api/test/create-backend-outbound-data', methods=['POST'])
@csrf.exempt
def api_create_backend_outbound_data():
    """创建后端仓出库测试数据"""
    try:
        # 获取后端仓库
        backend_warehouse = Warehouse.query.filter_by(warehouse_type='backend').first()
        if not backend_warehouse:
            return jsonify({'success': False, 'message': '未找到后端仓库'}), 400

        # 获取前端仓库
        pinghu_warehouse = Warehouse.query.filter_by(warehouse_name='平湖仓').first()
        kunshan_warehouse = Warehouse.query.filter_by(warehouse_name='昆山仓').first()

        if not pinghu_warehouse or not kunshan_warehouse:
            return jsonify({'success': False, 'message': '未找到前端仓库'}), 400

        # 创建测试出库记录
        sample_outbound_records = [
            {
                'outbound_time': datetime.now() - timedelta(hours=2),
                'batch_no': 'BT20250703001',
                'batch_sequence': 1,
                'batch_total': 3,
                'plate_number': '桂A12345',
                'customer_name': '深圳市ABC贸易有限公司',
                'order_type': '一般贸易',
                'customs_broker': '深圳报关行',
                'pallet_count': 5.5,
                'package_count': 80,
                'weight': 1200.5,
                'volume': 6.8,
                'documents': '2',
                'service_staff': '张三',
                'destination_warehouse_id': pinghu_warehouse.id,
                'remark1': '发往平湖仓',
                'remark2': '第一批货物'
            },
            {
                'outbound_time': datetime.now() - timedelta(hours=2),
                'batch_no': 'BT20250703001',
                'batch_sequence': 2,
                'batch_total': 3,
                'plate_number': '桂A12345',
                'customer_name': '广州XYZ物流公司',
                'order_type': '跨境电商',
                'customs_broker': '广州通关服务',
                'pallet_count': 3.0,
                'package_count': 45,
                'weight': 680.2,
                'volume': 4.2,
                'documents': '1',
                'service_staff': '李四',
                'destination_warehouse_id': pinghu_warehouse.id,
                'remark1': '发往平湖仓',
                'remark2': '第二批货物'
            },
            {
                'outbound_time': datetime.now() - timedelta(hours=2),
                'batch_no': 'BT20250703001',
                'batch_sequence': 3,
                'batch_total': 3,
                'plate_number': '桂A12345',
                'customer_name': '东莞制造企业',
                'order_type': '加工贸易',
                'customs_broker': '东莞报关公司',
                'pallet_count': 4.5,
                'package_count': 60,
                'weight': 950.8,
                'volume': 5.5,
                'documents': '3',
                'service_staff': '王五',
                'destination_warehouse_id': pinghu_warehouse.id,
                'remark1': '发往平湖仓',
                'remark2': '第三批货物'
            },
            {
                'outbound_time': datetime.now() - timedelta(hours=1),
                'batch_no': 'BT20250703002',
                'batch_sequence': 1,
                'batch_total': 2,
                'plate_number': '桂B67890',
                'customer_name': '佛山进出口公司',
                'order_type': '一般贸易',
                'customs_broker': '佛山报关行',
                'pallet_count': 8.0,
                'package_count': 120,
                'weight': 1800.0,
                'volume': 9.5,
                'documents': '4',
                'service_staff': '赵六',
                'destination_warehouse_id': kunshan_warehouse.id,
                'remark1': '发往昆山仓',
                'remark2': '急件处理'
            },
            {
                'outbound_time': datetime.now() - timedelta(hours=1),
                'batch_no': 'BT20250703002',
                'batch_sequence': 2,
                'batch_total': 2,
                'plate_number': '桂B67890',
                'customer_name': '中山电子科技',
                'order_type': '跨境电商',
                'customs_broker': '中山通关公司',
                'pallet_count': 2.5,
                'package_count': 35,
                'weight': 420.5,
                'volume': 3.2,
                'documents': '1',
                'service_staff': '孙七',
                'destination_warehouse_id': kunshan_warehouse.id,
                'remark1': '发往昆山仓',
                'remark2': '电子产品'
            }
        ]

        created_count = 0
        for record_data in sample_outbound_records:
            record = OutboundRecord(
                operated_warehouse_id=backend_warehouse.id,
                **record_data
            )
            db.session.add(record)
            created_count += 1

        db.session.commit()

        return jsonify({
            'success': True,
            'message': f'成功创建 {created_count} 条后端仓出库测试数据',
            'warehouse': backend_warehouse.warehouse_name
        })

    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"创建后端仓出库测试数据失败: {str(e)}")
        return jsonify({'success': False, 'message': f'创建测试数据失败: {str(e)}'}), 500


@bp.route('/test/dropdown')
def test_dropdown():
    """测试下拉框功能的页面"""
    return render_template('test_dropdown.html', title='下拉框测试')

@bp.route('/api/test/update-database-schema', methods=['POST'])
@csrf.exempt
def api_update_database_schema():
    """更新数据库结构，添加目的仓库字段"""
    try:
        # 检查是否已经存在destination_warehouse_id字段
        from sqlalchemy import text

        # 检查字段是否存在
        result = db.session.execute(text("PRAGMA table_info(outbound_record)"))
        columns = [row[1] for row in result.fetchall()]

        if 'destination_warehouse_id' not in columns:
            # 添加新字段
            db.session.execute(text("ALTER TABLE outbound_record ADD COLUMN destination_warehouse_id INTEGER"))
            db.session.commit()
            return jsonify({'success': True, 'message': '数据库结构更新成功，已添加目的仓库字段'})
        else:
            return jsonify({'success': True, 'message': '目的仓库字段已存在，无需更新'})

    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"更新数据库结构失败: {str(e)}")
        return jsonify({'success': False, 'message': f'更新数据库结构失败: {str(e)}'}), 500


@bp.route('/test-api')
def test_api_page():
    """API测试页面"""
    return render_template('test_api.html')

@bp.route('/api/test/fix-receive-records', methods=['POST'])
@csrf.exempt
def api_fix_receive_records():
    """修复接收记录车牌字段"""
    try:
        receive_records = ReceiveRecord.query.all()
        updated_count = 0

        for record in receive_records:
            if record.identification_code:
                outbound_record = OutboundRecord.query.filter_by(
                    identification_code=record.identification_code
                ).first()

                if outbound_record:
                    record.delivery_plate_number = outbound_record.plate_number
                    record.inbound_plate = outbound_record.inbound_plate
                    updated_count += 1

        db.session.commit()

        return jsonify({
            'success': True,
            'message': f'成功更新 {updated_count} 条接收记录'
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'message': f'修复失败: {str(e)}'
        }), 500

@bp.route('/api/test/clear-outbound-data', methods=['POST'])
@csrf.exempt
def api_clear_outbound_data():
    """清除出库测试数据"""
    try:
        # 删除所有出库记录
        deleted_count = OutboundRecord.query.delete()
        db.session.commit()

        return jsonify({
            'success': True,
            'message': f'成功清除 {deleted_count} 条出库记录'
        })

    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"清除出库数据失败: {str(e)}")
        return jsonify({'success': False, 'message': f'清除数据失败: {str(e)}'}), 500


# CSRF测试路由
@bp.route('/csrf-test', methods=['GET', 'POST'])
def csrf_test():
    """CSRF token测试页面"""
    if request.method == 'POST':
        test_input = request.form.get('test_input')
        flash(f'CSRF验证成功！接收到数据: {test_input}', 'success')
        return render_template('csrf_test.html')

@bp.route('/api/generate-identification-code', methods=['POST'])
@csrf.exempt
def generate_identification_code():
    """生成识别编码API"""
    try:
        data = request.get_json()

        # 获取参数
        customer_name = data.get('customer_name', '').strip()
        plate_number = data.get('plate_number', '').strip()
        operation_type = data.get('operation_type', 'inbound')

        # 验证必要参数
        if not customer_name:
            return jsonify({'success': False, 'message': '客户名称不能为空'}), 400

        if not plate_number:
            return jsonify({'success': False, 'message': '车牌号不能为空'}), 400

        # 获取当前用户的仓库ID
        warehouse_id = current_user.warehouse_id if hasattr(current_user, 'warehouse_id') else None
        if not warehouse_id:
            return jsonify({'success': False, 'message': '无法确定当前仓库，请联系管理员'}), 400

        # 生成识别编码
        identification_code = IdentificationCodeGenerator.generate_unique_identification_code(
            warehouse_id=warehouse_id,
            customer_name=customer_name,
            plate_number=plate_number,
            operation_type=operation_type
        )

        return jsonify({
            'success': True,
            'identification_code': identification_code,
            'message': '识别编码生成成功'
        })

    except Exception as e:
        return jsonify({'success': False, 'message': f'生成失败: {str(e)}'}), 500

    return render_template('csrf_test.html')

@bp.route('/api/check/identification-codes', methods=['GET'])
@csrf.exempt
@require_permission('ADMIN_MANAGE')
def api_check_identification_codes():
    """检查错误的识别编码（包含b2等后缀）"""
    try:
        error_suffixes = ['b2', '_b2', '.b2']
        inbound_errors = 0
        outbound_errors = 0
        inventory_errors = 0
        receive_errors = 0

        # 检查入库记录
        for suffix in error_suffixes:
            count = InboundRecord.query.filter(
                InboundRecord.identification_code.like(f'%{suffix}')
            ).count()
            inbound_errors += count

        # 检查出库记录
        for suffix in error_suffixes:
            count = OutboundRecord.query.filter(
                OutboundRecord.identification_code.like(f'%{suffix}')
            ).count()
            outbound_errors += count

        # 检查库存记录
        for suffix in error_suffixes:
            count = Inventory.query.filter(
                Inventory.identification_code.like(f'%{suffix}')
            ).count()
            inventory_errors += count

        # 检查接收记录
        for suffix in error_suffixes:
            count = ReceiveRecord.query.filter(
                ReceiveRecord.identification_code.like(f'%{suffix}')
            ).count()
            receive_errors += count

        total_errors = inbound_errors + outbound_errors + inventory_errors + receive_errors

        return jsonify({
            'success': True,
            'inbound_errors': inbound_errors,
            'outbound_errors': outbound_errors,
            'inventory_errors': inventory_errors,
            'receive_errors': receive_errors,
            'total_errors': total_errors
        })

    except Exception as e:
        current_app.logger.error(f"检查识别编码时出错: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'检查失败: {str(e)}'
        }), 500


@bp.route('/api/fix/identification-codes', methods=['POST'])
@csrf.exempt
@require_permission('ADMIN_MANAGE')
def api_fix_identification_codes():
    """修复错误的识别编码（移除b2等后缀）"""
    try:
        # 查找所有包含错误后缀的识别编码
        error_suffixes = ['b2', '_b2', '.b2']
        fixed_count = 0
        errors = []

        # 检查入库记录
        for suffix in error_suffixes:
            inbound_records = InboundRecord.query.filter(
                InboundRecord.identification_code.like(f'%{suffix}')
            ).all()

            for record in inbound_records:
                if record.identification_code and record.identification_code.endswith(suffix):
                    old_code = record.identification_code
                    new_code = old_code[:-len(suffix)]

                    # 检查新的识别编码是否已存在
                    existing = InboundRecord.query.filter_by(identification_code=new_code).first()
                    if not existing or existing.id == record.id:
                        record.identification_code = new_code
                        current_app.logger.info(f"修复入库记录识别编码: {old_code} -> {new_code}")
                        fixed_count += 1
                    else:
                        errors.append(f"入库记录ID {record.id}: 无法修复 {old_code}，因为 {new_code} 已存在")

        # 检查出库记录
        for suffix in error_suffixes:
            outbound_records = OutboundRecord.query.filter(
                OutboundRecord.identification_code.like(f'%{suffix}')
            ).all()

            for record in outbound_records:
                if record.identification_code and record.identification_code.endswith(suffix):
                    old_code = record.identification_code
                    new_code = old_code[:-len(suffix)]
                    record.identification_code = new_code
                    current_app.logger.info(f"修复出库记录识别编码: {old_code} -> {new_code}")
                    fixed_count += 1

        # 检查库存记录
        for suffix in error_suffixes:
            inventory_records = Inventory.query.filter(
                Inventory.identification_code.like(f'%{suffix}')
            ).all()

            for record in inventory_records:
                if record.identification_code and record.identification_code.endswith(suffix):
                    old_code = record.identification_code
                    new_code = old_code[:-len(suffix)]

                    # 检查新的识别编码是否已存在
                    existing = Inventory.query.filter_by(
                        identification_code=new_code,
                        operated_warehouse_id=record.operated_warehouse_id
                    ).first()

                    if not existing or existing.id == record.id:
                        record.identification_code = new_code
                        current_app.logger.info(f"修复库存记录识别编码: {old_code} -> {new_code}")
                        fixed_count += 1
                    else:
                        # 如果已存在，尝试合并库存
                        existing.pallet_count = (existing.pallet_count or 0) + (record.pallet_count or 0)
                        existing.package_count = (existing.package_count or 0) + (record.package_count or 0)
                        existing.weight = (existing.weight or 0) + (record.weight or 0)
                        existing.volume = (existing.volume or 0) + (record.volume or 0)
                        db.session.delete(record)
                        current_app.logger.info(f"合并库存记录: {old_code} -> {new_code}")
                        fixed_count += 1

        # 检查接收记录
        for suffix in error_suffixes:
            receive_records = ReceiveRecord.query.filter(
                ReceiveRecord.identification_code.like(f'%{suffix}')
            ).all()

            for record in receive_records:
                if record.identification_code and record.identification_code.endswith(suffix):
                    old_code = record.identification_code
                    new_code = old_code[:-len(suffix)]
                    record.identification_code = new_code
                    current_app.logger.info(f"修复接收记录识别编码: {old_code} -> {new_code}")
                    fixed_count += 1

        if fixed_count > 0:
            db.session.commit()

        return jsonify({
            'success': True,
            'message': f'成功修复 {fixed_count} 条记录的识别编码',
            'fixed_count': fixed_count,
            'errors': errors
        })

    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"修复识别编码时出错: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'修复失败: {str(e)}'
        }), 500


@bp.route('/api/test/check-backend-data', methods=['GET'])
@csrf.exempt
def api_check_backend_data():
    """检查后端仓数据状态"""
    try:
        # 检查后端仓库
        backend_warehouse = Warehouse.query.filter_by(warehouse_type='backend').first()
        backend_warehouse_info = {
            'exists': backend_warehouse is not None,
            'id': backend_warehouse.id if backend_warehouse else None,
            'name': backend_warehouse.warehouse_name if backend_warehouse else None
        }

        # 检查接收记录
        receive_records_count = ReceiveRecord.query.count()
        recent_receive_records = ReceiveRecord.query.order_by(ReceiveRecord.receive_time.desc()).limit(5).all()

        # 检查库存记录
        if backend_warehouse:
            inventory_count = Inventory.query.filter_by(operated_warehouse_id=backend_warehouse.id).count()
            inventory_with_stock = Inventory.query.filter(
                Inventory.operated_warehouse_id == backend_warehouse.id,
                (Inventory.pallet_count > 0) | (Inventory.package_count > 0)
            ).count()
            recent_inventory = Inventory.query.filter_by(operated_warehouse_id=backend_warehouse.id).order_by(Inventory.last_updated.desc()).limit(5).all()
        else:
            inventory_count = 0
            inventory_with_stock = 0
            recent_inventory = []

        return jsonify({
            'success': True,
            'data': {
                'backend_warehouse': backend_warehouse_info,
                'receive_records': {
                    'total_count': receive_records_count,
                    'recent': [
                        {
                            'id': r.id,
                            'customer_name': r.customer_name,
                            'identification_code': r.identification_code,
                            'pallet_count': r.pallet_count,
                            'package_count': r.package_count,
                            'receive_time': r.receive_time.strftime('%Y-%m-%d %H:%M:%S') if r.receive_time else None,
                            'operated_warehouse_id': r.operated_warehouse_id
                        } for r in recent_receive_records
                    ]
                },
                'inventory': {
                    'total_count': inventory_count,
                    'with_stock_count': inventory_with_stock,
                    'recent': [
                        {
                            'id': i.id,
                            'customer_name': i.customer_name,
                            'identification_code': i.identification_code,
                            'pallet_count': i.pallet_count,
                            'package_count': i.package_count,
                            'last_updated': i.last_updated.strftime('%Y-%m-%d %H:%M:%S') if i.last_updated else None,
                            'operated_warehouse_id': i.operated_warehouse_id
                        } for i in recent_inventory
                    ]
                }
            }
        })

    except Exception as e:
        current_app.logger.error(f"检查后端仓数据失败: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'检查失败: {str(e)}'
        }), 500


@bp.route('/api/parse_import_file', methods=['POST'])
@csrf.exempt
@require_permission('INBOUND_CREATE')
def api_parse_import_file():
    """解析导入文件API"""
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'message': '没有上传文件'}), 400

        file = request.files['file']
        if file.filename == '':
            return jsonify({'success': False, 'message': '没有选择文件'}), 400

        # 检查文件类型
        allowed_extensions = {'.xlsx', '.xls', '.csv'}
        file_ext = os.path.splitext(file.filename)[1].lower()
        if file_ext not in allowed_extensions:
            return jsonify({'success': False, 'message': '不支持的文件格式，请使用Excel或CSV文件'}), 400

        # 保存临时文件
        import tempfile
        with tempfile.NamedTemporaryFile(delete=False, suffix=file_ext) as tmp_file:
            file.save(tmp_file.name)
            temp_path = tmp_file.name

        try:
            # 解析文件
            if file_ext in ['.xlsx', '.xls']:
                import pandas as pd
                df = pd.read_excel(temp_path)
            else:  # CSV
                import pandas as pd
                df = pd.read_csv(temp_path, encoding='utf-8')

            # 转换为字典列表
            data = df.fillna('').to_dict('records')

            # 验证数据
            errors = []
            required_fields = ['货物名称', '发货人', '收货人', '板数', '件数']

            for i, row in enumerate(data):
                row_errors = []
                for field in required_fields:
                    if field not in row or not str(row[field]).strip():
                        row_errors.append(f"第{i+1}行缺少必填字段: {field}")

                # 验证数字字段
                for field in ['板数', '件数']:
                    if field in row and row[field]:
                        try:
                            value = float(row[field])
                            if value < 0:
                                row_errors.append(f"第{i+1}行{field}不能为负数")
                            if field in ['板数', '件数'] and value != int(value):
                                row_errors.append(f"第{i+1}行{field}必须为整数")
                        except (ValueError, TypeError):
                            row_errors.append(f"第{i+1}行{field}格式错误")

                errors.extend(row_errors)

            return jsonify({
                'success': True,
                'data': data,
                'errors': errors,
                'total_count': len(data),
                'valid_count': len(data) - len([e for e in errors if '行' in e])
            })

        finally:
            # 清理临时文件
            try:
                os.unlink(temp_path)
            except:
                pass

    except Exception as e:
        current_app.logger.error(f"解析导入文件失败: {str(e)}")
        return jsonify({'success': False, 'message': f'解析文件失败: {str(e)}'}), 500


@bp.route('/api/import_inbound_data', methods=['POST'])
@csrf.exempt
@require_permission('INBOUND_CREATE')
@log_operation('inbound', 'batch_import', 'inbound_record')
def api_import_inbound_data():
    """导入入库数据API"""
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'message': '没有上传文件'}), 400

        file = request.files['file']
        if file.filename == '':
            return jsonify({'success': False, 'message': '没有选择文件'}), 400

        # 检查用户权限
        if not current_user.warehouse:
            return jsonify({'success': False, 'message': '用户未分配仓库'}), 403

        # 检查文件类型
        allowed_extensions = {'.xlsx', '.xls', '.csv'}
        file_ext = os.path.splitext(file.filename)[1].lower()
        if file_ext not in allowed_extensions:
            return jsonify({'success': False, 'message': '不支持的文件格式'}), 400

        # 保存临时文件
        import tempfile
        with tempfile.NamedTemporaryFile(delete=False, suffix=file_ext) as tmp_file:
            file.save(tmp_file.name)
            temp_path = tmp_file.name

        try:
            # 解析文件
            if file_ext in ['.xlsx', '.xls']:
                import pandas as pd
                df = pd.read_excel(temp_path)
            else:  # CSV
                import pandas as pd
                df = pd.read_csv(temp_path, encoding='utf-8')

            # 转换为字典列表
            data = df.fillna('').to_dict('records')

            # 批量创建入库记录
            imported_count = 0
            from app.utils.batch_generator import BatchNumberGenerator
            batch_generator = BatchNumberGenerator()

            for row in data:
                try:
                    # 验证必填字段
                    required_fields = ['货物名称', '发货人', '收货人', '板数', '件数']
                    if not all(str(row.get(field, '')).strip() for field in required_fields):
                        continue

                    # 生成批次号
                    batch_number = batch_generator.generate_batch_number(
                        current_user.warehouse.warehouse_code
                    )

                    # 创建入库记录
                    record = InboundRecord(
                        batch_number=batch_number,
                        cargo_name=str(row['货物名称']).strip(),
                        shipper=str(row['发货人']).strip(),
                        receiver=str(row['收货人']).strip(),
                        pallet_count=int(float(row['板数'])),
                        package_count=int(float(row['件数'])),
                        warehouse_id=current_user.warehouse_id,
                        created_by=current_user.id,
                        inbound_time=datetime.now(),
                        # 可选字段
                        plate_number=str(row.get('车牌号', '')).strip() or None,
                        documents=str(row.get('单据号', '')).strip() or None,
                        remarks=str(row.get('备注', '')).strip() or None
                    )

                    db.session.add(record)
                    imported_count += 1

                except Exception as e:
                    current_app.logger.warning(f"导入第{imported_count+1}行数据失败: {str(e)}")
                    continue

            if imported_count > 0:
                db.session.commit()
                current_app.logger.info(f"批量导入成功，共导入 {imported_count} 条记录")
                return jsonify({
                    'success': True,
                    'message': f'导入成功，共导入 {imported_count} 条记录',
                    'imported_count': imported_count
                })
            else:
                db.session.rollback()
                return jsonify({'success': False, 'message': '没有有效数据可导入'}), 400

        finally:
            # 清理临时文件
            try:
                os.unlink(temp_path)
            except:
                pass

    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"导入入库数据失败: {str(e)}")
        return jsonify({'success': False, 'message': f'导入失败: {str(e)}'}), 500


@bp.route('/api/inventory/statistics', methods=['GET'])
@csrf.exempt
@require_permission('INVENTORY_VIEW')
def api_inventory_statistics():
    """获取正确的库存统计信息 - 区分原始库存和转移库存"""
    try:
        # 获取所有仓库
        warehouses = Warehouse.query.all()
        warehouse_stats = {}

        for warehouse in warehouses:
            # 查询该仓库的库存
            inventories = Inventory.query.filter_by(
                operated_warehouse_id=warehouse.id
            ).filter(
                (Inventory.pallet_count > 0) | (Inventory.package_count > 0)
            ).all()

            # 区分原始库存和转移库存
            original_inventory = []
            transfer_inventory = []

            for inv in inventories:
                # 判断是否为转移库存
                if warehouse.warehouse_type == 'backend':
                    # 后端仓库的库存都是转移库存
                    transfer_inventory.append(inv)
                else:
                    # 前端仓库的库存是原始库存
                    original_inventory.append(inv)

            # 计算统计数据
            original_pallet = sum(inv.pallet_count or 0 for inv in original_inventory)
            original_package = sum(inv.package_count or 0 for inv in original_inventory)
            transfer_pallet = sum(inv.pallet_count or 0 for inv in transfer_inventory)
            transfer_package = sum(inv.package_count or 0 for inv in transfer_inventory)

            warehouse_stats[warehouse.id] = {
                'warehouse_name': warehouse.warehouse_name,
                'warehouse_type': warehouse.warehouse_type,
                'original_inventory': {
                    'pallet_count': original_pallet,
                    'package_count': original_package,
                    'record_count': len(original_inventory)
                },
                'transfer_inventory': {
                    'pallet_count': transfer_pallet,
                    'package_count': transfer_package,
                    'record_count': len(transfer_inventory)
                },
                'total_records': len(inventories)
            }

        # 计算系统总库存（前端库存 + 后端库存）
        total_frontend_pallet = sum(
            stats['original_inventory']['pallet_count']
            for stats in warehouse_stats.values()
        )
        total_frontend_package = sum(
            stats['original_inventory']['package_count']
            for stats in warehouse_stats.values()
        )

        # 后端库存（转移库存，但仍属于系统总库存）
        total_backend_pallet = sum(
            stats['transfer_inventory']['pallet_count']
            for stats in warehouse_stats.values()
        )
        total_backend_package = sum(
            stats['transfer_inventory']['package_count']
            for stats in warehouse_stats.values()
        )

        # 系统总库存 = 前端库存 + 后端库存
        total_system_pallet = total_frontend_pallet + total_backend_pallet
        total_system_package = total_frontend_package + total_backend_package

        return jsonify({
            'success': True,
            'warehouse_statistics': warehouse_stats,
            'system_totals': {
                'frontend_inventory': {
                    'pallet_count': total_frontend_pallet,
                    'package_count': total_frontend_package,
                    'description': '前端仓库存（原始入库仓库的剩余库存）'
                },
                'backend_inventory': {
                    'pallet_count': total_backend_pallet,
                    'package_count': total_backend_package,
                    'description': '后端仓库存（从前端仓转移的库存）'
                },
                'total_system_inventory': {
                    'pallet_count': total_system_pallet,
                    'package_count': total_system_package,
                    'description': '系统总库存（前端库存 + 后端库存）'
                }
            },
            'explanation': {
                'frontend_inventory': '前端仓库存：各前端仓库自己入库的货物剩余量',
                'backend_inventory': '后端仓库存：从前端仓转移过来的货物，仍属于系统库存',
                'total_system_inventory': '系统总库存：所有仍在物流系统中的货物总量',
                'business_logic': '只有后端仓真正出库给客户时，才从系统总库存中扣除',
                'correct_total': f'系统总库存：板数 {total_system_pallet}，件数 {total_system_package}'
            }
        })

    except Exception as e:
        current_app.logger.error(f"获取库存统计失败: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'获取库存统计失败: {str(e)}'
        }), 500


@bp.route('/api/backend/inventory/search', methods=['GET'])
@csrf.exempt
@login_required
def api_backend_inventory_search():
    """后端仓库存搜索API"""
    try:
        # 获取搜索参数
        customer_name = request.args.get('customer_name', '').strip()
        identification_code = request.args.get('identification_code', '').strip()
        plate_number = request.args.get('plate_number', '').strip()

        # 构建查询 - 查询后端仓库的库存
        # 获取后端仓库ID
        backend_warehouses = Warehouse.query.filter_by(warehouse_type='backend').all()
        backend_warehouse_ids = [w.id for w in backend_warehouses]

        if not backend_warehouse_ids:
            return jsonify({
                'success': False,
                'message': '未找到后端仓库配置'
            }), 400

        query = Inventory.query.filter(
            Inventory.operated_warehouse_id.in_(backend_warehouse_ids)
        )

        # 添加搜索条件
        if customer_name:
            query = query.filter(Inventory.customer_name.like(f'%{customer_name}%'))
        if identification_code:
            query = query.filter(Inventory.identification_code.like(f'%{identification_code}%'))
        if plate_number:
            query = query.filter(Inventory.plate_number.like(f'%{plate_number}%'))

        # 只查询有库存的记录
        query = query.filter(
            db.or_(
                Inventory.pallet_count > 0,
                Inventory.package_count > 0
            )
        )

        # 执行查询
        inventory_records = query.order_by(Inventory.inbound_time.desc()).limit(100).all()

        # 转换为字典格式
        inventory_data = []
        for record in inventory_records:
            # 通过识别编码查找对应的前端仓出库记录，获取送货干线车信息
            delivery_plate_number = ''
            if record.identification_code:
                # 查找前端仓的出库记录（发运到后端仓的记录）
                frontend_warehouses = Warehouse.query.filter_by(warehouse_type='frontend').all()
                frontend_warehouse_ids = [w.id for w in frontend_warehouses]

                if frontend_warehouse_ids:
                    outbound_record = OutboundRecord.query.filter(
                        OutboundRecord.identification_code == record.identification_code,
                        OutboundRecord.operated_warehouse_id.in_(frontend_warehouse_ids)
                    ).first()

                    if outbound_record and outbound_record.plate_number:
                        delivery_plate_number = outbound_record.plate_number

            # 送货干线车是前端仓发运到后端仓的车牌号

            inventory_data.append({
                'id': record.id,
                'customer_name': record.customer_name or '',
                'identification_code': record.identification_code or '',
                'pallet_count': record.pallet_count or 0,
                'package_count': record.package_count or 0,
                'weight': float(record.weight or 0),
                'volume': float(record.volume or 0),
                'plate_number': record.plate_number or '',
                'inbound_date': record.inbound_time.strftime('%Y-%m-%d') if record.inbound_time else '',
                'service_staff': record.service_staff or '',
                'location': record.location or '',
                'export_mode': record.export_mode or '',
                'order_type': record.order_type or '',
                'customs_broker': record.customs_broker or '',
                'delivery_plate_number': delivery_plate_number,  # 只显示实际的送货干线车信息，直接入库的为空
                'delivery_truck': delivery_plate_number  # 备用字段，使用相同值
            })

        return jsonify({
            'success': True,
            'data': inventory_data,
            'total': len(inventory_data)
        })

    except Exception as e:
        current_app.logger.error(f"搜索后端仓库存失败: {str(e)}")
        return jsonify({'success': False, 'message': f'搜索失败: {str(e)}'}), 500


@bp.route('/api/backend/outbound/delete/<int:record_id>', methods=['DELETE'])
@login_required
@csrf.exempt
@require_permission('OUTBOUND_DELETE')
def api_backend_outbound_delete(record_id):
    """删除后端仓出库记录并回退库存"""
    try:
        # 查找出库记录
        outbound_record = OutboundRecord.query.get(record_id)
        if not outbound_record:
            return jsonify({'success': False, 'message': '出库记录不存在'}), 404

        # 检查权限 - 只有后端仓用户或管理员可以删除后端仓出库记录
        if not check_warehouse_permission('backend', 'delete'):
            return jsonify({'success': False, 'message': '您没有权限删除后端仓出库记录'}), 403

        # 记录要回退的数据
        customer_name = outbound_record.customer_name
        identification_code = outbound_record.identification_code
        pallet_count = outbound_record.pallet_count or 0
        package_count = outbound_record.package_count or 0
        weight = outbound_record.weight or 0
        volume = outbound_record.volume or 0

        current_app.logger.info(f"准备删除出库记录 ID={record_id}, 客户={customer_name}, 识别编码={identification_code}")
        current_app.logger.info(f"需要回退库存: 板数={pallet_count}, 件数={package_count}, 重量={weight}, 体积={volume}")

        # 查找对应的库存记录
        inventory = Inventory.query.filter_by(
            customer_name=customer_name,
            identification_code=identification_code
        ).first()

        if inventory:
            # 库存记录存在，增加库存
            inventory.pallet_count = (inventory.pallet_count or 0) + pallet_count
            inventory.package_count = (inventory.package_count or 0) + package_count
            inventory.weight = (inventory.weight or 0) + weight
            inventory.volume = (inventory.volume or 0) + volume

            # 对于PX开头的自行入库数据，需要恢复业务字段
            if identification_code and identification_code.startswith('PX/'):
                # 查找原始入库记录来恢复业务字段
                inbound_record = InboundRecord.query.filter_by(
                    identification_code=identification_code
                ).first()

                if inbound_record:
                    inventory.order_type = inbound_record.order_type
                    inventory.export_mode = inbound_record.export_mode
                    inventory.customs_broker = inbound_record.customs_broker
                    inventory.documents = inbound_record.documents

                    current_app.logger.info(f"批量删除时恢复PX自行入库数据的业务字段: {identification_code}, "
                                           f"订单类型: {inbound_record.order_type}, "
                                           f"出境模式: {inbound_record.export_mode}, "
                                           f"报关行: {inbound_record.customs_broker}")

            current_app.logger.info(f"更新现有库存记录: 新板数={inventory.pallet_count}, 新件数={inventory.package_count}")
        else:
            # 库存记录不存在，创建新的库存记录
            # 需要从出库记录中获取其他必要信息
            inventory = Inventory(
                customer_name=customer_name,
                identification_code=identification_code,
                plate_number=outbound_record.plate_number,
                pallet_count=pallet_count,
                package_count=package_count,
                weight=weight,
                volume=volume,
                order_type=outbound_record.order_type or '',
                export_mode=outbound_record.export_mode or '',
                customs_broker=outbound_record.customs_broker or '',
                service_staff=outbound_record.service_staff or '',
                remark1=outbound_record.remark1 or '',
                remark2=outbound_record.remark2 or '',
                inbound_date=outbound_record.outbound_time.date() if outbound_record.outbound_time else datetime.now().date()
            )
            db.session.add(inventory)
            current_app.logger.info(f"创建新库存记录: 板数={pallet_count}, 件数={package_count}")

        # 处理对应的在途货物记录回退（如果有的话）
        if identification_code and hasattr(outbound_record, 'batch_no') and outbound_record.batch_no:
            # 查找对应的在途货物记录
            transit_cargos = TransitCargo.query.filter_by(
                identification_code=identification_code,
                batch_no=outbound_record.batch_no,
                status='in_transit'
            ).all()

            if transit_cargos:
                for transit_cargo in transit_cargos:
                    current_app.logger.info(f"后端仓删除出库记录时同时删除在途货物记录: {transit_cargo.identification_code}, 批次: {transit_cargo.batch_no}")
                    db.session.delete(transit_cargo)

                current_app.logger.info(f"共删除 {len(transit_cargos)} 条在途货物记录")

        # 删除出库记录
        db.session.delete(outbound_record)

        # 提交事务
        db.session.commit()

        current_app.logger.info(f"成功删除出库记录 ID={record_id} 并回退库存")

        return jsonify({
            'success': True,
            'message': f'成功删除出库记录并回退库存：板数 {pallet_count}，件数 {package_count}'
        })

    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"删除后端仓出库记录失败: {str(e)}")
        current_app.logger.error(traceback.format_exc())
        return jsonify({'success': False, 'message': f'删除失败: {str(e)}'}), 500

@bp.route('/api/backend/outbound/batch_delete', methods=['POST'])
@login_required
@csrf.exempt
@require_permission('OUTBOUND_DELETE')
def api_backend_outbound_batch_delete():
    """批量删除后端仓出库记录并回退库存"""
    try:
        data = request.get_json()
        if not data or 'record_ids' not in data:
            return jsonify({'success': False, 'message': '请提供要删除的记录ID列表'}), 400

        record_ids = data['record_ids']
        if not record_ids or not isinstance(record_ids, list):
            return jsonify({'success': False, 'message': '记录ID列表不能为空'}), 400

        # 检查权限 - 只有后端仓用户或管理员可以删除后端仓出库记录
        if not check_warehouse_permission('backend', 'delete'):
            return jsonify({'success': False, 'message': '您没有权限删除后端仓出库记录'}), 403

        # 查找所有要删除的出库记录
        outbound_records = OutboundRecord.query.filter(OutboundRecord.id.in_(record_ids)).all()
        if not outbound_records:
            return jsonify({'success': False, 'message': '未找到要删除的出库记录'}), 404

        deleted_count = 0
        total_pallet_count = 0
        total_package_count = 0

        # 逐个处理删除和库存回退
        for outbound_record in outbound_records:
            try:
                # 记录要回退的数据
                customer_name = outbound_record.customer_name
                identification_code = outbound_record.identification_code
                pallet_count = outbound_record.pallet_count or 0
                package_count = outbound_record.package_count or 0
                weight = outbound_record.weight or 0
                volume = outbound_record.volume or 0

                current_app.logger.info(f"批量删除: 处理出库记录 ID={outbound_record.id}, 客户={customer_name}, 识别编码={identification_code}")

                # 查找对应的库存记录
                inventory = Inventory.query.filter_by(
                    customer_name=customer_name,
                    identification_code=identification_code
                ).first()

                if inventory:
                    # 库存记录存在，增加库存
                    inventory.pallet_count = (inventory.pallet_count or 0) + pallet_count
                    inventory.package_count = (inventory.package_count or 0) + package_count
                    inventory.weight = (inventory.weight or 0) + weight
                    inventory.volume = (inventory.volume or 0) + volume

                    # 对于PX开头的自行入库数据，需要恢复业务字段
                    if identification_code and identification_code.startswith('PX/'):
                        # 查找原始入库记录来恢复业务字段
                        inbound_record = InboundRecord.query.filter_by(
                            identification_code=identification_code
                        ).first()

                        if inbound_record:
                            inventory.order_type = inbound_record.order_type
                            inventory.export_mode = inbound_record.export_mode
                            inventory.customs_broker = inbound_record.customs_broker
                            inventory.documents = inbound_record.documents

                            current_app.logger.info(f"批量删除时恢复PX自行入库数据的业务字段: {identification_code}, "
                                                   f"订单类型: {inbound_record.order_type}, "
                                                   f"出境模式: {inbound_record.export_mode}, "
                                                   f"报关行: {inbound_record.customs_broker}")

                    current_app.logger.info(f"批量删除: 更新现有库存记录: 新板数={inventory.pallet_count}, 新件数={inventory.package_count}")
                else:
                    # 库存记录不存在，创建新的库存记录
                    inventory = Inventory(
                        customer_name=customer_name,
                        identification_code=identification_code,
                        plate_number=outbound_record.plate_number,
                        pallet_count=pallet_count,
                        package_count=package_count,
                        weight=weight,
                        volume=volume,
                        order_type=outbound_record.order_type or '',
                        export_mode=outbound_record.export_mode or '',
                        customs_broker=outbound_record.customs_broker or '',
                        service_staff=outbound_record.service_staff or '',
                        remark1=outbound_record.remark1 or '',
                        remark2=outbound_record.remark2 or '',
                        inbound_date=outbound_record.outbound_time.date() if outbound_record.outbound_time else datetime.now().date()
                    )
                    db.session.add(inventory)
                    current_app.logger.info(f"批量删除: 创建新库存记录: 板数={pallet_count}, 件数={package_count}")

                # 处理对应的在途货物记录回退（如果有的话）
                if identification_code and hasattr(outbound_record, 'batch_no') and outbound_record.batch_no:
                    # 查找对应的在途货物记录
                    transit_cargos = TransitCargo.query.filter_by(
                        identification_code=identification_code,
                        batch_no=outbound_record.batch_no,
                        status='in_transit'
                    ).all()

                    if transit_cargos:
                        for transit_cargo in transit_cargos:
                            current_app.logger.info(f"批量删除: 同时删除在途货物记录: {transit_cargo.identification_code}, 批次: {transit_cargo.batch_no}")
                            db.session.delete(transit_cargo)

                # 删除出库记录
                db.session.delete(outbound_record)

                # 累计统计
                deleted_count += 1
                total_pallet_count += pallet_count
                total_package_count += package_count

            except Exception as e:
                current_app.logger.error(f"批量删除: 处理记录 ID={outbound_record.id} 时出错: {str(e)}")
                continue

        # 提交事务
        db.session.commit()

        current_app.logger.info(f"批量删除成功: 共删除 {deleted_count} 条出库记录，回退库存板数 {total_pallet_count}，件数 {total_package_count}")

        return jsonify({
            'success': True,
            'message': f'成功删除 {deleted_count} 条出库记录并回退库存：板数 {total_pallet_count}，件数 {total_package_count}'
        })

    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"批量删除后端仓出库记录失败: {str(e)}")
        current_app.logger.error(traceback.format_exc())
        return jsonify({'success': False, 'message': f'批量删除失败: {str(e)}'}), 500

@bp.route('/api/backend/outbound/delete_batch/<batch_no>', methods=['DELETE'])
@login_required
@csrf.exempt
@require_permission('OUTBOUND_DELETE')
def api_backend_outbound_delete_batch(batch_no):
    """按批次删除后端仓出库记录并回退库存"""
    try:
        # 检查权限 - 只有后端仓用户或管理员可以删除后端仓出库记录
        if not check_warehouse_permission('backend', 'delete'):
            return jsonify({'success': False, 'message': '您没有权限删除后端仓出库记录'}), 403

        # 查找该批次的所有出库记录
        outbound_records = OutboundRecord.query.filter_by(batch_no=batch_no).all()
        if not outbound_records:
            return jsonify({'success': False, 'message': f'未找到批次号 {batch_no} 的出库记录'}), 404

        deleted_count = 0
        total_pallet_count = 0
        total_package_count = 0

        current_app.logger.info(f"开始删除批次 {batch_no}，共 {len(outbound_records)} 条记录")

        # 逐个处理删除和库存回退
        for outbound_record in outbound_records:
            try:
                # 记录要回退的数据
                customer_name = outbound_record.customer_name
                identification_code = outbound_record.identification_code
                pallet_count = outbound_record.pallet_count or 0
                package_count = outbound_record.package_count or 0
                weight = outbound_record.weight or 0
                volume = outbound_record.volume or 0

                current_app.logger.info(f"批次删除: 处理出库记录 ID={outbound_record.id}, 客户={customer_name}, 识别编码={identification_code}")

                # 查找对应的库存记录 - 必须在同一个仓库中查找
                inventory = Inventory.query.filter_by(
                    customer_name=customer_name,
                    identification_code=identification_code,
                    operated_warehouse_id=outbound_record.operated_warehouse_id
                ).first()

                if inventory:
                    # 库存记录存在，增加库存
                    inventory.pallet_count = (inventory.pallet_count or 0) + pallet_count
                    inventory.package_count = (inventory.package_count or 0) + package_count
                    inventory.weight = (inventory.weight or 0) + weight
                    inventory.volume = (inventory.volume or 0) + volume

                    # 对于PX开头的自行入库数据，需要恢复业务字段
                    if identification_code and identification_code.startswith('PX/'):
                        # 查找原始入库记录来恢复业务字段
                        inbound_record = InboundRecord.query.filter_by(
                            identification_code=identification_code
                        ).first()

                        if inbound_record:
                            inventory.order_type = inbound_record.order_type
                            inventory.export_mode = inbound_record.export_mode
                            inventory.customs_broker = inbound_record.customs_broker
                            inventory.documents = inbound_record.documents

                            current_app.logger.info(f"批次删除时恢复PX自行入库数据的业务字段: {identification_code}, "
                                                   f"订单类型: {inbound_record.order_type}, "
                                                   f"出境模式: {inbound_record.export_mode}, "
                                                   f"报关行: {inbound_record.customs_broker}")

                    current_app.logger.info(f"批次删除: 更新现有库存记录: 新板数={inventory.pallet_count}, 新件数={inventory.package_count}")
                else:
                    # 库存记录不存在，创建新的库存记录 - 必须设置正确的仓库ID
                    inventory = Inventory(
                        customer_name=customer_name,
                        identification_code=identification_code,
                        plate_number=outbound_record.plate_number,
                        pallet_count=pallet_count,
                        package_count=package_count,
                        weight=weight,
                        volume=volume,
                        order_type=outbound_record.order_type or '',
                        export_mode=outbound_record.export_mode or '',
                        customs_broker=outbound_record.customs_broker or '',
                        service_staff=outbound_record.service_staff or '',
                        operated_warehouse_id=outbound_record.operated_warehouse_id,  # 关键：设置正确的仓库ID
                        inbound_time=outbound_record.outbound_time if outbound_record.outbound_time else datetime.now()
                    )
                    db.session.add(inventory)
                    warehouse_name = outbound_record.operated_warehouse.warehouse_name if outbound_record.operated_warehouse else "未知仓库"
                    current_app.logger.info(f"批次删除: 创建新库存记录: 板数={pallet_count}, 件数={package_count}, 仓库={warehouse_name}")

                # 处理对应的在途货物记录回退（如果有的话）
                if identification_code and hasattr(outbound_record, 'batch_no') and outbound_record.batch_no:
                    # 查找对应的在途货物记录
                    transit_cargos = TransitCargo.query.filter_by(
                        identification_code=identification_code,
                        batch_no=outbound_record.batch_no,
                        status='in_transit'
                    ).all()

                    if transit_cargos:
                        for transit_cargo in transit_cargos:
                            current_app.logger.info(f"批次删除: 同时删除在途货物记录: {transit_cargo.identification_code}, 批次: {transit_cargo.batch_no}")
                            db.session.delete(transit_cargo)

                # 删除出库记录
                db.session.delete(outbound_record)

                # 累计统计
                deleted_count += 1
                total_pallet_count += pallet_count
                total_package_count += package_count

            except Exception as e:
                current_app.logger.error(f"批次删除: 处理记录 ID={outbound_record.id} 时出错: {str(e)}")
                current_app.logger.error(f"批次删除: 异常详情: {traceback.format_exc()}")
                # 不要continue，而是抛出异常让用户知道具体问题
                raise Exception(f"删除记录 ID={outbound_record.id} 时出错: {str(e)}")

        # 提交事务
        db.session.commit()

        current_app.logger.info(f"批次删除成功: 批次号 {batch_no}，共删除 {deleted_count} 条出库记录，回退库存板数 {total_pallet_count}，件数 {total_package_count}")

        return jsonify({
            'success': True,
            'message': f'成功删除批次 {batch_no} 的 {deleted_count} 条出库记录并回退库存：板数 {total_pallet_count}，件数 {total_package_count}'
        })

    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"删除批次 {batch_no} 失败: {str(e)}")
        current_app.logger.error(traceback.format_exc())
        return jsonify({'success': False, 'message': f'删除批次失败: {str(e)}'}), 500

@bp.route('/api/backend/outbound/save', methods=['POST'])
@login_required
@csrf.exempt
def api_backend_outbound_save():
    """后端仓出库保存API"""
    try:
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'message': '无效的请求数据'}), 400

        # 验证必填字段
        required_fields = ['departure_date', 'destination', 'plate_number', 'cargo_list']
        for field in required_fields:
            if not data.get(field):
                return jsonify({'success': False, 'message': f'缺少必填字段: {field}'}), 400

        cargo_list = data.get('cargo_list', [])
        if not cargo_list:
            return jsonify({'success': False, 'message': '请至少选择一项货物'}), 400

        # 生成批次号
        from app.utils.batch_generator import generate_batch_number
        batch_no = generate_batch_number(4)  # 4是凭祥北投仓的ID，不传递db_session避免事务冲突

        # 创建出库记录
        for cargo in cargo_list:
            # 检查库存是否足够
            inventory = Inventory.query.get(cargo['id'])
            if not inventory:
                raise Exception(f'库存记录不存在: {cargo["id"]}')

            if (inventory.pallet_count or 0) < (cargo.get('pallet_count', 0) or 0):
                raise Exception(f'库存板数不足: {cargo["customer_name"]}')

            if (inventory.package_count or 0) < (cargo.get('package_count', 0) or 0):
                raise Exception(f'库存件数不足: {cargo["customer_name"]}')

            # 使用用户输入的备注，不自动添加车挂和柜号信息
            remarks = cargo.get('remarks', '')

            # 查询该识别编码已有的最大batch_sequence，计算下一个序号
            max_sequence = db.session.query(db.func.max(OutboundRecord.batch_sequence)).filter(
                OutboundRecord.identification_code == inventory.identification_code
            ).scalar() or 0
            next_sequence = max_sequence + 1

            # 创建出库记录，使用库存记录的识别编码保持一致性
            outbound_record = OutboundRecord(
                outbound_time=datetime.strptime(data['departure_date'], '%Y-%m-%d'),
                customer_name=cargo['customer_name'],
                identification_code=inventory.identification_code,  # 使用库存记录的识别编码
                pallet_count=cargo.get('pallet_count', 0),
                package_count=cargo.get('package_count', 0),
                weight=cargo.get('weight', 0),
                volume=cargo.get('volume', 0),
                destination=data['destination'],
                plate_number=data['plate_number'],  # 出库车牌
                delivery_plate_number=cargo.get('delivery_plate_number', '') or cargo.get('delivery_truck', ''),  # 送货干线车（从货物数据获取）
                vehicle_type=data.get('vehicle_type', ''),
                transport_company=data.get('fleet', ''),  # 使用transport_company字段
                driver_phone=data.get('phone', ''),  # 使用driver_phone字段
                contact_window=data.get('contact_window', ''),
                trailer=data.get('trailer', ''),  # 车挂
                container_number=data.get('container_number', ''),  # 柜号
                order_type=cargo.get('order_type', ''),  # 添加订单类型字段
                export_mode=cargo.get('export_mode', ''),  # 添加出境模式字段
                customs_broker=cargo.get('customs_broker', ''),  # 添加报关行字段
                batch_no=batch_no,
                batch_sequence=next_sequence,  # 设置正确的批次序号
                remarks=remarks,
                operated_warehouse_id=4,  # 凭祥北投仓
                operated_by_user_id=current_user.id,
                service_staff=cargo.get('service_staff', ''),
                inbound_plate=cargo['plate_number']  # 入库车牌
            )

            db.session.add(outbound_record)

            # 更新库存
            inventory.pallet_count = (inventory.pallet_count or 0) - (cargo.get('pallet_count', 0) or 0)
            inventory.package_count = (inventory.package_count or 0) - (cargo.get('package_count', 0) or 0)

            # 如果库存为0，删除库存记录
            if (inventory.pallet_count or 0) <= 0 and (inventory.package_count or 0) <= 0:
                db.session.delete(inventory)

        # 提交事务
        db.session.commit()

        return jsonify({
            'success': True,
            'message': '出库保存成功',
            'batch_no': batch_no
        })

    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"后端仓出库保存失败: {str(e)}")
        return jsonify({'success': False, 'message': f'保存失败: {str(e)}'}), 500


@bp.route('/api/receiver/contact', methods=['GET'])
@login_required
@require_any_permission(['RECEIVER_LIST', 'OUTBOUND_VIEW', 'OUTBOUND_CREATE'])
def api_receiver_contact():
    """获取收货人联络窗口信息API"""
    try:
        warehouse_name = request.args.get('warehouse_name', '').strip()
        if not warehouse_name:
            return jsonify({'success': False, 'message': '缺少仓库名称参数'}), 400

        # 查询收货人信息
        receiver = Receiver.query.filter_by(warehouse_name=warehouse_name).first()

        if receiver:
            return jsonify({
                'success': True,
                'data': {
                    'warehouse_name': receiver.warehouse_name,
                    'contact': receiver.contact,
                    'address': receiver.address
                }
            })
        else:
            return jsonify({
                'success': False,
                'message': f'未找到仓库 {warehouse_name} 的联络窗口信息'
            })

    except Exception as e:
        current_app.logger.error(f"获取联络窗口信息失败: {str(e)}")
        return jsonify({'success': False, 'message': f'获取失败: {str(e)}'}), 500


@bp.route('/api/receiver/init', methods=['POST'])
@login_required
def api_receiver_init():
    """初始化收货人信息数据"""
    try:
        # 检查是否已经有数据
        existing_count = Receiver.query.count()
        if existing_count > 0:
            return jsonify({'success': True, 'message': f'收货人信息已存在 {existing_count} 条记录'})

        # 初始化数据
        receivers_data = [
            {
                'warehouse_name': '凭祥保税仓',
                'address': '广西壮族自治区崇左市凭祥市',
                'contact': '保税仓联络窗口'
            },
            {
                'warehouse_name': '春疆货场',
                'address': '新疆维吾尔自治区',
                'contact': '春疆货场联络窗口'
            }
        ]

        for data in receivers_data:
            receiver = Receiver(
                warehouse_name=data['warehouse_name'],
                address=data['address'],
                contact=data['contact']
            )
            db.session.add(receiver)

        db.session.commit()

        return jsonify({
            'success': True,
            'message': f'成功初始化 {len(receivers_data)} 条收货人信息'
        })

    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"初始化收货人信息失败: {str(e)}")
        import traceback
        current_app.logger.error(f"详细错误信息: {traceback.format_exc()}")
        return jsonify({'success': False, 'message': f'初始化失败: {str(e)}'}), 500


@bp.route('/export_backend_outbound')
@require_permission('OUTBOUND_VIEW')
def export_backend_outbound():
    """导出后端仓出库记录到Excel"""
    try:
        # 检查权限
        if not check_warehouse_permission('backend', 'view'):
            flash('您没有权限导出后端仓出库记录', 'error')
            return redirect(url_for('main.backend_outbound_list'))

        # 获取搜索参数
        date_start = request.args.get('date_start', '')
        date_end = request.args.get('date_end', '')

        # 如果没有指定日期范围，默认使用最近一周的日期范围
        if not date_start and not date_end:
            today = datetime.now().strftime('%Y-%m-%d')
            one_week_ago = (datetime.now() - timedelta(days=7)).strftime('%Y-%m-%d')
            date_start = one_week_ago
            date_end = today

        batch_no = request.args.get('batch_no', '')
        plate_number = request.args.get('plate_number', '')
        customer_name = request.args.get('customer_name', '')
        destination = request.args.get('destination', '')
        identification_code = request.args.get('identification_code', '')
        export_mode = request.args.get('export_mode', '')
        customs_broker = request.args.get('customs_broker', '')
        inbound_plate = request.args.get('inbound_plate', '')
        service_staff = request.args.get('service_staff', '')

        # 获取后端仓库
        backend_warehouse = Warehouse.query.filter_by(warehouse_type='backend').first()
        if not backend_warehouse:
            flash('未找到后端仓库', 'error')
            return redirect(url_for('main.backend_outbound_list'))

        # 构建查询，只查询后端仓的出库记录
        query = OutboundRecord.query.options(
            db.joinedload(OutboundRecord.operated_warehouse),
            db.joinedload(OutboundRecord.destination_warehouse)
        ).filter(OutboundRecord.operated_warehouse_id == backend_warehouse.id)

        # 根据用户权限过滤数据
        if hasattr(current_user, 'warehouse') and current_user.warehouse:
            if current_user.warehouse.warehouse_type == 'backend':
                # 后端仓用户只能导出自己仓库的数据
                query = query.filter_by(operated_warehouse_id=current_user.warehouse_id)

        # 日期范围过滤
        if date_start:
            try:
                start_date = datetime.strptime(date_start, '%Y-%m-%d')
                query = query.filter(OutboundRecord.outbound_time >= start_date)
            except ValueError:
                pass
        if date_end:
            try:
                end_date = datetime.strptime(date_end, '%Y-%m-%d')
                end_date = end_date.replace(hour=23, minute=59, second=59)
                query = query.filter(OutboundRecord.outbound_time <= end_date)
            except ValueError:
                pass

        # 其他搜索条件过滤
        if batch_no:
            query = query.filter(OutboundRecord.batch_no.like(f'%{batch_no}%'))
        if plate_number:
            query = query.filter(OutboundRecord.plate_number.like(f'%{plate_number}%'))
        if customer_name:
            query = query.filter(OutboundRecord.customer_name.like(f'%{customer_name}%'))
        if destination:
            query = query.filter(OutboundRecord.destination.like(f'%{destination}%'))
        if identification_code:
            query = query.filter(OutboundRecord.identification_code.like(f'%{identification_code}%'))
        if export_mode:
            query = query.filter(OutboundRecord.export_mode.like(f'%{export_mode}%'))
        if customs_broker:
            query = query.filter(OutboundRecord.customs_broker.like(f'%{customs_broker}%'))
        if inbound_plate:
            query = query.filter(OutboundRecord.inbound_plate.like(f'%{inbound_plate}%'))
        if service_staff:
            query = query.filter(OutboundRecord.service_staff.like(f'%{service_staff}%'))

        # 按出库时间降序排序
        records = query.order_by(OutboundRecord.outbound_time.desc()).all()

        # 如果没有记录，返回提示
        if not records:
            flash('没有找到符合条件的后端仓出库记录', 'warning')
            return redirect(url_for('main.backend_outbound_list'))

        # 准备导出数据
        data = []
        for record in records:
            data.append({
                '出库时间': record.outbound_time.strftime('%Y-%m-%d %H:%M:%S') if record.outbound_time else '',
                '批次号': record.batch_no or '',
                '送货干线车': record.delivery_plate_number or '',
                '出库/出境车牌': record.plate_number or '',
                '入库车牌': record.inbound_plate or '',
                '客户名称': record.customer_name or '',
                '识别编码': record.identification_code or '',
                '目的地': record.destination or '',
                '出境模式': record.export_mode or '',
                '报关行': record.customs_broker or '',
                '订单类型': record.order_type or '',
                '出库板数': record.pallet_count or 0,
                '出库件数': record.package_count or 0,
                '重量(KG)': record.weight or 0,
                '体积(CBM)': record.volume or 0,
                '单据': record.document_no or '',
                '跟单客服': record.service_staff or '',
                '备注': record.remarks or '',
                '入库日期': record.inbound_date.strftime('%Y-%m-%d') if record.inbound_date else '',
                '操作仓库': record.operated_warehouse.warehouse_name if record.operated_warehouse else '',
                '操作用户': record.operated_by_user.username if record.operated_by_user else '',
                '创建时间': record.outbound_time.strftime('%Y-%m-%d %H:%M:%S') if record.outbound_time else ''
            })

        # 使用pandas导出Excel
        try:
            import pandas as pd
            from io import BytesIO

            df = pd.DataFrame(data)
            output = BytesIO()

            with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
                df.to_excel(writer, sheet_name='后端仓出库记录', index=False)

                # 获取工作表和工作簿对象
                workbook = writer.book
                worksheet = writer.sheets['后端仓出库记录']

                # 设置列宽
                for i, col in enumerate(df.columns):
                    max_len = max(df[col].astype(str).map(len).max(), len(col)) + 2
                    worksheet.set_column(i, i, min(max_len, 50))

            # 设置文件指针到开始位置
            output.seek(0)

            # 生成下载文件名
            timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
            filename = f"后端仓出库记录导出_{timestamp}.xlsx"

            # 返回Excel文件
            return send_file(
                output,
                as_attachment=True,
                download_name=filename,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )

        except ImportError:
            flash("导出Excel需要pandas或xlsxwriter库，请安装: pip install pandas xlsxwriter 或 pip install openpyxl", "danger")
            return redirect(url_for('main.backend_outbound_list'))

    except Exception as e:
        current_app.logger.error(f"导出后端仓出库记录时出错: {str(e)}")
        flash(f"导出数据时出错: {str(e)}", "danger")
        return redirect(url_for('main.backend_outbound_list'))


@bp.route('/backend/outbound/print', methods=['GET'])
@require_permission('OUTBOUND_VIEW')
def backend_outbound_print_regular():
    """后端仓出库记录打印页面 - 常规出库记录"""
    try:
        # 检查页面权限
        from app.utils.permission_manager import PermissionManager
        if not PermissionManager.has_page_permission(current_user.id, 'BACKEND_OUTBOUND_PRINT'):
            flash('您没有权限访问后端仓出库单页面', 'error')
            return redirect(url_for('main.index'))

        # 检查权限
        if not check_warehouse_permission('backend', 'view'):
            flash('您没有权限访问后端仓出库记录', 'error')
            return redirect(url_for('main.index'))

        # 获取分页参数
        page = request.args.get('page', 1, type=int)
        per_page = 15  # 每页显示15个批次

        # 获取搜索参数
        date_start = request.args.get('date_start', '')
        date_end = request.args.get('date_end', '')
        search_field = request.args.get('search_field', 'customer_name')
        search_type = request.args.get('search_type', 'contains')
        search_value = request.args.get('search_value', '')

        # 构建查询，只查询后端仓的常规出库记录（排除返回前端仓的记录）
        query = OutboundRecord.query.join(
            Warehouse, OutboundRecord.operated_warehouse_id == Warehouse.id
        ).filter(
            db.and_(
                # 只显示后端仓的记录
                Warehouse.warehouse_type == 'backend',
                # 排除后端仓返回前端仓的记录
                ~OutboundRecord.remarks.like('%后端仓返回前端仓%')
            )
        )

        # 应用搜索条件
        if date_start:
            try:
                start_date = datetime.strptime(date_start, '%Y-%m-%d')
                query = query.filter(OutboundRecord.outbound_time >= start_date)
            except ValueError:
                flash('开始日期格式不正确', 'warning')

        if date_end:
            try:
                end_date = datetime.strptime(date_end, '%Y-%m-%d')
                end_date = end_date.replace(hour=23, minute=59, second=59)
                query = query.filter(OutboundRecord.outbound_time <= end_date)
            except ValueError:
                flash('结束日期格式不正确', 'warning')

        if search_value:
            if hasattr(OutboundRecord, search_field):
                field = getattr(OutboundRecord, search_field)
                if search_type == 'exact':
                    query = query.filter(field == search_value)
                elif search_type == 'startswith':
                    query = query.filter(field.like(f'{search_value}%'))
                elif search_type == 'endswith':
                    query = query.filter(field.like(f'%{search_value}'))
                else:  # contains
                    query = query.filter(field.like(f'%{search_value}%'))

        # 按批次号分组
        all_records = query.order_by(OutboundRecord.outbound_time.desc()).all()

        # 按批次号分组统计
        batch_stats = {}
        for record in all_records:
            batch_no = record.batch_no or '未分配批次'
            if batch_no not in batch_stats:
                batch_stats[batch_no] = {
                    'batch_no': batch_no,
                    'records': [],
                    'total_pallet': 0,
                    'total_package': 0,
                    'total_weight': 0,
                    'total_volume': 0,
                    'outbound_time': record.outbound_time,
                    'destination': record.destination,
                    'driver_name': record.driver_name,
                    'delivery_plate_number': record.delivery_plate_number
                }

            batch_stats[batch_no]['records'].append(record)
            batch_stats[batch_no]['total_pallet'] += record.pallet_count or 0
            batch_stats[batch_no]['total_package'] += record.package_count or 0
            batch_stats[batch_no]['total_weight'] += record.weight or 0
            batch_stats[batch_no]['total_volume'] += record.volume or 0

        # 转换为列表并分页
        batch_list = list(batch_stats.values())
        total_batches = len(batch_list)

        # 分页处理
        start_idx = (page - 1) * per_page
        end_idx = start_idx + per_page
        paginated_batches = batch_list[start_idx:end_idx]

        # 创建分页对象
        class BatchPagination:
            def __init__(self, items, page, per_page, total):
                self.items = items
                self.page = page
                self.per_page = per_page
                self.total = total
                self.pages = (total + per_page - 1) // per_page
                self.has_prev = page > 1
                self.has_next = page < self.pages
                self.prev_num = page - 1 if self.has_prev else None
                self.next_num = page + 1 if self.has_next else None

        batch_pagination = BatchPagination(paginated_batches, page, per_page, total_batches)

        # 准备搜索参数
        search_params = {
            'date_start': date_start,
            'date_end': date_end,
            'search_field': search_field,
            'search_type': search_type,
            'search_value': search_value
        }

        return render_template('backend/outbound_print.html',
                             batch_groups=batch_pagination,
                             search_params=search_params,
                             title='后端仓出库记录打印')

    except Exception as e:
        current_app.logger.error(f"查询后端仓出库记录时出错: {str(e)}")
        flash(f"查询数据时出错: {str(e)}", "danger")
        return redirect(url_for('main.backend_outbound_list'))


@bp.route('/backend/outbound/print_list', methods=['GET'])
@require_permission('OUTBOUND_VIEW')
def backend_outbound_print():
    """后端仓出库单打印列表 - 专门显示后端仓返回前端仓的数据"""
    # 检查页面权限
    from app.utils.permission_manager import PermissionManager
    if not PermissionManager.has_page_permission(current_user.id, 'BACKEND_OUTBOUND_PRINT'):
        flash('您没有权限访问后端仓出库单页面', 'error')
        return redirect(url_for('main.index'))
    # 获取分页参数
    page = request.args.get('page', 1, type=int)
    per_page = 15  # 每页显示15个批次记录

    # 获取搜索参数 - 默认显示最近7天的数据
    search_params = {
        'date_start': request.args.get('date_start', (datetime.now() - timedelta(days=7)).strftime('%Y-%m-%d')),
        'date_end': request.args.get('date_end', datetime.now().strftime('%Y-%m-%d')),
        'plate_number': request.args.get('plate_number', '').strip(),
        'batch_no': request.args.get('batch_no', '').strip(),
        'customer_name': request.args.get('customer_name', '').strip()
    }

    # 记录搜索参数
    current_app.logger.info(f"后端仓出库单打印列表搜索参数: {search_params}")

    # 检查是否有查询参数
    has_search_params = any([
        search_params['plate_number'],
        search_params['batch_no'],
        search_params['customer_name']
    ])

    # 如果没有额外的搜索参数，默认只取最近7天的数据以提高速度
    if not has_search_params:
        start_date = datetime.strptime(search_params['date_start'], '%Y-%m-%d')
        end_date = datetime.strptime(search_params['date_end'], '%Y-%m-%d')
        date_diff = (end_date - start_date).days

        if date_diff > 7:
            current_app.logger.info("日期范围超过7天且无其他筛选条件，默认只取最近7天的数据")
            search_params['date_start'] = (datetime.now() - timedelta(days=7)).strftime('%Y-%m-%d')

    # 构建查询 - 专门查询后端仓返回前端仓的数据
    try:
        # 限制最大记录数以防止加载过多数据导致页面卡顿
        max_records = 1000

        # 构建查询 - 显示后端仓转前端仓的出库记录
        # 获取所有前端仓库ID
        frontend_warehouses = Warehouse.query.filter_by(warehouse_type='frontend').all()
        frontend_warehouse_ids = [w.id for w in frontend_warehouses]
        frontend_warehouse_names = [w.warehouse_name for w in frontend_warehouses]

        query = OutboundRecord.query.join(
            Warehouse, OutboundRecord.operated_warehouse_id == Warehouse.id
        ).filter(
            db.and_(
                # 只显示后端仓的记录
                Warehouse.warehouse_type == 'backend',
                # 目标是前端仓库的记录
                db.or_(
                    OutboundRecord.destination_warehouse_id.in_(frontend_warehouse_ids),
                    OutboundRecord.destination.in_(frontend_warehouse_names)
                )
            )
        )

        # 添加仓库权限过滤
        if current_user.is_authenticated:
            is_admin = getattr(current_user, 'is_admin', False) or getattr(current_user, 'user_type', '') == 'admin'

            if not is_admin and hasattr(current_user, 'warehouse') and current_user.warehouse:
                if current_user.warehouse.warehouse_type == 'frontend':
                    # 前端仓用户不能访问后端仓出库单
                    current_app.logger.warning(f"前端仓用户 {current_user.username} 尝试访问后端仓出库单")
                    flash('您没有权限查看后端仓出库单', 'warning')
                    return redirect(url_for('main.index'))
                # 后端仓用户可以查看所有后端仓的出库记录
                elif current_user.warehouse.warehouse_type == 'backend':
                    current_app.logger.info(f"后端仓用户 {current_user.username} 查看后端仓出库单")
            elif is_admin:
                current_app.logger.info(f"管理员用户 {current_user.username} 查看后端仓出库单")
            else:
                current_app.logger.warning(f"用户 {current_user.username} 没有关联仓库，无法查看后端仓出库单")
                # 如果用户没有关联仓库，返回空结果
                query = query.filter(OutboundRecord.id == -1)  # 永远不会匹配的条件

        # 应用过滤条件
        if search_params['date_start'] and search_params['date_end']:
            start_date = datetime.strptime(search_params['date_start'], '%Y-%m-%d')
            end_date = datetime.strptime(search_params['date_end'], '%Y-%m-%d')
            end_date = end_date.replace(hour=23, minute=59, second=59)
            query = query.filter(OutboundRecord.outbound_time.between(start_date, end_date))

        if search_params['plate_number']:
            query = query.filter(OutboundRecord.plate_number.like(f"%{search_params['plate_number']}%"))

        if search_params['batch_no']:
            query = query.filter(OutboundRecord.batch_no.like(f"%{search_params['batch_no']}%"))

        if search_params['customer_name']:
            query = query.filter(OutboundRecord.customer_name.like(f"%{search_params['customer_name']}%"))

        # 按出库时间倒序排列
        query = query.order_by(OutboundRecord.outbound_time.desc())

        # 限制记录数量
        query = query.limit(max_records)

        # 获取所有记录
        all_records = query.all()
        current_app.logger.info(f"查询到 {len(all_records)} 条后端仓出库记录")

        if not all_records:
            current_app.logger.info("没有找到符合条件的后端仓出库记录")
            return render_template('backend_outbound_print_list.html',
                                batch_records=[],
                                batch_groups=[],
                                batch_pagination=None,
                                title='后端仓出库单',
                                search_params=search_params,
                                date_rowspans={},
                                plate_rowspans={},
                                batch_rowspans={})

        # 按批次号分组记录
        batch_groups = {}
        for record in all_records:
            batch_key = record.batch_no
            if batch_key not in batch_groups:
                batch_groups[batch_key] = {
                    'batch_no': batch_key,
                    'records': [],
                    'total_pallet_count': 0,
                    'total_package_count': 0,
                    'total_weight': 0,
                    'total_volume': 0,
                    'outbound_time': record.outbound_time,
                    'plate_number': record.plate_number,
                    'destination': record.destination
                }

            batch_groups[batch_key]['records'].append(record)
            batch_groups[batch_key]['total_pallet_count'] += record.pallet_count or 0
            batch_groups[batch_key]['total_package_count'] += record.package_count or 0
            batch_groups[batch_key]['total_weight'] += record.weight or 0
            batch_groups[batch_key]['total_volume'] += record.volume or 0

        # 转换为列表并按时间排序
        batch_stats = list(batch_groups.values())
        batch_stats.sort(key=lambda x: x['outbound_time'], reverse=True)

        # 获取所有批次号用于分页
        batch_nos = [batch['batch_no'] for batch in batch_stats]
        total_batches = len(batch_nos)

        # 分页处理
        start_idx = (page - 1) * per_page
        end_idx = start_idx + per_page
        paginated_batch_nos = batch_nos[start_idx:end_idx]
        paginated_batch_stats = [batch for batch in batch_stats if batch['batch_no'] in paginated_batch_nos]

        # 获取分页批次的所有记录
        paginated_records = []
        for batch in paginated_batch_stats:
            paginated_records.extend(batch['records'])

        # 计算rowspan
        date_rowspans = {}
        plate_rowspans = {}
        batch_rowspans = {}

        for batch in paginated_batch_stats:
            batch_no = batch['batch_no']
            record_count = len(batch['records'])
            batch_rowspans[batch_no] = record_count

            # 计算日期和车牌的rowspan
            outbound_date = batch['outbound_time'].strftime('%Y-%m-%d')
            plate_number = batch['plate_number']

            if outbound_date not in date_rowspans:
                date_rowspans[outbound_date] = 0
            date_rowspans[outbound_date] += record_count

            plate_key = f"{outbound_date}_{plate_number}"
            if plate_key not in plate_rowspans:
                plate_rowspans[plate_key] = 0
            plate_rowspans[plate_key] += record_count

        # 创建分页对象
        class BatchPagination:
            def __init__(self, items, page, per_page, total):
                self.items = items
                self.page = page
                self.per_page = per_page
                self.total = total
                self.pages = (total + per_page - 1) // per_page
                self.has_prev = page > 1
                self.has_next = page < self.pages
                self.prev_num = page - 1 if self.has_prev else None
                self.next_num = page + 1 if self.has_next else None

        batch_pagination = BatchPagination(paginated_batch_nos, page, per_page, total_batches)

        current_app.logger.debug("开始渲染后端仓出库单模板")

        return render_template('backend_outbound_print_list.html',
                            batch_records=paginated_records,
                            batch_groups=paginated_batch_stats,
                            batch_pagination=batch_pagination,
                            title='后端仓出库单',
                            search_params=search_params,
                            date_rowspans=date_rowspans,
                            plate_rowspans=plate_rowspans,
                            batch_rowspans=batch_rowspans)

    except Exception as e:
        current_app.logger.error(f"查询后端仓出库记录时出错: {str(e)}")
        flash(f"查询数据时出错: {str(e)}", "danger")
        return render_template('backend_outbound_print_list.html',
                            batch_records=[],
                            batch_groups=[],
                            batch_pagination=None,
                            title='后端仓出库单',
                            search_params=search_params,
                            date_rowspans={},
                            plate_rowspans={},
                            batch_rowspans={})


@bp.route('/backend/outbound/print_document', methods=['POST'])
@require_permission('OUTBOUND_VIEW')
def backend_outbound_print_document():
    """后端仓出库单打印文档"""
    record_ids = request.form.getlist('record_ids[]')
    current_app.logger.info(f"选中的后端仓出库记录ID: {record_ids}")

    if not record_ids:
        flash('请至少选择一条记录进行打印', 'warning')
        return redirect(url_for('main.backend_outbound_print'))

    # 查询选中的记录，包含关联数据
    records = OutboundRecord.query.options(
        db.joinedload(OutboundRecord.operated_warehouse),
        db.joinedload(OutboundRecord.destination_warehouse),
        db.joinedload(OutboundRecord.receiver)  # 添加收货人信息关联
    ).filter(
        OutboundRecord.id.in_(record_ids)
    ).all()
    current_app.logger.info(f"查询到的后端仓出库记录数: {len(records)}")

    if not records:
        flash('没有找到符合条件的记录', 'warning')
        return redirect(url_for('main.backend_outbound_print'))

    # 按批次号分组
    batch_groups = {}
    for record in records:
        batch_no = record.batch_no
        if batch_no not in batch_groups:
            batch_groups[batch_no] = {
                'batch_no': batch_no,
                'records': [],
                'total_pallet_count': 0,
                'total_package_count': 0,
                'total_weight': 0,
                'total_volume': 0,
                'outbound_time': record.outbound_time,
                'plate_number': record.plate_number,
                'destination': record.destination,
                'operated_warehouse': record.operated_warehouse,
                'driver_name': record.driver_name,
                'driver_phone': record.driver_phone,
                'warehouse_address': record.warehouse_address,
                'detailed_address': record.detailed_address,
                'contact_window': record.contact_window,
                'destination_warehouse': record.destination_warehouse
            }

        batch_groups[batch_no]['records'].append(record)
        batch_groups[batch_no]['total_pallet_count'] += record.pallet_count or 0
        batch_groups[batch_no]['total_package_count'] += record.package_count or 0
        batch_groups[batch_no]['total_weight'] += record.weight or 0
        batch_groups[batch_no]['total_volume'] += record.volume or 0

    # 转换为列表并按时间排序
    batch_groups = list(batch_groups.values())
    batch_groups.sort(key=lambda x: x['outbound_time'])

    # 获取操作仓库信息
    source_warehouse = None
    if records:
        source_warehouse = records[0].operated_warehouse

    # 渲染后端仓出库单打印模板
    return render_template('backend_outbound_print_template.html',
                          batch_groups=batch_groups,
                          records=records,
                          source_warehouse=source_warehouse,
                          now=datetime.now())


# ==================== 系统状态检查路由 ====================

@bp.route('/health')
def health_check():
    """健康检查路由 - 根路径版本"""
    try:
        from app.startup_manager import startup_manager
        from flask import jsonify

        if startup_manager.is_ready:
            return jsonify({
                'status': 'ready',
                'message': '系统运行正常',
                'timestamp': startup_manager.startup_time.isoformat() if startup_manager.startup_time else None
            })
        else:
            return jsonify({
                'status': 'starting',
                'message': '系统正在启动中',
                'timestamp': startup_manager.startup_time.isoformat() if startup_manager.startup_time else None
            }), 202
    except Exception as e:
        current_app.logger.error(f"健康检查失败: {e}")
        return jsonify({
            'status': 'ready',  # 如果出错，假设系统已就绪
            'message': '健康检查出错，假设系统正常',
            'error': str(e)
        })


@bp.route('/startup-status')
def startup_status():
    """启动状态检查路由 - 根路径版本"""
    try:
        from app.startup_manager import startup_manager
        from flask import jsonify

        status = startup_manager.get_startup_status()
        return jsonify(status)
    except Exception as e:
        current_app.logger.error(f"获取启动状态失败: {e}")
        return jsonify({
            'is_ready': True,  # 如果出错，假设系统已就绪
            'elapsed_time': 0,
            'components': {
                'database': True,
                'cache': True,
                'services': True,
                'optimization': True
            },
            'messages': [
                {
                    'component': 'system',
                    'message': '启动状态检查出错，假设系统已就绪',
                    'timestamp': None
                }
            ],
            'error': str(e)
        })